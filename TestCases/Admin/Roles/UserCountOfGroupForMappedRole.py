'''
Created on 23-Jul-2018

@author: Sheethu C
'''
import os.path
import time
import traceback

from BaseTestClass import BaseTestClass
from BaseTestClass import driver
from openpyxl.reader.excel import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import xlrd
from BaseTestClass import projectPath
from CampaignPageElements import CampPage
from CreateLearnerNew import CreateLearner
from CreateLessonDifferentLessons import CreateLessonDifferentLessons
from DeleteLesson import DeleteLesson
from PlanAssignmentForPageElements import PlanAssignmentForPageElements
from GroupsPageElements import GroupsPageElements
from AddTriggerForPageElements import AddTriggerPage
from BaseTestClass import projectPath
from RoleXpathElements import RoleXpathElements
class UserCountOfGroupForMappedRole:
    def groupCreation(self,groupName,FirstName1,FirstName2,FirstName3):
        driver.refresh()
        wait=WebDriverWait(driver, 60)
        
        group=GroupsPageElements()
        
        group.adminSideMenuUnexpanded()
        print "Clicked on admin icon"
         
        group.groupSideMenuExpanded()  
        print "Clicked on Group icon"
        
        
        print "Checking Group page is displayed"
        if driver.find_element_by_xpath(group.groupPageHeader()).is_displayed():
            print "Group page is displayed successfully"
        else:
            print "Group page is not displayed"
            raise Exception
        
        
        group.createGroupButton()
        print "Clicked on Create Group button"
        
        group.enteringGroupname(groupName)
        
        print "Group Name entered....."
        
        group.nextButton()
        print "Clicked on Next button"
        
        wait.until(EC.visibility_of_element_located((By.XPATH,group.groupDetailPageHeader())))
        print "Checking group is created"
        
        time.sleep(4)
        headerText=driver.find_element_by_xpath(group.groupDetailPageHeader()).text
        
        if headerText==groupName:
            print "Group created successfully"
            
        else:
            print "Group not created"
            raise Exception
        
        
        print "Adding user to Group by name"
        group.addByNameButton()
        group.addByName(FirstName1)
        group.addByName(FirstName2)
        group.addByName(FirstName3)
        
        
        print "Verifying user is added to group"
        group.groupAddedInList(FirstName1)
        group.groupAddedInList(FirstName2)
        group.groupAddedInList(FirstName3)
       
        
        group.saveButton()
        print "Clicked on Save button"

        group.saveButtonPopup()
        print "Clicked on Save from popup"
        
        
        print "Checking Created Group is displayed in Grid"
        group.groupsLinkFromBreadCrumb()
        
        print "Searching Group"
        group.createdGroupDisplayedInGrid(groupName)
         
        
        driver.refresh()
    def updateUser(self):
        
        print "Updating Details"
        book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        second_sheet = book.sheet_by_name('UserAssignToRole')
        #First User
        cell = second_sheet.cell(8,1)
        FirstName = cell.value
        FirstNameId = FirstName.split("_")
        emp = FirstNameId[0]+"_"
        ids = FirstNameId[1]
        empId1 = int(ids)+1
        FirstNameUpdated1= emp+str(empId1)
        
        
        cell2 = second_sheet.cell(11,1)
        LearnId = cell2.value
        LearnsId = LearnId.split("#")
        emp = LearnsId[0]+"#"
        ids = LearnsId[1]
        empId = int(ids)+1
        LearnerEmpId1 = emp+str(empId)
        
        
        cell1 = second_sheet.cell(10,1)
        LearnEmailId = cell1.value
        LearnerEmail = LearnEmailId.split("@")
        spitValue = LearnerEmail[0][:4]
        email = spitValue+str(empId)
        id = "@"+LearnerEmail[1]
        LearnerEmailId1 = email+id
        
        
        
        #second User
        cell = second_sheet.cell(14,1)
        FirstName = cell.value
        FirstNameId = FirstName.split("_")
        emp = FirstNameId[0]+"_"
        ids = FirstNameId[1]
        empId1 = int(ids)+1
        FirstNameUpdated2= emp+str(empId1)
        
        cell2 = second_sheet.cell(17,1)
        LearnId = cell2.value
        LearnsId = LearnId.split("#")
        emp = LearnsId[0]+"#"
        ids = LearnsId[1]
        empId = int(ids)+1
        LearnerEmpId2 = emp+str(empId)
        
        
        cell1 = second_sheet.cell(16,1)
        LearnEmailId = cell1.value
        LearnerEmail = LearnEmailId.split("@")
        spitValue = LearnerEmail[0][:4]
        email = spitValue+str(empId)
        id = "@"+LearnerEmail[1]
        LearnerEmailId2 = email+id
        
        
        #sThird User
        cell = second_sheet.cell(20,1)
        FirstName = cell.value
        FirstNameId = FirstName.split("_")
        emp = FirstNameId[0]+"_"
        ids = FirstNameId[1]
        empId1 = int(ids)+1
        FirstNameUpdated3= emp+str(empId1)
        
        
        cell2 = second_sheet.cell(23,1)
        LearnId = cell2.value
        LearnsId = LearnId.split("#")
        emp = LearnsId[0]+"#"
        ids = LearnsId[1]
        empId = int(ids)+1
        LearnerEmpId3 = emp+str(empId)
        
        
        cell1 = second_sheet.cell(22,1)
        LearnEmailId = cell1.value
        LearnerEmail = LearnEmailId.split("@")
        spitValue = LearnerEmail[0][:4]
        email = spitValue+str(empId)
        id = "@"+LearnerEmail[1]
        LearnerEmailId3 = email+id
        
        wb = load_workbook(os.path.join('Test_Data/TestData.xlsx'))
        
        #print (wb.sheetnames)
    
        sheet = wb['UserAssignToRole']
        
        sheet.cell(row=9, column=2).value = FirstNameUpdated1    
        sheet.cell(row=11, column=2).value = LearnerEmailId1
        sheet.cell(row=12, column=2).value = LearnerEmpId1
        
        sheet.cell(row=15, column=2).value = FirstNameUpdated2    
        sheet.cell(row=17, column=2).value = LearnerEmailId2
        sheet.cell(row=18, column=2).value = LearnerEmpId2
        
        sheet.cell(row=21, column=2).value = FirstNameUpdated3    
        sheet.cell(row=23, column=2).value = LearnerEmailId3
        sheet.cell(row=24, column=2).value = LearnerEmpId3
        
       
        wb.save(os.path.join('Test_Data/TestData.xlsx'))
    def roleCreation(self,RoleName,Description,GroupName):    
        createrole =RoleXpathElements()
        wait=WebDriverWait(driver, 80)
        driver.refresh()
        wait.until(EC.visibility_of_element_located((By.XPATH,createrole.adminSideMenu())))
        driver.find_element_by_xpath(createrole.adminSideMenu()).click()
        print "Clicked on admin page"
        wait.until(EC.visibility_of_element_located((By.XPATH,createrole.roleSideMenu())))
        driver.find_element_by_xpath(createrole.roleSideMenu()).click()
        print "Clicked on Role icon"
        wait.until(EC.visibility_of_element_located((By.XPATH,createrole.createRole())))
        driver.find_element_by_xpath(createrole.createRole()).click()
        print "Clicked on Create role Button"
        time.sleep(4)
        createrole.roleCreation(RoleName,Description)
        print "Creating Role with Description"
        createrole.adminClick()
        print "Clicked on Administrator "
        ch=RoleXpathElements()
        ch.checkAdministrator()
        time.sleep(4)
        print "Checking Check box selection"
        chk=RoleXpathElements()
        chk.checkboxUser()
        print "User is selected"
        chk.checkboxContent()
        print "Content Manger is selected"
        chk.checkboxRoles()
        print "Roles is selected"
        chk.checkboxTag()
        print "Tags is selected"
        chk.checkboxBranding()
        print "Branding is selected"
        chk.checkboxIntegration()
        print "Integration is selected"
        
        #adding group to the role
        wait.until(EC.visibility_of_element_located((By.XPATH,createrole.usertab())))
        driver.find_element_by_xpath(createrole.usertab()).click()
        wait.until(EC.visibility_of_element_located((By.XPATH,createrole.searchbox())))
        driver.find_element_by_xpath(createrole.searchbox()).click()
        time.sleep(4)
        wait.until(EC.visibility_of_element_located((By.XPATH,"//div[@class='Select-placeholder']")))
        names=driver.find_element_by_xpath("//div[@class='Select-placeholder']")
        time.sleep(2)
        webdriver.ActionChains(driver).move_to_element(names).click()
        webdriver.ActionChains(driver).move_to_element(names).send_keys(GroupName).perform()
        userDisplayed=wait.until(EC.visibility_of_element_located((By.XPATH,"//div[@role='option']/span[contains(.,'"+GroupName+"')]")))
        webdriver.ActionChains(driver).move_to_element(userDisplayed).click().perform()
        wait.until(EC.visibility_of_element_located((By.XPATH,"//tbody/tr/td[2]/a[.='"+GroupName+"']")))
        #checking added group is in the list
        ExpectedGroupName= driver.find_element_by_xpath(createrole.groupDispaly(GroupName)).text
        if ExpectedGroupName == GroupName:
            print "Group is displayed in the Grid"
        else:
            print "Group is not displayed in the Grid"
            raise Exception
        createrole.roleSave()
        print "Clicked on Save Role Button"
        print "Searching for the Created Role in the List"
        createrole.roleSearch(RoleName)
    def editRole(self,RoleName):
        driver.refresh()
        createrole =RoleXpathElements()
        wait=WebDriverWait(driver, 80)
        driver.refresh()
        wait.until(EC.visibility_of_element_located((By.XPATH,createrole.adminSideMenu())))
        driver.find_element_by_xpath(createrole.adminSideMenu()).click()
        print "Clicked on admin page"
        wait.until(EC.visibility_of_element_located((By.XPATH,createrole.roleSideMenu())))
        driver.find_element_by_xpath(createrole.roleSideMenu()).click()
        print "Clicked on Role icon"
        time.sleep(6)
        createrole.roleSearch(RoleName)
        #clicking on Role
        driver.find_element_by_xpath(createrole.roleclick()).click()
        wait.until(EC.visibility_of_element_located((By.XPATH,createrole.editrolebreadcrump())))
        wait.until(EC.visibility_of_element_located((By.XPATH,createrole.usertab())))
        driver.find_element_by_xpath(createrole.usertab()).click()
        #checking count of users in Group
        wait.until(EC.visibility_of_element_located((By.XPATH,createrole.usercount())))
        usercount = driver.find_element_by_xpath(createrole.usercount()).text
        
        if  usercount == 3:
            print "Actual user count is matching with the expected user count"
        else:
            print "Actual user count is not matching with the expected user count"
            raise Exception
    def mainUserCountOfGroupForMappedRole(self):
        try:
            book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
            first_sheet = book.sheet_by_name('UserAssignToRole')
            
            cell = first_sheet.cell(7,1)
            groupName = cell.value
            
            cell = first_sheet.cell(8,1)
            FirstName1 = cell.value
        
            cell = first_sheet.cell(9,1)
            LastName1 = cell.value
        
            cell = first_sheet.cell(10,1)
            Email1 = cell.value
        
            cell = first_sheet.cell(11,1)
            EmployeeId1 = cell.value
        
            cell = first_sheet.cell(12,1)
            Password1 = cell.value
        
            cell = first_sheet.cell(13,1)
            NewPassword1 = cell.value
            
            #Second User
            cell = first_sheet.cell(14,1)
            FirstName2 = cell.value
        
            cell = first_sheet.cell(15,1)
            LastName2 = cell.value
        
            cell = first_sheet.cell(16,1)
            Email2 = cell.value
        
            cell = first_sheet.cell(17,1)
            EmployeeId2 = cell.value
        
            cell = first_sheet.cell(18,1)
            Password2 = cell.value
        
            cell = first_sheet.cell(19,1)
            NewPassword2 = cell.value
            
            
            #Third User
            cell = first_sheet.cell(20,1)
            FirstName3 = cell.value
        
            cell = first_sheet.cell(21,1)
            LastName3 = cell.value
        
            cell = first_sheet.cell(22,1)
            Email3 = cell.value
        
            cell = first_sheet.cell(23,1)
            EmployeeId3 = cell.value
        
            cell = first_sheet.cell(24,1)
            Password3 = cell.value
        
            cell = first_sheet.cell(25,1)
            NewPassword3 = cell.value
            
            #Role Creation
            
            cell = first_sheet.cell(26,1)
            RoleName = cell.value
            
            cell = first_sheet.cell(27,1)
            Description = cell.value
        
            book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
            second_sheet = book.sheet_by_name('Login_Credentials')
            cell = second_sheet.cell(1,1)
            url = cell.value
        
            cell = second_sheet.cell(3,1)
            username = cell.value
        
            cell = second_sheet.cell(3,2)
            password = cell.value
        
            ob=CreateLearner()
            #ob.createNewLearnerMain(FirstName1,LastName1,Email1,EmployeeId1,Password1,NewPassword1,url,username,password)
            #ob.createNewLearnerMain(FirstName2,LastName2,Email2,EmployeeId2,Password2,NewPassword2,url,username,password)
            #ob.createNewLearnerMain(FirstName3,LastName3,Email3,EmployeeId3,Password3,NewPassword3,url,username,password)
            grp= UserCountOfGroupForMappedRole()
            #grp.groupCreation(groupName,FirstName1,FirstName2,FirstName3)
            grp.roleCreation(RoleName,Description,groupName)
        except Exception as e:
            traceback.print_exc()
            print (e)
            raise Exception   
          
        finally:
            ob=UserCountOfGroupForMappedRole()
            ob.updateUser() 
            second_sheet = book.sheet_by_name('Login_Credentials')
            cell = second_sheet.cell(1,1)
            url = cell.value
            driver.get(url)
            try:
                WebDriverWait(driver, 5).until(EC.alert_is_present())

                alert = driver.switch_to.alert
                alert.accept()
                print("alert accepted")
            except Exception:
                print("no alert")
    
        
        
        
if __name__=='__main__':
    
    btc=BaseTestClass()
    btc.userLogin() 

    gr=UserCountOfGroupForMappedRole()
    gr.mainUserCountOfGroupForMappedRole()   
        
    
    
    
    
    
    
    
    