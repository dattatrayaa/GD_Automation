'''
Created on 15-May-2018

@author: dattatraya
'''

import os
import time
import traceback

from BaseTestClass import BaseTestClass
from BaseTestClass import driver
from openpyxl.reader.excel import load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import xlrd

from CreateUserWithRole import CreateUserWithRole
from GroupsPageElements import GroupsPageElements
from UsersPageElements import UsersPageElements
from BaseTestClass import projectPath

class GroupWithAnyAttributesThree:
    
    def createOne(self,FirstName,LastName,Email,EmployeeId,Password,reportsTo,attributeName1):
        wait=WebDriverWait(driver, 60)
        driver.refresh()
        user=UsersPageElements()
        
        user.AdminFromSideMenuClick()
        print "Clicked on Admin"
        
        user.UsersFromSideMenuClick()
        print "clicked on Users"
        
        user.AddOrEditUserButtonClick()
        print "Clicked on Add or editUser"

        
        user.AddAnIndividualUserButtonClick()
        print "Clicked on Add An individual User"
        time.sleep(1)
        wait.until(EC.visibility_of_element_located((By.XPATH,user.AddUserPageHeaderXpath())))
        
        print "Verifying Add user Page"
        if driver.find_element_by_xpath(user.AddUserPageHeaderXpath()).is_displayed():
            print("Add user Page is displayed")
        else:
            print ""
            raise Exception
        
        
        print "Verifying First Name field"
        fn=wait.until(EC.visibility_of_element_located((By.ID,user.FirstNameFieldID())))
        if fn.is_displayed():
            print(" First Name field displayed")
        else:
            print ""
            raise Exception
        fn.send_keys(FirstName)
        print "FirstName is Entered ::"+FirstName
        
        
        
        print "Last NAme verifying"
        if driver.find_element_by_id(user.LastNameFieldID()).is_displayed():
            print(" Last Name field displayed")
        else:
            print ""
            raise Exception
        driver.find_element_by_id(user.LastNameFieldID()).send_keys(LastName)
        print "Last Name is Entered ::"+LastName
        
        
        print "Email verifying"
        if driver.find_element_by_id(user.EmailFieldID()).is_displayed():
            print("Email field displayed")
        else:
            print ""
            raise Exception
        driver.find_element_by_id(user.EmailFieldID()).send_keys(Email)
        print "Email is Entered ::"+Email
        
        
        print "Employee ID verifying"
        if driver.find_element_by_id(user.EmployeeIDFieldID()).is_displayed():
            print("Employee ID field displayed")
        else:
            print ""
            raise Exception
        driver.find_element_by_id(user.EmployeeIDFieldID()).send_keys(EmployeeId)
        print "Employee ID  is Entered ::"+EmployeeId
        
        
        print "Adding additional Attribute"
        user.AdditionalAttributesViewClick()
        
        
        
        print "Creating Option for city"
        user.ReportsToAttributeFieldEnterData(reportsTo)
        print "Selcting Option"
        user.SelectCreateOption(attributeName1)
        print "Option Selected"
        
        
        
        print "Inherited Role Verifying"
        if driver.find_element_by_xpath(user.InHeritedRoleFieldXpath()).is_displayed():
            print("Inherited Role field displayed")
        else:
            print ""
            raise Exception
        
        print "Password Field is Verifying"
        if driver.find_element_by_id(user.PassWordFieldID()).is_displayed():
            print("Password field displayed")
        else:
            print ""
            raise Exception
        driver.find_element_by_id(user.PassWordFieldID()).send_keys(Password)
        print "Password is Entered ::"+Password
        
        
        user.AddButtonClick()
        print "Clicked on add button"
        time.sleep(2)

        user.SaveButtonClick()        
        print "Clicked on Save"
        
        wait.until(EC.visibility_of_element_located((By.XPATH,user.UsersPageHeader())))
        
        
        print "Searching for the Created User"
        user.SearchForUserInGrid(FirstName)
        
        
        user.CheckUserDisplayedInGrid(FirstName) 
    
    
    def createTwo(self,FirstName,LastName,Email,EmployeeId,Password,state,attributeName2):
        wait=WebDriverWait(driver, 60)
        driver.refresh()
        user=UsersPageElements()
        
        user.AdminFromSideMenuClick()
        print "Clicked on Admin"
        
        user.UsersFromSideMenuClick()
        print "clicked on Users"
        
        user.AddOrEditUserButtonClick()
        print "Clicked on Add or editUser"

        
        user.AddAnIndividualUserButtonClick()
        print "Clicked on Add An individual User"
        time.sleep(1)
        wait.until(EC.visibility_of_element_located((By.XPATH,user.AddUserPageHeaderXpath())))
        
        print "Verifying Add user Page"
        if driver.find_element_by_xpath(user.AddUserPageHeaderXpath()).is_displayed():
            print("Add user Page is displayed")
        else:
            print ""
            raise Exception
        
        
        print "Verifying First Name field"
        fn=wait.until(EC.visibility_of_element_located((By.ID,user.FirstNameFieldID())))
        if fn.is_displayed():
            print(" First Name field displayed")
        else:
            print ""
            raise Exception
        fn.send_keys(FirstName)
        print "FirstName is Entered ::"+FirstName
        
        
        
        print "Last NAme verifying"
        if driver.find_element_by_id(user.LastNameFieldID()).is_displayed():
            print(" Last Name field displayed")
        else:
            print ""
            raise Exception
        driver.find_element_by_id(user.LastNameFieldID()).send_keys(LastName)
        print "Last Name is Entered ::"+LastName
        
        
        print "Email verifying"
        if driver.find_element_by_id(user.EmailFieldID()).is_displayed():
            print("Email field displayed")
        else:
            print ""
            raise Exception
        driver.find_element_by_id(user.EmailFieldID()).send_keys(Email)
        print "Email is Entered ::"+Email
        
        
        print "Employee ID verifying"
        if driver.find_element_by_id(user.EmployeeIDFieldID()).is_displayed():
            print("Employee ID field displayed")
        else:
            print ""
            raise Exception
        driver.find_element_by_id(user.EmployeeIDFieldID()).send_keys(EmployeeId)
        print "Employee ID  is Entered ::"+EmployeeId
        
        
        print "Adding additional Attribute"
        user.AdditionalAttributesViewClick()
        
        
        
        print "Creating Option for city"
        user.StateAttributeFieldEnterData(state)
        print "Selecting Option"
        user.SelectCreateOption(attributeName2)
        print "Option Selected "+attributeName2
        
        
        
        print "Inherited Role Verifying"
        if driver.find_element_by_xpath(user.InHeritedRoleFieldXpath()).is_displayed():
            print("Inherited Role field displayed")
        else:
            print ""
            raise Exception
        
        print "Password Field is Verifying"
        if driver.find_element_by_id(user.PassWordFieldID()).is_displayed():
            print("Password field displayed")
        else:
            print ""
            raise Exception
        driver.find_element_by_id(user.PassWordFieldID()).send_keys(Password)
        print "Password is Entered ::"+Password
        
        
        user.AddButtonClick()
        print "Clicked on add button"
        time.sleep(2)

        user.SaveButtonClick()        
        print "Clicked on Save"
        
        wait.until(EC.visibility_of_element_located((By.XPATH,user.UsersPageHeader())))
        
        
        print "Searching for the Created User"
        user.SearchForUserInGrid(FirstName)
        
        
        user.CheckUserDisplayedInGrid(FirstName) 
        
    
    
    def groupCreateForCampaign(self,groupName,FirstName,FirstName1,attributeName1,reportsTo,state,attributeName2):
        driver.refresh()
        wait=WebDriverWait(driver, 60)
        
        group=GroupsPageElements()
        
        group.adminSideMenuUnexpanded()
        print "Clicked on admin icon"
         
        group.groupSideMenuExpanded()  
        print "Clicked on Group icon"
        
        
        print "Checking Group page is displayed"
        if driver.find_element_by_xpath(group.groupPageHeader()).is_displayed():
            print "Group page is displayed successfully"
        else:
            print "Group page is not displayed"
            raise Exception
        
        
        group.createGroupButton()
        print "Clicked on Create Group button"
        
        group.enteringGroupname(groupName)
        
        print "Group Name entered....."
        
        group.nextButton()
        print "Clicked on Next button"
        
        wait.until(EC.visibility_of_element_located((By.XPATH,group.groupDetailPageHeader())))
        print "Checking group is created"
        
        time.sleep(4)
        headerText=driver.find_element_by_xpath(group.groupDetailPageHeader()).text
        
        if headerText==groupName:
            print "Group created successfully"
            
        else:
            print "Group not created"
            raise Exception
        
        
        print "Adding user to Group by name"
        
        group.AddByAttributes()
        
        print "Checking all option is selected"
        if group.AllOptionSelected()=="all":
            print "all is selected"
        else:
            print "all is not selected"
            raise Exception
        
        
        print "Changing user match to 'Any'"
        group.SelectAnyOFTheFollowingOption()
        
        
        
        print "Adding User by attribute"
        group.AttributeNameEnter(attributeName1)
        group.ValueSelect(reportsTo)
        print "Selected "+attributeName1
        
        group.addAttributePopupButtonClick()
        print "Adding user by attribute"
        group.AttributeNameEnter(attributeName2)
        group.ValueSelect(state)
        print "Selected "+attributeName2
        
        
        
        print "Clicking on Preview button"
        group.PreviewGroupButtonClick()
        
        print "Checking Filter is applied"
        group.OverViewContainer(reportsTo)
        
        
        print "Verifying user is added to group"
        group.groupAddedInList(FirstName)
        group.groupAddedInList(FirstName1)
        
        group.saveButton()
        print "Clicked on Save button"

        group.saveButtonPopup()
        print "Clicked on Save from popup"
        
        
        print "Checking Created Group is displayed in Grid"
        group.groupsLinkFromBreadCrumb()
        
        print "Searching Group"
        group.createdGroupDisplayedInGrid(groupName)
         
        
        driver.refresh()
        
        
        
    def CreateGroupWithAnyAttributeThree(self):
        book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        first_sheet = book.sheet_by_name('Groups')
        
        cell = first_sheet.cell(176,1)
        FirstName = cell.value
        
        FirstNameId = FirstName.split("_")
        emp = FirstNameId[0]+"_"
        ids = FirstNameId[1]
        empId = int(ids)+1
        FirstNameUpdated= emp+str(empId)
        
        cell = first_sheet.cell(177,1)
        LastName = cell.value
        
        LastNameID = LastName.split("_")
        emp = LastNameID[0]+"_"
        ids = LastNameID[1]
        empId = int(ids)+1
        LastNameUpdated= emp+str(empId)
        
        cell = first_sheet.cell(178,1)
        Email = cell.value
        
        EmailId = Email.split("_")
        emp = EmailId[0]+"_"
        ids = EmailId[1]
        remaining="_"+EmailId[2]
        empId = int(ids)+1
        EmailIdUpdated= emp+str(empId)+remaining
        
        cell = first_sheet.cell(179,1)
        EmployeeId = cell.value
        
        Employee = EmployeeId.split("_")
        emp = Employee[0]+"_"
        ids = Employee[1]
        empId = int(ids)+1
        EmployeeIdUpdated= emp+str(empId)
        
        cell1 = first_sheet.cell(180,1)
        Password = cell1.value
        
        cell1 = first_sheet.cell(192,1)
        reportsTo = cell1.value
        cell1 = first_sheet.cell(182,1)
        attributeName1 = cell1.value
        
        
        
        
        cell1 = first_sheet.cell(183,1)
        groupName = cell1.value
        groupNameId = groupName.split("_")
        emp = groupNameId[0]+"_"
        ids = groupNameId[1]
        empId = int(ids)+1
        groupNameUpdated= emp+str(empId)
        
        
        cell = first_sheet.cell(185,1)
        FirstName1 = cell.value
        FirstNameId = FirstName1.split("_")
        emp = FirstNameId[0]+"_"
        ids = FirstNameId[1]
        empId = int(ids)+1
        FirstNameUpdated1= emp+str(empId)
        
        cell = first_sheet.cell(186,1)
        LastName1 = cell.value
        LastNameID = LastName1.split("_")
        emp = LastNameID[0]+"_"
        ids = LastNameID[1]
        empId = int(ids)+1
        LastNameUpdated1= emp+str(empId)
        
        cell = first_sheet.cell(187,1)
        Email1 = cell.value
        EmailId = Email1.split("_")
        emp = EmailId[0]+"_"
        ids = EmailId[1]
        remaining="_"+EmailId[2]
        empId = int(ids)+1
        EmailIdUpdated1= emp+str(empId)+remaining
        
        cell = first_sheet.cell(188,1)
        EmployeeId1 = cell.value
        Employee = EmployeeId1.split("_")
        emp = Employee[0]+"_"
        ids = Employee[1]
        empId = int(ids)+1
        EmployeeIdUpdated1= emp+str(empId)
        
        cell1 = first_sheet.cell(189,1)
        Password1 = cell1.value
        
        cell1 = first_sheet.cell(190,1)
        state = cell1.value
        stateId = state.split("_")
        emp = stateId[0]+"_"
        ids = stateId[1]
        empId = int(ids)+1
        stateIdUpdated= emp+str(empId)
        cell1 = first_sheet.cell(191,1)
        attributeName2 = cell1.value
        
        
        
        
        
        cell = first_sheet.cell(192,1)
        FirstName2 = cell.value
        
        FirstNameId = FirstName2.split("_")
        emp = FirstNameId[0]+"_"
        ids = FirstNameId[1]
        empId = int(ids)+1
        FirstNameUpdated2= emp+str(empId)
        
        cell = first_sheet.cell(193,1)
        LastName2 = cell.value
        
        LastNameID = LastName2.split("_")
        emp = LastNameID[0]+"_"
        ids = LastNameID[1]
        empId = int(ids)+1
        LastNameUpdated2= emp+str(empId)
        
        
        
        cell = first_sheet.cell(194,1)
        Email2 = cell.value
        
        EmailId = Email2.split("_")
        emp = EmailId[0]+"_"
        ids = EmailId[1]
        remaining="_"+EmailId[2]
        empId = int(ids)+1
        EmailIdUpdated2= emp+str(empId)+remaining
        
        cell = first_sheet.cell(195,1)
        EmployeeId2 = cell.value
        
        Employee = EmployeeId2.split("_")
        emp = Employee[0]+"_"
        ids = Employee[1]
        empId = int(ids)+1
        EmployeeIdUpdated2= emp+str(empId)
        
        cell1 = first_sheet.cell(196,1)
        Password2 = cell1.value
        
        cell1 = first_sheet.cell(197,1)
        role = cell1.value
        
        cell1 = first_sheet.cell(196,1)
        NewPassword = cell1.value
        
        book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        s_sheet = book.sheet_by_name('Login_Credentials')
        
        cell = s_sheet.cell(1,1)
        url = cell.value
        
        cell = s_sheet.cell(3,1)
        username = cell.value
        
        cell = s_sheet.cell(3,2)
        password = cell.value
        
        
        
        
        
        
        
        #updating user values
        wb = load_workbook(os.path.join('Test_Data/TestData.xlsx'))
            #print (wb.sheetnames)
        
        sheet = wb['Groups']
        
        sheet.cell(row=177, column=2).value = FirstNameUpdated
        sheet.cell(row=178, column=2).value = LastNameUpdated
        sheet.cell(row=179, column=2).value = EmailIdUpdated
        sheet.cell(row=180, column=2).value = EmployeeIdUpdated
        
        sheet.cell(row=184, column=2).value = groupNameUpdated
        
        sheet.cell(row=186, column=2).value = FirstNameUpdated1
        sheet.cell(row=187, column=2).value = LastNameUpdated1
        sheet.cell(row=188, column=2).value = EmailIdUpdated1
        sheet.cell(row=189, column=2).value = EmployeeIdUpdated1
        sheet.cell(row=191, column=2).value = stateIdUpdated
        
        sheet.cell(row=193, column=2).value = FirstNameUpdated2
        sheet.cell(row=194, column=2).value = LastNameUpdated2
        sheet.cell(row=195, column=2).value = EmailIdUpdated2
        sheet.cell(row=196, column=2).value = EmployeeIdUpdated2
        
        
        
        wb.save(os.path.join('Test_Data/TestData.xlsx'))
       
        
        try:
            
            cr=CreateUserWithRole()
            cr.createUserWithRoleMain(FirstName2, LastName2, Email2, EmployeeId2, Password2, role, NewPassword, url, username, password)
            t=GroupWithAnyAttributesThree()
            print "\nCreating New User\n"
            t.createOne(FirstName, LastName, Email, EmployeeId, Password, reportsTo, attributeName1)
            t.createTwo(FirstName1, LastName1, Email1, EmployeeId1, Password1, state, attributeName2)
            
            t.groupCreateForCampaign(groupName, FirstName, FirstName1, attributeName1, reportsTo, state, attributeName2)
            
            
        except Exception as e:
            traceback.print_exc()
            print (e)
            raise Exception   
          
        finally:  
            second_sheet = book.sheet_by_name('Login_Credentials')
            cell = second_sheet.cell(1,1)
            url = cell.value
            driver.get(url)
            try:
                WebDriverWait(driver, 5).until(EC.alert_is_present())

                alert = driver.switch_to.alert
                alert.accept()
                print("alert accepted")
            except Exception:
                print("no alert")
                
                
            wait=WebDriverWait(driver, 60)
            unDropDown=wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")))
            driver.execute_script('arguments[0].click()',unDropDown)
            
            signOut=driver.find_element_by_xpath("html/body/div/div/div[1]/div[2]/div[2]/a")
            driver.execute_script('arguments[0].click()',signOut)
            
            
            
            element = wait.until(EC.presence_of_element_located((By.ID, "password")))
            
            if driver.title == "Grovo":
                print("Grovo Application URL Opened")
            else:
                raise Exception.message
    
            print "Grovo Sign-In page is displayed"
            
            print "Entering User name"
            driver.find_element_by_xpath(".//*[@id='username']").send_keys(username)
           
            print "Entering Password"
            element.send_keys(password)
            
            element.send_keys(Keys.TAB)
            print "Clicking on Sign_In button"
            driver.find_element_by_xpath("//*[@id='submitButton']").click()
            
            print "Successfully Logged Into Users account"
            time.sleep(3)
            
    
        
        
        
if __name__=='__main__':
    
    btc=BaseTestClass()
    btc.UserLogin() 

    gr=GroupWithAnyAttributesThree()
    gr.CreateGroupWithAnyAttributeThree()



