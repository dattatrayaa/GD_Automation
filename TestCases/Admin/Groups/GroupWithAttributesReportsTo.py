'''
Created on 11-May-2018

@author: dattatraya
'''

import os
import time
import traceback

from BaseTestClass import BaseTestClass
from BaseTestClass import driver
from openpyxl.reader.excel import load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import xlrd

from BaseTestClass import projectPath
from GroupsPageElements import GroupsPageElements
from UsersPageElements import UsersPageElements
from CreateUserWithRole import CreateUserWithRole


class GroupWithAttributesReportsTo:
    
    def create(self,FirstName,LastName,Email,EmployeeId,Password,reportsTo,attributeName):
        wait=WebDriverWait(driver, 60)
        driver.refresh()
        user=UsersPageElements()
        
        user.AdminFromSideMenuClick()
        print "Clicked on Admin"
        
        user.UsersFromSideMenuClick()
        print "clicked on Users"
        
        user.AddOrEditUserButtonClick()
        print "Clicked on Add or editUser"

        
        user.AddAnIndividualUserButtonClick()
        print "Clicked on Add An individual User"
        time.sleep(1)
        wait.until(EC.visibility_of_element_located((By.XPATH,user.AddUserPageHeaderXpath())))
        
        print "Verifying Add user Page"
        if driver.find_element_by_xpath(user.AddUserPageHeaderXpath()).is_displayed():
            print("Add user Page is displayed")
        else:
            print ""
            raise Exception
        
        
        print "Verifying First Name field"
        fn=wait.until(EC.visibility_of_element_located((By.ID,user.FirstNameFieldID())))
        if fn.is_displayed():
            print(" First Name field displayed")
        else:
            print ""
            raise Exception
        fn.send_keys(FirstName)
        print "FirstName is Entered ::"+FirstName
        
        
        
        print "Last NAme verifying"
        if driver.find_element_by_id(user.LastNameFieldID()).is_displayed():
            print(" Last Name field displayed")
        else:
            print ""
            raise Exception
        driver.find_element_by_id(user.LastNameFieldID()).send_keys(LastName)
        print "Last Name is Entered ::"+LastName
        
        
        print "Email verifying"
        if driver.find_element_by_id(user.EmailFieldID()).is_displayed():
            print("Email field displayed")
        else:
            print ""
            raise Exception
        driver.find_element_by_id(user.EmailFieldID()).send_keys(Email)
        print "Email is Entered ::"+Email
        
        
        print "Employee ID verifying"
        if driver.find_element_by_id(user.EmployeeIDFieldID()).is_displayed():
            print("Employee ID field displayed")
        else:
            print ""
            raise Exception
        driver.find_element_by_id(user.EmployeeIDFieldID()).send_keys(EmployeeId)
        print "Employee ID  is Entered ::"+EmployeeId
        
        
        print "Adding additional Attribute"
        user.AdditionalAttributesViewClick()
        
        
        
        print "Creating Option for city"
        user.ReportsToAttributeFieldEnterData(reportsTo)
        print "Selcting Option"
        user.SelectCreateOption(attributeName)
        print "Option Selected"
        
        
        
        print "Inherited Role Verifying"
        if driver.find_element_by_xpath(user.InHeritedRoleFieldXpath()).is_displayed():
            print("Inherited Role field displayed")
        else:
            print ""
            raise Exception
        
        print "Password Field is Verifying"
        if driver.find_element_by_id(user.PassWordFieldID()).is_displayed():
            print("Password field displayed")
        else:
            print ""
            raise Exception
        driver.find_element_by_id(user.PassWordFieldID()).send_keys(Password)
        print "Password is Entered ::"+Password
        
        
        user.AddButtonClick()
        print "Clicked on add button"
        time.sleep(2)

        user.SaveButtonClick()        
        print "Clicked on Save"
        
        wait.until(EC.visibility_of_element_located((By.XPATH,user.UsersPageHeader())))
        
        
        print "Searching for the Created User"
        user.SearchForUserInGrid(FirstName)
        
        
        user.CheckUserDisplayedInGrid(FirstName) 
        
        
        
    def groupCreateForCampaign(self,groupName,FirstName,attributeName,reportsTo):
        driver.refresh()
        wait=WebDriverWait(driver, 60)
        
        group=GroupsPageElements()
        
        group.adminSideMenuUnexpanded()
        print "Clicked on admin icon"
         
        group.groupSideMenuExpanded()  
        print "Clicked on Group icon"
        
        
        print "Checking Group page is displayed"
        if driver.find_element_by_xpath(group.groupPageHeader()).is_displayed():
            print "Group page is displayed successfully"
        else:
            print "Group page is not displayed"
            raise Exception
        
        
        group.createGroupButton()
        print "Clicked on Create Group button"
        
        group.enteringGroupname(groupName)
        
        print "Group Name entered....."
        
        group.nextButton()
        print "Clicked on Next button"
        
        wait.until(EC.visibility_of_element_located((By.XPATH,group.groupDetailPageHeader())))
        print "Checking group is created"
        
        time.sleep(4)
        headerText=driver.find_element_by_xpath(group.groupDetailPageHeader()).text
        
        if headerText==groupName:
            print "Group created successfully"
            
        else:
            print "Group not created"
            raise Exception
        
        
        print "Adding user to Group by name"
        
        group.AddByAttributes()
        
        print "Checking all option is selected"
        if group.AllOptionSelected()=="all":
            print "all is selected"
        else:
            print "all is not selected"
            raise Exception
        
        print "Adding User by attribute"
        group.AttributeNameEnter(attributeName)
        
        print "Selecting value"
        group.ValueSelect(reportsTo)
        
        print "Clicking on Preview button"
        group.PreviewGroupButtonClick()
        
        print "Checking Filter is applied"
        group.OverViewContainer(reportsTo)
        
        
        print "Verifying user is added to group"
        group.groupAddedInList(FirstName)
        
        group.saveButton()
        print "Clicked on Save button"

        group.saveButtonPopup()
        print "Clicked on Save from popup"
        
        
        print "Checking Created Group is displayed in Grid"
        group.groupsLinkFromBreadCrumb()
        
        print "Searching Group"
        group.createdGroupDisplayedInGrid(groupName)
         
        
        driver.refresh()
    
    
    def CreateGroupWithAttributeReportsTo(self):
        book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        first_sheet = book.sheet_by_name('Groups')
        
        cell = first_sheet.cell(88,1)
        FirstName = cell.value
        
        FirstNameId = FirstName.split("_")
        emp = FirstNameId[0]+"_"
        ids = FirstNameId[1]
        empId = int(ids)+1
        FirstNameUpdated= emp+str(empId)
        
        cell = first_sheet.cell(89,1)
        LastName = cell.value
        
        LastNameID = LastName.split("_")
        emp = LastNameID[0]+"_"
        ids = LastNameID[1]
        empId = int(ids)+1
        LastNameUpdated= emp+str(empId)
        
        
        
        cell = first_sheet.cell(90,1)
        Email = cell.value
        
        EmailId = Email.split("_")
        emp = EmailId[0]+"_"
        ids = EmailId[1]
        remaining="_"+EmailId[2]
        empId = int(ids)+1
        EmailIdUpdated= emp+str(empId)+remaining
        
        cell = first_sheet.cell(91,1)
        EmployeeId = cell.value
        
        Employee = EmployeeId.split("_")
        emp = Employee[0]+"_"
        ids = Employee[1]
        empId = int(ids)+1
        EmployeeIdUpdated= emp+str(empId)
        
        cell1 = first_sheet.cell(92,1)
        Password = cell1.value
        
        cell1 = first_sheet.cell(96,1)
        reportsTo = cell1.value
        
        
        cell1 = first_sheet.cell(94,1)
        groupName = cell1.value
        groupNameId = groupName.split("_")
        emp = groupNameId[0]+"_"
        ids = groupNameId[1]
        empId = int(ids)+1
        groupNameUpdated= emp+str(empId)
        
        cell1 = first_sheet.cell(95,1)
        attributeName = cell1.value
        
        
        #Learner 2
        
        cell = first_sheet.cell(96,1)
        FirstName1 = cell.value
        
        FirstNameId = FirstName1.split("_")
        emp = FirstNameId[0]+"_"
        ids = FirstNameId[1]
        empId = int(ids)+1
        FirstNameUpdated1= emp+str(empId)
        
        cell = first_sheet.cell(97,1)
        LastName1 = cell.value
        
        LastNameID = LastName1.split("_")
        emp = LastNameID[0]+"_"
        ids = LastNameID[1]
        empId = int(ids)+1
        LastNameUpdated1= emp+str(empId)
        
        
        
        cell = first_sheet.cell(98,1)
        Email1 = cell.value
        
        EmailId = Email1.split("_")
        emp = EmailId[0]+"_"
        ids = EmailId[1]
        remaining="_"+EmailId[2]
        empId = int(ids)+1
        EmailIdUpdated1= emp+str(empId)+remaining
        
        cell = first_sheet.cell(99,1)
        EmployeeId1 = cell.value
        
        Employee = EmployeeId1.split("_")
        emp = Employee[0]+"_"
        ids = Employee[1]
        empId = int(ids)+1
        EmployeeIdUpdated1= emp+str(empId)
        
        cell1 = first_sheet.cell(100,1)
        Password1 = cell1.value
        
        cell1 = first_sheet.cell(101,1)
        role = cell1.value
        
        cell1 = first_sheet.cell(100,1)
        NewPassword = cell1.value
        
        book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        s_sheet = book.sheet_by_name('Login_Credentials')
        
        cell = s_sheet.cell(1,1)
        url = cell.value
        
        cell = s_sheet.cell(3,1)
        username = cell.value
        
        cell = s_sheet.cell(3,2)
        password = cell.value
        
        #updating user values
        wb = load_workbook(os.path.join('Test_Data/TestData.xlsx'))
            #print (wb.sheetnames)
        
        sheet = wb['Groups']
        
        sheet.cell(row=89, column=2).value = FirstNameUpdated
        sheet.cell(row=90, column=2).value = LastNameUpdated
        sheet.cell(row=91, column=2).value = EmailIdUpdated
        sheet.cell(row=92, column=2).value = EmployeeIdUpdated
        
        sheet.cell(row=95, column=2).value = groupNameUpdated
        
        sheet.cell(row=97, column=2).value = FirstNameUpdated1
        sheet.cell(row=98, column=2).value = LastNameUpdated1
        sheet.cell(row=99, column=2).value = EmailIdUpdated1
        sheet.cell(row=100, column=2).value = EmployeeIdUpdated1
        
        
        
        wb.save(os.path.join('Test_Data/TestData.xlsx'))
       
        
        try:
            print "\nCreating USer with Role '"+role+"'\n"
            cr=CreateUserWithRole()
            cr.createUserWithRoleMain(FirstName1, LastName1, Email1, EmployeeId1, Password1, role, NewPassword, url, username, password)
            
            t=GroupWithAttributesReportsTo()
            print "\nCreating New User\n"
            t.create(FirstName, LastName, Email, EmployeeId, Password, reportsTo, attributeName)
            
            
            print "\nCreating Group with '"+reportsTo+"' Attribute\n"
            t.groupCreateForCampaign(groupName, FirstName, attributeName, reportsTo)
            
            
        except Exception as e:
            traceback.print_exc()
            print (e)
            raise Exception   
          
        finally:  
            second_sheet = book.sheet_by_name('Login_Credentials')
            cell = second_sheet.cell(1,1)
            url = cell.value
            driver.get(url)
            try:
                WebDriverWait(driver, 5).until(EC.alert_is_present())

                alert = driver.switch_to.alert
                alert.accept()
                print("alert accepted")
            except Exception:
                print("no alert")
                
                
            wait=WebDriverWait(driver, 60)
            unDropDown=wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")))
            driver.execute_script('arguments[0].click()',unDropDown)
            
            signOut=driver.find_element_by_xpath("html/body/div/div/div[1]/div[2]/div[2]/a")
            driver.execute_script('arguments[0].click()',signOut)
            
            
            
            element = wait.until(EC.presence_of_element_located((By.ID, "password")))
            
            if driver.title == "Grovo":
                print("Grovo Application URL Opened")
            else:
                raise Exception.message
    
            print "Grovo Sign-In page is displayed"
            
            print "Entering User name"
            driver.find_element_by_xpath(".//*[@id='username']").send_keys(username)
           
            print "Entering Password"
            element.send_keys(password)
            
            element.send_keys(Keys.TAB)
            print "Clicking on Sign_In button"
            driver.find_element_by_xpath("//*[@id='submitButton']").click()
            
            print "Successfully Logged Into Users account"
            time.sleep(3)
            
        
        
        
if __name__=='__main__':
    
    btc=BaseTestClass()
    btc.UserLogin() 

    gr=GroupWithAttributesReportsTo()
    gr.CreateGroupWithAttributeReportsTo()


    
