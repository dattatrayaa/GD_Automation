'''
Created on 24-Jul-2018

@author: Sheethu C
'''
import os.path
import time
import traceback

from BaseTestClass import BaseTestClass
from BaseTestClass import driver
from openpyxl.reader.excel import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import xlrd
from BaseTestClass import projectPath
from CampaignPageElements import CampPage
from CreateLearnerNew import CreateLearner
from CreateLessonDifferentLessons import CreateLessonDifferentLessons
from DeleteLesson import DeleteLesson
from PlanAssignmentForPageElements import PlanAssignmentForPageElements
from GroupsPageElements import GroupsPageElements
from AddTriggerForPageElements import AddTriggerPage
from BaseTestClass import projectPath
from GroupDelete import GroupDelete
from CreateUserXpath import CreateUserXpath
class GroupDeletedUsersDisplay:
    def groupCreation(self,groupName,FirstName1):
        driver.refresh()
        wait=WebDriverWait(driver, 60)
        
        group=GroupsPageElements()
        
        group.adminSideMenuUnexpanded()
        print "Clicked on admin icon"
         
        group.groupSideMenuExpanded()  
        print "Clicked on Group icon"
        
        
        print "Checking Group page is displayed"
        if driver.find_element_by_xpath(group.groupPageHeader()).is_displayed():
            print "Group page is displayed successfully"
        else:
            print "Group page is not displayed"
            raise Exception
        
        
        group.createGroupButton()
        print "Clicked on Create Group button"
        
        group.enteringGroupname(groupName)
        
        print "Group Name entered....."
        
        group.nextButton()
        print "Clicked on Next button"
        
        wait.until(EC.visibility_of_element_located((By.XPATH,group.groupDetailPageHeader())))
        print "Checking group is created"
        
        time.sleep(4)
        headerText=driver.find_element_by_xpath(group.groupDetailPageHeader()).text
        
        if headerText==groupName:
            print "Group created successfully"
            
        else:
            print "Group not created"
            raise Exception
        
        
        print "Adding user to Group by name"
        group.addByNameButton()
        group.addByName(FirstName1)
        
        
        print "Verifying user is added to group"
        group.groupAddedInList(FirstName1)
        
    
        group.saveButton()
        print "Clicked on Save button"

        group.saveButtonPopup()
        print "Clicked on Save from popup"
        
        
        print "Checking Created Group is displayed in Grid"
        group.groupsLinkFromBreadCrumb()
        
        print "Searching Group"
        group.createdGroupDisplayedInGrid(groupName)
         
        
        driver.refresh()
        
    def updateUser(self):
        print "Updating Details"
        book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        second_sheet = book.sheet_by_name('Groups')
        #First User
        cell = second_sheet.cell(201,1)
        FirstName = cell.value
        FirstNameId = FirstName.split("_")
        emp = FirstNameId[0]+"_"
        ids = FirstNameId[1]
        empId1 = int(ids)+1
        FirstNameUpdated1= emp+str(empId1)
        
        
        cell2 = second_sheet.cell(204,1)
        LearnId = cell2.value
        LearnsId = LearnId.split("#")
        emp = LearnsId[0]+"#"
        ids = LearnsId[1]
        empId = int(ids)+1
        LearnerEmpId1 = emp+str(empId)
        
        
        cell1 = second_sheet.cell(203,1)
        LearnEmailId = cell1.value
        LearnerEmail = LearnEmailId.split("@")
        spitValue = LearnerEmail[0][:4]
        email = spitValue+str(empId)
        id = "@"+LearnerEmail[1]
        LearnerEmailId1 = email+id
        
        
        
        #second User
        cell = second_sheet.cell(207,1)
        FirstName = cell.value
        FirstNameId = FirstName.split("_")
        emp = FirstNameId[0]+"_"
        ids = FirstNameId[1]
        empId1 = int(ids)+1
        FirstNameUpdated2= emp+str(empId1)
        
        cell2 = second_sheet.cell(210,1)
        LearnId = cell2.value
        LearnsId = LearnId.split("#")
        emp = LearnsId[0]+"#"
        ids = LearnsId[1]
        empId = int(ids)+1
        LearnerEmpId2 = emp+str(empId)
        
        
        cell1 = second_sheet.cell(209,1)
        LearnEmailId = cell1.value
        LearnerEmail = LearnEmailId.split("@")
        spitValue = LearnerEmail[0][:4]
        email = spitValue+str(empId)
        id = "@"+LearnerEmail[1]
        LearnerEmailId2 = email+id
        
        wb = load_workbook(os.path.join('Test_Data/TestData.xlsx'))
        
        #print (wb.sheetnames)
    
        sheet = wb['Groups']
        
        sheet.cell(row=202, column=2).value = FirstNameUpdated1    
        sheet.cell(row=204, column=2).value = LearnerEmailId1
        sheet.cell(row=205, column=2).value = LearnerEmpId1
        
        sheet.cell(row=208, column=2).value = FirstNameUpdated2    
        sheet.cell(row=210, column=2).value = LearnerEmailId2
        sheet.cell(row=211, column=2).value = LearnerEmpId2
       
        wb.save(os.path.join('Test_Data/TestData.xlsx'))
    def deactivateUser(self,FirstName1,emailID):
        wait=WebDriverWait(driver,80)
        driver.refresh()
        user =CreateUserXpath()
        wait.until(EC.visibility_of_element_located((By.XPATH,user.adminSideMenu())))
        driver.find_element_by_xpath(user.adminSideMenu()).click()
        print "Clicked on admin page"
        wait.until(EC.visibility_of_element_located((By.XPATH,user.usersideMenu())))
        driver.find_element_by_xpath(user.usersideMenu()).click()
        user.userdeactivate(FirstName1,emailID)
    def addUserToGroup(self,GroupName,FirstName2,FirstName1):
        wait=WebDriverWait(driver,80)
        driver.refresh()
        user =CreateUserXpath()
        wait.until(EC.visibility_of_element_located((By.XPATH,user.adminSideMenu())))
        driver.find_element_by_xpath(user.adminSideMenu()).click()
        print "Clicked on admin page"
        wait.until(EC.visibility_of_element_located((By.XPATH,user.groupsidemenu())))
        driver.find_element_by_xpath(user.groupsidemenu()).click()
        wait.until(EC.visibility_of_element_located((By.ID,user.searchGroup())))
        driver.find_element_by_id(user.searchGroup()).send_keys(GroupName)
        wait.until(EC.visibility_of_element_located((By.XPATH,user.groupgrid(GroupName))))
        driver.find_element_by_xpath(user.groupgrid(GroupName)).click()
        group=GroupsPageElements()
        group.addByNameButton()
        group.addByName(FirstName2)
        print "Verifying user is added to group"
        group.groupAddedInList(FirstName2)
        time.sleep(5)
        if driver.find_element_by_xpath("//tbody/tr/td[.='"+FirstName1+"']"):
            raise Exception
        else:
            print "deactivated user is not in the list"
        
        group.saveButton()
        print "Clicked on Save button"
        group.saveButtonPopup()
        print "Clicked on Save from popup"
        
    def mainGroupDeletedUsersDisplay(self):
        try:
            book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
            first_sheet = book.sheet_by_name("Groups")
            
            cell = first_sheet.cell(200,1)
            groupName = cell.value
            
            cell = first_sheet.cell(201,1)
            FirstName1 = cell.value
        
            cell = first_sheet.cell(202,1)
            LastName1 = cell.value
        
            cell = first_sheet.cell(203,1)
            Email1 = cell.value
        
            cell = first_sheet.cell(204,1)
            EmployeeId1 = cell.value
        
            cell = first_sheet.cell(205,1)
            Password1 = cell.value
        
            cell = first_sheet.cell(206,1)
            NewPassword1 = cell.value
            
            #Second User
            cell = first_sheet.cell(207,1)
            FirstName2 = cell.value
        
            cell = first_sheet.cell(208,1)
            LastName2 = cell.value
        
            cell = first_sheet.cell(209,1)
            Email2 = cell.value
        
            cell = first_sheet.cell(210,1)
            EmployeeId2 = cell.value
        
            cell = first_sheet.cell(211,1)
            Password2 = cell.value
        
            cell = first_sheet.cell(212,1)
            NewPassword2 = cell.value
        
            book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
            second_sheet = book.sheet_by_name('Login_Credentials')
            cell = second_sheet.cell(1,1)
            url = cell.value
        
            cell = second_sheet.cell(3,1)
            username = cell.value
        
            cell = second_sheet.cell(3,2)
            password = cell.value
        
            ob=CreateLearner()
            ob.createNewLearnerMain(FirstName1,LastName1,Email1,EmployeeId1,Password1,NewPassword1,url,username,password)
            ob.createNewLearnerMain(FirstName2,LastName2,Email2,EmployeeId2,Password2,NewPassword2,url,username,password)
            #ob.createNewLearnerMain(FirstName3,LastName3,Email3,EmployeeId3,Password3,NewPassword3,url,username,password)
            grp=GroupDeletedUsersDisplay()
            grp.groupCreation(groupName,FirstName1)
            grp.deactivateUser(FirstName1,Email1)
            grp.addUserToGroup(groupName,FirstName2,FirstName1)
        except Exception as e:
            traceback.print_exc()
            print (e)
            raise Exception   
          
        finally:
            ob=GroupDeletedUsersDisplay()
            grpd= GroupDelete()
            ob.updateUser()
            grpd.deleteGroup(groupName) 
            second_sheet = book.sheet_by_name('Login_Credentials')
            cell = second_sheet.cell(1,1)
            url = cell.value
            driver.get(url)
            try:
                WebDriverWait(driver, 5).until(EC.alert_is_present())

                alert = driver.switch_to.alert
                alert.accept()
                print("alert accepted")
            except Exception:
                print("no alert")
    
        
        
        
if __name__=='__main__':
    
    btc=BaseTestClass()
    btc.userLogin() 

    gr=GroupDeletedUsersDisplay()
    gr.mainGroupDeletedUsersDisplay()   
    
    
    
    
    
    
    
    