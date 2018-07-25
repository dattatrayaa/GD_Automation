'''
Created on 24-Jul-2018

@author: Sheethu C
'''

import os.path
import time
import traceback
from BaseTestClass import BaseTestClass
from BaseTestClass import driver
from openpyxl.reader.excel import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import xlrd
from BaseTestClass import projectPath
from CampaignPageElements import CampPage
from CreateLearnerNew import CreateLearner
from CreateLessonDifferentLessons import CreateLessonDifferentLessons
from DeleteLesson import DeleteLesson
from PlanAssignmentForPageElements import PlanAssignmentForPageElements
from GroupsPageElements import GroupsPageElements
from AddTriggerForPageElements import AddTriggerPage
from BaseTestClass import projectPath
from GroupDelete import GroupDelete
from CreateUserXpath import CreateUserXpath
class GroupCreationUserAutoSuggestion:
    def groupCreation(self,groupName,FirstName,LastName):
        driver.refresh()
        wait=WebDriverWait(driver, 60)
        
        group=GroupsPageElements()
        
        group.adminSideMenuUnexpanded()
        print "Clicked on admin icon"
         
        group.groupSideMenuExpanded()  
        print "Clicked on Group icon"
        
        
        print "Checking Group page is displayed"
        if driver.find_element_by_xpath(group.groupPageHeader()).is_displayed():
            print "Group page is displayed successfully"
        else:
            print "Group page is not displayed"
            raise Exception
        
        
        group.createGroupButton()
        print "Clicked on Create Group button"
        
        group.enteringGroupname(groupName)
        
        print "Group Name entered....."
        
        group.nextButton()
        print "Clicked on Next button"
        
        wait.until(EC.visibility_of_element_located((By.XPATH,group.groupDetailPageHeader())))
        print "Checking group is created"
        
        time.sleep(4)
        headerText=driver.find_element_by_xpath(group.groupDetailPageHeader()).text
        
        if headerText==groupName:
            print "Group created successfully"
            
        else:
            print "Group not created"
            raise Exception
        
        
        print "Adding user to Group by name"
        addByName=wait.until(EC.element_to_be_clickable((By.XPATH,".//*[@id='content']/div/div[3]/div[2]/div/div[2]/div/div/div[1]/div/div/div[2]/div/button")))
        addByName.click()
        wait.until(EC.visibility_of_element_located((By.XPATH,"//div[@class='Select-placeholder']")))
        wait.until(EC.visibility_of_element_located((By.XPATH,"//div[@class='Select-placeholder']")))
        names=driver.find_element_by_xpath("//div[@class='Select-placeholder']")
        time.sleep(2)
        webdriver.ActionChains(driver).move_to_element(names).click()
        
        lName=LastName.title()
        webdriver.ActionChains(driver).move_to_element(names).send_keys(FirstName+" "+lName).perform()
        userDisplayed=wait.until(EC.visibility_of_element_located((By.XPATH,"//div[@role='option']/span[contains(.,'"+FirstName+"')]")))
        if userDisplayed.is_displayed():
            print "Searched user is dispalyed"
            webdriver.ActionChains(driver).move_to_element(userDisplayed).click().perform()
            print "User selected and added to group"
            time.sleep(2)
            wait.until(EC.invisibility_of_element_located((By.XPATH,"html/body/div[4]/div/div/div/div")))
            print "Verifying user is added to group"
            group.groupAddedInList(FirstName)
            group.saveButton()
            print "Clicked on Save button"
            group.saveButtonPopup()
            print "Clicked on Save from popup"
            print "Checking Created Group is displayed in Grid"
            group.groupsLinkFromBreadCrumb()
            print "Searching Group"
            group.createdGroupDisplayedInGrid(groupName)
        else:
            raise Exception
        
        driver.refresh()
    def updateUser(self):
        print "Updating Details"
        book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        second_sheet = book.sheet_by_name('Groups')
        #First User
        cell = second_sheet.cell(216,1)
        FirstName = cell.value
        FirstNameId = FirstName.split("_")
        emp = FirstNameId[0]+"_"
        ids = FirstNameId[1]
        empId1 = int(ids)+1
        FirstNameUpdated1= emp+str(empId1)
        
        
        cell2 = second_sheet.cell(219,1)
        LearnId = cell2.value
        LearnsId = LearnId.split("#")
        emp = LearnsId[0]+"#"
        ids = LearnsId[1]
        empId = int(ids)+1
        LearnerEmpId1 = emp+str(empId)
        
        
        cell1 = second_sheet.cell(218,1)
        LearnEmailId = cell1.value
        LearnerEmail = LearnEmailId.split("@")
        spitValue = LearnerEmail[0][:4]
        email = spitValue+str(empId)
        id = "@"+LearnerEmail[1]
        LearnerEmailId1 = email+id
        wb = load_workbook(os.path.join('Test_Data/TestData.xlsx'))
        
        #print (wb.sheetnames)
    
        sheet = wb['Groups']
        
        sheet.cell(row=217, column=2).value = FirstNameUpdated1    
        sheet.cell(row=219, column=2).value = LearnerEmailId1
        sheet.cell(row=220, column=2).value = LearnerEmpId1
        
        wb.save(os.path.join('Test_Data/TestData.xlsx'))
        
        
    def mainGroupCreationUserAutoSuggestion(self):
        try:
            book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
            first_sheet = book.sheet_by_name("Groups")
            
            cell = first_sheet.cell(215,1)
            groupName = cell.value
            
            cell = first_sheet.cell(216,1)
            FirstName= cell.value
        
            cell = first_sheet.cell(217,1)
            LastName = cell.value
        
            cell = first_sheet.cell(218,1)
            Email = cell.value
        
            cell = first_sheet.cell(219,1)
            EmployeeId = cell.value
        
            cell = first_sheet.cell(220,1)
            Password = cell.value
        
            cell = first_sheet.cell(221,1)
            NewPassword = cell.value
            
           
        
            book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
            second_sheet = book.sheet_by_name('Login_Credentials')
            cell = second_sheet.cell(1,1)
            url = cell.value
        
            cell = second_sheet.cell(3,1)
            username = cell.value
        
            cell = second_sheet.cell(3,2)
            password = cell.value
        
            ob=CreateLearner()
            ob.createNewLearnerMain(FirstName,LastName,Email,EmployeeId,Password,NewPassword,url,username,password)
            grp=GroupCreationUserAutoSuggestion()
            grp.groupCreation(groupName,FirstName,LastName)
            
            
        except Exception as e:
            traceback.print_exc()
            print (e)
            raise Exception   
          
        finally:
            ob=GroupCreationUserAutoSuggestion()
            grpd= GroupDelete()
            ob.updateUser()
            grpd.deleteGroup(groupName) 
            second_sheet = book.sheet_by_name('Login_Credentials')
            cell = second_sheet.cell(1,1)
            url = cell.value
            driver.get(url)
            try:
                WebDriverWait(driver, 5).until(EC.alert_is_present())

                alert = driver.switch_to.alert
                alert.accept()
                print("alert accepted")
            except Exception:
                print("no alert")
    
        
        
        
if __name__=='__main__':
    
    btc=BaseTestClass()
    btc.userLogin() 

    gr=GroupCreationUserAutoSuggestion()
    gr.mainGroupCreationUserAutoSuggestion() 
        
        
        