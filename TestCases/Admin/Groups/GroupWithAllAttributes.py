'''
Created on 14-May-2018

@author: dattatraya
'''

import os
import time
import traceback

from BaseTestClass import BaseTestClass
from BaseTestClass import driver
from openpyxl.reader.excel import load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import xlrd

from GroupsPageElements import GroupsPageElements
from UsersPageElements import UsersPageElements
from CreateUserWithRole import CreateUserWithRole
from BaseTestClass import projectPath
class GroupWithAllAttributes:
    def create(self,FirstName,LastName,Email,EmployeeId,Password,cityName,attributeName1,countryName,attributeName2,deptName,attributeName3,
               dayOfCurrentmonth,attributeName4,jobtitle,attributeName5,location,attributeName6,region,attributeName7,reportsTo,attributeName8,state,attributeName9):
        wait=WebDriverWait(driver, 60)
        driver.refresh()
        user=UsersPageElements()
        
        user.AdminFromSideMenuClick()
        print "Clicked on Admin"
        
        user.UsersFromSideMenuClick()
        print "clicked on Users"
        
        user.AddOrEditUserButtonClick()
        print "Clicked on Add or editUser"

        
        user.AddAnIndividualUserButtonClick()
        print "Clicked on Add An individual User"
        time.sleep(1)
        wait.until(EC.visibility_of_element_located((By.XPATH,user.AddUserPageHeaderXpath())))
        
        print "Verifying Add user Page"
        if driver.find_element_by_xpath(user.AddUserPageHeaderXpath()).is_displayed():
            print("Add user Page is displayed")
        else:
            print ""
            raise Exception
        
        
        print "Verifying First Name field"
        fn=wait.until(EC.visibility_of_element_located((By.ID,user.FirstNameFieldID())))
        if fn.is_displayed():
            print(" First Name field displayed")
        else:
            print ""
            raise Exception
        fn.send_keys(FirstName)
        print "FirstName is Entered ::"+FirstName
        
        
        
        print "Last NAme verifying"
        if driver.find_element_by_id(user.LastNameFieldID()).is_displayed():
            print(" Last Name field displayed")
        else:
            print ""
            raise Exception
        driver.find_element_by_id(user.LastNameFieldID()).send_keys(LastName)
        print "Last Name is Entered ::"+LastName
        
        
        print "Email verifying"
        if driver.find_element_by_id(user.EmailFieldID()).is_displayed():
            print("Email field displayed")
        else:
            print ""
            raise Exception
        driver.find_element_by_id(user.EmailFieldID()).send_keys(Email)
        print "Email is Entered ::"+Email
        
        
        print "Employee ID verifying"
        if driver.find_element_by_id(user.EmployeeIDFieldID()).is_displayed():
            print("Employee ID field displayed")
        else:
            print ""
            raise Exception
        driver.find_element_by_id(user.EmployeeIDFieldID()).send_keys(EmployeeId)
        print "Employee ID  is Entered ::"+EmployeeId
        
        
        print "Adding additional Attribute"
        user.AdditionalAttributesViewClick()
        
        
        
        print "Creating Option for city"
        user.CityAttributeFieldEnterData(cityName)
        print "Selecting Option"
        user.SelectCreateOption(attributeName1)
        print "Option Selected "+attributeName1
        
        
        
        print "Creating Option for Country"
        user.CountryAttributeFieldEnterData(countryName)
        print "Selecting Option"
        user.SelectCreateOption(attributeName2)
        print "Option Selected :"+attributeName2
        
        
        
        print "Creating Option for Department"
        user.DepartmentAttributeFieldEnterData(deptName)
        print "Selecting Option"
        user.SelectCreateOption(attributeName3)
        print "Option Selected :"+attributeName3
        
        
        
        print "Creating Option for Hiredate"
        user.HireDateAttributeFieldSelectDay(dayOfCurrentmonth)
        print "Selecting Option"
               
        
        print "Creating Option for Job Title"
        user.jobTitleAttributeFieldEnterData(jobtitle)
        print "Selecting Option"
        user.SelectCreateOption(attributeName5)
        print "Option Selected :"+attributeName5
        
        
        
        print "Creating Option for Location"
        user.LocationAttributeFieldEnterData(location)
        print "Selecting Option"
        user.SelectCreateOption(attributeName6)
        print "Option Selected :"+attributeName6
        
        
        
        print "Creating Option for Region"
        user.RegionAttributeFieldEnterData(region)
        print "Selecting Option"
        user.SelectCreateOption(attributeName7)
        print "Option Selected :"+attributeName7
        
        
        
        print "Creating Option for Reports to"
        user.ReportsToAttributeFieldEnterData(reportsTo)
        print "Selecting Option"
        user.SelectCreateOption(attributeName8)
        print "Option Selected :"+attributeName8
        
        
        
        print "Creating Option for State"
        user.StateAttributeFieldEnterData(state)
        print "Selecting Option"
        user.SelectCreateOption(attributeName9)
        print "Option Selected :"+attributeName9
        
        
        
        
        
        
        
        print "Inherited Role Verifying"
        if driver.find_element_by_xpath(user.InHeritedRoleFieldXpath()).is_displayed():
            print("Inherited Role field displayed")
        else:
            print ""
            raise Exception
        
        print "Password Field is Verifying"
        if driver.find_element_by_id(user.PassWordFieldID()).is_displayed():
            print("Password field displayed")
        else:
            print ""
            raise Exception
        driver.find_element_by_id(user.PassWordFieldID()).send_keys(Password)
        print "Password is Entered ::"+Password
        
        
        user.AddButtonClick()
        print "Clicked on add button"
        time.sleep(2)

        user.SaveButtonClick()        
        print "Clicked on Save"
        
        wait.until(EC.visibility_of_element_located((By.XPATH,user.UsersPageHeader())))
        
        
        print "Searching for the Created User"
        user.SearchForUserInGrid(FirstName)
        
        
        user.CheckUserDisplayedInGrid(FirstName) 
        
        
        
    def groupCreateForCampaign(self,groupName,FirstName,cityName,attributeName1,countryName,attributeName2,deptName,attributeName3,
               dayOfCurrentmonth,attributeName4,jobtitle,attributeName5,location,attributeName6,region,attributeName7,reportsTo,attributeName8,state,attributeName9):
        driver.refresh()
        wait=WebDriverWait(driver, 60)
        
        group=GroupsPageElements()
        
        group.adminSideMenuUnexpanded()
        print "Clicked on admin icon"
         
        group.groupSideMenuExpanded()  
        print "Clicked on Group icon"
        
        
        print "Checking Group page is displayed"
        if driver.find_element_by_xpath(group.groupPageHeader()).is_displayed():
            print "Group page is displayed successfully"
        else:
            print "Group page is not displayed"
            raise Exception
        
        
        group.createGroupButton()
        print "Clicked on Create Group button"
        
        group.enteringGroupname(groupName)
        
        print "Group Name entered....."
        
        group.nextButton()
        print "Clicked on Next button"
        
        wait.until(EC.visibility_of_element_located((By.XPATH,group.groupDetailPageHeader())))
        print "Checking group is created"
        
        time.sleep(4)
        headerText=driver.find_element_by_xpath(group.groupDetailPageHeader()).text
        
        if headerText==groupName:
            print "Group created successfully"
            
        else:
            print "Group not created"
            raise Exception
        
        
        print "Adding user to Group by name"
        
        group.AddByAttributes()
        
        print "Checking all option is selected"
        if group.AllOptionSelected()=="all":
            print "all is selected"
        else:
            print "all is not selected"
            raise Exception
        
        
        print "Adding User by attribute"
        group.AttributeNameEnter(attributeName1)
        group.ValueSelect(cityName)
        print "Selected "+attributeName1
        
        
        
        group.addAttributePopupButtonClick()
        print "Adding User by attribute"
        group.AttributeNameEnter(attributeName2)
        group.ValueSelect(countryName)
        print "Selected "+attributeName2
        
        
        group.addAttributePopupButtonClick()
        print "Adding User by attribute"
        group.AttributeNameEnter(attributeName3)
        group.ValueSelect(deptName)
        print "Selected "+attributeName3
        
        
        group.addAttributePopupButtonClick()
        print "Adding User by attribute"
        group.AttributeNameEnter(attributeName4)
        group.DateSelect(dayOfCurrentmonth)
        print "Selected "+attributeName4
        
        
        group.addAttributePopupButtonClick()
        print "Adding User by attribute"
        group.AttributeNameEnter(attributeName5)
        group.ValueSelect(jobtitle)
        print "Selected "+attributeName5
        
        
        group.addAttributePopupButtonClick()
        print "Adding User by attribute"
        group.AttributeNameEnter(attributeName6)
        group.ValueSelect(location)
        print "Selected "+attributeName6
        
        
        group.addAttributePopupButtonClick()
        print "Adding User by attribute"
        group.AttributeNameEnter(attributeName7)
        group.ValueSelect(region)
        print "Selected "+attributeName7
        
        
        group.addAttributePopupButtonClick()
        print "Adding User by attribute"
        group.AttributeNameEnter(attributeName8)
        group.ValueSelect(reportsTo)
        print "Selected "+attributeName8
        
        
        group.addAttributePopupButtonClick()
        print "Adding User by attribute"
        group.AttributeNameEnter(attributeName9)
        group.ValueSelect(state)
        print "Selected "+attributeName9
        
        
        
        
        
        
        
        
        print "Clicking on Preview button"
        group.PreviewGroupButtonClick()
        
               
        
        print "Verifying user is added to group"
        group.groupAddedInList(FirstName)
        
        group.saveButton()
        print "Clicked on Save button"

        group.saveButtonPopup()
        print "Clicked on Save from popup"
        
        
        print "Checking Created Group is displayed in Grid"
        group.groupsLinkFromBreadCrumb()
        
        print "Searching Group"
        group.createdGroupDisplayedInGrid(groupName)
         
        
        driver.refresh()
    
    
    def CreateGroupWithAllAttributes(self):
        book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        first_sheet = book.sheet_by_name('Groups')
        
        cell = first_sheet.cell(105,1)
        FirstName = cell.value
        
        FirstNameId = FirstName.split("_")
        emp = FirstNameId[0]+"_"
        ids = FirstNameId[1]
        empId = int(ids)+1
        FirstNameUpdated= emp+str(empId)
        
        cell = first_sheet.cell(106,1)
        LastName = cell.value
        
        LastNameID = LastName.split("_")
        emp = LastNameID[0]+"_"
        ids = LastNameID[1]
        empId = int(ids)+1
        LastNameUpdated= emp+str(empId)
        
        
        
        cell = first_sheet.cell(107,1)
        Email = cell.value
        
        EmailId = Email.split("_")
        emp = EmailId[0]+"_"
        ids = EmailId[1]
        remaining="_"+EmailId[2]
        empId = int(ids)+1
        EmailIdUpdated= emp+str(empId)+remaining
        
        cell = first_sheet.cell(108,1)
        EmployeeId = cell.value
        
        Employee = EmployeeId.split("_")
        emp = Employee[0]+"_"
        ids = Employee[1]
        empId = int(ids)+1
        EmployeeIdUpdated= emp+str(empId)
        
        cell1 = first_sheet.cell(109,1)
        Password = cell1.value
        
        
        
        cell1 = first_sheet.cell(110,1)
        groupName = cell1.value
        groupNameId = groupName.split("_")
        emp = groupNameId[0]+"_"
        ids = groupNameId[1]
        empId = int(ids)+1
        groupNameUpdated= emp+str(empId)
        
        
        #Attributes
        
        cell1 = first_sheet.cell(112,1)
        cityName = cell1.value
        cityNameId = cityName.split("_")
        emp = cityNameId[0]+"_"
        ids = cityNameId[1]
        empId = int(ids)+1
        cityNameIdUpdated= emp+str(empId)
        
        cell1 = first_sheet.cell(113,1)
        attributeName1 = cell1.value
        
        cell1 = first_sheet.cell(114,1)
        countryName = cell1.value
        countryNameId = countryName.split("_")
        emp = countryNameId[0]+"_"
        ids = countryNameId[1]
        empId = int(ids)+1
        countryNameUpdated= emp+str(empId)
        cell1 = first_sheet.cell(115,1)
        attributeName2 = cell1.value
        
        cell1 = first_sheet.cell(116,1)
        deptName = cell1.value
        deptNameId = deptName.split("_")
        emp = deptNameId[0]+"_"
        ids = deptNameId[1]
        empId = int(ids)+1
        deptNameIdUpdated= emp+str(empId)
        cell1 = first_sheet.cell(117,1)
        attributeName3 = cell1.value
        
        cell1 = first_sheet.cell(118,1)
        dayOfCurrentmonth = cell1.value
        cell1 = first_sheet.cell(119,1)
        attributeName4 = cell1.value

        
        cell1 = first_sheet.cell(120,1)
        jobtitle = cell1.value
        jobtitleId = jobtitle.split("_")
        emp = jobtitleId[0]+"_"
        ids = jobtitleId[1]
        empId = int(ids)+1
        jobtitleIdUpdated= emp+str(empId)
        cell1 = first_sheet.cell(121,1)
        attributeName5 = cell1.value
        
        cell1 = first_sheet.cell(122,1)
        location = cell1.value
        locationId = location.split("_")
        emp = locationId[0]+"_"
        ids = locationId[1]
        empId = int(ids)+1
        locationIdUpdated= emp+str(empId)
        cell1 = first_sheet.cell(123,1)
        attributeName6 = cell1.value
        
        cell1 = first_sheet.cell(124,1)
        region = cell1.value
        regionId = region.split("_")
        emp = regionId[0]+"_"
        ids = regionId[1]
        empId = int(ids)+1
        regionIdUpdated= emp+str(empId)
        cell1 = first_sheet.cell(125,1)
        attributeName7 = cell1.value
        
        
        cell1 = first_sheet.cell(126,1)
        state = cell1.value
        stateId = state.split("_")
        emp = stateId[0]+"_"
        ids = stateId[1]
        empId = int(ids)+1
        stateIdUpdated= emp+str(empId)
        cell1 = first_sheet.cell(127,1)
        attributeName9 = cell1.value
        
        #For Reports to
        cell = first_sheet.cell(128,1)
        FirstName1 = cell.value
        FirstNameId = FirstName1.split("_")
        emp = FirstNameId[0]+"_"
        ids = FirstNameId[1]
        empId = int(ids)+1
        FirstNameUpdated1= emp+str(empId)
        
        cell = first_sheet.cell(129,1)
        LastName1 = cell.value
        LastNameID = LastName1.split("_")
        emp = LastNameID[0]+"_"
        ids = LastNameID[1]
        empId = int(ids)+1
        LastNameUpdated1= emp+str(empId)
        
        cell = first_sheet.cell(130,1)
        Email1 = cell.value
        EmailId = Email1.split("_")
        emp = EmailId[0]+"_"
        ids = EmailId[1]
        remaining="_"+EmailId[2]
        empId = int(ids)+1
        EmailIdUpdated1= emp+str(empId)+remaining
        
        cell = first_sheet.cell(131,1)
        EmployeeId1 = cell.value
        Employee = EmployeeId1.split("_")
        emp = Employee[0]+"_"
        ids = Employee[1]
        empId = int(ids)+1
        EmployeeIdUpdated1= emp+str(empId)
        
        cell1 = first_sheet.cell(132,1)
        Password1 = cell1.value
        
        cell1 = first_sheet.cell(133,1)
        role = cell1.value
        
        cell1 = first_sheet.cell(132,1)
        NewPassword = cell1.value
        
        cell1 = first_sheet.cell(134,1)
        attributeName8 = cell1.value
        
        
        book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        s_sheet = book.sheet_by_name('Login_Credentials')
        cell = s_sheet.cell(1,1)
        url = cell.value
        cell = s_sheet.cell(3,1)
        username = cell.value
        cell = s_sheet.cell(3,2)
        password = cell.value
        
        
        
        
        
        
       
        
        
        #updating user values
        wb = load_workbook(os.path.join('Test_Data/TestData.xlsx'))
            #print (wb.sheetnames)
        
        sheet = wb['Groups']
        
        sheet.cell(row=106, column=2).value = FirstNameUpdated
        sheet.cell(row=107, column=2).value = LastNameUpdated
        sheet.cell(row=108, column=2).value = EmailIdUpdated
        sheet.cell(row=109, column=2).value = EmployeeIdUpdated
        sheet.cell(row=111, column=2).value = groupNameUpdated
        
        sheet.cell(row=113, column=2).value = cityNameIdUpdated
        sheet.cell(row=115, column=2).value = countryNameUpdated
        sheet.cell(row=117, column=2).value = deptNameIdUpdated
        sheet.cell(row=121, column=2).value = jobtitleIdUpdated
        sheet.cell(row=123, column=2).value = locationIdUpdated
        sheet.cell(row=125, column=2).value = regionIdUpdated
        sheet.cell(row=127, column=2).value = stateIdUpdated
        
        sheet.cell(row=129, column=2).value = FirstNameUpdated1
        sheet.cell(row=130, column=2).value = LastNameUpdated1
        sheet.cell(row=131, column=2).value = EmailIdUpdated1
        sheet.cell(row=132, column=2).value = EmployeeIdUpdated1
        
        
        
        
        wb.save(os.path.join('Test_Data/TestData.xlsx'))
       
        
        try:
            t=GroupWithAllAttributes()
            
            cr=CreateUserWithRole()
            cr.createUserWithRoleMain(FirstName1, LastName1, Email1, EmployeeId1, Password1, role, NewPassword, url, username, password)
            
            print "\nCreating New User\n"
            t.create(FirstName, LastName, Email, EmployeeId, Password, cityName, attributeName1, countryName, attributeName2, deptName, attributeName3, dayOfCurrentmonth, 
                     attributeName4, jobtitle, attributeName5, location, attributeName6, region, attributeName7, FirstName1, attributeName8, state, attributeName9)
            
            
            print "\nCreating Group with '"+region+"' Attribute\n"
            t.groupCreateForCampaign(groupName, FirstName, cityName, attributeName1, countryName, attributeName2, deptName, attributeName3, dayOfCurrentmonth, attributeName4, 
                                     jobtitle, attributeName5, location, attributeName6, region, attributeName7, FirstName1, attributeName8, state, attributeName9)
            
            
        except Exception as e:
            traceback.print_exc()
            print (e)
            raise Exception   
          
        finally:  
            second_sheet = book.sheet_by_name('Login_Credentials')
            cell = second_sheet.cell(1,1)
            url = cell.value
            driver.get(url)
            try:
                WebDriverWait(driver, 5).until(EC.alert_is_present())

                alert = driver.switch_to.alert
                alert.accept()
                print("alert accepted")
            except Exception:
                print("no alert")
    
        
        
        
if __name__=='__main__':
    
    btc=BaseTestClass()
    btc.UserLogin() 

    gr=GroupWithAllAttributes()
    gr.CreateGroupWithAllAttributes()

    

    
    


