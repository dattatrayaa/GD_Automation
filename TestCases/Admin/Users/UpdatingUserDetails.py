'''
Created on 12-Mar-2018

@author: QA
'''

from operator import contains
import os.path
import time
import traceback

from openpyxl.reader.excel import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import xlrd
from BaseTestClass import driver
from BaseTestClass import BaseTestClass
from BaseTestClass import projectPath
from BaseTestClass import excelPath
class UpdatingUserDetails():

    def mainUpdateExcelWithUserDetails(self):
        try:
            book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
            second_sheet = book.sheet_by_name('CreateuserfromUI')
            cell2 = second_sheet.cell(1,4)
            LId = cell2.value
            cell1 = second_sheet.cell(1,3)
            LEmail = cell1.value
            cell4 = second_sheet.cell(2,4)
            CId = cell4.value
            cell3 = second_sheet.cell(2,3)
            CEmail = cell3.value
            cell6 = second_sheet.cell(3,4)
            MId = cell6.value
            cell5 = second_sheet.cell(3,3)
            MEmail = cell5.value
            cell8 = second_sheet.cell(4,4)
            LAId = cell8.value
            cell7 = second_sheet.cell(4,3)
            LAEmail = cell7.value
            
            
            
            
            
            
            cell2 = second_sheet.cell(1,4)
            LearnId = cell2.value
            LearnsId = LearnId.split("#")
            emp = LearnsId[0]+"#"
            ids = LearnsId[1]
            empId = int(ids)+1
            LearnerEmpId = emp+str(empId)
            print LearnerEmpId
            
            cell1 = second_sheet.cell(1,3)
            LearnEmailId = cell1.value
            LearnerEmail = LearnEmailId.split("@")
            spitValue = LearnerEmail[0][:4]
            email = spitValue+str(empId)
            id = "@"+LearnerEmail[1]
            LearnerEmailId = email+id
            print LearnerEmailId
            
            
            cell4 = second_sheet.cell(2,4)
            CreatorId = cell4.value
            CreatorsId = CreatorId.split("#")
            emp1 = CreatorsId[0]+"#"
            ids1 = CreatorsId[1]
            empId1 = int(ids1)+1
            CreatorsEmpId = emp1+str(empId1)
            print CreatorsEmpId
            
            
            cell3 = second_sheet.cell(2,3)
            CreateEmailId = cell3.value
            CreatorEmail = CreateEmailId.split("@")
            spitValue1 = CreatorEmail[0][:4]
            email1 = spitValue1+str(empId1)
            id1 = "@"+CreatorEmail[1]
            CreatorEmailId = email1+id1
            print CreatorEmailId
            
            cell6 = second_sheet.cell(3,4)
            MasterAdminId = cell6.value
            MasterAdminsId = MasterAdminId.split("#")
            emp2 = MasterAdminsId[0]+"#"
            ids2 = MasterAdminsId[1]
            empId2 = int(ids2)+1
            MasterAdminsEmpId = emp2+str(empId2)
            print MasterAdminsEmpId
            
            cell5 = second_sheet.cell(3,3)
            MasterEmail = cell5.value
            MasterAdminEmail = MasterEmail.split("@")
            spitValue2 = MasterAdminEmail[0][:4]
            email2 = spitValue2+str(empId2)
            id2 = "@"+MasterAdminEmail[1]
            MasterAdminEmailId = email2+id2
            print MasterAdminEmailId
            
            cell8 = second_sheet.cell(4,4)
            LearningAdminId = cell8.value
            LearningAdminsId = LearningAdminId.split("#")
            emp3 = LearningAdminsId[0]+"#"
            ids3 = LearningAdminsId[1]
            empId3 = int(ids3)+1
            LearningAdminsEmpId = emp3+str(empId3)
            print LearningAdminsEmpId
            
            cell7 = second_sheet.cell(4,3)
            LearningEmail = cell7.value
            LearningAdminEmail = LearningEmail.split("@")
            spitValue3 = LearningAdminEmail[0][:4]
            email3 = spitValue3+str(empId3)
            id3 = "@"+LearningAdminEmail[1]
            LearningAdminEmailId = email3+id3
            print LearningAdminEmailId
            
            
           
            
            wb = load_workbook(os.path.join('Test_Data/TestData.xlsx'))
            
            #print (wb.sheetnames)
        
            sheet = wb['CreateuserfromUI']
            
            
            sheet.cell(row=2, column=4).value = LearnerEmailId
            sheet.cell(row=2, column=5).value = LearnerEmpId
            
            sheet.cell(row=3, column=4).value = CreatorEmailId
            sheet.cell(row=3, column=5).value = CreatorsEmpId
            
            sheet.cell(row=4, column=4).value = MasterAdminEmailId
            sheet.cell(row=4, column=5).value = MasterAdminsEmpId
            
            sheet.cell(row=5, column=4).value = LearningAdminEmailId
            sheet.cell(row=5, column=5).value = LearningAdminsEmpId
            
            wb.save(os.path.join('Test_Data/TestData.xlsx'))
            cell1 = sheet.cell(2,4)
            ActualLearnerEmailId = cell1.value
            
            cell1 = sheet.cell(2,5)
            ActualLearnerEmpId = cell1.value
            
            cell1 = sheet.cell(3,4)
            ActualCreatorEmailId = cell1.value
            
            cell1 = sheet.cell(3,5)
            ActualCreatorsEmpId = cell1.value
            
            cell1 = sheet.cell(4,4)
            ActualMasterAdminEmailId = cell1.value
            
            cell1 = sheet.cell(4,5)
            ActualMasterAdminsEmpId = cell1.value
            
            cell1 = sheet.cell(5,4)
            ActualLearningAdminEmailId = cell1.value
            
            cell1 = sheet.cell(5,5)
            ActualLearningAdminsEmpId = cell1.value
            
            print "ActualLearnerEmailId :"+ActualLearnerEmailId
            print "LEmail  :"+LEmail 
            if (LEmail  != ActualLearnerEmailId):
                print "Updated"
            else:
                raise Exception
            print "LId :"+LId
            print "ActualLearnerEmpId :"+ActualLearnerEmpId
            if (LId !=ActualLearnerEmpId):
                print "Updated"
            else:
                raise Exception
            print "CEmail  :"+CEmail 
            print "ActualCreatorEmailId :"+ActualCreatorEmailId
            
            if (CEmail  !=ActualCreatorEmailId):
                print "Updated"
            else:
                raise Exception
            print "CId  :"+CId 
            print "ActualCreatorsEmpId :"+ActualCreatorsEmpId
            
            if (CId  !=ActualCreatorsEmpId):
                print "Updated"
            else:
                raise Exception
            
            print "LAEmail  :"+LAEmail 
            print "ActualLearningAdminEmailId :"+ActualLearningAdminEmailId
            
            if (LAEmail  !=ActualLearningAdminEmailId):
                print "Updated"
            else:
                raise Exception
            print "LAId  :"+LAId 
            print "ActualLearningAdminsEmpId :"+ActualLearningAdminsEmpId
            if (LAId   !=ActualLearningAdminsEmpId):
                print "Updated"
            else:
                raise Exception
            
            print "MEmail  :"+MEmail 
            print "ActualMasterAdminEmailId :"+ActualMasterAdminEmailId
            
            if (MEmail  !=ActualMasterAdminEmailId):
                print "Updated"
            else:
                raise Exception
            print "MId  :"+MId 
            print "ActualMasterAdminsEmpId :"+ActualMasterAdminsEmpId
            if (MId  !=ActualMasterAdminsEmpId):
                print "Updated"
            else:
                raise Exception
            
            print "All User Data Updated in Excel" 
        
        except Exception as e:
            traceback.print_exc()
            print (e)
            raise Exception
       
    def UpdationOfExcelValues(self):
        try:    
            
            update=UpdatingUserDetails()
            update.mainUpdateExcelWithUserDetails()
           
        finally: 
            book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
            second_sheet = book.sheet_by_name('Login_Credentials')
            cell = second_sheet.cell(1,1)
            url = cell.value
            driver.get(url)    
           
if __name__ == '__main__':
    
    obj11= UpdatingUserDetails()
    obj12= BaseTestClass()
    obj12.UserLogin()
    obj11.UpdationOfExcelValues()           
        
