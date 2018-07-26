'''
Created on 08-Jun-2018

@author: dattatraya
'''
import os

from openpyxl.reader.excel import load_workbook
import xlrd
import traceback


class ExcelFunctions:
    
        
        
    def OpenExcelFile(self,ExcelFileName):
        try:
            self.book=xlrd.open_workbook(os.path.abspath(os.path.join(os.path.dirname(__file__),ExcelFileName)))
        except Exception:
            print traceback
            raise Exception('Not able to Open Excel File or check excel file name and type')
        
    def OpenSheet(self,SheetName):
        try:
            self.first_sheet=self.book.sheet_by_name(SheetName) 
            
        except Exception:
            print traceback
            raise Exception('No sheet named '+SheetName)
        
    def getCellData(self,RowNum,ColNum):
        cell1 = self.first_sheet.cell(RowNum,ColNum)
        CellData = cell1.value
        return CellData
    
    def updateExcelFileName(self,ExcelFileName):
        self.wb = load_workbook(os.path.abspath(os.path.join(os.path.dirname(__file__),ExcelFileName)))
    
    def updateExcelSheetName(self,sheetname):
        self.sheet = self.wb[sheetname]
    
    def updateCellData(self,celldata,RowNum,ColNum):
        self.sheet.cell(row=RowNum, column=ColNum).value = celldata
        
    def SaveExcelFile(self,ExcelFileName):
        self.wb.save(os.path.abspath(os.path.join(os.path.dirname(__file__),ExcelFileName)))
        
        
        
        
        
        
        
        
        
        