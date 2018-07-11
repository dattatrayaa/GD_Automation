'''
Created on 11-Jul-2018

@author: geethukn
'''
from operator import contains
import os.path
import time
import traceback

from openpyxl.reader.excel import load_workbook
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import wait, expected_conditions as EC
from selenium.webdriver.support.select import Select
import xlrd

from ADPElements import ADPXpath
from BaseTestClass import BaseTestClass
from BaseTestClass import WebDriverWait
from BaseTestClass import driver


class ADPDisconnect:
    def adpdisconnecting(self):
        ADPDisconnect = ADPXpath()
        
        wait=WebDriverWait(driver,90)
        wait.until(EC.visibility_of_element_located((By.XPATH,ADPDisconnect.adminMenu())))
        # Click on Admin from side menu
        driver.find_element_by_xpath(ADPDisconnect.adminMenu()).click()
        print "Click on Admin from side menu"
        
        
        wait=WebDriverWait(driver,90)
        wait.until(EC.visibility_of_element_located((By.XPATH,ADPDisconnect.integrationMenu())))
        review = driver.find_element_by_xpath(ADPDisconnect.integrationMenu())
        webdriver.ActionChains(driver).move_to_element(review).click(review).perform()
        
        
        wait=WebDriverWait(driver,90)
        wait.until(EC.visibility_of_element_located((By.XPATH,ADPDisconnect.editsetting())))
        # Click on Configure button
        driver.find_element_by_xpath(ADPDisconnect.editsetting()).click()
        print "Click on Edit button"
        
       
        wait.until(EC.visibility_of_element_located((By.XPATH,ADPDisconnect.disconnectlink())))
        # Click on Go To ADP Button
        driver.find_element_by_xpath(ADPDisconnect.disconnectlink()).click()
        print "Click on Disconnect link"
        
        wait.until(EC.visibility_of_element_located((By.XPATH,ADPDisconnect.disconnectbutton())))
        # Click on Go To ADP Button
        driver.find_element_by_xpath(ADPDisconnect.disconnectbutton()).click()
        print "Click on disconnect button"
        
        #time.sleep(6)
        #window_after = driver.window_handles[1]
        #driver.switch_to.window(window_after)
        #wait.until(EC.visibility_of_element_located((By.XPATH,ADPDisconnect.adpBuyNow())))
        #wait.until(EC.element_to_be_clickable((By.XPATH,ADPDisconnect.adpBuyNow())))
        # Click on Go To ADP Button
        #review = driver.find_element_by_xpath(ADPDisconnect.adpBuyNow())
        #webdriver.ActionChains(driver).move_to_element(review).click(review).perform()
        #print "Click on Buy now button button
        
        #login to ADPMarket place
        
        window_after = driver.window_handles[1]
        driver.switch_to.window(window_after)
        wait.until(EC.visibility_of_element_located((By.XPATH,ADPDisconnect.adpLoginusename())))
        print "Reading data from excel sheet"
        book=xlrd.open_workbook(os.path.join('TestData.xlsx'))
        
        first_sheet = book.sheet_by_name('ADP User details')

        # read a cell
        cell = first_sheet.cell(1,0)
        ADPUserID = cell.value
        print "UserId : "+ADPUserID
        cell = first_sheet.cell(1,1)
        Password = cell.value
        print "Password : "+Password
        
        
       
        userid =driver.find_element_by_xpath(ADPDisconnect.adpLoginusename())        
        webdriver.ActionChains(driver).move_to_element(userid).click(userid).perform()
        userid.send_keys(ADPUserID)
        password =driver.find_element_by_xpath(ADPDisconnect.adpLoginpassword())        
        webdriver.ActionChains(driver).move_to_element(password).click(password).perform()
        password.send_keys(Password)
        driver.find_element_by_xpath(ADPDisconnect.adpsigninbutton()).click()
        print "Click on Sign_In button"
        
        wait.until(EC.visibility_of_element_located((By.XPATH,ADPDisconnect.updatesubscription())))
        driver.find_element_by_xpath(ADPDisconnect.updatesubscription()).click()
        print "Click on Update subscription button"
        
        wait.until(EC.visibility_of_element_located((By.XPATH,ADPDisconnect.cancelsubscription())))
        driver.find_element_by_xpath(ADPDisconnect.cancelsubscription()).click()
        print "Click on Cancel subscription button"
        
        wait.until(EC.visibility_of_element_located((By.XPATH,ADPDisconnect.donecancelsubscription())))
        driver.find_element_by_xpath(ADPDisconnect.donecancelsubscription()).click()
        print "Click on Done from Cancel subscription popup"
        print"Your subscription to Grovo Microlearning Solution was successfully canceled."
        
    
if __name__ == '__main__':
    
    
    ob= BaseTestClass()
    ob.UserLogin()
    obj1= ADPDisconnect()
    obj1.adpdisconnecting()
    
  
   
    print "Test executed successfully"
    
    