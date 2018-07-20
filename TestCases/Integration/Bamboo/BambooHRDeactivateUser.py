'''
Created on Jul 17, 2018

@author: Shavinlal E
'''

from Tkinter import Tk
from __builtin__ import Exception
from os.path import os
import time
import traceback

from openpyxl.reader.excel import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import xlrd

from BambooHR_Elements import BambooHR_Elements
from BaseTestClass import BaseTestClass
from BaseTestClass import driver


class BambooHRDeactivaete:
    
    def deactivateuser(self,user_firstname,subdomain,sandbox_username,sandbox_password,suggested_attribute_name,selected_attribute_count,expected_connected_status):
        
        bamboohr= BambooHR_Elements()
        
        wait=WebDriverWait(driver, 60)
        wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.admin_tab())))
        
        # Click on Admin from side menu
        driver.find_element_by_xpath(bamboohr.admin_tab()).click()
        print "Clicking on Admin from side menu"
        
        # wait for the users to be loaded
        wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.userslabel_in_userlist())))
        
        print "Enter the First name of synced user in search users field"
        
        print "The First name of synced user is "+" "+user_firstname
        
        webdriver.ActionChains(driver).move_to_element(driver.find_element_by_xpath(bamboohr.search_users_field())).click( driver.find_element_by_xpath(bamboohr.search_users_field())).perform()
        
        driver.find_element_by_xpath(bamboohr.search_users_field()).send_keys(user_firstname)
        
        time.sleep(10)
        
        wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.wait_searcheduser())))
        
        #Going to Click on Deactivate link for the user 
        driver.find_element_by_xpath(bamboohr.searchuser_deactivatelink()).click()
        print "Clicking on Deactivate link for the searched user"
        
        #wait for the deactivate pop up
        
        print "waiting for the deactivate pop up screen"
        wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.deactivatebutton_in_popup())))
        
        # Click on the Deactivate user button
        driver.find_element_by_xpath(bamboohr.deactivatebutton_in_popup()).click()
        print "Clicking on Deactivate user button from the deactivate pop up"
        
        
        #Clicking on Clear search link
        driver.find_element_by_xpath(bamboohr.clearsearch_link()).click()
        print "Clicking on clear search link"
        
        # wait for the user page to load again
        wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.userslabel_in_userlist())))
        
        # search for the deleted user
        
        print "Enter the First name of deactivated user in search users field"
        
        print "The First name of deactivated user is "+" "+user_firstname
        
        webdriver.ActionChains(driver).move_to_element(driver.find_element_by_xpath(bamboohr.search_users_field())).click( driver.find_element_by_xpath(bamboohr.search_users_field())).perform()
        
        driver.find_element_by_xpath(bamboohr.search_users_field()).send_keys(user_firstname)
        
        time.sleep(10)
        
        if(wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.sorry_noresultfound())))):
           
            print "successfully deactivated the user"
            
        else:
            print "Failed to deactivate the user"
            raise Exception
        
        #Going to click on Integrations
        driver.find_element_by_xpath(bamboohr.integrations_subtab()).click()
        print "Clicking on Integrations"
        
        # wait for Configure button
        wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.configure_button())))
        
        # Clicking on Configure button
        driver.find_element_by_xpath(bamboohr.configure_button()).click()
        print "Clicking on Configure button"
        
        #Enter the valid sub domain 
        wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.subdomain_field())))
        driver.find_element_by_xpath(bamboohr.subdomain_field()).send_keys(subdomain)
        print "Entering the sub domain in BambooHR authentication page"

        #Clicking on Connect button
        wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.connect_button())))
        
        if (driver.find_element_by_xpath(bamboohr.connect_button()).is_displayed()):
            driver.find_element_by_xpath(bamboohr.connect_button()).click()
            print "Clicking on Connect button"
            
        else:
            print "Failed to find the Connect button"
            raise Exception
        
        time.sleep(10)
        #wait for the login

        wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.sandbox_username())))
        driver.find_element_by_xpath(bamboohr.sandbox_username()).send_keys(sandbox_username)
        print "Enter sandbox email for login"
        
        
        driver.find_element_by_xpath(bamboohr.sandbox_password()).send_keys(sandbox_password)
        print "Enter sandbox password for login"
        
        
        driver.find_element_by_xpath(bamboohr.sandbox_login()).click()
        print "Clicking on Login button"
        
        try:
            wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.api_field())))
            wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.connect_button())))
            connectbutton=wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.connect_button())))
            driver.execute_script("arguments[0].click();",connectbutton)
            #driver.find_element_by_xpath(bamboohr.connect_button()).click()
            time.sleep(4)
            
        except Exception:
            wait=WebDriverWait(driver, 20)
            wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.subdomain_field())))
            driver.find_element_by_xpath(bamboohr.subdomain_field()).send_keys(subdomain)
            time.sleep(2)
            #click on Connect BUtton
            connectbutton=wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.connect_button())))
            driver.execute_script("arguments[0].click();",connectbutton)
            time.sleep(4)
            #driver.find_element_by_xpath(bamboohr.connect_button()).click()
            
            #Waiting for the API key and Click on Connect
            wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.api_field())))
            connectbutton=wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.connect_button())))
            driver.execute_script("arguments[0].click();",connectbutton)
            time.sleep(4) 
        
        # wait for the select screen
        wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.select_scree_summary())))
        
        select_screen_element = driver.find_element_by_xpath(bamboohr.select_scree_summary())
        
        if(select_screen_element.is_displayed()):
            
            print 'Authorization is successful with valid sub domain and valid api key and select screen is displayed'
        else:
            print "Failed to authenticate with valid sub domain and valid api key"
            raise Exception
        
 
        
        # select check box for user field name
        wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.userfield_checkbox())))
        driver.find_element_by_xpath(bamboohr.userfield_checkbox()).click()
 
        
        time.sleep(2)
        driver.find_element_by_xpath(bamboohr.userfield_checkbox()).click()
        print "Select only required user fields by double selecting the check box next to USER FIELD NAME"
        
        time.sleep(2)
        
        # Going to enter a suggested attribute value in search attribute field
        print "Going to enter a suggested attribute value in search attribute field"
        

        
        print "Suggested attribute name is"+" "+suggested_attribute_name
        
        
        driver.find_element_by_xpath(bamboohr.search_attributes_field()).send_keys(suggested_attribute_name)
        
        print "Entered the suggested attribute name is the search attribute field"
        
        wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.search_attributes_field_checkbox())))
        driver.find_element_by_xpath(bamboohr.search_attributes_field_checkbox()).click()
        
        print "Clicking on the check box for the suggested attribute"
        time.sleep(6)
        
        # Going to validate the user attribute value displayed in the summary text
        
        print "Going to get the count of user attributes selected from summary text"
        wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.select_scree_summary())))
        summary_text= driver.find_element_by_xpath(bamboohr.select_scree_summary()).text
        
        
        splitted_summary= summary_text.split(" ")
        splitted_summary_count = splitted_summary[0]
        
        

        print "str(splitted_summary_count" +str(splitted_summary_count)
        print "str(selected_attribute_count" +str(selected_attribute_count) 
                                                     
                                               
        if int(splitted_summary_count) == int(selected_attribute_count):
            
            print "The count displayed in summary is matching with the selected attribute count"
            
        else:
            print "The count displayed in select summary is not matching with the actual selection of attributes"
            raise Exception
            
        time.sleep(2)
        # Clicking on Next Button
        
        if (driver.find_element_by_xpath(bamboohr.nextbutton_selectscreen()).is_displayed()):
            
            driver.find_element_by_xpath(bamboohr.nextbutton_selectscreen()).click()
        else:
            
            print 'Failed to find the Sync button'
            raise Exception
        
        
        
        print "Clicking on Next button from select screen"
        
        #wait for Sync button
        wait.until(EC.element_to_be_clickable((By.XPATH, bamboohr.syncbutton_confirmscreen())))
        
        # Going to validate the count of selected attributes 
        
        print "Going to validate the count of attributes displayed in the confirm pop up screen"
        confirm_screen_text= driver.find_element_by_xpath(bamboohr.confirm_attributes_count()).text
        
        confirm_screen_text_split = confirm_screen_text.split(" ")
        attribute_count_in_confirm_page = confirm_screen_text_split[5]
        
        if (int(attribute_count_in_confirm_page) == int (selected_attribute_count)):
            
            print "The attribute count displayed in the confirm page is matching with count of selected attributes"
            
        else:
            
            print "The count of attributes in conform page is not matching with the attribute count in select screen"
            raise Exception
        
        
        
        
        #Going to check if the edit link is displayed and is clickable
        
        print "Going to check if the edit button is displaying in the confirm screen"
        
        edit_button_select= driver.find_element_by_xpath(bamboohr.edit_button_in_confirm_screen())
        
        if (edit_button_select.is_displayed()):
            if (edit_button_select.is_enabled()): 
                print "Edit button is displaying in confirm screen and its clickable"
            
        else:
            
            print "Unable to find the click button in confirm screen"
            raise Exception

        time.sleep(2)
        # Clicking on Sync button from Confirm screen
        if (driver.find_element_by_xpath(bamboohr.syncbutton_confirmscreen()).is_displayed()):
            
            driver.find_element_by_xpath(bamboohr.syncbutton_confirmscreen()).click()
            print "Clicking on Sync button from select screen"
            
        else:
            
            print 'Failed to find the Sync button in Confirm pop up screen'
            raise Exception
        
        time.sleep(2)
        
        # Clicking on Check now link
        
        while True:
            try:
                
                wait.until(EC.visibility_of_element_located((By.XPATH, bamboohr.connected_status())))
                # time.sleep(2)
                webdriver.ActionChains(driver).move_to_element(driver.find_element_by_xpath(bamboohr.checknow_link())).click(driver.find_element_by_xpath(bamboohr.checknow_link())).perform()
                
            except:
                print "Failed to find check now link"
                break                  
        #driver.find_element_by_xpath(bamboohr.checknow_link()).click()
        print "Clicking on Check now link"
        
        wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.connected_status())))
                   
        actual_connected_status = driver.find_element_by_xpath(bamboohr.connected_status()).text
        
        

        print actual_connected_status
        print expected_connected_status
        
        if (actual_connected_status == expected_connected_status):
        
            print "The HRIS Sync is successful" 
        else:
            
            print "Failed to connect to HRIS System"
            raise Exception
        
        time.sleep(5)
        
        # Clicking on Users link in side menu tab
        
        print "Going to view the users managed by BambooHRIS"
        wait.until(EC.visibility_of_element_located((By.XPATH,"(//a[@href='/admin/users'])[1]")))
        driver.find_element_by_xpath("(//a[@href='/admin/users'])[1]").click()
        driver.find_element_by_xpath(bamboohr.users_tab()).click()
        print "Clicking on Users link"
        
        # wait for the search user field
        wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.search_users_field())))
        
        print "Enter the First name of synced user in search users field"
        print "The First name of synced user is "+" "+user_firstname
        time.sleep(15)
        webdriver.ActionChains(driver).move_to_element(driver.find_element_by_xpath(bamboohr.search_users_field())).click( driver.find_element_by_xpath(bamboohr.search_users_field())).perform()
        
        driver.find_element_by_xpath(bamboohr.search_users_field()).send_keys(user_firstname)
        time.sleep(15)
        
        if(wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr. wait_searcheduser())))):
           
            print "The deactivated user got reactivated after the re sync"
            
        else:
            print "Failed to reactivate the deactivated user"
            raise Exception
       



    def deactivateuserMain(self):    
        
        try:
            book=xlrd.open_workbook(os.path.join('E:\NewWorkspace\FirstProjectInPython\TestData.xlsx'))
            first_sheet = book.sheet_by_name('BambooHR')
            cell3= first_sheet.cell(1,8)
            user_firstname = cell3.value 
            
            cell1= first_sheet.cell(1,0)
            subdomain = cell1.value
            
            cell5= first_sheet.cell(1,4)
            sandbox_username = cell5.value 
            
            cell6= first_sheet.cell(1,5)
            sandbox_password = cell6.value
            
            cell10= first_sheet.cell(4,3)
            suggested_attribute_name = cell10.value
            
            cell9= first_sheet.cell(4,2)
            selected_attribute_count = cell9.value
            
            cell11= first_sheet.cell(4,4)
            expected_connected_status = cell11.value 
            
            obj1 = BambooHRDeactivaete()
            obj1.deactivateuser(user_firstname,subdomain,sandbox_username,sandbox_password,suggested_attribute_name,selected_attribute_count,expected_connected_status)
    
            print "TEST CASE EXECUTION SUCCESSFULLY COMPLETED"
        except Exception as e:
            traceback.print_exc()
            print (e)
            raise Exception
            
        finally:
            #print "clicking on Home"
            book=xlrd.open_workbook(os.path.join('E:\NewWorkspace\FirstProjectInPython\TestData.xlsx'))
            first_sheet = book.sheet_by_name('Login_Credentials')
            #print("Fetching the Attribute Name from Excel Sheet\n")
            # read a cell
            cell = first_sheet.cell(1,1)
            HomeURL = cell.value
            #print HomeURL
            driver.get(HomeURL)
            #print "Home Page Loaded"


if __name__ == '__main__':
    
    obj11= BambooHRDeactivaete()
    obj12= BaseTestClass()
    obj12.UserLogin()
    obj11.deactivateuserMain()

    

    print "Test executed successfully"
           
