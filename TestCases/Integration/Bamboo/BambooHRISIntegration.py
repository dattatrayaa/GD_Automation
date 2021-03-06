'''
Created on Jul 2, 2018

@author: Shavinlal E
'''
from Tkinter import Tk
from os.path import os
import time
import traceback
from openpyxl.reader.excel import load_workbook
from selenium import webdriver

from selenium.webdriver.common.by import By

from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import xlrd
import traceback
from BambooHR_Elements import BambooHR_Elements
from BaseTestClass import BaseTestClass
from BaseTestClass import driver
from __builtin__ import Exception
from BaseTestClass import projectPath


class BambooHRISIntegration:
    

    
    def settingup_bamboohr_integration(self):
        
        bamboohr= BambooHR_Elements()
        book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        first_sheet = book.sheet_by_name('BambooHR')
        
        cell4= first_sheet.cell(1,3)
        bamboohr_sandbox_url = cell4.value 
        
        cell5= first_sheet.cell(1,4)
        sandbox_username = cell5.value 
        
        cell6= first_sheet.cell(1,5)
        sandbox_password = cell6.value
        
        cell1= first_sheet.cell(1,0)
        subdomain = cell1.value
        
        cell1= first_sheet.cell(6,1)
        bambboURL = cell1.value

        book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        second_sheet = book.sheet_by_name('Login_Credentials')
        cell = second_sheet.cell(1,1)
        url = cell.value
        
        wait=WebDriverWait(driver, 60)
        wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.admin_tab())))
        
        book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        first_sheet = book.sheet_by_name('BambooHR')
        
        # Click on Admin from side menu
        driver.find_element_by_xpath(bamboohr.admin_tab()).click()
        print "Clicking on Admin from side menu"
        
        # Waiting for Integrations link
        wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.integrations_subtab())))
        
        review = driver.find_element_by_xpath(bamboohr.integrations_subtab())
        webdriver.ActionChains(driver).move_to_element(review).click(review).perform()
        
        # Click on Integrations
        wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.integrations_subtab())))
        driver.find_element_by_xpath(bamboohr.integrations_subtab()).click()
        print "Clicking on Integrations"
        
        
        #wait for Configure button
        wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.configure_button())))
        #Going to validate the UI elements in the BambooHR configuration screen
                   
        #print "Going to verify the display of View instructions link in configure screen"
        
        cell7= first_sheet.cell(4,0)
        expected_viewinstructrion_link_name= cell7.value 
        
        actual_viewinstructrion_link_name = driver.find_element_by_xpath(bamboohr.view_instructions_link()).text
        
        if (expected_viewinstructrion_link_name == actual_viewinstructrion_link_name):
            
            print "The view instruction link is displaying in BambooHR configure screen"
            
        else:
            print "Failed to find the view instructions link in Bamboohr configuration screen"
            raise Exception 

        
        # Clicking on Configure button
        time.sleep(4)
        wait.until(EC.element_to_be_clickable((By.XPATH,bamboohr.configure_button())))
        ele=driver.find_element_by_xpath(bamboohr.configure_button())
        driver.execute_script("arguments[0].click();", ele)
        print "Clicking on Configure button"

        # wait for sand box fiedl and enter invalid subdomain
        wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.subdomain_field())))
        driver.find_element_by_xpath(bamboohr.subdomain_field()).send_keys("invalid_subdomain")
        print "Entering invalid sub domain in BambooHR authentication page"
        wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.connect_button())))
        driver.find_element_by_xpath(bamboohr.connect_button()).click()
        time.sleep(4)
        currentURL = driver.current_url
        if bambboURL == currentURL :
            print "Bamboo URL is loaded"
        else:
            print "Bamboo URL is not loaded"
            raise Exception
        driver.get(url+"admin/integrations")
        time.sleep(4)
        
        # Clicking on Configure button
        time.sleep(4)
        wait.until(EC.element_to_be_clickable((By.XPATH,bamboohr.configure_button())))
        ele=driver.find_element_by_xpath(bamboohr.configure_button())
        driver.execute_script("arguments[0].click();", ele)
        print "Clicking on Configure button"
        
        # Enter the valid sub domain 
        wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.subdomain_field())))
        driver.find_element_by_xpath(bamboohr.subdomain_field()).clear()
        driver.find_element_by_xpath(bamboohr.subdomain_field()).send_keys(subdomain)
        print "Entering the sub domain in BambooHR authentication page"

        #Clicking on Connect button
        wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.connect_button())))
        
        if (driver.find_element_by_xpath(bamboohr.connect_button()).is_displayed()):
            driver.find_element_by_xpath(bamboohr.connect_button()).click()
            print "Clicking on Connect button"
            
        else:
            print "Failed to find the Connect button"
            raise Exception
        
        time.sleep(10)
        #wait for the login
        time.sleep(6)
        cell5= first_sheet.cell(1,4)
        sandbox_username = cell5.value 
        
        cell6= first_sheet.cell(1,5)
        sandbox_password = cell6.value
        wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.sandbox_username())))
        driver.find_element_by_xpath(bamboohr.sandbox_username()).send_keys(sandbox_username)
        print "Enter sandbox email for login"
        
        
        driver.find_element_by_xpath(bamboohr.sandbox_password()).send_keys(sandbox_password)
        print "Enter sandbox password for login"
        
        
        driver.find_element_by_xpath(bamboohr.sandbox_login()).click()
        print "Clicking on Login button"
        

        # Getting the proper sub domain name from excel
        
        cell1= first_sheet.cell(1,0)
        subdomain = cell1.value 
        
        cell2= first_sheet.cell(1,1)
        api_key = cell2.value 
        
        print "The bamboo hr sub-domain is"+" "+subdomain
        print "The bamboo hr API key is"+" "+api_key
        
        
        # Enter the valid sub domain
        #Clicking on Connect button
        if (wait.until(EC.visibility_of_element_located((By.XPATH,'//*[@id="api-key"]')))):
            
            wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.subdomain_field())))
            driver.find_element_by_xpath(bamboohr.subdomain_field()).click()
            
            driver.find_element_by_xpath(bamboohr.subdomain_field()).clear()
            driver.find_element_by_xpath(bamboohr.subdomain_field()).send_keys(subdomain)
            time.sleep(4)
            print "Entering the sub domain in BambooHR authentication page"
            wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.connect_button())))
            connectbutton=wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.connect_button())))
            driver.execute_script("arguments[0].click();",connectbutton)
            time.sleep(4)
        else:
            wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.subdomain_field())))
            #driver.find_element_by_xpath(bamboohr.subdomain_field()).click()
        
            #driver.find_element_by_xpath(bamboohr.subdomain_field()).clear()
            driver.find_element_by_xpath(bamboohr.subdomain_field()).send_keys(subdomain)
            #click on Connect BUtton
            wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.connect_button())))
            connectbutton=wait.until(EC.element_to_be_clickable((By.XPATH,bamboohr.connect_button())))
            driver.execute_script("arguments[0].click();",connectbutton)
            time.sleep(4)
            #Waiting for the API key and Click on Connect
            wait.until(EC.visibility_of_element_located((By.XPATH,'//*[@id="api-key"]')))
            connectbutton=wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.connect_button())))
            driver.execute_script("arguments[0].click();",connectbutton)
            time.sleep(4) 
        
        # wait for the select screen
        wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.select_scree_summary())))
        
        select_screen_element = driver.find_element_by_xpath(bamboohr.select_scree_summary())
        
        if(select_screen_element.is_displayed()):
            
            print 'Authorization is successful with valid sub domain and valid api key and select screen is displayed'
        else:
            print "Failed to authenticate with valid sub domain and valid api key"
            raise Exception
        
 
        
                
        
        # select check box for user field name
        wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.userfield_checkbox())))
        driver.find_element_by_xpath(bamboohr.userfield_checkbox()).click()
 
        
        time.sleep(2)
        driver.find_element_by_xpath(bamboohr.userfield_checkbox()).click()
        print "Select only required user fields by double selecting the check box next to USER FIELD NAME"
        
        time.sleep(2)
        
        # Going to enter a suggested attribute value in search attribute field
        print "Going to enter a suggested attribute value in search attribute field"
        
        
        cell10= first_sheet.cell(4,3)
        suggested_attribute_name = cell10.value
        
        print "Suggested attribute name is"+" "+suggested_attribute_name
        
        
        driver.find_element_by_xpath(bamboohr.search_attributes_field()).send_keys(suggested_attribute_name)
        
        print "Entered the suggested attribute name is the search attribute field"
        
        wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.search_attributes_field_checkbox())))
        driver.find_element_by_xpath(bamboohr.search_attributes_field_checkbox()).click()
        
        print "Clicking on the check box for the suggested attribute"
        time.sleep(6)
        
        # Going to validate the user attribute value displayed in the summary text
        
        print "Going to get the count of user attributes selected from summary text"
        wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.select_scree_summary())))
        summary_text= driver.find_element_by_xpath(bamboohr.select_scree_summary()).text
        
        
        splitted_summary= summary_text.split(" ")
        splitted_summary_count = splitted_summary[0]
        
        
        cell9= first_sheet.cell(4,2)
        selected_attribute_count = cell9.value
        print "str(splitted_summary_count" +str(splitted_summary_count)
        print "str(selected_attribute_count" +str(selected_attribute_count) 
                                                     
                                               
        if int(splitted_summary_count) == int(selected_attribute_count):
            
            print "The count displayed in summary is matching with the selected attribute count"
            
        else:
            print "The count displayed in select summary is not matching with the actual selection of attributes"
            raise Exception
            
        time.sleep(2)
        # Clicking on Next Button
        
        if (driver.find_element_by_xpath(bamboohr.nextbutton_selectscreen()).is_displayed()):
            
            driver.find_element_by_xpath(bamboohr.nextbutton_selectscreen()).click()
        else:
            
            print 'Failed to find the Sync button'
            raise Exception
        
        
        
        print "Clicking on Next button from select screen"
        
        #wait for Sync button
        wait.until(EC.element_to_be_clickable((By.XPATH, bamboohr.syncbutton_confirmscreen())))
        
        # Going to validate the count of selected attributes 
        
        print "Going to validate the count of attributes displayed in the confirm pop up screen"
        confirm_screen_text= driver.find_element_by_xpath(bamboohr.confirm_attributes_count()).text
        
        confirm_screen_text_split = confirm_screen_text.split(" ")
        attribute_count_in_confirm_page = confirm_screen_text_split[5]
        
        if (int(attribute_count_in_confirm_page) == int (selected_attribute_count)):
            
            print "The attribute count displayed in the confirm page is matching with count of selected attributes"
            
        else:
            
            print "The count of attributes in conform page is not matching with the attribute count in select screen"
            raise Exception
        
        
        
        
        #Going to check if the edit link is displayed and is clickable
        
        print "Going to check if the edit button is displaying in the confirm screen"
        
        edit_button_select= driver.find_element_by_xpath(bamboohr.edit_button_in_confirm_screen())
        
        if (edit_button_select.is_displayed()):
            if (edit_button_select.is_enabled()): 
                print "Edit button is displaying in confirm screen and its clickable"
            
        else:
            
            print "Unable to find the click button in confirm screen"
            raise Exception

        time.sleep(2)
        # Clicking on Sync button from Confirm screen
        if (driver.find_element_by_xpath(bamboohr.syncbutton_confirmscreen()).is_displayed()):
            
            driver.find_element_by_xpath(bamboohr.syncbutton_confirmscreen()).click()
            print "Clicking on Sync button from select screen"
            
        else:
            
            print 'Failed to find the Sync button in Confirm pop up screen'
            raise Exception
        
        time.sleep(2)
        
        # Clicking on Check now link
        
        while True:
            try:
                
                wait.until(EC.visibility_of_element_located((By.XPATH, bamboohr.connected_status())))
                # time.sleep(2)
                webdriver.ActionChains(driver).move_to_element(driver.find_element_by_xpath(bamboohr.checknow_link())).click(driver.find_element_by_xpath(bamboohr.checknow_link())).perform()
                
            except:
                print "Failed to find check now link"
                break                  
        #driver.find_element_by_xpath(bamboohr.checknow_link()).click()
        print "Clicking on Check now link"
        
        wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.connected_status())))
                   
        actual_connected_status = driver.find_element_by_xpath(bamboohr.connected_status()).text
        
        
        cell11= first_sheet.cell(4,4)
        expected_connected_status = cell11.value 
        print actual_connected_status
        print expected_connected_status
        
        if (actual_connected_status == expected_connected_status):
        
            print "The HRIS Sync is successful" 
        else:
            
            print "Failed to connect to HRIS System"
            raise Exception
        
        time.sleep(5)
        
        # Clicking on Users link in side menu tab
        
        print "Going to view the users managed by BambooHRIS"
        wait.until(EC.visibility_of_element_located((By.XPATH,"(//a[@href='/admin/users'])[1]")))
        driver.find_element_by_xpath("(//a[@href='/admin/users'])[1]").click()
        driver.find_element_by_xpath(bamboohr.users_tab()).click()
        print "Clicking on Users link"
        
        # wait for the search user field
        wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.search_users_field())))
        
        print "Enter the First name of synced user in search users field"
        
        cell3= first_sheet.cell(1,8)
        user_firstname = cell3.value 
        
        print "The First name of synced user is "+" "+user_firstname
        time.sleep(15)
        webdriver.ActionChains(driver).move_to_element(driver.find_element_by_xpath(bamboohr.search_users_field())).click( driver.find_element_by_xpath(bamboohr.search_users_field())).perform()
        
        driver.find_element_by_xpath(bamboohr.search_users_field()).send_keys(user_firstname)
        
        # Going to Click and open the created users
        #time.sleep(15)
        
        wait.until(EC.visibility_of_element_located((By.XPATH,"//table/tbody/tr[1]/td[2]/a[.='"+user_firstname+"']")))
        driver.find_element_by_xpath("//table/tbody/tr[1]/td[2]/a[.='"+user_firstname+"']").click()
        print "Clicking on the synced user"
        
        # Going to get the First name
        
        wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='create-edit-user-search-firstName']")))
        actual_firstname= driver.find_element_by_xpath(".//*[@id='create-edit-user-search-firstName']").get_attribute('value')
        print "The displayed first name is "+actual_firstname
        
        actual_lastname= driver.find_element_by_xpath(".//*[@id='create-edit-user-search-lastName']").get_attribute('value')
        print "The displayed last name is "+actual_lastname
        
        
        actual_emailidd= driver.find_element_by_xpath(".//*[@id='create-edit-user-search-username']").get_attribute('value')
        print "The displayed email id is "+actual_emailidd
         
                
        actual_employeeid= driver.find_element_by_xpath(".//*[@id='create-edit-user-search-employeeId']").get_attribute('value')
        print "The displayed employee id is "+actual_employeeid
        
        
        cell3= first_sheet.cell(1,8)
        user_firstname = cell3.value 
        
        cell4= first_sheet.cell(1,9)
        expected_lastname = cell4.value 
        
        cell5= first_sheet.cell(1,10)
        expected_emailid = cell5.value 
        
        cell6= first_sheet.cell(1,7)
        expected_employeeid = cell6.value 
        
        #print expected_lastname
        # print expected_emailid
        # print expected_employeeid
        
        if(actual_firstname == user_firstname):
            
            if(actual_lastname == expected_lastname):
                
                if(actual_emailidd == expected_emailid):
                    
                    if(actual_employeeid == expected_employeeid):
                    
                        print "The synced user details are matching as in BambooHR sandbox"
           
        else:
            print "Failed to validate the synced user details"
            raise Exception
        
        # Going to Click on User Attributes tab from side menu
        
        time.sleep(2)
        
        driver.find_element_by_xpath(bamboohr.user_attributes_tab()).click()
        print "Clicking on User attributes tab"
        
        # Going to Click on Grovo attributes tab
        driver.find_element_by_xpath(bamboohr.grovo_attributres_subtab()).click()
        print "Clicking on Grovo Attributes tab"
        
        # Going to fetch the HSRIS Integrations text for the synced attributes
        
        #get the text for First Name.Last name, Primary email, Employee id
        
        wait.until(EC.visibility_of_element_located((By.XPATH,".//table/tbody/tr/td[1]/a[.='First Name']/../../td[2]")))
        
        First_name_source= driver.find_element_by_xpath(".//table/tbody/tr/td[1]/a[.='First Name']/../../td[2]").text
        Last_name_source= driver.find_element_by_xpath(".//table/tbody/tr/td[1]/a[.='Last Name']/../../td[2]").text
        Employeeid_source= driver.find_element_by_xpath(".//table/tbody/tr/td[1]/a[.='Employee ID']/../../td[2]").text
        Primary_email= driver.find_element_by_xpath(".//table/tbody/tr/td[1]/a[.='Primary Email']/../../td[2]").text
        City= driver.find_element_by_xpath(".//table/tbody/tr/td[1]/a[.='City']/../../td[2]").text
        
        
        
        if (First_name_source == "HRIS Integration"):
            
            if(Last_name_source == "HRIS Integration"):
                
                if(Employeeid_source == "HRIS Integration"):
                    
                    if(Primary_email == "HRIS Integration"):
                        
                        if(City == "HRIS Integration"):
                        
                            print "The source for synced attributes are displaying as HRIS Integration"
                        
        else:
                         
            print "Failed to validate the attribute source details"
            raise Exception
        
        # Click on Integrations
        wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.integrations_subtab())))
        driver.find_element_by_xpath(bamboohr.integrations_subtab()).click()
        print "Clicking on Integrations"
        
        wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.connected_status())))
        #Click on edit settings link
        driver.find_element_by_xpath(bamboohr.edit_settings_link()).click()
        print "Clicking on edit settings link"  
        
        # Going to click on Disconnect link
        wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.disconnect_link())))
        driver.find_element_by_xpath(bamboohr.disconnect_link()).click()
        print "Clicking on disconnect link"
        
        # Going to click on Disconnect button from the displayed pop up
        wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.disconnect_button())))
        driver.find_element_by_xpath(bamboohr.disconnect_button()).click()
        print "Clicking on disconnect button"
        
        #wait for Configure link
        wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.configure_button())))
        if (bamboohr.configure_button().is_displayed()):
            
            print "Successfully disconnected BambooHR"
            
        else:
            print "Failed to disconnect BambooHR"
            raise Exception
            
        
        
    def createuser_in_bamboohr(self,Employee_Number,Employee_FirstName,Employee_LastName,Work_Email):
        
         
         
        bamboohr= BambooHR_Elements()
        wait=WebDriverWait(driver, 60)
        print "Going to create a user in bamboohr sand box account"
        
        
        book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        first_sheet = book.sheet_by_name('BambooHR')
        
        cell4= first_sheet.cell(1,3)
        bamboohr_sandbox_url = cell4.value 
        
        cell5= first_sheet.cell(1,4)
        sandbox_username = cell5.value 
        
        cell6= first_sheet.cell(1,5)
        sandbox_password = cell6.value
        
        
        
        print "Going to access BambooHR sand box account url"+" "+bamboohr_sandbox_url
        driver.get(bamboohr_sandbox_url)
        driver.maximize_window()
        
        
        driver.find_element_by_xpath(bamboohr.sandbox_username()).send_keys(sandbox_username)
        print "Enter sandbox email for login"
        
        
        driver.find_element_by_xpath(bamboohr.sandbox_password()).send_keys(sandbox_password)
        print "Enter sandbox password for login"
        
        
        driver.find_element_by_xpath(bamboohr.sandbox_login()).click()
        print "Clicking on Login button"
        
        # Waiting untill the Dash board is displaying
        
        wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.sandbox_dashboard())))
        
 
        # Going to create a new Employee with Required attrubute values
        print "Going to create a new employee with required attribute values"
        
        
        #Clicking on Employees link
        wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.employees_link())))
        driver.find_element_by_xpath(bamboohr.employees_link()).click()
        print "Clicking on Employees link"
        
        
        #wait for add_employee_button and ten click on add employee button
        
        wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.add_employee_button())))
        driver.find_element_by_xpath(bamboohr.add_employee_button()).click()
        print"Clicking on Add Employee Button"
        time.sleep(2)
        
        
        # Enter employee ID
        

        wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.employee_number_field())))
        driver.find_element_by_xpath(bamboohr.employee_number_field()).send_keys(Employee_Number)
        
        print "Entering employee number as "+ Employee_Number
        
        
        # Enter First name
       
        wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.employee_firstname_field())))
        driver.find_element_by_xpath(bamboohr.employee_firstname_field()).send_keys(Employee_FirstName)
        
        print "Entering employee first name as "+Employee_FirstName
        
        
        # Enter Last name
       
        wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.employee_lastname_field())))
        driver.find_element_by_xpath(bamboohr.employee_lastname_field()).send_keys(Employee_LastName)
       
        print "Entering employee last name as "+Employee_LastName
        
        
                
        # Enter Work email
        element= wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.employee_workemail_field())))
        element.send_keys(Work_Email)
        print "Entering employee work email as "+ Work_Email
        
        # Clicking on SAVE button
        driver.find_element_by_xpath(bamboohr.employee_form_save()).click()
        time.sleep(3)
        
    
        """wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.employee_personal_tab())))
        driver.find_element_by_xpath(bamboohr.employee_personal_tab()).click()"""
        
        
        
        
    
    
    def updating_the_employee_values_and_startmain(self):
        
        
        book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        first_sheet = book.sheet_by_name('BambooHR')
        
        
        #Learner 1
        cell = first_sheet.cell(1,8)
        Employee_FirstName = cell.value
        
        FirstNameId = Employee_FirstName.split("_")
        emp = FirstNameId[0]+"_"
        ids = FirstNameId[1]
        empId = int(ids)+1
        FirstNameUpdated= emp+str(empId)
        
        cell = first_sheet.cell(1,9)
        Employee_LastName = cell.value
        
        LastNameID = Employee_LastName.split("_")
        emp = LastNameID[0]+"_"
        ids = LastNameID[1]
        empId = int(ids)+1
        LastNameUpdated= emp+str(empId)
        
        
        
        cell = first_sheet.cell(1,10)
        Work_Email = cell.value
        
        EmailId = Work_Email.split("_")
        emp = EmailId[0]+"_"
        ids = EmailId[1]
        remaining="_"+EmailId[2]
        empId = int(ids)+1
        EmailIdUpdated= emp+str(empId)+remaining
        
        cell = first_sheet.cell(1,7)
        Employee_Number = cell.value
        
        Employee = Employee_Number.split("_")
        emp = Employee[0]+"_"
        ids = Employee[1]
        empId = int(ids)+1
        EmployeeIdUpdated= emp+str(empId)
   
        
        
        #For Original User
        book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        s_sheet = book.sheet_by_name('BambooHR')
        
        
        #updating user values
        wb = load_workbook(os.path.join('Test_Data/TestData.xlsx'))
        #print (wb.sheetnames)
        
        sheet = wb['BambooHR']
        
        sheet.cell(row=2, column=9).value = FirstNameUpdated
        sheet.cell(row=2, column=10).value = LastNameUpdated
        sheet.cell(row=2, column=11).value =EmailIdUpdated
        sheet.cell(row=2, column=8).value = EmployeeIdUpdated
        
        
        
        wb.save(os.path.join('Test_Data/TestData.xlsx'))
            
        print "All User Data Updated in Excel"
        
        try:
            book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
            second_sheet = book.sheet_by_name('Login_Credentials')
            cell = second_sheet.cell(1,1)
            url = cell.value
            obj13= BambooHRISIntegration()
            #obj13.createuser_in_bamboohr(Employee_Number, Employee_FirstName, Employee_LastName, Work_Email)
            obj13.createuser_in_bamboohr(EmployeeIdUpdated, FirstNameUpdated, LastNameUpdated, EmailIdUpdated)
            driver.get(url)
            obj13.settingup_bamboohr_integration()
        except Exception as e:
            traceback.print_exc()
            print (e)
            raise Exception 
        finally:
            book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
            second_sheet = book.sheet_by_name('Login_Credentials')
            cell = second_sheet.cell(1,1)
            url = cell.value
            driver.get(url)
           
    
     
if __name__ == '__main__':
    
    obj11= BambooHR()
    obj12= BaseTestClass()
        
    #obj11.updating_the_employee_values_and_startmain()
    obj12.UserLogin()
    obj11.settingup_bamboohr_integration()
    

    print "Test executed successfully"
           
