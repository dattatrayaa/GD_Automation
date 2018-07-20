'''
Created on Jul 17, 2018

@author: Shavinlal E
'''

from Tkinter import Tk
from __builtin__ import Exception
from os.path import os
import time
import traceback

from openpyxl.reader.excel import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import xlrd

from BambooHR_Elements import BambooHR_Elements
from BaseTestClass import BaseTestClass
from BaseTestClass import driver


class BambooHRUpdateUserDetails:
    
    def bamboohrUpdateuser(self,bamboohr_sandbox_url,sandbox_username,sandbox_password,user_firstname,Employee_Number,Employee_LastName,Work_Email):
        
        bamboohr= BambooHR_Elements()
        
        wait=WebDriverWait(driver, 60)

        
        print "Going to access BambooHR sand box account url"+" "+bamboohr_sandbox_url
        driver.get(bamboohr_sandbox_url)
        driver.maximize_window()
        
        
        driver.find_element_by_xpath(bamboohr.sandbox_username()).send_keys(sandbox_username)
        print "Enter sandbox email for login"
        
        
        driver.find_element_by_xpath(bamboohr.sandbox_password()).send_keys(sandbox_password)
        print "Enter sandbox password for login"
        
        
        driver.find_element_by_xpath(bamboohr.sandbox_login()).click()
        print "Clicking on Login button"
        
        # Waiting untill the Dash board is displaying
        
        wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.sandbox_dashboard())))
        
        #search for the employees
        driver.find_element_by_xpath(bamboohr.searchemployee_field()).send_keys(user_firstname)
        print"Searching for the user already created in the application"
        
        #click on search button
        driver.find_element_by_xpath(bamboohr.search_icon()).click()
        
        #wait for the user grid after the search
        wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.search_result_grid())))
        
        
        driver.find_element_by_xpath(bamboohr.clickon_searcheduser()).click()
        print "Clicking on the searched user and opening the user details page"
        
        #wait for the employee first name field
        wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.employee_number_field())))
        driver.find_element_by_xpath(bamboohr.employee_number_field()).clear()
    
        #entering the updated employee number
        driver.find_element_by_xpath(bamboohr.employee_number_field()).send_keys(Employee_Number+"Updated")
        print "Entering employee number as "+ Employee_Number+"Updated"
        
        
        """ Enter First name
       
        wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.employee_firstname_field())))
        driver.find_element_by_xpath(bamboohr.employee_firstname_field()).clear()
        driver.find_element_by_xpath(bamboohr.employee_firstname_field()).send_keys(Employee_FirstName+"Updated")
        
        print "Entering employee first name as "+Employee_FirstName+"Updated" 
        """
        
        
        # Enter Last name
       
        wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.employee_lastname_field())))
        driver.find_element_by_xpath(bamboohr.employee_lastname_field()).clear()
        driver.find_element_by_xpath(bamboohr.employee_lastname_field()).send_keys(Employee_LastName+"Updated")
       
        print "Entering employee last name as "+Employee_LastName+"Updated"
        
        
                
        # Enter Work email
        driver.find_element_by_xpath(bamboohr.employee_workemail_field()).clear()
        driver.find_element_by_xpath(bamboohr.employee_workemail_field()).send_keys(Work_Email+"Updated")
        print "Entering employee work email as "+ Work_Email+"Updated"
        
        # Clicking on SAVE button
        driver.find_element_by_xpath(bamboohr.employee_form_save()).click()
        
        time.sleep(6)
        
        #click on the account
        wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.sandbox_accounticon())))
        driver.find_element_by_xpath(bamboohr.sandbox_accounticon()).click()
        print "Clicking on account icon in the top right corner of sandbox dashboard header"
        
        time.sleep(2)
        
        wait.until(EC.visibility_of_element_located((By.XPATH,"//*[@id='infoLinks']/li[6]/a")))
        time.sleep(2)
        driver.find_element_by_xpath("//*[@id='infoLinks']/li[6]/a").click()
        
        
        
    def verifyUpdateduser(self,Employee_FirstName,suggested_attribute_name,subdomain,sandbox_username,sandbox_password,Employee_LastName,Work_Email,Employee_Number,expected_connected_status):
        
        bamboohr= BambooHR_Elements()
        
        wait=WebDriverWait(driver, 60)
        
        # Click on Admin from side menu
        driver.find_element_by_xpath(bamboohr.admin_tab()).click()
        print "Clicking on Admin from side menu"
        
        # Waiting for Integrations link
        wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.integrations_subtab())))
        
        review = driver.find_element_by_xpath(bamboohr.integrations_subtab())
        webdriver.ActionChains(driver).move_to_element(review).click(review).perform()
        
        # Click on Integrations
        wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.integrations_subtab())))
        driver.find_element_by_xpath(bamboohr.integrations_subtab()).click()
        print "Clicking on Integrations"
        
        
        # wait for Configure button
        wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.configure_button())))
        
        # Clicking on Configure button
        driver.find_element_by_xpath(bamboohr.configure_button()).click()
        print "Clicking on Configure button"
        
        #Enter the valid sub domain 
        wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.subdomain_field())))
        driver.find_element_by_xpath(bamboohr.subdomain_field()).send_keys(subdomain)
        print "Entering the sub domain in BambooHR authentication page"

        #Clicking on Connect button
        wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.connect_button())))
        
        if (driver.find_element_by_xpath(bamboohr.connect_button()).is_displayed()):
            driver.find_element_by_xpath(bamboohr.connect_button()).click()
            print "Clicking on Connect button"
            
        else:
            print "Failed to find the Connect button"
            raise Exception
        
        #wait for the login

        wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.sandbox_username())))
        driver.find_element_by_xpath(bamboohr.sandbox_username()).send_keys(sandbox_username)
        print "Enter sandbox email for login"
        
        
        driver.find_element_by_xpath(bamboohr.sandbox_password()).send_keys(sandbox_password)
        print "Enter sandbox password for login"
        
        
        driver.find_element_by_xpath(bamboohr.sandbox_login()).click()
        print "Clicking on Login button"
        
        wait=WebDriverWait(driver, 20)
        
        try:
            wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.api_field())))
            wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.connect_button())))
            connectbutton=wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.connect_button())))
            driver.execute_script("arguments[0].click();",connectbutton)
            #driver.find_element_by_xpath(bamboohr.connect_button()).click()
            time.sleep(4)
            
        except Exception:
            wait=WebDriverWait(driver, 20)
            wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.subdomain_field())))
            driver.find_element_by_xpath(bamboohr.subdomain_field()).send_keys(subdomain)
            time.sleep(2)
            #click on Connect BUtton
            connectbutton=wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.connect_button())))
            driver.execute_script("arguments[0].click();",connectbutton)
            time.sleep(4)
            #driver.find_element_by_xpath(bamboohr.connect_button()).click()
            
            #Waiting for the API key and Click on Connect
            wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.api_field())))
            connectbutton=wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.connect_button())))
            driver.execute_script("arguments[0].click();",connectbutton)
            time.sleep(4) 
            
        time.sleep(10)
        
        """
        # wait for the select screen
        wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.nextbutton_selectscreen())))
        
        select_screen_element = driver.find_element_by_xpath(bamboohr.nextbutton_selectscreen())
        
        if(select_screen_element.is_displayed()):
            
            print 'Authorization is successful with valid sub domain and valid api key and select screen is displayed'
        else:
            print "Failed to authenticate with valid sub domain and valid api key"
            raise Exception"""
        
 
        
        # select check box for user field name
        wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.userfield_checkbox())))
        driver.find_element_by_xpath(bamboohr.userfield_checkbox()).click()
 
        
        time.sleep(2)
        driver.find_element_by_xpath(bamboohr.userfield_checkbox()).click()
        print "Select only required user fields by double selecting the check box next to USER FIELD NAME"
        
        time.sleep(2)
        
        # Going to enter a suggested attribute value in search attribute field
        print "Going to enter a suggested attribute value in search attribute field"
        

        
        print "Suggested attribute name is"+" "+suggested_attribute_name
        
        
        driver.find_element_by_xpath(bamboohr.search_attributes_field()).send_keys(suggested_attribute_name)
        
        print "Entered the suggested attribute name is the search attribute field"
        
        wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.search_attributes_field_checkbox())))
        driver.find_element_by_xpath(bamboohr.search_attributes_field_checkbox()).click()
        
        print "Clicking on the check box for the suggested attribute"
        time.sleep(6)
        
        # Clicking on Next Button
        
        if (driver.find_element_by_xpath(bamboohr.nextbutton_selectscreen()).is_displayed()):
            
            driver.find_element_by_xpath(bamboohr.nextbutton_selectscreen()).click()
        else:
            
            print 'Failed to find the Sync button'
            raise Exception
        
        
        
        print "Clicking on Next button from select screen"
        
        #wait for Sync button
        wait.until(EC.element_to_be_clickable((By.XPATH, bamboohr.syncbutton_confirmscreen())))

        time.sleep(2)
        # Clicking on Sync button from Confirm screen
        if (driver.find_element_by_xpath(bamboohr.syncbutton_confirmscreen()).is_displayed()):
            
            driver.find_element_by_xpath(bamboohr.syncbutton_confirmscreen()).click()
            print "Clicking on Sync button from select screen"
            
        else:
            
            print 'Failed to find the Sync button in Confirm pop up screen'
            raise Exception
        
        time.sleep(2)
        
        # Clicking on Check now link
        
        while True:
            try:
                
                wait.until(EC.visibility_of_element_located((By.XPATH, bamboohr.connected_status())))
                # time.sleep(2)
                webdriver.ActionChains(driver).move_to_element(driver.find_element_by_xpath(bamboohr.checknow_link())).click(driver.find_element_by_xpath(bamboohr.checknow_link())).perform()
                
            except:
                print "Failed to find check now link"
                break                  
        #driver.find_element_by_xpath(bamboohr.checknow_link()).click()
        print "Clicking on Check now link"
        
        wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.connected_status())))
                   
        actual_connected_status = driver.find_element_by_xpath(bamboohr.connected_status()).text
      
        
        if (actual_connected_status == expected_connected_status):
        
            print "The HRIS Sync is successful" 
        else:
            
            print "Failed to connect to HRIS System"
            raise Exception
        
        time.sleep(5)
        
        #click on admin tab
        driver.find_element_by_xpath(bamboohr.admin_tab()).click()
    
        
        time.sleep(2)
        
        # wait for the users to be loaded
        wait.until(EC.visibility_of_element_located((By.XPATH,bamboohr.userslabel_in_userlist())))
        
        print "Enter the First name of synced user in search users field"
        
        print "The First name of synced user is "+" "+Employee_FirstName
        
        webdriver.ActionChains(driver).move_to_element(driver.find_element_by_xpath(bamboohr.search_users_field())).click( driver.find_element_by_xpath(bamboohr.search_users_field())).perform()
        
        driver.find_element_by_xpath(bamboohr.search_users_field()).send_keys(Employee_FirstName)
        
        time.sleep(10)
        
        # Going to Click and open the created users
        
        
        wait.until(EC.visibility_of_element_located((By.XPATH,"//table/tbody/tr[1]/td[1]/a[.='"+Employee_FirstName+"']")))
        driver.find_element_by_xpath("//table/tbody/tr[1]/td[1]/a[.='"+Employee_FirstName+"']").click()
        print "Clicking on the user whom the details has been updated"
        
        
        # Going to get the First name
        
        wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='create-edit-user-search-firstName']")))
        actual_firstname= driver.find_element_by_xpath(".//*[@id='create-edit-user-search-firstName']").get_attribute('value')
        print "The displayed first name is "+actual_firstname
        
        actual_lastname= driver.find_element_by_xpath(".//*[@id='create-edit-user-search-lastName']").get_attribute('value')
        print "The displayed last name is "+actual_lastname
        
        
        actual_emailidd= driver.find_element_by_xpath(".//*[@id='create-edit-user-search-username']").get_attribute('value')
        print "The displayed email id is "+actual_emailidd
         
                
        actual_employeeid= driver.find_element_by_xpath(".//*[@id='create-edit-user-search-employeeId']").get_attribute('value')
        print "The displayed employee id is "+actual_employeeid
        
      
            
        if(actual_lastname == Employee_LastName+"Updated"):
                
            if(actual_emailidd == Work_Email+"Updated"):
                    
                if(actual_employeeid == Employee_Number+"Updated"):
                    
                    print "The synced user details are matching as in BambooHR sandbox"
           
        else:
            print "Failed to validate the synced user details"
            raise Exception
        
        
        
        
        
        
        
        

       
    def bamboohrUpdateuserMain(self):    
        
        try:
            book=xlrd.open_workbook(os.path.join('E:\NewWorkspace\FirstProjectInPython\TestData.xlsx'))
            first_sheet = book.sheet_by_name('BambooHR')
        
            cell4= first_sheet.cell(1,3)
            bamboohr_sandbox_url = cell4.value 
        
            cell5= first_sheet.cell(1,4)
            sandbox_username = cell5.value 
        
            cell6= first_sheet.cell(1,5)
            sandbox_password = cell6.value
            
            cell3= first_sheet.cell(1,8)
            user_firstname = cell3.value 
            
            cell = first_sheet.cell(1,7)
            Employee_Number = cell.value
            
            cell = first_sheet.cell(1,8)
            Employee_FirstName = cell.value
            
            cell = first_sheet.cell(1,9)
            Employee_LastName = cell.value
            
            cell = first_sheet.cell(1,10)
            Work_Email = cell.value
            
            cell10= first_sheet.cell(4,3)
            suggested_attribute_name = cell10.value
            
            cell1= first_sheet.cell(1,0)
            subdomain = cell1.value
            
            cell5= first_sheet.cell(1,4)
            sandbox_username = cell5.value 
            
            cell6= first_sheet.cell(1,5)
            sandbox_password = cell6.value

            
            cell11= first_sheet.cell(4,4)
            expected_connected_status = cell11.value 

        
            
            obj1 = BambooHRUpdateUserDetails()
            obj1.bamboohrUpdateuser(bamboohr_sandbox_url,sandbox_username,sandbox_password,user_firstname,Employee_Number,Employee_LastName,Work_Email)
            obj12= BaseTestClass()
            obj12.UserLogin()
            obj1.verifyUpdateduser(Employee_FirstName,suggested_attribute_name,subdomain,sandbox_username,sandbox_password,Employee_LastName,Work_Email,Employee_Number,expected_connected_status)
            
            print "TEST CASE EXECUTION SUCCESSFULLY COMPLETED"
        except Exception as e:
            traceback.print_exc()
            print (e)
            raise Exception
            
        finally:
            #print "clicking on Home"
            book=xlrd.open_workbook(os.path.join('E:\NewWorkspace\FirstProjectInPython\TestData.xlsx'))
            first_sheet = book.sheet_by_name('Login_Credentials')
            #print("Fetching the Attribute Name from Excel Sheet\n")
            # read a cell
            cell = first_sheet.cell(1,1)
            HomeURL = cell.value
            #print HomeURL
            driver.get(HomeURL)
            #print "Home Page Loaded"


if __name__ == '__main__':
    
    obj11= BambooHRUpdateUserDetails()
    obj11.bamboohrUpdateuserMain()

    

    

    print "Test executed successfully"
           
