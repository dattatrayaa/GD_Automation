'''
Created on 19-Jun-2018

@author: Sheethu C
'''
from operator import contains
import os.path
from time import gmtime, strftime
import time
import traceback
from openpyxl.reader.excel import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import xlrd
from BaseTestClass import BaseTestClass
from BaseTestClass import driver
from DeleteLesson import DeleteLesson
from CampaignPageElements import CampPage
from CreateLessonDifferentCards import CreateLessonDifferentCards
from UsersPageElements import UsersPageElements
from LessonsPageElements import LessonsPageElements
from BaseTestClass import projectPath
class ReviewWhatYouveLearned():
    
    
    def allCardsOneTime(self,lessonName,questionCard, ans1, ans2):
        
        print "\nCreating lesson with one card"
        wait=WebDriverWait(driver, 60)
        driver.refresh()
        print "Clicking on Lessons button from side menu"
        
        lessons=LessonsPageElements()
        lessons.lessonsButtonSideMenuUnexpanded()
     
        print "Click on Create lesson button"
        lessons.createLessonButton()
        
        print "Verifying Create new lesson tab is displayed"
        
        
        if lessons.createANewLessonPopupHeader()== "Create a new lesson":
            print("Create a new lesson tab is displayed")
        else:
            print ""
            raise Exception
        
               
        
        lessons.clickOnBlankLesson()
        print "Clicked on Blank lesson"
        
        print "Creating New lesson"
        lessons.settingLessonName(lessonName)
        print "Entered lesson name ::"+lessonName
      
      
        #Text Card
        objfore=CreateLessonDifferentCards()
         
        
        #objfore.docCard(documentPath, timetoUploadDoc)
        objfore.quesCard(questionCard, ans1, ans2)
       
        print "All Cards inserted"
        
        
        print "Publishing lesson"
        lessons.readyToPublishButtonClick()
        
        print "Clicked on publish button"
        lessons.publishButtonClick()

        print "Lesson published"
        
        
        lessons.backButton()
        
        #Verifying created lesson is displayed in list
        
        wait.until(EC.visibility_of_element_located((By.XPATH,"(//tbody/tr/td[2]/a[.='"+lessonName+"'])[1]")))

        if driver.find_element_by_xpath("(//tbody/tr/td[2]/a[.='"+lessonName+"'])[1]").is_displayed():
            
            print "\nLesson is displayed in Grid ::"+lessonName
            
        else:
            print "Lesson not displaying in grid"
            raise Exception
        
        
        lessons.expandSideMenu()
    
    def campaignTolearnerAllcards(self,campaignTitle,campDescription,lessonName,nameOFuser):
        
        elements=CampPage()
        
        wait=WebDriverWait(driver, 60)
        
        
        print "\n\nCreating Campaign"
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.campaignButtonFromSideMenuXpath())))
        elements.campaignButtonFromSideMenu()
        
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.createCampaignButtonXpath())))

        if elements.campaignsPageHeaderText()=="Campaigns":
            print "Campaigns page displayed"
        else:
            print "Campaigns page is not displayed"
            raise Exception
        
        
        print "Clicking on Create Campaign button"
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.createCampaignButtonXpath())))
        elements.createCampaignButton()
        
        
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.Camp_titleXpath())))
        print "Create Campaign page is displayed"
        
                  
        elements.titleTextField(campaignTitle)
        print "Title entered ::campTitle"
        
        elements.descriptionField(campDescription)
        print "Description entered ::campDescription"
        
        print "Adding Lesson"
        #Add lesson button
        elements.addlessonButton()
        
        #Waiting until first lesson in pop is displayed
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.FirstLessonWaitXpath())))
        
        #Searching lesson by its name
        elements.searchLesson(lessonName)
        
        #Waiting until lesson displayed
        elements.waitUntilSearchedLessonDisplayed(lessonName)
        
        #selecting searched lesson
        elements.selectSearchedLesson(lessonName)
        
        #waiting until add to campaign button is click able
        wait.until(EC.element_to_be_clickable((By.XPATH,elements.AddToCampaign_ButtonXpath())))
        
        #Clicking on Add to Campaign button
        elements.addToCampaignButton()
        
        
        
        #Verifying Added lesson is displayed in Grid
        print "\nVerifying Added lesson is displayed in Grid"
        if elements.firstLessonInGrid()==lessonName:
            print "Lesson displayed in grid"
        else:
            print "Lesson not displayed in grid"
        
        wait.until(EC.element_to_be_clickable((By.XPATH,elements.SaveAndExit_ButtonXpath())))
        #Clicking on save & exit button
        print "Clicking on Save and plan assignmet button"
        elements.saveAndPlanAssignmentbutton()
        
        
        #Verifying Plan assignment for page is displayed
        print "Verifying Plan assignment for page is displayed"
        if campaignTitle in elements.planAssignementForHeaderText():
            print "Page '"+elements.planAssignementForHeaderText()+"' is displayed"
        else:
            print "Plan assignment page is not displayed"
            raise Exception
        
        
        #Checking One time triggered radio button is selected by Default
        if driver.find_element_by_xpath("//input[@id='assignment-one-time']").is_selected():
            print "Radio button One Time triggered is selected"
        else:
            print "Radio button is not selected"
            raise Exception
        
        
        '''print "Difference between two dates"
        from datetime import datetime
        date_format = "%m/%d/%Y"
        a = datetime.strptime(elements.assignDateValue(), date_format)
        b = datetime.strptime(elements.dueDateValue(), date_format)
        delta = b - a
        print "The difference between two dates is ::"
        print delta.days'''
        
        
        #Adding user to this campaign
        elements.addingUser(nameOFuser)
        
        
        print "Verifying User is displayed in Grid"
        if nameOFuser in elements.userInGrid():
            print "User is displayed in Grid"
        else:
            print "user is not displayed in grid"
            raise Exception        
        
        
        
        elements.sendAssignment()
        print "Clicked on Send assignment button"
        
        
        print "Verifying Confirm pop up is displayed"
        if elements.confirmPopupsendingAssignement().is_displayed():
            print "Confirm assignment pop up is displayed"
        else:
            print "Confirm assignment pop up is not displayed"
            raise Exception
        
        
        elements.sendAssignmentFromPopup()
        print "Clicked on Send Assignment button from pop up"
        
        
        
        #Verifying Campaign Detail page is displayed
        print "\nVerifying campaign detail page is displayed"
        
        if elements.campaignDetailPageHeaderText()==campaignTitle:
            print "Campaign detail page is displayed"
        else:
            print "Campaign detail page is not displayed"
            raise Exception
        
        #verifying in Campaigns displayed in Campaigns grid
        elements.searchingForlesson(campaignTitle)
        
        if elements.actualCampTitleINGrid()==campaignTitle:
            print "Campaign '"+campaignTitle+"' displayed in Grid"
        
        else:
            print "Campaign is not displayed in Grid"
            raise Exception
        
        print "\n----Text Execution Completed----\n"
    
    def create(self,FirstName,LastName,Email,EmployeeId,Password):
        
        wait=WebDriverWait(driver, 60)
        driver.refresh()
        user=UsersPageElements()
        
        user.AdminFromSideMenuClick()
        print "Clicked on Admin"
        
        user.UsersFromSideMenuClick()
        print "clicked on Users"
        
        user.AddOrEditUserButtonClick()
        print "Clicked on Add or editUser"

        
        user.AddAnIndividualUserButtonClick()
        print "Clicked on Add An individual User"
        time.sleep(1)
        wait.until(EC.visibility_of_element_located((By.XPATH,user.AddUserPageHeaderXpath())))
        
        print "Verifying Add user Page"
        if driver.find_element_by_xpath(user.AddUserPageHeaderXpath()).is_displayed():
            print("Add user Page is displayed")
        else:
            print ""
            raise Exception
        
        
        print "Verifying First Name field"
        fn=wait.until(EC.visibility_of_element_located((By.ID,user.FirstNameFieldID())))
        if fn.is_displayed():
            print(" First Name field displayed")
        else:
            print ""
            raise Exception
        fn.send_keys(FirstName)
        print "FirstName is Entered ::"+FirstName
        
        
        
        print "Last NAme verifying"
        if driver.find_element_by_id(user.LastNameFieldID()).is_displayed():
            print(" Last Name field displayed")
        else:
            print ""
            raise Exception
        driver.find_element_by_id(user.LastNameFieldID()).send_keys(LastName)
        print "Last Name is Entered ::"+LastName
        
        
        print "Email verifying"
        if driver.find_element_by_id(user.EmailFieldID()).is_displayed():
            print("Email field displayed")
        else:
            print ""
            raise Exception
        driver.find_element_by_id(user.EmailFieldID()).send_keys(Email)
        print "Email is Entered ::"+Email
        
        
        print "Employee ID verifying"
        if driver.find_element_by_id(user.EmployeeIDFieldID()).is_displayed():
            print("Employee ID field displayed")
        else:
            print ""
            raise Exception
        driver.find_element_by_id(user.EmployeeIDFieldID()).send_keys(EmployeeId)
        print "Employee ID  is Entered ::"+EmployeeId
        
        
        print "Inherited Role Verifying"
        if driver.find_element_by_xpath(user.InHeritedRoleFieldXpath()).is_displayed():
            print("Inherited Role field displayed")
        else:
            print ""
            raise Exception
        
        print "Password Field is Verifying"
        if driver.find_element_by_id(user.PassWordFieldID()).is_displayed():
            print("Password field displayed")
        else:
            print ""
            raise Exception
        driver.find_element_by_id(user.PassWordFieldID()).send_keys(Password)
        print "Password is Entered ::"+Password
        
        
        user.AddButtonClick()
        print "Clicked on add button"
        time.sleep(2)

        user.SaveButtonClick()        
        print "Clicked on Save"
        
        wait.until(EC.visibility_of_element_located((By.XPATH,user.UsersPageHeader())))
        
        
        print "Searching for the Created User"
        user.SearchForUserInGrid(FirstName)
        
        
        user.CheckUserDisplayedInGrid(FirstName) 
        driver.refresh()
        
    def createuserLogin(self,Email,Password,NewPassword):
        ele =driver.find_element_by_xpath(".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")
        driver.execute_script('arguments[0].click()',ele)
        elem=driver.find_element_by_xpath("html/body/div/div/div[1]/div[2]/div[2]/a")
        driver.execute_script('arguments[0].click()',elem)
        wait=WebDriverWait(driver, 60)
        wait.until(EC.visibility_of_element_located((By.ID,"username")))
    
       
        print "Grovo Sign-In page is displayed"
        
        print "Entering User name"
        driver.find_element_by_xpath(".//*[@id='username']").send_keys(Email)
       
        print "Entering Password"
        element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "password")))
        element.send_keys(Password)
        
        element.send_keys(Keys.TAB)
        print "Clicking on Sign_In button"
        driver.find_element_by_xpath("//*[@id='submitButton']").click()
        '''wait.until(EC.visibility_of_element_located((By.ID,"currentPassword")))
        driver.find_element_by_id("currentPassword").send_keys(Password)
        print "Current Password is entered :"+Password
        driver.find_element_by_id("newPassword").send_keys(NewPassword)
        print "New Password is entered :"+NewPassword
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div[2]/div/div/div[2]/div[1]/div[2]/button")))
        driver.find_element_by_xpath("html/body/div[2]/div/div/div[2]/div[1]/div[2]/button").click()
        
        
        wait.until(EC.invisibility_of_element_located((By.XPATH,"html/body/div[2]/div/div/div[2]/div[1]/div[2]/button")))'''
        wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='global-header-search']")))
        print "Home Page is Loaded"
        
        print "Learner Home Page Verification"
        actualresult = driver.find_element_by_xpath("//*[@id='content']/div/div[1]/div/nav/div[1]/a[2]/span")
        if actualresult.is_displayed():
            print"User is able to login and Dashboard is displayed.."
        else:
            print"User not able to login.."
            raise Exception    
        print "Sign out "
    
    def againLoginUser(self,url,username,password):
        ele =driver.find_element_by_xpath(".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")
        driver.execute_script('arguments[0].click()',ele)
        elem=driver.find_element_by_xpath("html/body/div/div/div[1]/div[2]/div[2]/a")
        driver.execute_script('arguments[0].click()',elem)
        
        element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "password")))
        if driver.title == "Grovo":
            print("Grovo Application URL Opened")
        else:
            raise Exception.message

        print "Grovo Sign-In page is displayed"
        
        print "Entering User name"
        driver.find_element_by_xpath(".//*[@id='username']").send_keys(username)
       
        print "Entering Password"
        element.send_keys(password)
        
        element.send_keys(Keys.TAB)
        print "Clicking on Sign_In button"
        driver.find_element_by_xpath("//*[@id='submitButton']").click()
        
        print "Successfully Logged Into Grovo Application"
        time.sleep(5)
        
        
    def updateUser(self):
        
        print "Updating Details"
        book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        second_sheet = book.sheet_by_name('PersonalSettings')
        
        cell = second_sheet.cell(50,1)
        FirstName = cell.value
        FirstNameId = FirstName.split("_")
        emp = FirstNameId[0]+"_"
        ids = FirstNameId[1]
        empId1 = int(ids)+1
        FirstNameUpdated= emp+str(empId1)
        
        
        cell2 = second_sheet.cell(53,1)
        LearnId = cell2.value
        LearnsId = LearnId.split("#")
        emp = LearnsId[0]+"#"
        ids = LearnsId[1]
        empId = int(ids)+1
        LearnerEmpId = emp+str(empId)
        
        
        cell1 = second_sheet.cell(52,1)
        LearnEmailId = cell1.value
        LearnerEmail = LearnEmailId.split("@")
        spitValue = LearnerEmail[0][:4]
        email = spitValue+str(empId)
        id = "@"+LearnerEmail[1]
        LearnerEmailId = email+id
       
       
        
       
        
        wb = load_workbook(os.path.join('Test_Data/TestData.xlsx'))
        
        #print (wb.sheetnames)
    
        sheet = wb['UserAssignToRole']
        
        sheet.cell(row=52, column=2).value = FirstNameUpdated    
        sheet.cell(row=53, column=2).value = LearnerEmailId
        sheet.cell(row=54, column=2).value = LearnerEmpId
        
       
        wb.save(os.path.join('Test_Data/TestData.xlsx'))
    def reviewVerification(self,campaignTitle):
        wait=WebDriverWait(driver, 60)
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div[1]/div/div[1]/div/nav/div[1]/a[2]")))
        ele =driver.find_element_by_xpath(".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")
        driver.execute_script('arguments[0].click()',ele)
        elem=driver.find_element_by_xpath("html/body/div[1]/div/div[1]/div[2]/div[1]/a[1]")
        driver.execute_script('arguments[0].click()',elem)
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div[1]/div/div[3]/div[2]/div/div[1]/div/div/div/div[1]/div/h1")))
        wait.until(EC.visibility_of_element_located((By.XPATH,"(//div[2]/div[1]/div/div/a[1]/span[contains(.,'"+campaignTitle+"')])[1]")))
        startButton=wait.until(EC.visibility_of_element_located((By.XPATH,"(//a/span[contains(.,'"+campaignTitle+"')]/../../a[2])[1]")))
        startButton.click()
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div[1]/div/div/div/div[4]/div/div/div/div[1]/div/span[.='"+campaignTitle+"']")))
        time.sleep(4)
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div[1]/div/div/div/div[4]/div/div/div/div[2]/button")))
        driver.find_element_by_xpath("html/body/div[1]/div/div/div/div[4]/div/div/div/div[2]/button").click()
        time.sleep(4)
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div[1]/div/div/div/div[4]/div/div/div/div/div/div[2]/div/div/div[1]/div/div/p")))
        driver.find_element_by_xpath("html/body/div[1]/div/div/div/div[4]/div/div/div/div/div/div[2]/div/div/div[1]/div/div/p").click()
        time.sleep(4)
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div[1]/div/div/div/div[5]/div/span/button")))
        driver.find_element_by_xpath("html/body/div[1]/div/div/div/div[5]/div/span/button").click()
        time.sleep(4)
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div[1]/div/div/div/div[4]/div/div/div/div/div[1]/h1")))
        time.sleep(4)
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div[1]/div/div/div/div[1]/span/button")))
        driver.find_element_by_xpath("html/body/div[1]/div/div/div/div[1]/span/button").click()
        time.sleep(4)
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div[1]/div/div[1]/div/nav/div[1]/a[2]/span")))
        driver.find_element_by_xpath("html/body/div[1]/div/div[1]/div/nav/div[1]/a[2]/span").click()
        review =driver.find_element_by_xpath("html/body/div[1]/div/div[3]/div[2]/div/div[5]/div[1]/div[1]/h2") 
        webdriver.ActionChains(driver).move_to_element(review).perform()
        rcampaign =driver.find_element_by_xpath("html/body/div[1]/div/div[3]/div[2]/div/div[5]/div[1]/div[2]/div/div/div[1]/a/div[2]/span[1]")  
        reviewcampaign=rcampaign.text
        if reviewcampaign==campaignTitle:
            print "review campaign is displaying proper"
        else:
            print "review campaign is not displaying proper"
            raise Exception
        ele =driver.find_element_by_xpath(".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")
        driver.execute_script('arguments[0].click()',ele)
        elem=driver.find_element_by_xpath("html/body/div/div/div[1]/div[2]/div[2]/a")
        driver.execute_script('arguments[0].click()',elem)
        
        
    def mainReviewWhatyouHaveLearned(self):
        book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        first_sheet = book.sheet_by_name('PersonalSettings')
        
        cell1 = first_sheet.cell(46,1)
        lessonName1 = cell1.value
        
        cell2 = first_sheet.cell(47,1)
        questionCard= cell2.value
        
        cell2 = first_sheet.cell(48,1)
        ans1= cell2.value
        
        cell2 = first_sheet.cell(49,1)
        ans2= cell2.value
        
        cell = first_sheet.cell(50,1)
        FirstName = cell.value
        
        cell = first_sheet.cell(51,1)
        LastName = cell.value
        
        cell = first_sheet.cell(52,1)
        Email = cell.value
        
        cell = first_sheet.cell(53,1)
        EmployeeId = cell.value
        
        cell = first_sheet.cell(54,1)
        Password = cell.value
        
        cell = first_sheet.cell(55,1)
        NewPassword = cell.value
        
        cell1 = first_sheet.cell(56,1)
        campaignTitle = cell1.value
        
        cell1 = first_sheet.cell(57,1)
        campDescription = cell1.value
        
        
        cell1 = first_sheet.cell(46,1)
        lessonName = cell1.value
        
        try:
            
           
            
            
            
            obj=ReviewWhatYouveLearned()
            #obj.create(FirstName, LastName, Email, EmployeeId, Password)
            obj.allCardsOneTime(lessonName,questionCard, ans1, ans2)
            obj.campaignTolearnerAllcards(campaignTitle,campDescription,lessonName,FirstName)
            obj.createuserLogin(Email,Password,NewPassword)
            obj.reviewVerification(campaignTitle)
            
        except Exception as e:
            traceback.print_exc()
            print (e)
            raise Exception   
        
if __name__ == '__main__':
     btc=BaseTestClass()
     btc.UserLogin()
     obj11=ReviewWhatYouveLearned()
     obj11.mainReviewWhatyouHaveLearned()     
        