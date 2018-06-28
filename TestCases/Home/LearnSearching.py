'''
Created on 15-Jun-2018

@author: Sheethu C
'''
from operator import contains
import os.path
import time
import traceback

from CreateLearnerNew import CreateLearner
from openpyxl.reader.excel import load_workbook
from selenium import webdriver
from selenium.webdriver.common import keys
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import xlrd
from CreateLessonDifferentLessons import CreateLessonDifferentLessons
from CreateTrackComman import CreateTrackComman
from BaseTestClass import BaseTestClass
from BaseTestClass import driver
from DeleteLesson import DeleteLesson
from BaseTestClass import projectPath
class LearnSearching():
    def serachingLearn(self,SearchItem):
        wait=WebDriverWait(driver,80)
        wait.until(EC.visibility_of_element_located((By.ID("global-header-search"))))
        if driver.find_element_by_id("global-header-search").is_displayed():
            print "Search box found"
        else:
            print "search box Not found "
        driver.find_element_by_id("global-header-search").send_keys(SearchItem)
        driver.find_element_by_id("global-header-search").send_keys(Keys.ENTER)
        wait.until(EC.visibility_of_element_located((By.XPATH("(//span[contains(.,'"+SearchItem+"')])[1]"))))
        searchtrack=driver.find_element_by_xpath("(//span[contains(.,'"+SearchItem+"')])[1]")
        if searchtrack.is_displayed():
            print "Track found"
        else:
            print "Track Not found"
            raise Exception
        lessontime=driver.find_element_by_xpath("(//time/span[2])[1]")
        if lessontime.is_displayed():
            print "time  found"
        else:
            print "time Not found"
            raise Exception
        wait.until(EC.visibility_of_element_located((By.XPATH("html/body/div[1]/div/div[3]/div[2]/div/div/div[2]/div/ul/li[2]"))))
        driver.find_element_by_xpath("html/body/div[1]/div/div[3]/div[2]/div/div/div[2]/div/ul/li[2]").click()
        searchlesson=driver.find_element_by_xpath("(//span[contains(.,'"+SearchItem+"')])[1]")
        if searchlesson.is_displayed():
            print "Lesson found"
        else:
            print "Lesson Not found"
            raise Exception
        print "Verifying Play button in the lesson"
        elehover =driver.find_element_by_xpath("html/body/div[1]/div/div[3]/div[2]/div/div/div[2]/div/div/div/ul/div[1]/a/div[1]/div/span")
        hover=webdriver.ActionChains(driver).move_to_element(elehover)
        hover.perform()
        if elehover.is_displayed():
            print "Play button found"
        else:
            print "Play button  Not found"
            raise Exception
        lessontime=driver.find_element_by_xpath("(//time/span[2])[1]")
        if lessontime.is_displayed():
            print "time  found"
        else:
            print "time Not found"
            raise Exception
    def mainhomeDisplay(self):
        book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        first_sheet =book.sheet_by_name("PersonalSettings")
        
        cell =first_sheet.cell(18,1)
        FirstName =cell.value
        
        cell = first_sheet.cell(18,1)
        LastName = cell.value
        
        cell = first_sheet.cell(20,1)
        Email = cell.value
        
        cell = first_sheet.cell(21,1)
        EmployeeId = cell.value
        
        cell = first_sheet.cell(22,1)
        Password = cell.value
        
        cell = first_sheet.cell(23,1)
        NewPassword = cell.value
        
        cell = first_sheet.cell(24,1)
        SearchItem = cell.value
        
        cell = first_sheet.cell(24,1)
        lessonName = cell.value
        
        cell = first_sheet.cell(25,1)
        textCard = cell.value
        
        cell = first_sheet.cell(16,1)
        titleOfTrack = cell.value
        
        cell = first_sheet.cell(16,1)
        Imagefilepath = cell.value
        
        cell = first_sheet.cell(16,1)
        description = cell.value
        
        cell = first_sheet.cell(16,1)
        tagName = cell.value
        
        cell = first_sheet.cell(16,1)
        expectedSuccessText = cell.value
        
        book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        second_sheet = book.sheet_by_name('Login_Credentials')
        cell = second_sheet.cell(1,1)
        url = cell.value
        
        cell = first_sheet.cell(3,1)
        username = cell.value
        
        cell = first_sheet.cell(3,2)
        loginpassword = cell.value
        
        
        try:
            lesson =CreateLessonDifferentLessons()
            lesson.lessonWithText(lessonName,textCard)
            track=CreateTrackComman()
            track.createTrack(titleOfTrack, Imagefilepath, description, tagName, lessonName, expectedSuccessText)
            createUser=CreateLearner()
            createUser.create(FirstName, LastName, Email, EmployeeId, Password)
            createUser.createLearnerLogin(Email, Password, NewPassword)
            search=LearnSearching()
            search.serachingLearn(lessonName)
            createUser.againLoginUser(url,username, loginpassword)
            
        except Exception as e:
            traceback.print_exc()
            print (e)
            raise Exception
        finally:
            lesdel= DeleteLesson()
            lesdel.lessonDelete(lessonName) 
            book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
            first_sheet = book.sheet_by_name('Login_Credentials')
            print("Fetching the Attribute Name from Excel Sheet\n")
            cell = first_sheet.cell(1,1)
            HomeURL = cell.value
            print HomeURL
            driver.get(HomeURL)
        
if __name__ == '__main__':
     btc=BaseTestClass()
     btc.UserLogin()
     obj11=LearnSearching()
     obj11.serachingLearn()       