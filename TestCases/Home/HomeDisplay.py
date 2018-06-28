'''
Created on 15-Jun-2018

@author: Sheethu C
'''

from operator import contains
import os.path
import time
import traceback

from CreateLearnerNew import CreateLearner
from openpyxl.reader.excel import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import xlrd
from BaseTestClass import BaseTestClass
from BaseTestClass import driver
from BaseTestClass import projectPath
class HomeDisplay():
    def displayHome(self,FirstName):
        wait=WebDriverWait(driver,80)
        wait.until(EC.visibility_of_element_located((By.XPATH("html/body/div[1]/div/div[1]/div/nav/div[1]/a[2]"))))
        if driver.find_element_by_xpath("html/body/div[1]/div/div[1]/div/nav/div[1]/a[2]/span").is_displayed():
            print "Home found"
        else:
            print "Not found Hoem icon"
            raise Exception
        Home=driver.find_element_by_xpath("html/body/div[1]/div/div[1]/div/nav/div[1]/a[2]/span")
        if Home.text == "Home":
            print "Home icon found"
        else:
            print "Home icon is not found"
            raise Exception
        Library=driver.find_element_by_xpath("html/body/div[1]/div/div[1]/div/nav/div[1]/a[3]/span")
        if Library.text == "Library":
            print "Home icon found"
        else:
            print "Home icon is not found"
            raise Exception
        Mycalendar=driver.find_element_by_xpath("html/body/div[1]/div/div[3]/div[2]/div/div[2]/div/div/div/div[2]/h2")
        webdriver.ActionChains(driver).move_to_element(Mycalendar).perform()
        if Mycalendar.text == "My calendar":
            print "Home icon found"
        else:
            print "Home icon is not found"
            raise Exception
        FeaturedLessons=driver.find_element_by_xpath("html/body/div[1]/div/div[3]/div[2]/div/div[3]/div/div[1]/div/div/h2")
        webdriver.ActionChains(driver).move_to_element(FeaturedLessons).perform()
        if FeaturedLessons.text == "Featured lessons":
            print "Featured lessons  found"
        else:
            print "Featured lessons  is not found"   
            raise Exception
        BrowsebyTag=driver.find_element_by_xpath("html/body/div[1]/div/div[3]/div[2]/div/div[4]/div/div[1]/div/div/h2")
        webdriver.ActionChains(driver).move_to_element(BrowsebyTag).perform()
        if BrowsebyTag.text == "Browse by tag":
            print "BrowsebyTag  found"
        else:
            print "BrowsebyTag is not found"      
            raise Exception
        RevieWhatYouveLearned=driver.find_element_by_xpath("html/body/div[1]/div/div[3]/div[2]/div/div[5]/div[1]/div[1]/h2")
        webdriver.ActionChains(driver).move_to_element(RevieWhatYouveLearned).perform()
        if RevieWhatYouveLearned.text == "Review what you've learned":
            print "Review what you've learned  found"
        else:
            print "Review what you've learned is not found"      
            raise Exception
        if driver.find_element_by_xpath("html/body/div[1]/div/div[1]/div/nav/div[1]/a[1]/div").is_displayed():
            print "Grovo icon found"
        else:
            print "Not found grovo icon"
            raise Exception
        
        print "Verifying What You Have to Learn"
        if driver.find_element_by_id("global-header-search").is_displayed():
            print "What You Have to Learn found"
        else:
            print "Not found What You Have to Learn"
            raise Exception
        
        print "Verifying HELP"
        
        if driver.find_element_by_xpath("html/body/div/div/div[2]/div/span[2]").is_displayed():
            print "HELP  found"
        else:
            print "Not found HELP"
            raise Exception
        
        HolderName =driver.find_element_by_xpath("html/body/div[1]/div/div[1]/div/nav/div[2]/a/span[2]").text  
        if HolderName == FirstName:
            print "Holder name is correct"
        else:
            print "Holder name is not correct"   
            
        wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")))
        ele =driver.find_element_by_xpath(".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")
        driver.execute_script('arguments[0].click()',ele)
            
        MyLearning =driver.find_element_by_xpath("html/body/div[1]/div/div[1]/div/nav/div[2]/a/span[2]").text  
        if MyLearning =="My learning":
            print "my-learning is displaying"
        else:
            print "my-learning name is not displaying"      
            raise Exception
        
        PersonalSettings =driver.find_element_by_xpath("html/body/div[1]/div/div[1]/div[2]/div[1]/a[2]").text  
        if PersonalSettings =="Personal settings":
            print "PersonalSettings is displaying"
        else:
            print "PersonalSettings is not displaying"      
            raise Exception
        
        Support =driver.find_element_by_xpath("html/body/div[1]/div/div[1]/div[2]/div[1]/a[3]").text  
        if Support =="Support":
            print "Support is displaying"
        else:
            print "Support is not displaying"      
            raise Exception
        SignOut =driver.find_element_by_xpath("html/body/div[1]/div/div[1]/div[2]/div[2]/a").text  
        if SignOut =="Sign out":
            print "Sign out is displaying"
        else:
            print "Sign out is not displaying"      
            raise Exception
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div/div/div[1]/div[2]/div[2]/a")))
        elem=driver.find_element_by_xpath("html/body/div/div/div[1]/div[2]/div[2]/a")
        driver.execute_script('arguments[0].click()',elem)
    def mainhomeDisplay(self):
        book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        first_sheet =book.sheet_by_name("PersonalSettings")
        
        cell =first_sheet.cell(10,1)
        FirstName =cell.value
        
        cell = first_sheet.cell(11,1)
        LastName = cell.value
        
        cell = first_sheet.cell(12,1)
        Email = cell.value
        
        cell = first_sheet.cell(13,1)
        EmployeeId = cell.value
        
        cell = first_sheet.cell(14,1)
        Password = cell.value
        
        cell = first_sheet.cell(15,1)
        NewPassword = cell.value
        
        book=xlrd.open_workbook(os.path.join('TestData.xlsx'))
        second_sheet = book.sheet_by_name('Login_Credentials')
        cell = second_sheet.cell(1,1)
        url = cell.value
        
        cell = first_sheet.cell(3,1)
        username = cell.value
        
        cell = first_sheet.cell(3,2)
        loginpassword = cell.value
        
        try:
            createUser=CreateLearner()
            createUser.create(FirstName, LastName, Email, EmployeeId, Password)
            createUser.createLearnerLogin(Email, Password, NewPassword)
            hme =HomeDisplay()
            hme.displayHome(FirstName)
            createUser.againLoginUser(url,username, loginpassword)
            
        except Exception as e:
            traceback.print_exc()
            print (e)
            raise Exception
        finally:
            book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
            first_sheet = book.sheet_by_name('Login_Credentials')
            print("Fetching the Attribute Name from Excel Sheet\n")
            cell = first_sheet.cell(1,1)
            HomeURL = cell.value
            print HomeURL
            driver.get(HomeURL)
        
if __name__ == '__main__':
     btc=BaseTestClass()
     btc.UserLogin()
     obj11=HomeDisplay()
     obj11.mainhomeDisplay()       
    