'''
Created on 18-Jun-2018

@author: Sheethu C
'''
from operator import contains
import os.path
from time import gmtime, strftime
import time
import traceback
from openpyxl.reader.excel import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import xlrd

from BaseTestClass import BaseTestClass
from BaseTestClass import driver
from CreateLearnerNew import CreateLearner
from DeleteLesson import DeleteLesson
from UsersPageElements import UsersPageElements
from CampaignPageElements import CampPage
from CreateLessonDifferentLessons import CreateLessonDifferentLessons
from PlanAssignmentForPageElements import PlanAssignmentForPageElements
from BaseTestClass import projectPath
class MyCalendar():
    
    def createuser(self,FirstName,LastName,Email,EmployeeId,Password):
        
        wait=WebDriverWait(driver, 60)
        driver.refresh()
        user=UsersPageElements()
        
        user.AdminFromSideMenuClick()
        print "Clicked on Admin"
        
        user.UsersFromSideMenuClick()
        print "clicked on Users"
        
        user.AddOrEditUserButtonClick()
        print "Clicked on Add or editUser"

        
        user.AddAnIndividualUserButtonClick()
        print "Clicked on Add An individual User"
        time.sleep(1)
        wait.until(EC.visibility_of_element_located((By.XPATH,user.AddUserPageHeaderXpath())))
        
        print "Verifying Add user Page"
        if driver.find_element_by_xpath(user.AddUserPageHeaderXpath()).is_displayed():
            print("Add user Page is displayed")
        else:
            print ""
            raise Exception
        
        
        print "Verifying First Name field"
        fn=wait.until(EC.visibility_of_element_located((By.ID,user.FirstNameFieldID())))
        if fn.is_displayed():
            print(" First Name field displayed")
        else:
            print ""
            raise Exception
        fn.send_keys(FirstName)
        print "FirstName is Entered ::"+FirstName
        
        
        
        print "Last NAme verifying"
        if driver.find_element_by_id(user.LastNameFieldID()).is_displayed():
            print(" Last Name field displayed")
        else:
            print ""
            raise Exception
        driver.find_element_by_id(user.LastNameFieldID()).send_keys(LastName)
        print "Last Name is Entered ::"+LastName
        
        
        print "Email verifying"
        if driver.find_element_by_id(user.EmailFieldID()).is_displayed():
            print("Email field displayed")
        else:
            print ""
            raise Exception
        driver.find_element_by_id(user.EmailFieldID()).send_keys(Email)
        print "Email is Entered ::"+Email
        
        
        print "Employee ID verifying"
        if driver.find_element_by_id(user.EmployeeIDFieldID()).is_displayed():
            print("Employee ID field displayed")
        else:
            print ""
            raise Exception
        driver.find_element_by_id(user.EmployeeIDFieldID()).send_keys(EmployeeId)
        print "Employee ID  is Entered ::"+EmployeeId
        
        
        print "Inherited Role Verifying"
        if driver.find_element_by_xpath(user.InHeritedRoleFieldXpath()).is_displayed():
            print("Inherited Role field displayed")
        else:
            print ""
            raise Exception
        
        print "Password Field is Verifying"
        if driver.find_element_by_id(user.PassWordFieldID()).is_displayed():
            print("Password field displayed")
        else:
            print ""
            raise Exception
        driver.find_element_by_id(user.PassWordFieldID()).send_keys(Password)
        print "Password is Entered ::"+Password
        
        
        user.AddButtonClick()
        print "Clicked on add button"
        time.sleep(2)

        user.SaveButtonClick()        
        print "Clicked on Save"
        
        wait.until(EC.visibility_of_element_located((By.XPATH,user.UsersPageHeader())))
        
        
        print "Searching for the Created User"
        user.SearchForUserInGrid(FirstName)
        
        
        user.CheckUserDisplayedInGrid(FirstName) 
    
        driver.refresh()
    
    
    def campaignToCheckDueToday(self,campaignTitle,campDescription,lessonName,FirstName,minPassingScore,numberOfAttempts):
        
        elements=CampPage()
        
        wait=WebDriverWait(driver, 60)
        
        driver.refresh()
        print "\n\nCreating Campaign"
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.campaignButtonFromSideMenuXpath())))
        elements.campaignButtonFromSideMenu()
        
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.createCampaignButtonXpath())))

        if elements.campaignsPageHeaderText()=="Campaigns":
            print "Campaigns page displayed"
        else:
            print "Campaigns page is not displayed"
            raise Exception
        
        
        print "Clicking on Create Campaign button"
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.createCampaignButtonXpath())))
        elements.createCampaignButton()
        
        
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.Camp_titleXpath())))
        print "Create Campaign page is displayed"
        
                  
        elements.titleTextField(campaignTitle)
        print "Title entered"
        
        elements.descriptionField(campDescription)
        print "Description entered"
        
        print "Adding Lesson"
        #Add lesson button
        elements.addlessonButton()
        
        #Waiting until first lesson in pop is displayed
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.FirstLessonWaitXpath())))
        
        #Searching lesson by its name
        elements.searchLesson(lessonName)
        
        #Waiting until lesson displayed
        elements.waitUntilSearchedLessonDisplayed(lessonName)
        
        #selecting searched lesson
        elements.selectSearchedLesson(lessonName)
        
        #waiting until add to campaign button is click able
        wait.until(EC.element_to_be_clickable((By.XPATH,elements.AddToCampaign_ButtonXpath())))
        
        #Clicking on Add to Campaign button
        elements.addToCampaignButton()
        
        
        time.sleep(2)
        #Verifying Added lesson is displayed in Grid
        print "\nVerifying Added lesson is displayed in Grid"
        if elements.firstLessonInGrid()==lessonName:
            print "Lesson displayed in grid"
        else:
            print "Lesson not displayed in grid"
            raise Exception
        
        
        print "Setting minimum passing score"
        elements.makeThisAsAGradedCampaign()
        elements.setMinimumPassingScore(minPassingScore)
        print "Setting maximum number of attempts"
        elements.setAMaxNoOfAttempts(numberOfAttempts)
        
        
        
        wait.until(EC.element_to_be_clickable((By.XPATH,elements.SaveAndExit_ButtonXpath())))
        print "Clicking on Save & Exit button"
        elements.saveAndPlanAssignmentbutton()
        
        planAssign=PlanAssignmentForPageElements()
        if campaignTitle in planAssign.planAssignmentHeader():
            print "'"+planAssign.planAssignmentHeader()+"' is displayed"
        else:
            print "Plan Assignment page is not displayed"
            raise Exception
        
        
        print "Assigning assignment to a user"
        selectUser=wait.until(EC.visibility_of_element_located((By.XPATH,planAssign.AddGroupTextField())))
        webdriver.ActionChains(driver).move_to_element(selectUser).click(selectUser).send_keys(FirstName).perform()
        print "'"+FirstName+"' is enetered"
        groupdisplayed=wait.until(EC.visibility_of_element_located((By.XPATH,"(//div[@role='option']/span)[1]")))
        webdriver.ActionChains(driver).move_to_element(groupdisplayed).click().perform()
        print "'"+FirstName+" is selected"
        
        
        print "Checking User is displayed in Grid"
        userInGrid=wait.until(EC.visibility_of_element_located((By.XPATH,"//table/tbody/tr/td[1]")))
        if FirstName in userInGrid.text:
            print "'"+userInGrid.text+"' is displayed"
        else:
            print "User in grid is not displayed"
            raise Exception


        print "Changing Due date to todays date"
        driver.find_element_by_xpath("//input[@name='campaign-due-date']").click()
        datePick=wait.until(EC.visibility_of_element_located((By.XPATH,"//div[@class='react-datepicker__day react-datepicker__day--today react-datepicker__day--outside-month' or @class='react-datepicker__day react-datepicker__day--today']")))
        print "'"+str(datePick.text)+"' is selected"
        datePick.click()
        
        
        
        print "Click on Send Assignment"
        wait.until(EC.visibility_of_element_located((By.XPATH,planAssign.assignButton())))
        sendAssign=wait.until(EC.element_to_be_clickable((By.XPATH,planAssign.assignButton())))
        sendAssign.click()
        print "Clicked on Send Assignment button"
        
        
        planAssign.sendAssignmentButtonPopup()
        print "Clicked on Yes,Send assignment button from pop up"
    
        
        
        
        #Verifying Campaign Detail page is displayed
        print "\nVerifying campaign detail page is displayed"
        
        if elements.campaignDetailPageHeaderText()==campaignTitle:
            print "Campaign detail page is displayed"
        else:
            print "Campaign detail page is not displayed"
            raise Exception
        
        #verifying in Campaigns displayed in Campaigns grid
        elements.searchingForlesson(campaignTitle)
        
        if elements.actualCampTitleINGrid()==campaignTitle:
            print "Campaign '"+campaignTitle+"' displayed in Grid"
        
        else:
            print "Campaign is not displayed in Grid"
            raise Exception
        
    def createuserrLogin(self,Email,Password,NewPassword):
        
        wait=WebDriverWait(driver, 60)
        driver.refresh()
        wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")))
        
        
        print "Logging out"
        print "Clicking on Username Dropdown"
        ele =driver.find_element_by_xpath(".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")
        driver.execute_script('arguments[0].click()',ele)
        elem=driver.find_element_by_xpath("html/body/div/div/div[1]/div[2]/div[2]/a")
        driver.execute_script('arguments[0].click()',elem)
        
        print "Clicked on Sign Out option"
        
        element = wait.until(EC.presence_of_element_located((By.ID, "password")))
    
        wait=WebDriverWait(driver, 60)
        print "Grovo Sign-In page is displayed"
        
        print "Entering User name"
        driver.find_element_by_xpath(".//*[@id='username']").send_keys(Email)
       
        print "Entering Password"
        element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "password")))
        element.send_keys(Password)
        
        element.send_keys(Keys.TAB)
        print "Clicking on Sign_In button"
        driver.find_element_by_xpath("//*[@id='submitButton']").click()
        wait.until(EC.visibility_of_element_located((By.ID,"currentPassword")))
        driver.find_element_by_id("currentPassword").send_keys(Password)
        print "Current Password is entered :"+Password
        driver.find_element_by_id("newPassword").send_keys(NewPassword)
        print "New Password is entered :"+NewPassword
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div[2]/div/div/div[2]/div[1]/div[2]/button")))
        driver.find_element_by_xpath("html/body/div[2]/div/div/div[2]/div[1]/div[2]/button").click()
        
        
        wait.until(EC.invisibility_of_element_located((By.XPATH,"html/body/div[2]/div/div/div[2]/div[1]/div[2]/button")))
        wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='global-header-search']")))
        print "Home Page is Loaded"
        
        print "Learner Home Page Verification"
        actualresult = driver.find_element_by_xpath("//*[@id='content']/div/div[1]/div/nav/div[1]/a[2]/span")
        if actualresult.is_displayed():
            print"User is able to login and Dashboard is displayed.."
        else:
            print"User not able to login.."
            raise Exception
    
        
        
        
    def mycalendarCheckStatusToday(self,campaignTitle):
    
        wait =WebDriverWait(driver,80)
        driver.refresh()
        time.sleep(6)
        driver.refresh()
        time.sleep(6)
        driver.refresh()
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div[1]/div/div[3]/div[2]/div/div[2]/div/div/div/div[2]/h2")))
        mycal =driver.find_element_by_xpath("html/body/div[1]/div/div[3]/div[2]/div/div[2]/div/div/div/div[2]/h2")
        if mycal.is_displayed():
            print "My Calendar section is displaying"
        else:
            print "My Calendar section is not displaying"
            raise Exception
        print "Verifying the My Calendar section."
        print "Verifying Current Time"
        showtime = strftime("%d-%B-%y ", gmtime())
        mn =(showtime.split('-'))
        ActualMonth=mn[1]
        print "ActualMonth"+ActualMonth
        ActualDate =mn[0].strip()
        print "ActualDate"+ActualDate
        dateString =driver.find_element_by_xpath("html/body/div[1]/div/div[3]/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/table/tbody/tr[1]/td/h3").text  
        dtstr =(dateString.split(' '))
        
        ExpectedDate=dtstr[4]
        print "ExpectedDate"+ExpectedDate
        ExpectedMonth=dtstr[3]
        print "ExpectedMonth"+ExpectedMonth
         
        if ActualMonth == ExpectedMonth:
            print "Current month name in the my calendar is correct"
        else:
            print "Current month Name is displaying incorrect"
            raise Exception
        
        if ActualDate == ExpectedDate:
            print "Current date  in the my calendar is correct"
        else:
            print "Current Date is displaying incorrect"
            raise Exception
        
        print "Verifying the assignment in the calendar"
       
        wait.until(EC.visibility_of_element_located((By.XPATH,"//td[@class='rc-calendar-cell rc-calendar-today rc-calendar-selected-date rc-calendar-selected-day']/div/div")))
        
        dt=driver.find_element_by_xpath("//td[@class='rc-calendar-cell rc-calendar-today rc-calendar-selected-date rc-calendar-selected-day']/div/div")
        webdriver.ActionChains(driver).move_to_element(dt).click(dt).perform()
        if driver.find_element_by_xpath("(//table/tbody/tr/td/a/h3[.='"+campaignTitle+"'])[1]"):
            print "Todays assignment verified in the mycalendar"
        else:
            print "Todays assignment is not verified in the mycalendar"
            raise Exception
       
        wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")))
        
        
        print "Logging out"
        print "Clicking on Username Dropdown"
        ele =driver.find_element_by_xpath(".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")
        driver.execute_script('arguments[0].click()',ele)
        elem=driver.find_element_by_xpath("html/body/div/div/div[1]/div[2]/div[2]/a")
        driver.execute_script('arguments[0].click()',elem)
        
        print "Clicked on Sign Out option"
        
        

        
    
        
    def signinMaster(self,url,username,password):
        driver.get(url)
        element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "password")))
        if driver.title == "Grovo":
            print("Grovo Application URL Opened")
        else:
            raise Exception.message

        print "Grovo Sign-In page is displayed"
        
        print "Entering User name"
        driver.find_element_by_xpath(".//*[@id='username']").send_keys(username)
       
        print "Entering Password"
        element.send_keys(password)
        
        element.send_keys(Keys.TAB)
        print "Clicking on Sign_In button"
        driver.find_element_by_xpath("//*[@id='submitButton']").click()
        
        print "Successfully Logged Into Grovo Application"
        time.sleep(5)  
        
    def mainCalendar(self):   
        
        book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        first_sheet = book.sheet_by_name('My calendar')
        
        
        #Campaign
        cell1 = first_sheet.cell(1,1)
        campaignTitle = cell1.value
        
        cell1 = first_sheet.cell(2,1)
        campDescription = cell1.value
        
        
        
        cell1 = first_sheet.cell(3,1)
        minPassingScore = cell1.value
        
        cell1 = first_sheet.cell(4,1)
        numberOfAttempts = cell1.value
        
        cell1 = first_sheet.cell(5,1)
        lessonName = cell1.value
        
        cell2 = first_sheet.cell(6,1)
        questionCard= cell2.value
        
        cell2 = first_sheet.cell(7,1)
        ans1= cell2.value
        
        cell2 = first_sheet.cell(8,1)
        ans2= cell2.value
        
        
        
        
        
        #Learner 1
        cell = first_sheet.cell(9,1)
        FirstName = cell.value
        
        FirstNameId = FirstName.split("_")
        emp = FirstNameId[0]+"_"
        ids = FirstNameId[1]
        empId = int(ids)+1
        FirstNameUpdated= emp+str(empId)
        
        cell = first_sheet.cell(10,1)
        LastName = cell.value
        
        LastNameID = LastName.split("_")
        emp = LastNameID[0]+"_"
        ids = LastNameID[1]
        empId = int(ids)+1
        LastNameUpdated= emp+str(empId)
        print "LastNameUpdated"+LastNameUpdated
        
        
        cell = first_sheet.cell(12,1)
        Email = cell.value
        
       
        
        LearnerEmail = Email.split("@")
        spitValue = LearnerEmail[0][:4]
        email = spitValue+str(empId)
        id = "@"+LearnerEmail[1]
        EmailIdUpdated = email+id
        
        cell = first_sheet.cell(13,1)
        EmployeeId = cell.value
        
        Employee = EmployeeId.split("#")
        emp = Employee[0]+"#"
        ids = Employee[1]
        empId = int(ids)+1
        EmployeeIdUpdated= emp+str(empId)
        
        
        cell = first_sheet.cell(14,1)
        Password = cell.value
        
        cell = first_sheet.cell(15,1)
        NewPassword = cell.value
        
        cell = first_sheet.cell(16,1)
        role = cell.value
        
        cell = first_sheet.cell(17,1)
        status = cell.value
        
        #For Original User
        book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        s_sheet = book.sheet_by_name('Login_Credentials')
        
        cell = s_sheet.cell(1,1)
        url = cell.value
        
        cell = s_sheet.cell(3,1)
        username = cell.value
        
        cell = s_sheet.cell(3,2)
        password = cell.value 
        
        
        
        
        #updating user values
        wb = load_workbook(os.path.join('Test_Data/TestData.xlsx'))
            #print (wb.sheetnames)
        
        sheet = wb['My calendar']
        sheet.cell(row=10, column=2).value = FirstNameUpdated
        sheet.cell(row=11, column=2).value = LastNameUpdated
        sheet.cell(row=13, column=2).value = EmailIdUpdated
        sheet.cell(row=14, column=2).value = EmployeeIdUpdated
        wb.save(os.path.join('Test_Data/TestData.xlsx'))
            
        print "All User Data Updated in Excel"
        
        showtime = strftime("%d-%B-%y ", gmtime())
        mn =(showtime.split('-'))
        ActualMonth=mn[1]
        print "ActualMonth"+ActualMonth
        ActualDate =mn[2].strip()
        try:
            print "\nCreating a New Learner\n"
            ik=MyCalendar()
            #ik.createuser(FirstName, LastName, Email, EmployeeId, Password)
            ot=CreateLessonDifferentLessons()
            ot.lessonWithQuestion(lessonName, questionCard, ans1, ans2)
            print "\nCreating a lesson\n"
            print "\nCreating Campaign\n"
            
            
            ik.campaignToCheckDueToday(campaignTitle, campDescription, lessonName, FirstName, minPassingScore, numberOfAttempts)
            ik.createuserrLogin(Email, Password, NewPassword)
           
            
            print "Verifying MyCalendar"
            ik.mycalendarCheckStatusToday(campaignTitle)
            ik.signinMaster(url,username,password)
           
            
        except Exception as e:
            traceback.print_exc()
            print (e)
            raise Exception 
         
        finally:  
            try:
                lesdel= DeleteLesson()
                lesdel.lessonDelete(lessonName)
            except Exception:
                print("no alert")
            second_sheet = book.sheet_by_name('Login_Credentials')
            cell = second_sheet.cell(1,1)
            url = cell.value
            driver.get(url)
            try:
                WebDriverWait(driver, 5).until(EC.alert_is_present())

                alert = driver.switch_to.alert
                alert.accept()
                print("alert accepted")
            except Exception:
                print("no alert")
            book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
            s_sheet = book.sheet_by_name('Login_Credentials')
        
            cell = s_sheet.cell(1,1)
            url = cell.value
        
            cell = s_sheet.cell(3,1)
            username = cell.value
        
            cell = s_sheet.cell(3,2)
            password = cell.value 
            wait=WebDriverWait(driver, 60)
            unDropDown=wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")))
            driver.execute_script('arguments[0].click()',unDropDown)
            
            signOut=driver.find_element_by_xpath("html/body/div/div/div[1]/div[2]/div[2]/a")
            driver.execute_script('arguments[0].click()',signOut)
            
            
            
            element = wait.until(EC.presence_of_element_located((By.ID, "password")))
            
            
            time.sleep(3)
            if driver.title == "Grovo":
                print("Grovo Application URL Opened")
            else:
                raise Exception.message
    
            print "Grovo Sign-In page is displayed"
            
            print "Entering User name"
            driver.find_element_by_xpath(".//*[@id='username']").send_keys(username)
           
            print "Entering Password"
            element.send_keys(password)
            
            element.send_keys(Keys.TAB)
            print "Clicking on Sign_In button"
            driver.find_element_by_xpath("//*[@id='submitButton']").click()
            
            print "Successfully Logged Into Users account"
            time.sleep(3)
             
if __name__ == '__main__':
     btc=BaseTestClass()
     btc.UserLogin()
     obj11=MyCalendar()
     obj11.mainCalendar()               
          