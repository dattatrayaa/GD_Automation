'''
Created on 15-Jun-2018

@author: Sheethu C
'''
from operator import contains
import os.path
import time
import traceback
from openpyxl.reader.excel import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import xlrd
from BaseTestClass import BaseTestClass
from BaseTestClass import driver
from CreateLearnerNew import CreateLearner
from BaseTestClass import projectPath
class PersonalSettingsDisplay():
    def displayAndChangePassword(self,NewPassword,currentPassword):
        wait =WebDriverWait(driver)
        wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")))
        ele =driver.find_element_by_xpath(".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")
        driver.execute_script('arguments[0].click()',ele)
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div[1]/div/div[1]/div[2]/div[1]/a[2]")))
        ele=driver.find_element_by_xpath("html/body/div[1]/div/div[1]/div[2]/div[1]/a[2]")
        driver.execute_script('arguments[0].click()',ele)
        print "Loading Personal Settings Page"
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div[1]/div/div[3]/div[2]/div/header/h1")))
        personalHeader=driver.find_element_by_xpath("html/body/div[1]/div/div[3]/div[2]/div/header/h1")
        if personalHeader.text =="Personal Settings":
            print "Personal Settings page displayed"
        else:
            print "Personal Settings page is not displayed"
            raise Exception
    
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div[1]/div/div[3]/div[2]/div/section/h2")))
        changePassword=driver.find_element_by_xpath("html/body/div[1]/div/div[3]/div[2]/div/section/h2")
        if changePassword.text =="Personal Settings":
            print "changePassword text is displayed"
        else:
            print "changePassword text is not displayed"
            raise Exception
        
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div[1]/div/div[3]/div[2]/div/section/div/div[1]/label/span[1]")))
        resetPassword=driver.find_element_by_xpath("html/body/div[1]/div/div[3]/div[2]/div/section/div/div[1]/label/span[1]")
        if resetPassword.text =="Personal Settings":
            print "resetPassword text is displayed"
        else:
            print "resetPassword text is not displayed"
            raise Exception
        if driver.find_element_by_xpath("html/body/div[1]/div/div[3]/div[2]/div/section/div/div[2]").is_enabled():
            print "Before clicking reset Password,Password field and New password are disabled"
        else:
            raise Exception
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div[1]/div/div[3]/div[2]/div/section/div/div[1]/label/span[2]")))
        driver.find_element_by_xpath("html/body/div[1]/div/div[3]/div[2]/div/section/div/div[1]/label/span[2]").click()
        wait.until(EC.visibility_of_element_located((By.ID,"currentPassword")))
        if driver.find_element_by_id("currentPassword").is_displayed():
            print "Current Password filed is displaying"
        else:
            print "Current Password is not displaying"
            raise Exception
        driver.find_element_by_id("currentPassword").send_keys(currentPassword)
        time.sleep(4)
        if driver.find_element_by_id("newPassword").is_displayed():
            print "newPassword  filed is displaying"
        else:
            print "newPassword  is not displaying"
            raise Exception
        driver.find_element_by_id("newPassword").send_keys(NewPassword)
        
        character8=driver.find_element_by_xpath("//table/tbody/tr[1]/td[1]")
        if character8.text =="At least 8 characters":
            print "At least 8 characters text is displayed"
        else:
            print "At least 8 characters text is not displayed"
            raise Exception
        
        characterupper=driver.find_element_by_xpath("//tbody/tr[1]/td[2]")
        if characterupper.text =="One uppercase letter":
            print "One uppercase letter text is displayed"
        else:
            print "One uppercase letter text is not displayed"
            raise Exception
        
        characterNumorletr=driver.find_element_by_xpath("//table/tbody/tr[2]/td[1]")
        if characterNumorletr.text =="One number or symbol":
            print "One number or symbol text is not displayed"
            raise Exception
        
        characterlower=driver.find_element_by_xpath("//table/tbody/tr[2]/td[2]")
        if characterlower.text =="One number or symbol":
            print "One lowercase letter text is displayed"
        else:
            print "One lowercase letter text is not displayed"
            raise Exception
        
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div[1]/div/div[3]/div[2]/div/section/div/div[2]/div[2]/button[1]")))
        wait.until(EC.element_to_be_clickable((By.XPATH,"html/body/div[1]/div/div[3]/div[2]/div/section/div/div[2]/div[2]/button[1]")))
        driver.find_element_by_xpath("html/body/div[1]/div/div[3]/div[2]/div/section/div/div[2]/div[2]/button[1]").click()
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div[1]/div/div[3]/div[2]/div/section/div/div[2]/div[1]/div[1]/div")))
        time.sleep(4)
        wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")))
        ele =driver.find_element_by_xpath(".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")
        driver.execute_script('arguments[0].click()',ele)
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div/div/div[1]/div[2]/div[2]/a")))
        elem=driver.find_element_by_xpath("html/body/div/div/div[1]/div[2]/div[2]/a")
        driver.execute_script('arguments[0].click()',elem)
        
    def updateUser(self):
        
        print "Updating Details"
        book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        second_sheet = book.sheet_by_name('UserAssignToRole')
        cell = second_sheet.cell(1,1)
        FirstName = cell.value
        
        FirstNameId = FirstName.split("_")
        emp = FirstNameId[0]+"_"
        ids = FirstNameId[1]
        empId1 = int(ids)+1
        FirstNameUser3= emp+str(empId1)
        
        cell6 = second_sheet.cell(4,1)
        LearnId = cell6.value
        LearnsId = LearnId.split("#")
        emp = LearnsId[0]+"#"
        ids = LearnsId[1]
        empId = int(ids)+1
        MasterAdminsEmpId = emp+str(empId)
        
        
        
        cell5 = second_sheet.cell(3,1)
        LearnEmailId = cell5.value
        LearnerEmail = LearnEmailId.split("@")
        spitValue = LearnerEmail[0][:4]
        email = spitValue+str(empId)
        id2 = "@"+LearnerEmail[1]
        MasterAdminEmailId = email+id2
        
        
        
        
        
        
       
        
        wb = load_workbook(os.path.join('Test_Data/TestData.xlsx'))
        
        #print (wb.sheetnames)
    
        sheet = wb['UserAssignToRole']
        
        
        
        sheet.cell(row=1, column=2).value = FirstNameUser3    
        sheet.cell(row=1, column=4).value = MasterAdminEmailId
        sheet.cell(row=1, column=5).value = MasterAdminsEmpId
        
        
        
        wb.save(os.path.join('Test_Data/TestData.xlsx'))   
        
        
        
        
    def mainPersonalsettingsDisplay(self):
        book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        first_sheet =book.sheet_by_name("PersonalSettings")
        
        cell =first_sheet.cell(1,1)
        FirstName =cell.value
        
        cell = first_sheet.cell(2,1)
        LastName = cell.value
        
        cell = first_sheet.cell(3,1)
        Email = cell.value
        
        cell = first_sheet.cell(4,1)
        EmployeeId = cell.value
        
        cell = first_sheet.cell(5,1)
        Password = cell.value
        
        
        cell = first_sheet.cell(6,1)
        NewPassword = cell.value
        
        cell = first_sheet.cell(7,1)
        ResetPassword = cell.value
        
        book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        second_sheet = book.sheet_by_name('Login_Credentials')
        cell = second_sheet.cell(1,1)
        url = cell.value
        cell = second_sheet.cell(3,1)
        username = cell.value
        
        cell = second_sheet.cell(3,2)
        loginpassword = cell.value
        
        try:
            createUser=CreateLearner()
            createUser.create(FirstName, LastName, Email, EmployeeId, Password)
            createUser.createLearnerLogin(Email, Password, NewPassword)
            personal=PersonalSettingsDisplay()
            personal.displayAndChangePassword(ResetPassword,Password)
            createUser.createLearnerLogin(url,Email, ResetPassword)
            createUser.againLoginUser(url,username, loginpassword)
        except Exception as e:
            traceback.print_exc()
            print (e)
            raise Exception
        finally:
            personal=PersonalSettingsDisplay()
            personal.updateUser()
            book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
            first_sheet = book.sheet_by_name('Login_Credentials')
            print("Fetching the Attribute Name from Excel Sheet\n")
            # read a cell
            cell = first_sheet.cell(1,1)
            HomeURL = cell.value
            print HomeURL
            driver.get(HomeURL)
        
if __name__ == '__main__':
     btc=BaseTestClass()
     btc.UserLogin()
     obj11=PersonalSettingsDisplay()
     obj11.mainPersonalsettingsDisplay()       
    