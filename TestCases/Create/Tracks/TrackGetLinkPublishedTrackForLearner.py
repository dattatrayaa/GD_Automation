'''
Created on 18-Jul-2018

@author: OptisLabs
'''
from Tkinter import Tk
import os
import time
import traceback

from BaseTestClass import BaseTestClass
from BaseTestClass import driver
from openpyxl.reader.excel import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import xlrd

from CreateLessonDifferentLessons import CreateLessonDifferentLessons
from DeleteLesson import DeleteLesson
from DeleteTrack import DeleteTrack
from TracksPageElements import TracksPageElements
from CreateLearnerNew import CreateLearner


class TrackGetLinkPublishedTrackForLearner:
    
    def createTrack(self,titleOfTrack,Imagefilepath,description,tagName,lessonname,expectedSuccessText):
        
        
        print "\nCreating track with one lesson contains Text Card"
        driver.refresh()
        wait=WebDriverWait(driver, 180)
        track=TracksPageElements()
        
        print "Clicking on Lessons button from side menu"
        track.lessonsButton()
        
        print "Clicking on Track button from side menu"
        track.tracksButton()
        
        track.createTracksButton()
        
        print "Entering title"
        titlefield=wait.until(EC.visibility_of_element_located((By.XPATH,track.trackTitle())))
        titlefield.send_keys(titleOfTrack)
        print "Title entered ::"+titleOfTrack
        
        driver.find_element_by_css_selector('input[type="file"]').send_keys(Imagefilepath)
        print "waiting to upload image"
        wait.until(EC.visibility_of_element_located((By.XPATH,track.imageDisplayed())))
        print "Image uploaded"
        
        print "Entering Description"
        driver.find_element_by_xpath(track.trackDescription()).send_keys(description)
        print "Description entered ::"+description 
        
        
        print "Adding tag"
        addTags=driver.find_element_by_xpath(track.addingTag())
        webdriver.ActionChains(driver).move_to_element(addTags).click().send_keys(tagName).perform()
        
        option=wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='react-select-2--option-0']")))
        webdriver.ActionChains(driver).move_to_element(option).click(option).perform()
        
        driver.find_element_by_xpath(track.trackDescription()).send_keys(" ")
        
        
        print "\nAdding created two lessons"
        
        print "Clicking on Add lessons button"
       
        driver.execute_script("window.scrollTo(0, 0);")
        
        track.addLessonButton()
                
        wait.until(EC.visibility_of_element_located((By.XPATH,track.serchLesson())))
        
        searchLesson=driver.find_element_by_xpath(track.serchLesson())
        print "Searching for first lesson in Add lessons pop up"
        searchLesson.send_keys(lessonname)
        track.searchedLessonAdd(lessonname)
        time.sleep(2)
        
        print "Adding to Track"
        track.addingToTrack()
        
        print "\nChecking added lesson is selected lesson from Pop up"
        
        
        lessonTextAddedToGrid=driver.find_element_by_xpath("//li[1]/div[2]/div/h4/div").text
        if lessonTextAddedToGrid==lessonname:
            print "Selected Lesson '"+lessonname+"' is displayed in grid of tracks page"
        
        else:
            print "Lesson is not displayed"
            raise Exception
        
        print "Clicking on Publish Track button"
        
        print "Checking Get Link button is disabled before publishing track"
        
        try:
            ele=wait.until(EC.visibility_of_element_located((By.XPATH,track.GetLinkButton())))
            if "is-disabled" in ele.get_attribute("class"):
                print "Get link button is disabled"
            else:
                raise Exception
        except Exception:
            raise Exception("Get Link button is not disabled")
        
        
        #Publishing Track
        print "Clicking on Publish Track button"
        
        track.publishButton()        
        print "\nVerifying Success message is displaying"
        
        
        track.successmessage(expectedSuccessText, titleOfTrack)
        
        print "Clicking on Get Link button"
        try:
            ele.click()
            print "Clicked on Get Link button"
        except Exception as e:
            print e
            traceback.print_exc()
            raise Exception("Failed to click on Get link button")
        
        print "Selecting Value get link option"
        try:
            ele=wait.until(EC.visibility_of_element_located((By.XPATH,track.GetLinkOption())))
            ele.click()
            print "Clicked on Get Link Options"
        except Exception as e:
            print e
            traceback.print_exc()
            raise Exception("Failed to click on Get link Options")
        
        print "Clicking on Copy button"
        track.GetLinkCopyButton()
        print "Clicked on COPY button"

        clip = Tk() 
        self.copiedLink = clip.selection_get(selection = "CLIPBOARD")
    
    
    def CheckForLearnerTrackDisplayed(self,Email,Password,titleOfTrack,lessonName,username,password):
        
        wait=WebDriverWait(driver, 60)
        driver.refresh()
        wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")))
        track=TracksPageElements()
        
        print "Logging out"
        print "Clicking on Username Dropdown"
        ele =driver.find_element_by_xpath(".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")
        driver.execute_script('arguments[0].click()',ele)
        elem=driver.find_element_by_xpath("html/body/div/div/div[1]/div[2]/div[2]/a")
        driver.execute_script('arguments[0].click()',elem)
        
        print "Clicked on Sign Out option"
        
        wait.until(EC.presence_of_element_located((By.ID, "password")))
        
        print "Pasting link in the URL bar"
        driver.get(self.copiedLink)
        
        element=wait.until(EC.presence_of_element_located((By.ID, "password")))
        
        if driver.title == "Grovo":
            print("Grovo Application URL Opened")
        else:
            raise Exception.message

        print "Grovo Sign-In page is displayed"
        
        print "Entering User name"
        driver.find_element_by_xpath(".//*[@id='username']").send_keys(Email)
       
        print "Entering Password"
        element.send_keys(Password)
        
        element.send_keys(Keys.TAB)
        print "Clicking on Sign_In button"
        driver.find_element_by_xpath("//*[@id='submitButton']").click()
        
        print "Successfully Logged Into Users account"  

        print "Checking User lands in Published Track page of '"+titleOfTrack+"'"
        
        try:
            title=wait.until(EC.visibility_of_element_located((By.XPATH, track.titleInTrackDetailPage())))
        except Exception:
            traceback.print_exc()
            raise Exception("Title not displayed")
        
        
        if title.text==titleOfTrack:
            print "Title of Track '"+title.text+"' is displayed"
        else:
            raise Exception("Track title not displayed")
        
        
        try:
            title=wait.until(EC.visibility_of_element_located((By.XPATH, track.lessonInTrackDetailPage(lessonName))))
        except Exception:
            traceback.print_exc()
            raise Exception("Lesson Title not displayed in Track details page")
        
        print "Lesson '"+lessonName+"'displayed in Track details page"
        
        
        print "Verified published track displayed for other user"
        
        print "Logging in as Original User"
        
        print "Logging out of current user"
        unDropDown=wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")))
        driver.execute_script('arguments[0].click()',unDropDown)
        
        signOut=driver.find_element_by_xpath("html/body/div/div/div[1]/div[2]/div[2]/a")
        driver.execute_script('arguments[0].click()',signOut)
        print "Clicked on  Sign out"
        
        
        element = wait.until(EC.presence_of_element_located((By.ID, "password")))
        
        if driver.title == "Grovo":
            print("Grovo Application URL Opened")
        else:
            raise Exception.message

        print "Grovo Sign-In page is displayed"
        
        print "Entering User name"
        driver.find_element_by_xpath(".//*[@id='username']").send_keys(username)
       
        print "Entering Password"
        element.send_keys(password)
        
        element.send_keys(Keys.TAB)
        print "Clicking on Sign_In button"
        driver.find_element_by_xpath("//*[@id='submitButton']").click()
        
        print "Successfully Logged Into Users account"
        
        time.sleep(5)    


    def TrackGetLinkPublishedForLearnerMain(self):
        
        book=xlrd.open_workbook(os.path.join('TestData.xlsx'))
        first_sheet = book.sheet_by_name('TrackCreate')
        
        cell1 = first_sheet.cell(181,3)
        titleOfTrack = cell1.value
        
        cell2 = first_sheet.cell(182,3)
        Imagefilepath = cell2.value
        
        cell2 = first_sheet.cell(183,3)
        description = cell2.value
        
        cell2 = first_sheet.cell(184,3)
        tagName = cell2.value
        
        cell2 = first_sheet.cell(185,3)
        lessonName= cell2.value
        
        cell2 = first_sheet.cell(186,3)
        textCard= cell2.value
        
        cell2 = first_sheet.cell(187,3)
        expectedSuccessText= cell2.value
        
        
        
        cell = first_sheet.cell(188,3)
        FirstName = cell.value
        
        FirstNameId = FirstName.split("_")
        emp = FirstNameId[0]+"_"
        ids = FirstNameId[1]
        empId = int(ids)+1
        FirstNameUpdated= emp+str(empId)
        
        cell = first_sheet.cell(189,3)
        LastName = cell.value
        
        LastNameID = LastName.split("_")
        emp = LastNameID[0]+"_"
        ids = LastNameID[1]
        empId = int(ids)+1
        LastNameUpdated= emp+str(empId)
        
        
        
        cell = first_sheet.cell(190,3)
        Email = cell.value
        
        EmailId = Email.split("_")
        emp = EmailId[0]+"_"
        ids = EmailId[1]
        remaining="_"+EmailId[2]
        empId = int(ids)+1
        EmailIdUpdated= emp+str(empId)+remaining
        
        cell = first_sheet.cell(191,3)
        EmployeeId = cell.value
        
        Employee = EmployeeId.split("_")
        emp = Employee[0]+"_"
        ids = Employee[1]
        empId = int(ids)+1
        EmployeeIdUpdated= emp+str(empId)
        
        cell = first_sheet.cell(192,3)
        Password = cell.value
        
        cell = first_sheet.cell(192,3)
        NewPassword = cell.value
        

        #For Original User
        book=xlrd.open_workbook(os.path.join('TestData.xlsx'))
        s_sheet = book.sheet_by_name('Login_Credentials')
        
        cell = s_sheet.cell(1,1)
        url = cell.value
        
        cell = s_sheet.cell(3,1)
        username = cell.value
        
        cell = s_sheet.cell(3,2)
        password = cell.value
        
        #updating user values
        wb = load_workbook(os.path.abspath(os.path.join(os.path.dirname(__file__),'TestData.xlsx')))
            #print (wb.sheetnames)
        
        sheet = wb['TrackCreate']
        
        sheet.cell(row=189, column=4).value = FirstNameUpdated
        sheet.cell(row=190, column=4).value = LastNameUpdated
        sheet.cell(row=191, column=4).value = EmailIdUpdated
        sheet.cell(row=192, column=4).value = EmployeeIdUpdated
        
        
        wb.save(os.path.abspath(os.path.join(os.path.dirname(__file__),'TestData.xlsx')))
            
        print "All User Data Updated in Excel"
        
        
        
        
        try:
            print "\nCreating user with Learner credentials"
            create=CreateLearner()
            create.createNewLearnerMain(FirstName, LastName, Email, EmployeeId, Password, NewPassword, url, username, password)
            
            print "\nCreating Lesson\n"
            cd=CreateLessonDifferentLessons()
            cd.lessonWithText(lessonName, textCard)
            
            print "\nCreating Track"
            cr=TrackGetLinkPublishedTrackForLearner()
            cr.createTrack(titleOfTrack, Imagefilepath, description, tagName, lessonName, expectedSuccessText)
                       
            print "\nNow Check for other user track get link is working"
            cr.CheckForLearnerTrackDisplayed(Email, Password, titleOfTrack, lessonName, username, password)
            
          
        finally: 
            second_sheet = book.sheet_by_name('Login_Credentials')
            cell = second_sheet.cell(1,1)
            url = cell.value
            driver.get(url)
            try:
                WebDriverWait(driver, 5).until(EC.alert_is_present(),
                                   'Timed out waiting for PA creation ' +
                                   'confirmation popup to appear.')

                alert = driver.switch_to.alert
                alert.accept()
                print("alert accepted")
            except Exception:
                print("no alert")
                
            try:
                d=DeleteLesson()
                d.lessonDelete(lessonName)
                t=DeleteTrack()
                t.mainDelete(titleOfTrack)
            except Exception:
                driver.get(url)



if __name__ == '__main__':
    u1=BaseTestClass()
    u1.UserLogin()
    
    tr=TrackGetLinkPublishedTrackForLearner()
    tr.TrackGetLinkPublishedForLearnerMain()
    
    
    