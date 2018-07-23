'''
Created on 12-Jul-2018

@author: OptisLabs
'''
from Tkinter import Tk
import os
import time
import traceback

from BaseTestClass import BaseTestClass
from BaseTestClass import driver
from openpyxl.reader.excel import load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import xlrd

from CreateLessonDifferentCards import CreateLessonDifferentCards
from DeleteLesson import DeleteLesson
from LessonsPageElements import LessonsPageElements
from CreateLearnerNew import CreateLearner


class LessonGetLinkBuilderToLearner:
    
    def LessongetLink(self,lessonName,textCard,url):
        
        print "Creating Lesson"
    
        print "\nCreating lesson with one card"
        wait=WebDriverWait(driver, 60)
        driver.refresh()
        print "Clicking on Lessons button from side menu"
        
        lessons=LessonsPageElements()
        lessons.lessonsButtonSideMenuUnexpanded()
     
        print "Click on Create lesson button"
        lessons.createLessonButton()
        
        print "Verifying Create new lesson tab is displayed"
        
        
        if lessons.createANewLessonPopupHeader()== "Create a new lesson":
            print("Create a new lesson tab is displayed")
        else:
            print ""
            raise Exception
        
               
        
        lessons.clickOnBlankLesson()
        print "Clicked on Blank lesson"
        
        print "Creating New lesson"
        lessons.settingLessonName(lessonName)
        print "Entered lesson name ::"+lessonName
        
        
        
        e=CreateLessonDifferentCards()
        e.textCard(textCard)
        
        print "Clicking on Get Link Button"
        lessons.GetLinkButton()
        print "Clicked on Get link button"
        
        print "Verifying Get link is displayed with status Unpublished Draft"
        if lessons.getlinKUnpublishedDraft()=="UNPUBLISHED DRAFT":
            print "By default Unpublished Draft status is displayed"
        else:
            print "Unpublished Draft status is not displayed"
            raise Exception
        
        print "Changing status from unpublished draft to Lesson builder"
        try:
            status=wait.until(EC.visibility_of_element_located((By.XPATH,lessons.GetLinkStatusButton())))
            status.click()
        except Exception:
            print "Failed to click on Get link status button"
            raise Exception("Failed to click on Get link status button")
        print "Clicked on Unpublished draft button"
        
        print "Changing status"
        try:
            builder=wait.until(EC.visibility_of_element_located((By.XPATH,lessons.lessonBuilderOption())))
            builder.click()
            print "Clicked on Lesson builder option"
        except Exception:
            print "Failed to click on lesson builder status button"
            raise Exception("Failed to click on lesson builder status button")
        
        
        print "Checking link is displayed in popup"
        try:
            linkDisplayed=wait.until(EC.visibility_of_element_located((By.XPATH,lessons.linkDisplayed())))
            print "link is displayed in popup" +linkDisplayed.text
        except Exception as e:
            traceback.print_exc()
            print e
            print "Link is not displayed in pop up"
        
        print "Verifying Copy link is displayed"
        lessons.GetLinkCopyButton()
        print "Clicked on Copy button, link is copied to clip-board"
        
        clip = Tk() 
        self.copiedLink = clip.selection_get(selection = "CLIPBOARD")
        
        lessons.backButton()
        print "Clicked on Back button"
        driver.get(url)
        
        
        
    def LogoutAndcheckLesssonBuilderLinkForLearner(self,Email,Password,username,password,expectedValidationMsg):
        
        print "\n\nChecking sent assignment is displayed for User"
        wait=WebDriverWait(driver, 60)
        driver.refresh()
        wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")))
        
        
        print "Logging out"
        print "Clicking on Username Dropdown"
        ele =driver.find_element_by_xpath(".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")
        driver.execute_script('arguments[0].click()',ele)
        elem=driver.find_element_by_xpath("html/body/div/div/div[1]/div[2]/div[2]/a")
        driver.execute_script('arguments[0].click()',elem)
        
        print "Clicked on Sign Out option"
        
        wait.until(EC.presence_of_element_located((By.ID, "password")))
        
        print "Pasting link in the URL bar"
        driver.get(self.copiedLink)
        
        element=wait.until(EC.presence_of_element_located((By.ID, "password")))
        
        if driver.title == "Grovo":
            print("Grovo Application URL Opened")
        else:
            raise Exception.message

        print "Grovo Sign-In page is displayed"
        
        print "Entering User name"
        driver.find_element_by_xpath(".//*[@id='username']").send_keys(Email)
       
        print "Entering Password"
        element.send_keys(Password)
        
        element.send_keys(Keys.TAB)
        print "Clicking on Sign_In button"
        driver.find_element_by_xpath("//*[@id='submitButton']").click()
        
        print "Successfully Logged Into Users account"  
        
        print "Checking any validation message is displayed"
        try:
            validationMessageActual=wait.until(EC.visibility_of_element_located((By.XPATH, "html/body/div[1]/div/div[3]/div[2]/div/div/h2")))
            print "validation message is displayed"
        except Exception:
            print "Draft status for lessons is not displayed"
            raise Exception
        
        if validationMessageActual.text==expectedValidationMsg:
            print "Validation message '"+validationMessageActual.text+"' is displayed"
        else:
            print "Validation message is not displayed"
            raise Exception("Validation message is not displayed")
        
        print "Links 'Back' and 'homepage' is displayed in message"
        
        try:
            wait.until(EC.visibility_of_element_located((By.XPATH, "//button[.='back']")))
            print "'back' link is displayed"
        except Exception:
            print "Back link is not displayed"
            raise Exception
        try:
            homepage=wait.until(EC.visibility_of_element_located((By.XPATH, "//a[.='homepage']")))
            print "'homepage' link is displayed"
        except Exception:
            print "homepage link is not displayed"
            raise Exception
        
        print "Clicking on homePage link"
        try:
            homepage.click()
            print "Clicked on 'homepage' link"
        except Exception:
            print "Failed to click on homepage link"
            raise Exception
        
        
        
        
        print "Logging in as Original User"
        
        print "Logging out of current user"
        unDropDown=wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")))
        driver.execute_script('arguments[0].click()',unDropDown)
        
        signOut=driver.find_element_by_xpath("html/body/div/div/div[1]/div[2]/div[2]/a")
        driver.execute_script('arguments[0].click()',signOut)
        print "Clicked on  Sign out"
        
        
        element = wait.until(EC.presence_of_element_located((By.ID, "password")))
        
        if driver.title == "Grovo":
            print("Grovo Application URL Opened")
        else:
            raise Exception.message

        print "Grovo Sign-In page is displayed"
        
        print "Entering User name"
        driver.find_element_by_xpath(".//*[@id='username']").send_keys(username)
       
        print "Entering Password"
        element.send_keys(password)
        
        element.send_keys(Keys.TAB)
        print "Clicking on Sign_In button"
        driver.find_element_by_xpath("//*[@id='submitButton']").click()
        
        print "Successfully Logged Into Users account"
        
        time.sleep(5)    
        
        
        
        
    def LessonGetlinkLessonBuilderForLearnerMain(self):
        
        book=xlrd.open_workbook(os.path.abspath(os.path.join(os.path.dirname(__file__),'TestData.xlsx')))
        first_sheet = book.sheet_by_name('LessonCreate')
        
        cell1 = first_sheet.cell(75,1)
        lessonName = cell1.value
        
        cell2 = first_sheet.cell(76,1)
        textCard= cell2.value
        
        
        #Learner 1
        cell = first_sheet.cell(77,1)
        FirstName = cell.value
        
        FirstNameId = FirstName.split("_")
        emp = FirstNameId[0]+"_"
        ids = FirstNameId[1]
        empId = int(ids)+1
        FirstNameUpdated= emp+str(empId)
        
        cell = first_sheet.cell(78,1)
        LastName = cell.value
        
        LastNameID = LastName.split("_")
        emp = LastNameID[0]+"_"
        ids = LastNameID[1]
        empId = int(ids)+1
        LastNameUpdated= emp+str(empId)
        
        
        
        cell = first_sheet.cell(79,1)
        Email = cell.value
        
        EmailId = Email.split("_")
        emp = EmailId[0]+"_"
        ids = EmailId[1]
        remaining="_"+EmailId[2]
        empId = int(ids)+1
        EmailIdUpdated= emp+str(empId)+remaining
        
        cell = first_sheet.cell(80,1)
        EmployeeId = cell.value
        
        Employee = EmployeeId.split("_")
        emp = Employee[0]+"_"
        ids = Employee[1]
        empId = int(ids)+1
        EmployeeIdUpdated= emp+str(empId)
        
        cell = first_sheet.cell(81,1)
        Password = cell.value
        
        cell = first_sheet.cell(81,1)
        NewPassword = cell.value
        
        cell = first_sheet.cell(82,1)
        expectedValidationMsg = cell.value
        
        
        
        #For Original User
        book=xlrd.open_workbook(os.path.join('TestData.xlsx'))
        s_sheet = book.sheet_by_name('Login_Credentials')
        
        cell = s_sheet.cell(1,1)
        url = cell.value
        
        cell = s_sheet.cell(3,1)
        username = cell.value
        
        cell = s_sheet.cell(3,2)
        password = cell.value
        
            
        
        
        #updating user values
        wb = load_workbook(os.path.abspath(os.path.join(os.path.dirname(__file__),'TestData.xlsx')))
            #print (wb.sheetnames)
        
        sheet = wb['LessonCreate']
        
        sheet.cell(row=78, column=2).value = FirstNameUpdated
        sheet.cell(row=79, column=2).value = LastNameUpdated
        sheet.cell(row=80, column=2).value = EmailIdUpdated
        sheet.cell(row=81, column=2).value = EmployeeIdUpdated
        
        
        wb.save(os.path.abspath(os.path.join(os.path.dirname(__file__),'TestData.xlsx')))
            
        print "All User Data Updated in Excel"
        
        
        
        
        try:
            print "\nCreating a New Learner\n"
            lt=CreateLearner()
            lt.createNewLearnerMain(FirstName, LastName, Email, EmployeeId, Password, NewPassword, url, username, password)
            
            print "\nCreating a lesson with coping get link\n"
            un=LessonGetLinkBuilderToLearner()
            un.LessongetLink(lessonName, textCard, url)
            print "\nChecking for Learner lesson builder is not displayed\n"
            un.LogoutAndcheckLesssonBuilderLinkForLearner(Email, Password, username, password, expectedValidationMsg)
            print "\n----Test Execution Completed----\n"
            
            
        except Exception as e:
            traceback.print_exc()
            print (e)
            raise Exception 
         
        finally:  
            second_sheet = book.sheet_by_name('Login_Credentials')
            cell = second_sheet.cell(1,1)
            url = cell.value
            driver.get(url)
            try:
                WebDriverWait(driver, 5).until(EC.alert_is_present())

                alert = driver.switch_to.alert
                alert.accept()
                print("alert accepted")
            except Exception:
                print("no alert")
            
            wait=WebDriverWait(driver, 60)
            unDropDown=wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")))
            driver.execute_script('arguments[0].click()',unDropDown)
            
            signOut=driver.find_element_by_xpath("html/body/div/div/div[1]/div[2]/div[2]/a")
            driver.execute_script('arguments[0].click()',signOut)
            
            
            
            element = wait.until(EC.presence_of_element_located((By.ID, "password")))
            
            
            time.sleep(3)
            if driver.title == "Grovo":
                print("Grovo Application URL Opened")
            else:
                raise Exception.message
    
            print "Grovo Sign-In page is displayed"
            
            print "Entering User name"
            driver.find_element_by_xpath(".//*[@id='username']").send_keys(username)
           
            print "Entering Password"
            element.send_keys(password)
            
            element.send_keys(Keys.TAB)
            print "Clicking on Sign_In button"
            driver.find_element_by_xpath("//*[@id='submitButton']").click()
            
            print "Successfully Logged Into Users account"
            time.sleep(3)
            
            try:
                d=DeleteLesson()
                d.lessonDelete(lessonName)
            except Exception:
                driver.get(url)
            
            
            
            
if __name__ == '__main__':
    
    btc=BaseTestClass()
    btc.UserLogin()
    
    les=LessonGetLinkBuilderToLearner()
    les.LessonGetlinkLessonBuilderForLearnerMain()
        
        
        
        
    