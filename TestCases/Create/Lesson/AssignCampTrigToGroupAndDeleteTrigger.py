'''
Created on 29-Mar-2018

@author: dattatraya
'''

import os.path
import time
import traceback

from BaseTestClass import BaseTestClass
from BaseTestClass import driver
from openpyxl.reader.excel import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import xlrd
from BaseTestClass import projectPath
from CampaignPageElements import CampPage
from CreateLearnerNew import CreateLearner
from CreateLessonDifferentLessons import CreateLessonDifferentLessons
from DeleteLesson import DeleteLesson
from PlanAssignmentForPageElements import PlanAssignmentForPageElements
from GroupsPageElements import GroupsPageElements
from AddTriggerForPageElements import AddTriggerPage
from BaseTestClass import projectPath


class AssignCampTrigToGroupAndDeleteTrigger:
    
    def campaignTrigToGroup(self,campaignTitle,campDescription,lessonName,groupName,minPassingScore,numberOfAttempts):
        
        elements=CampPage()
        
        wait=WebDriverWait(driver, 60)
        
        driver.refresh()
        print "\n\nCreating Campaign"
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.campaignButtonFromSideMenuXpath())))
        elements.campaignButtonFromSideMenu()
        
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.createCampaignButtonXpath())))

        if elements.campaignsPageHeaderText()=="Campaigns":
            print "Campaigns page displayed"
        else:
            print "Campaigns page is not displayed"
            raise Exception
        
        
        print "Clicking on Create Campaign button"
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.createCampaignButtonXpath())))
        elements.createCampaignButton()
        
        
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.Camp_titleXpath())))
        print "Create Campaign page is displayed"
        
                  
        elements.titleTextField(campaignTitle)
        print "Title entered"
        
        elements.descriptionField(campDescription)
        print "Description entered"
        
        print "Adding Lesson"
        #Add lesson button
        elements.addlessonButton()
        
        #Waiting until first lesson in pop is displayed
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.FirstLessonWaitXpath())))
        
        #Searching lesson by its name
        elements.searchLesson(lessonName)
        
        #Waiting until lesson displayed
        elements.waitUntilSearchedLessonDisplayed(lessonName)
        
        #selecting searched lesson
        elements.selectSearchedLesson(lessonName)
        
        #waiting until add to campaign button is click able
        wait.until(EC.element_to_be_clickable((By.XPATH,elements.AddToCampaign_ButtonXpath())))
        
        #Clicking on Add to Campaign button
        elements.addToCampaignButton()
        
        
        time.sleep(2)
        #Verifying Added lesson is displayed in Grid
        print "\nVerifying Added lesson is displayed in Grid"
        if elements.firstLessonInGrid()==lessonName:
            print "Lesson displayed in grid"
        else:
            print "Lesson not displayed in grid"
            raise Exception
        
        
        print "Making this as a Graded campaign"
        elements.makeThisAsAGradedCampaign()
        print "Settng minimum passing score"
        elements.setMinimumPassingScore(minPassingScore)
        print "Setting maximum no of attempts"
        elements.setAMaxNoOfAttempts(numberOfAttempts)
        
        
        wait.until(EC.element_to_be_clickable((By.XPATH,elements.SaveAndExit_ButtonXpath())))
        print "Clicking on Save and plan assignmet button"
        elements.saveAndPlanAssignmentbutton()
        
        
        
        #Verifying Plan assignment for page is displayed
        print "Verifying Plan assignment for page is displayed"
        if campaignTitle in elements.planAssignementForHeaderText():
            print "Page '"+elements.planAssignementForHeaderText()+"' is displayed"
        else:
            print "Plan assignment page is not displayed"
            raise Exception
        
        
        #Checking One time triggered radio button is selected by Default
        if driver.find_element_by_xpath("//input[@id='assignment-one-time']").is_selected():
            print "Radio button One Time triggered is selected"
        else:
            print "Radio button is not selected"
            raise Exception
        
        
        
        print "Clicking on Triggered radio button"
        
        planAssign=PlanAssignmentForPageElements()
        planAssign.triggredRadioButton()
        wait.until(EC.visibility_of_element_located((By.XPATH,planAssign.NewHireOnBoarding())))
        
        print "Clicked on Triggered"
        
        #Checking New hire on boarding is selected 
        if driver.find_element_by_xpath("//input[@id='trigger-new-hire']").is_selected():
            print "Radio button New Hire Onboarding is selected"
        else:
            print "Radio button New Hire Onboarding is not selected"
            raise Exception
        
        

        
        print "Selecting New to group"
        driver.find_element_by_xpath(planAssign.NewtoGroup()).click()
        wait.until(EC.visibility_of_element_located((By.XPATH,planAssign.AddGroupTextField()))) 
        
        searchGroups=driver.find_element_by_xpath(planAssign.AddGroupTextField())
        
        webdriver.ActionChains(driver).move_to_element(searchGroups).click().send_keys(groupName).perform()
        groupdisplayed=wait.until(EC.visibility_of_element_located((By.XPATH,"(//div[@role='option']/span)[1]")))
        
        webdriver.ActionChains(driver).move_to_element(groupdisplayed).click().perform()
        
        
        print "Checking Group is displayed in Grid"
        
        if driver.find_element_by_xpath("//table/tbody/tr/td[1]/a").text==groupName:
            print "Group '"+groupName+"' is displayed in grid"
        else:
            print "Group is not displayed in Grid"
            raise Exception
        
        
        print "Clicking on checkbox 'Also assign to learners who currently match this criteria' "
        
        driver.find_element_by_xpath(planAssign.CheckBoxCurrentUserMatching()).click()
        wait.until(EC.visibility_of_element_located((By.XPATH,planAssign.YourCriteriaMatchesBox())))
        
        
        print "Saving Trigger"
        planAssign.saveTrigger()
        
        
        wait.until(EC.visibility_of_element_located((By.XPATH,planAssign.confirmPopupYesSaveButton())))
        driver.find_element_by_xpath(planAssign.confirmPopupYesSaveButton()).click()
        
        print "Clicked on Yes,Save button from pop up"
        

        print "Checking for In Campaign details page Trigger is displayed"
        try:
            elements.TriggerDisplayedInGridForGroup(groupName)
            print "Trigger with group name is displayed in Campaign details page"
        except Exception:
            print "Trigger is not displayed"
            raise Exception
        
        
        #Verifying Campaign Detail page is displayed
        print "\nVerifying campaign detail page is displayed"
        
        if elements.campaignDetailPageHeaderText()==campaignTitle:
            print "Campaign detail page is displayed"
        else:
            print "Campaign detail page is not displayed"
            raise Exception
        
        #verifying in Campaigns displayed in Campaigns grid
        elements.searchingForlesson(campaignTitle)
        
        if elements.actualCampTitleINGrid()==campaignTitle:
            print "Campaign '"+campaignTitle+"' displayed in Grid"
        
        else:
            print "Campaign is not displayed in Grid"
            raise Exception
        
         
         
        print "\nVerifying 'Y' is displayed for created triggred campaign"
        hasTrigger=driver.find_element_by_xpath("(//tr/td[1]/a[.='"+campaignTitle+"']/../../td[2])[1]").text
        
        if hasTrigger=="Y":
            print "'Y' displayed in Has trigger column"
        else:
            print "Invalid data in Hastrigger column"
            
            
    
    def groupCreateForCampaign(self,groupName,FirstName):
        driver.refresh()
        wait=WebDriverWait(driver, 60)
        
        group=GroupsPageElements()
        
        group.adminSideMenuUnexpanded()
        print "Clicked on admin icon"
         
        group.groupSideMenuExpanded()  
        print "Clicked on Group icon"
        
        
        print "Checking Group page is displayed"
        if driver.find_element_by_xpath(group.groupPageHeader()).is_displayed():
            print "Group page is displayed successfully"
        else:
            print "Group page is not displayed"
            raise Exception
        
        
        group.createGroupButton()
        print "Clicked on Create Group button"
        
        group.enteringGroupname(groupName)
        
        print "Group Name entered....."
        
        group.nextButton()
        print "Clicked on Next button"
        
        wait.until(EC.visibility_of_element_located((By.XPATH,group.groupDetailPageHeader())))
        print "Checking group is created"
        
        time.sleep(4)
        headerText=driver.find_element_by_xpath(group.groupDetailPageHeader()).text
        
        if headerText==groupName:
            print "Group created successfully"
            
        else:
            print "Group not created"
            raise Exception
        
        
        print "Adding user to Group by name"
        group.addByNameButton()
        group.addByName(FirstName)
        
        print "Verifying user is added to group"
        group.groupAddedInList(FirstName)
        
        group.saveButton()
        print "Clicked on Save button"

        group.saveButtonPopup()
        print "Clicked on Save from popup"
        
        
        print "Checking Created Group is displayed in Grid"
        group.groupsLinkFromBreadCrumb()
        
        print "Searching Group"
        group.createdGroupDisplayedInGrid(groupName)
         
        
        driver.refresh()
    
        
       
    
    
    def deleteTrigger(self,campaignTitle,groupName):
        elements=CampPage()
        driver.refresh()
        wait=WebDriverWait(driver, 120)
        
        
        print "\n\nCreating Campaign"
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.campaignButtonFromSideMenuXpath())))
        elements.campaignButtonFromSideMenu()
        
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.createCampaignButtonXpath())))

        if elements.campaignsPageHeaderText()=="Campaigns":
            print "Campaigns page displayed"
        else:
            print "Campaigns page is not displayed"
            raise Exception
        
        print "Clicking on Created Campaign from Grid"
    
    
        elements.CampaignLinkFromCampaignGrid(campaignTitle)
        
        elements.TriggerDisplayedInGridForGroup(groupName)
        
        print "Click on Edit button"
        elements.editTriggerButtonForGroup(groupName)
        
        print "Clicking on Delete trigger"
        trig=AddTriggerPage()
        trig.deleteTriggerButtonEditTriggerPage()
        
        print "Clicking on Delete trigger from popup"
        trig.deleteTriggerFromPopup()
        elements.campaignDetailPageHeaderText()
        print "'"+elements.campaignDetailPageHeaderText()+"' is displayed"
        
        print "Trigger deleted"
        wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[3]/div[2]/div/div[3]/div/h3/span[2]")))
        
        
        driver.refresh()
        
        
        
        
    def newUserAddingToGroup(self,groupName,FirstName1):
        
        driver.refresh()
        wait=WebDriverWait(driver, 60)
        
        group=GroupsPageElements()
        
        group.adminSideMenuUnexpanded()
        print "Clicked on admin icon"
         
        group.groupSideMenuExpanded()  
        print "Clicked on Group icon"
        
        
        
        print "Checking Group page is displayed"
        if driver.find_element_by_xpath(group.groupPageHeader()).is_displayed():
            print "Group page is displayed successfully"
        else:
            print "Group page is not displayed"
            raise Exception
        
        wait.until(EC.visibility_of_element_located((By.XPATH,group.searchGroupTextFieldXpath())))
        driver.find_element_by_xpath(".//*[@id='search-groups']").send_keys(groupName)
        ele=wait.until(EC.visibility_of_element_located((By.XPATH,group.groupInGrid(groupName))))
        print "Group found"
        ele.click()
        
        
        print "Adding user to Group by name"
        group.addByNameButton()
        group.addByName(FirstName1)
        
        
        print "Checking user is added"
        group.groupAddedInList(FirstName1)
        
        print "Clicking on Save button"
        group.saveButton()


        group.saveButtonPopup()
        print "Clicked on Save from popup"
        
        
        print "New User is successfully added to group"
        
        driver.refresh()
        
     
     
    def checkDeletedTrigger(self,Email,Password,username,password,campaignTitle):
        wait=WebDriverWait(driver, 60)
        driver.refresh()
        print "Logging out as current user"
        print "Clicking on Username Drop down"
        
        ele=wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")))
        driver.execute_script('arguments[0].click()',ele)
        elem=driver.find_element_by_xpath("html/body/div/div/div[1]/div[2]/div[2]/a")
        driver.execute_script('arguments[0].click()',elem)
        
        print "Clicked on Sign Out option"
        
        element = wait.until(EC.presence_of_element_located((By.ID, "password")))
        
        if driver.title == "Grovo":
            print("Grovo Application URL Opened")
        else:
            raise Exception.message

        print "Grovo Sign-In page is displayed"
        
        print "Entering User name"
        driver.find_element_by_xpath(".//*[@id='username']").send_keys(Email)
       
        print "Entering Password"
        element.send_keys(Password)
        
        element.send_keys(Keys.TAB)
        print "Clicking on Sign_In button"
        driver.find_element_by_xpath("//*[@id='submitButton']").click()
        
        print "Successfully Logged Into Users account"
        time.sleep(3)
        
        ele3=wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")))
        driver.execute_script('arguments[0].click()',ele3)
        
        ele4 =driver.find_element_by_xpath(".//*[@id='content']/div/div[1]/div[2]/div[1]/a[1]")
        driver.execute_script('arguments[0].click()',ele4)
        time.sleep(6)
        listsCamps=wait.until(EC.visibility_of_all_elements_located((By.XPATH,"//div/div[1]/div/div/a[1]/span")))
        
        count = len(listsCamps)
        for count in range (0,count):
            if listsCamps[count].text!=campaignTitle:
                print "Campaign is not displayed"
            else:
                print "Campaign is displayed"
                raise Exception
                
        
        
        print "Logging in as a Original user"
        
        unDropDown=wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")))
        driver.execute_script('arguments[0].click()',unDropDown)
        
        signOut=driver.find_element_by_xpath("html/body/div/div/div[1]/div[2]/div[2]/a")
        driver.execute_script('arguments[0].click()',signOut)
        
        
        
        element = wait.until(EC.presence_of_element_located((By.ID, "password")))
        
        if driver.title == "Grovo":
            print("Grovo Application URL Opened")
        else:
            raise Exception.message

        print "Grovo Sign-In page is displayed"
        
        print "Entering User name"
        driver.find_element_by_xpath(".//*[@id='username']").send_keys(username)
       
        print "Entering Password"
        element.send_keys(password)
        
        element.send_keys(Keys.TAB)
        print "Clicking on Sign_In button"
        driver.find_element_by_xpath("//*[@id='submitButton']").click()
        
        print "Successfully Logged Into Users account"
        
        time.sleep(5) 
        
        
        
    def assignCampTriggredToGroupDeleteTriggerAndCheckForUser(self):
        
        book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        first_sheet = book.sheet_by_name('CampAssign')
        
        
        #Campaign
        cell1 = first_sheet.cell(342,1)
        campaignTitle = cell1.value
        
        cell1 = first_sheet.cell(343,1)
        campDescription = cell1.value
        
        
        #Group 1
        cell1 = first_sheet.cell(364,1)
        groupName = cell1.value
        
        groupNameId = groupName.split("_")
        emp = groupNameId[0]+"_"
        ids = groupNameId[1]
        empId = int(ids)+1
        GroupNameUpdated= emp+str(empId)
        
        
        
        cell1 = first_sheet.cell(344,1)
        minPassingScore = cell1.value
        
        cell1 = first_sheet.cell(345,1)
        numberOfAttempts = cell1.value
        
        cell1 = first_sheet.cell(346,1)
        lessonName = cell1.value
        
        cell2 = first_sheet.cell(347,1)
        questionCard= cell2.value
        
        cell2 = first_sheet.cell(348,1)
        ans1= cell2.value
        
        cell2 = first_sheet.cell(349,1)
        ans2= cell2.value
        
        
        
        
        
        #Learner 1
        cell = first_sheet.cell(351,1)
        FirstName = cell.value
        
        FirstNameId = FirstName.split("_")
        emp = FirstNameId[0]+"_"
        ids = FirstNameId[1]
        empId = int(ids)+1
        FirstNameUpdated= emp+str(empId)
        
        cell = first_sheet.cell(352,1)
        LastName = cell.value
        
        LastNameID = LastName.split("_")
        emp = LastNameID[0]+"_"
        ids = LastNameID[1]
        empId = int(ids)+1
        LastNameUpdated= emp+str(empId)
        
        
        
        cell = first_sheet.cell(353,1)
        Email = cell.value
        
        EmailId = Email.split("_")
        emp = EmailId[0]+"_"
        ids = EmailId[1]
        remaining="_"+EmailId[2]
        empId = int(ids)+1
        EmailIdUpdated= emp+str(empId)+remaining
        
        cell = first_sheet.cell(354,1)
        EmployeeId = cell.value
        
        Employee = EmployeeId.split("_")
        emp = Employee[0]+"_"
        ids = Employee[1]
        empId = int(ids)+1
        EmployeeIdUpdated= emp+str(empId)
        
        
        
        
        #Learner 2
        #Learner 1
        cell = first_sheet.cell(356,1)
        FirstName1 = cell.value
        
        FirstNameId = FirstName1.split("_")
        emp = FirstNameId[0]+"_"
        ids = FirstNameId[1]
        empId = int(ids)+1
        FirstNameUpdated1= emp+str(empId)
        
        cell = first_sheet.cell(357,1)
        LastName1 = cell.value
        
        LastNameID = LastName1.split("_")
        emp = LastNameID[0]+"_"
        ids = LastNameID[1]
        empId = int(ids)+1
        LastNameUpdated1= emp+str(empId)
        
        
        
        cell = first_sheet.cell(358,1)
        Email1 = cell.value
        
        EmailId = Email1.split("_")
        emp = EmailId[0]+"_"
        ids = EmailId[1]
        remaining="_"+EmailId[2]
        empId = int(ids)+1
        EmailIdUpdated1= emp+str(empId)+remaining
        
        cell = first_sheet.cell(359,1)
        EmployeeId1 = cell.value
        
        Employee = EmployeeId1.split("_")
        emp = Employee[0]+"_"
        ids = Employee[1]
        empId = int(ids)+1
        EmployeeIdUpdated1= emp+str(empId)
        
        
        cell = first_sheet.cell(360,1)
        Password = cell.value
        
        cell = first_sheet.cell(361,1)
        NewPassword = cell.value
        
        
        #For Original User
        book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        s_sheet = book.sheet_by_name('Login_Credentials')
        
        cell = s_sheet.cell(1,1)
        url = cell.value
        
        cell = s_sheet.cell(3,1)
        username = cell.value
        
        cell = s_sheet.cell(3,2)
        password = cell.value
        
        
        
        
        #updating user values
        wb = load_workbook(os.path.join('Test_Data/TestData.xlsx'))
            #print (wb.sheetnames)
        
        sheet = wb['CampAssign']
        
        sheet.cell(row=352, column=2).value = FirstNameUpdated
        sheet.cell(row=353, column=2).value = LastNameUpdated
        sheet.cell(row=354, column=2).value = EmailIdUpdated
        sheet.cell(row=355, column=2).value = EmployeeIdUpdated
        
        sheet.cell(row=365, column=2).value = GroupNameUpdated
        
        
        sheet.cell(row=357, column=2).value = FirstNameUpdated1
        sheet.cell(row=358, column=2).value = LastNameUpdated1
        sheet.cell(row=359, column=2).value = EmailIdUpdated1
        sheet.cell(row=360, column=2).value = EmployeeIdUpdated1
        
        
        wb.save(os.path.join('Test_Data/TestData.xlsx'))
            
        print "All User Data Updated in Excel"
        
        
        
        
        try:
            print "\nCreating a New Learner\n"
            lt=CreateLearner()
            lt.createNewLearnerMain(FirstName, LastName, Email, EmployeeId, Password, NewPassword, url, username, password)
            lt.createNewLearnerMain(FirstName1, LastName1, Email1, EmployeeId1, Password, NewPassword, url, username, password)
            
            
            print "\nCreating Group\n"
            lk=AssignCampTrigToGroupAndDeleteTrigger()
            lk.groupCreateForCampaign(groupName, FirstName)
            
            print "\nCreating a lesson\n"
            ot=CreateLessonDifferentLessons()
            ot.lessonWithQuestion(lessonName, questionCard, ans1, ans2)
            
            
            print "\nCreating Campaign\n"
            lk.campaignTrigToGroup(campaignTitle, campDescription, lessonName, groupName, minPassingScore, numberOfAttempts)
            
            print "\nDelete Trigger\n"
            lk.deleteTrigger(campaignTitle,groupName)
            
            print "\nNew User adding to group\n"
            lk.newUserAddingToGroup(groupName, FirstName1)
            
            print "\nCheck for New User Assignment is not displayed"
            lk.checkDeletedTrigger(Email, Password, username, password, campaignTitle)
            
            
            
            print "\n----Test Execution Completed----\n"
            
            
        except Exception as e:
            traceback.print_exc()
            print (e)
            raise Exception 
         
        finally: 
            driver.save_screenshot("ScreenShots/AssignCampTrigToGroupAndDeleteTrigger.png")
             
            second_sheet = book.sheet_by_name('Login_Credentials')
            cell = second_sheet.cell(1,1)
            url = cell.value
            driver.get(url)
            try:
                WebDriverWait(driver, 5).until(EC.alert_is_present())

                alert = driver.switch_to.alert
                alert.accept()
                print("alert accepted")
            except Exception:
                print("no alert")
                
            
            
            wait=WebDriverWait(driver, 60)
            unDropDown=wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")))
            driver.execute_script('arguments[0].click()',unDropDown)
            
            signOut=driver.find_element_by_xpath("html/body/div/div/div[1]/div[2]/div[2]/a")
            driver.execute_script('arguments[0].click()',signOut)
            
            
            
            element = wait.until(EC.presence_of_element_located((By.ID, "password")))
            
            if driver.title == "Grovo":
                print("Grovo Application URL Opened")
            else:
                raise Exception.message
    
            print "Grovo Sign-In page is displayed"
            
            print "Entering User name"
            driver.find_element_by_xpath(".//*[@id='username']").send_keys(username)
           
            print "Entering Password"
            element.send_keys(password)
            
            element.send_keys(Keys.TAB)
            print "Clicking on Sign_In button"
            driver.find_element_by_xpath("//*[@id='submitButton']").click()
            
            print "Successfully Logged Into Users account"
            time.sleep(3)
            
            try:
                oj=DeleteLesson()
                oj.lessonDelete(lessonName)
            except Exception:
                driver.get(url)
                
if __name__ == '__main__':
    
    btc=BaseTestClass()
    btc.UserLogin()
    
    ne=AssignCampTrigToGroupAndDeleteTrigger()
    ne.assignCampTriggredToGroupDeleteTriggerAndCheckForUser()
            
            
