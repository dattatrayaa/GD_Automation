'''
Created on 10-Apr-2018

@author: dattatraya
'''

import os.path
import time
import traceback

from BaseTestClass import BaseTestClass
from BaseTestClass import driver
from openpyxl.reader.excel import load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import xlrd
from BaseTestClass import projectPath
from CampaignPageElements import CampPage
from CreateLearnerNew import CreateLearner
from CreateLessonDifferentLessons import CreateLessonDifferentLessons
from DeleteLesson import DeleteLesson
from ManageAssignMentForPageElements import ManageAssignMentForPageElements
from BaseTestClass import projectPath

class AssignCampVerifyManageAssignmentForPage:
    
    def campaignToVerifyManageAssignPage(self,campaignTitle,campDescription,lessonName,nameOFuser,minPassingScore,numberOfAttempts):
        
        elements=CampPage()
        
        wait=WebDriverWait(driver, 60)
        
        driver.refresh()
        print "\n\nCreating Campaign"
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.campaignButtonFromSideMenuXpath())))
        elements.campaignButtonFromSideMenu()
        
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.createCampaignButtonXpath())))

        if elements.campaignsPageHeaderText()=="Campaigns":
            print "Campaigns page displayed"
        else:
            print "Campaigns page is not displayed"
            raise Exception
        
        
        print "Clicking on Create Campaign button"
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.createCampaignButtonXpath())))
        elements.createCampaignButton()
        
        
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.Camp_titleXpath())))
        print "Create Campaign page is displayed"
        
                  
        elements.titleTextField(campaignTitle)
        print "Title entered"
        
        elements.descriptionField(campDescription)
        print "Description entered"
        
        print "Adding Lesson"
        #Add lesson button
        elements.addlessonButton()
        
        #Waiting until first lesson in pop is displayed
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.FirstLessonWaitXpath())))
        
        #Searching lesson by its name
        elements.searchLesson(lessonName)
        
        #Waiting until lesson displayed
        elements.waitUntilSearchedLessonDisplayed(lessonName)
        
        #selecting searched lesson
        elements.selectSearchedLesson(lessonName)
        
        #waiting until add to campaign button is click able
        wait.until(EC.element_to_be_clickable((By.XPATH,elements.AddToCampaign_ButtonXpath())))
        
        #Clicking on Add to Campaign button
        elements.addToCampaignButton()
        
        
        time.sleep(2)
        #Verifying Added lesson is displayed in Grid
        print "\nVerifying Added lesson is displayed in Grid"
        if elements.firstLessonInGrid()==lessonName:
            print "Lesson displayed in grid"
        else:
            print "Lesson not displayed in grid"
            raise Exception
        
        
        print "Setting minimum passing score"
        elements.makeThisAsAGradedCampaign()
        elements.setMinimumPassingScore(minPassingScore)
        print "Setting maximum number of attempts"
        elements.setAMaxNoOfAttempts(numberOfAttempts)
        
        
        
        wait.until(EC.element_to_be_clickable((By.XPATH,elements.SaveAndExit_ButtonXpath())))
        print "Clicking on Save and plan assignmet button"
        elements.saveAndPlanAssignmentbutton()
        
        
        
        #Verifying Plan assignment for page is displayed
        print "Verifying Plan assignment for page is displayed"
        if campaignTitle in elements.planAssignementForHeaderText():
            print "Page '"+elements.planAssignementForHeaderText()+"' is displayed"
        else:
            print "Plan assignment page is not displayed"
            raise Exception
        
        
        #Checking One time triggered radio button is selected by Default
        if driver.find_element_by_xpath("//input[@id='assignment-one-time']").is_selected():
            print "Radio button One Time triggered is selected"
        else:
            print "Radio button is not selected"
            raise Exception
        
        time.sleep(3)
        print "Difference between two dates"
        from datetime import datetime
        date_format = "%m/%d/%Y"
        a = datetime.strptime(elements.assignDateValue(), date_format)
        b = datetime.strptime(elements.dueDateValue(), date_format)
        delta = b - a
        print "The difference between two dates is ::"
        print delta.days
        
        
        #Adding user to this campaign
        elements.addingUser(nameOFuser)
        
        
        print "Verifying User is displayed in Grid"
        if nameOFuser in elements.userInGrid():
            print "User is displayed in Grid"
        else:
            print "User is not displayed in grid"
            raise Exception        
        
        
        
        elements.sendAssignment()
        print "Clicked on Send assignment button"
        
        
        print "Verifying Confirm pop up is displayed"
        if elements.confirmPopupsendingAssignement().is_displayed():
            print "Confirm assignment pop up is displayed"
        else:
            print "Confirm assignment pop up is not displayed"
            raise Exception
        
        
        elements.sendAssignmentFromPopup()
        print "Clicked on Send Assignment button from pop up"
        
        
        
        #Verifying Campaign Detail page is displayed
        print "\nVerifying campaign detail page is displayed"
        
        if elements.campaignDetailPageHeaderText()==campaignTitle:
            print "Campaign detail page is displayed"
        else:
            print "Campaign detail page is not displayed"
            raise Exception
        
        #verifying in Campaigns displayed in Campaigns grid
        elements.searchingForlesson(campaignTitle)
        
        if elements.actualCampTitleINGrid()==campaignTitle:
            print "Campaign '"+campaignTitle+"' displayed in Grid"
        
        else:
            print "Campaign is not displayed in Grid"
            raise Exception
        
        
        
        
    def verifyingManageAssignmentForPage(self,campaignTitle,FirstName,LastName,manageAssignText,actualCancelAssignButtonTxt,actualCampaignText,actualAsssignedByText,actualAssignText,
                                         actualDueDateText,actualAllowedAttemptsText,actualAssignmentReportingText,actualGridHeaderText,actualAssignedToText,
                                         actualprogressText,actualCompletedOnText,actualAttemptNoText,actualAssignScoreText,actualActionsText):
        
        elements=CampPage()
        
        wait=WebDriverWait(driver, 60)
        
        driver.refresh()
        print "\n\nCreating Campaign"
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.campaignButtonFromSideMenuXpath())))
        elements.campaignButtonFromSideMenu()
        
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.createCampaignButtonXpath())))

        if elements.campaignsPageHeaderText()=="Campaigns":
            print "Campaigns page displayed"
        else:
            print "Campaigns page is not displayed"
            raise Exception
        
        print "Clicking on link of campaign created"
        elements.CampaignLinkFromCampaignGrid(campaignTitle)
        
        
        print "Verifying Campaign details page is displayed"
        if elements.campaignDetailPageHeaderText()==campaignTitle:
            print "Header Text '"+elements.campaignDetailPageHeaderText()+"' is displayed"
        else:
            print "Header Text not displayed"
            raise Exception
        
        
        time.sleep(2)
        print "Clicking on Learner name present in Assigned to column"
        learnerName=wait.until(EC.visibility_of_element_located((By.XPATH,"//tbody/tr/td[1]/a/span/span[contains(.,'"+FirstName+"')]")))
        learnerName.click()
        
        wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[3]/div[2]/div/div[2]/div/div/div/div/div[1]/span[1]/a[contains(.,'"+campaignTitle+"')]")))
        
        print "Verifying Manage Assignment for page is displayed"
        
        m=ManageAssignMentForPageElements()
        header=wait.until(EC.visibility_of_element_located((By.XPATH,m.HeaderTextManageAssign())))
        print header.text
        if "Manage assignment for: 1 learner" ==header.text:
            print "'"+header.text+"' is displayed in Header"
        else:
            print "Header Text is not displayed is not valid"
            raise Exception
        
        
        print "Verifying BreadCrumb"
        try:
            wait.until(EC.visibility_of_element_located((By.XPATH,"//div/div[1]/span[.='"+manageAssignText+"']/../a[1 and contains(.,'Campaigns')]/../a[2 and contains(.,'"+campaignTitle+"')]")))
            print "Breadcrumb 'Campaigns>"+campaignTitle+">"+manageAssignText+"' is displayed in header section"
        except Exception:
            print "Breadcrumb is not displayed properly"
            raise Exception
        
        
        
        
        
        print "\nVerifying Cancel Assignment button is displayed"
        cancelAssign=driver.find_element_by_xpath(".//*[@id='content']/div/div[3]/div[2]/div/header/div[2]/button")
        if cancelAssign.text==actualCancelAssignButtonTxt:
            print "'"+cancelAssign.text+"' is displayed"
        else:
            print "Cancel Assignment button is not displayed properly"
            raise Exception
        
        
        print "Verifying Campaign title is displayed in container"
        camp=driver.find_element_by_xpath(".//*[@id='content']/div/div[3]/div[2]/div/div[2]/div/div/div/div/div[1]/span[1]")
        if actualCampaignText in camp.text:
            print "'"+camp.text+"' is displayed"
        else:
            print "Campaign text is not displayed"
            raise Exception
        
        print "Verifying Assigned by text"
        assignedBy=wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[3]/div[2]/div/div[2]/div/div/div/div/div[1]/span[2]")))
        print assignedBy.text
        if actualAsssignedByText in assignedBy.text:
            print "'"+assignedBy.text+"' is displayed"
        else:
            print "Assigned by text is not displayed"
            raise Exception
        
        
        print "Verifying Assign date text is displayed"
        assignDate=driver.find_element_by_xpath("//ul/li[1]/small")
        if actualAssignText in assignDate.text:
            print "'"+assignDate.text+"' is displayed"
        else:
            print "Assign date text is not displayed"
            raise Exception
        
        
        print "Verifying Due date text is displayed"
        dueDate=driver.find_element_by_xpath("//ul/li[2]/small")
        if actualDueDateText in dueDate.text:
            print "'"+dueDate.text+"' is displayed"
        else:
            print "Due date text is not displayed"
            raise Exception
        
        
        print "Verifying Allowed attempts text is displayed"
        alAttempts=driver.find_element_by_xpath("//ul/li[3]/small")
        if actualAllowedAttemptsText in alAttempts.text:
            print "'"+alAttempts.text+"' is displayed"
        else:
            print "Allowed Attempts text is not displayed"
            raise Exception
        
        
        print "Verifying Assignment reporting text is displayed"
        assignmentReporting=driver.find_element_by_xpath("//ul/li[4]/small")
        if actualAssignmentReportingText in assignmentReporting.text:
            print "'"+assignmentReporting.text+"' is displayed"
        else:
            print "Assignment Reporting text is not displayed"
            raise Exception
        
        
        
        print "Verifying grid header Text is displayed"
        gridHeader=wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[3]/div[2]/div/div[3]/div/div/h3")))
        if gridHeader.text==actualGridHeaderText:
            print "'"+gridHeader.text+"' is displayed"
        else:
            print "Grid header text is displayed is not valid"
            raise Exception
        
        
        print "Verifying table headers"
        assignedTo=driver.find_element_by_xpath("//table/thead/tr/th[1]")
        if assignedTo.text==actualAssignedToText:
            print "'"+assignedTo.text+"' is displayed"
        else:
            print "Assigned to text is not valid"
            raise Exception
        
        
        progress=driver.find_element_by_xpath("//table/thead/tr/th[2]")
        if progress.text==actualprogressText:
            print "'"+progress.text+"' is displayed"
        else:
            print "progress text is not valid"
            raise Exception
        
        
        completedOn=driver.find_element_by_xpath("//table/thead/tr/th[3]")
        if completedOn.text==actualCompletedOnText:
            print "'"+completedOn.text+"' is displayed"
        else:
            print "Completed On text is not valid"
            raise Exception
        
        
        attemptNo=driver.find_element_by_xpath("//table/thead/tr/th[4]")
        if attemptNo.text==actualAttemptNoText:
            print "'"+attemptNo.text+"' is displayed"
        else:
            print "Attempt no text is not valid"
            raise Exception
        
        
        assigScore=driver.find_element_by_xpath("//table/thead/tr/th[5]")
        if assigScore.text==actualAssignScoreText:
            print "'"+assigScore.text+"' is displayed"
        else:
            print "Assignment score text is not valid"
            raise Exception
        
        actions=driver.find_element_by_xpath("//table/thead/tr/th[6]")
        if actions.text==actualActionsText:
            print "'"+actions.text+"' is displayed"
        else:
            print "Actions text is not valid"
            raise Exception
        
        print "Verifying Cancel button is displayed in Actions column"
        cancelButton=driver.find_element_by_xpath("//table/tbody/tr/td[6]/button[.='Cancel']")
        if cancelButton.is_displayed():
            print "'"+cancelButton.text+"' is displayed in Actions column"
        else:
            print "Cancel button is not displayed"
            raise Exception
        
        
        print "All elements are displayed in Manage assignments for page"
        
        driver.refresh()
        
          
    
    
        
        
    def assignCampToVerifyManageAssignmentForPage(self):
        
        book = xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        first_sheet = book.sheet_by_name('CampAssign')
        
        
        #Campaign
        cell1 = first_sheet.cell(561,1)
        campaignTitle = cell1.value
        
        cell1 = first_sheet.cell(562,1)
        campDescription = cell1.value
        
        cell1 = first_sheet.cell(563,1)
        minPassingScore = cell1.value
        
        cell1 = first_sheet.cell(564,1)
        numberOfAttempts = cell1.value
        
        cell1 = first_sheet.cell(565,1)
        lessonName = cell1.value
        
        cell2 = first_sheet.cell(566,1)
        questionCard= cell2.value
        
        cell2 = first_sheet.cell(567,1)
        ans1= cell2.value
        
        cell2 = first_sheet.cell(568,1)
        ans2= cell2.value
        
        
        
        #Learner 1
        cell = first_sheet.cell(570,1)
        FirstName = cell.value
        
        FirstNameId = FirstName.split("_")
        emp = FirstNameId[0]+"_"
        ids = FirstNameId[1]
        empId = int(ids)+1
        FirstNameUpdated= emp+str(empId)
        
        cell = first_sheet.cell(571,1)
        LastName = cell.value
        
        LastNameID = LastName.split("_")
        emp = LastNameID[0]+"_"
        ids = LastNameID[1]
        empId = int(ids)+1
        LastNameUpdated= emp+str(empId)
        
        
        
        cell = first_sheet.cell(572,1)
        Email = cell.value
        
        EmailId = Email.split("_")
        emp = EmailId[0]+"_"
        ids = EmailId[1]
        remaining="_"+EmailId[2]
        empId = int(ids)+1
        EmailIdUpdated= emp+str(empId)+remaining
        
        cell = first_sheet.cell(573,1)
        EmployeeId = cell.value
        
        Employee = EmployeeId.split("_")
        emp = Employee[0]+"_"
        ids = Employee[1]
        empId = int(ids)+1
        EmployeeIdUpdated= emp+str(empId)
        
        
        cell = first_sheet.cell(574,1)
        Password = cell.value
        
        cell = first_sheet.cell(575,1)
        NewPassword = cell.value
        
        
        
        #Ui elements for Manage Assignment for page
        cell = first_sheet.cell(577,1)
        manageAssignText = cell.value
        cell = first_sheet.cell(578,1)
        actualCancelAssignButtonTxt = cell.value
        cell = first_sheet.cell(579,1)
        actualCampaignText = cell.value
        cell = first_sheet.cell(580,1)
        actualAsssignedByText = cell.value
        cell = first_sheet.cell(581,1)
        actualAssignText = cell.value
        cell = first_sheet.cell(582,1)
        actualDueDateText = cell.value
        cell = first_sheet.cell(583,1)
        actualAllowedAttemptsText = cell.value
        cell = first_sheet.cell(584,1)
        actualAssignmentReportingText = cell.value
        cell = first_sheet.cell(585,1)
        actualGridHeaderText = cell.value
        cell = first_sheet.cell(586,1)
        actualAssignedToText = cell.value
        cell = first_sheet.cell(587,1)
        actualprogressText = cell.value
        cell = first_sheet.cell(588,1)
        actualCompletedOnText = cell.value
        cell = first_sheet.cell(589,1)
        actualAttemptNoText = cell.value
        cell = first_sheet.cell(590,1)
        actualAssignScoreText = cell.value
        cell = first_sheet.cell(591,1)
        actualActionsText = cell.value
        
        
        
        
        
        
        
        #For Original User
        book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        s_sheet = book.sheet_by_name('Login_Credentials')
        
        cell = s_sheet.cell(1,1)
        url = cell.value
        
        cell = s_sheet.cell(3,1)
        username = cell.value
        
        cell = s_sheet.cell(3,2)
        password = cell.value
        
        
        
        
        #updating user values
        wb = load_workbook(os.path.join('Test_Data/TestData.xlsx'))
            #print (wb.sheetnames)
        
        sheet = wb['CampAssign']
        
        sheet.cell(row=571, column=2).value = FirstNameUpdated
        sheet.cell(row=572, column=2).value = LastNameUpdated
        sheet.cell(row=573, column=2).value = EmailIdUpdated
        sheet.cell(row=574, column=2).value = EmployeeIdUpdated
        
        
        
        wb.save(os.path.join('Test_Data/TestData.xlsx'))
            
        print "All User Data Updated in Excel"
        
        
        
        
        try:
            print "\nCreating a New Learner\n"
            lt=CreateLearner()
            lt.createNewLearnerMain(FirstName, LastName, Email, EmployeeId, Password, NewPassword, url, username, password)
            
            
            print "\nCreating Group\n"
            lk=AssignCampVerifyManageAssignmentForPage()
            
            print "\nCreating a lesson\n"
            ot=CreateLessonDifferentLessons()
            ot.lessonWithQuestion(lessonName, questionCard, ans1, ans2)
            
            
            print "\nCreating Campaign\n"
            lk.campaignToVerifyManageAssignPage(campaignTitle, campDescription, lessonName, FirstName, minPassingScore, numberOfAttempts)
            
            
            print "\nVerifying Manage Assignment for page\n"
            lk.verifyingManageAssignmentForPage(campaignTitle, FirstName, LastName, manageAssignText,
                                                 actualCancelAssignButtonTxt, actualCampaignText, 
                                                 actualAsssignedByText, actualAssignText, actualDueDateText, 
                                                 actualAllowedAttemptsText, actualAssignmentReportingText, actualGridHeaderText, 
                                                 actualAssignedToText, actualprogressText, actualCompletedOnText, actualAttemptNoText, 
                                                 actualAssignScoreText, actualActionsText)
            
            
            
            
            
            print "\n----Test Execution Completed----\n"
            
            
        except Exception as e:
            traceback.print_exc()
            print (e)
            raise Exception 
         
        finally:
            driver.save_screenshot("ScreenShots/AssignCampVerifyManageAssignmentForPage.png") 
            second_sheet = book.sheet_by_name('Login_Credentials')
            cell = second_sheet.cell(1,1)
            url = cell.value
            driver.get(url)
            try:
                WebDriverWait(driver, 5).until(EC.alert_is_present())

                alert = driver.switch_to.alert
                alert.accept()
                print("alert accepted")
            except Exception:
                print("no alert")
                
            
            
            wait=WebDriverWait(driver, 60)
            unDropDown=wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")))
            driver.execute_script('arguments[0].click()',unDropDown)
            
            signOut=driver.find_element_by_xpath("html/body/div/div/div[1]/div[2]/div[2]/a")
            driver.execute_script('arguments[0].click()',signOut)
            
            
            
            element = wait.until(EC.presence_of_element_located((By.ID, "password")))
            
            if driver.title == "Grovo":
                print("Grovo Application URL Opened")
            else:
                raise Exception.message
    
            print "Grovo Sign-In page is displayed"
            
            print "Entering User name"
            driver.find_element_by_xpath(".//*[@id='username']").send_keys(username)
           
            print "Entering Password"
            element.send_keys(password)
            
            element.send_keys(Keys.TAB)
            print "Clicking on Sign_In button"
            driver.find_element_by_xpath("//*[@id='submitButton']").click()
            
            print "Successfully Logged Into Users account"
            time.sleep(3)
            
            try:
                d=DeleteLesson()
                d.lessonDelete(lessonName)
            except Exception:
                driver.get(url) 
                
if __name__ == '__main__':
    
    btc=BaseTestClass()
    btc.UserLogin()
    
    ne=AssignCampVerifyManageAssignmentForPage()
    ne.assignCampToVerifyManageAssignmentForPage()
            
            
            
