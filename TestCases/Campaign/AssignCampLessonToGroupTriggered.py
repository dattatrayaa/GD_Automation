'''
Created on 21-Mar-2018

@author: dattatraya
'''
import os.path
import time
import traceback

from BaseTestClass import BaseTestClass
from BaseTestClass import driver
from openpyxl.reader.excel import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import xlrd

from CampaignPageElements import CampPage
from CreateLearnerNew import CreateLearner
from CreateLessonDifferentLessons import CreateLessonDifferentLessons
from PlanAssignmentForPageElements import PlanAssignmentForPageElements
from DeleteLesson import DeleteLesson
from GroupsPageElements import GroupsPageElements
from BaseTestClass import projectPath

class AssignCampLessonToGroupTriggered:
    
    def campaignToGroupTriggered(self,campaignTitle,campDescription,lessonName,groupName,minPassingScore,numberOfAttempts):
        
        elements=CampPage()
        
        wait=WebDriverWait(driver, 60)
        
        driver.refresh()
        print "\n\nCreating Campaign"
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.campaignButtonFromSideMenuXpath())))
        elements.campaignButtonFromSideMenu()
        
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.createCampaignButtonXpath())))

        if elements.campaignsPageHeaderText()=="Campaigns":
            print "Campaigns page displayed"
        else:
            print "Campaigns page is not displayed"
            raise Exception
        
        
        print "Clicking on Create Campaign button"
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.createCampaignButtonXpath())))
        elements.createCampaignButton()
        
        
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.Camp_titleXpath())))
        print "Create Campaign page is displayed"
        
                  
        elements.titleTextField(campaignTitle)
        print "Title entered"
        
        elements.descriptionField(campDescription)
        print "Description entered"
        
        print "Adding Lesson"
        #Add lesson button
        elements.addlessonButton()
        
        #Waiting until first lesson in pop is displayed
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.FirstLessonWaitXpath())))
        
        #Searching lesson by its name
        elements.searchLesson(lessonName)
        
        #Waiting until lesson displayed
        elements.waitUntilSearchedLessonDisplayed(lessonName)
        
        #selecting searched lesson
        elements.selectSearchedLesson(lessonName)
        
        #waiting until add to campaign button is click able
        wait.until(EC.element_to_be_clickable((By.XPATH,elements.AddToCampaign_ButtonXpath())))
        
        #Clicking on Add to Campaign button
        elements.addToCampaignButton()
        
        
        time.sleep(2)
        #Verifying Added lesson is displayed in Grid
        print "\nVerifying Added lesson is displayed in Grid"
        if elements.firstLessonInGrid()==lessonName:
            print "Lesson displayed in grid"
        else:
            print "Lesson not displayed in grid"
            raise Exception
        
        
        print "Making this as a Graded campaign"
        elements.makeThisAsAGradedCampaign()
        print "Settng minimum passing score"
        elements.setMinimumPassingScore(minPassingScore)
        print "Setting maximum no of attempts"
        elements.setAMaxNoOfAttempts(numberOfAttempts)
        
        
        wait.until(EC.element_to_be_clickable((By.XPATH,elements.SaveAndExit_ButtonXpath())))
        print "Clicking on Save and plan assignmet button"
        elements.saveAndPlanAssignmentbutton()
        
        
        
        #Verifying Plan assignment for page is displayed
        print "Verifying Plan assignment for page is displayed"
        if campaignTitle in elements.planAssignementForHeaderText():
            print "Page '"+elements.planAssignementForHeaderText()+"' is displayed"
        else:
            print "Plan assignment page is not displayed"
            raise Exception
        
        
        #Checking One time triggered radio button is selected by Default
        if driver.find_element_by_xpath("//input[@id='assignment-one-time']").is_selected():
            print "Radio button One Time triggered is selected"
        else:
            print "Radio button is not selected"
            raise Exception
        
        
        
        print "Clicking on Triggered radio button"
        planAssignPage=PlanAssignmentForPageElements()
        
        planAssignPage.triggredRadioButton()
        
        wait.until(EC.visibility_of_element_located((By.XPATH,planAssignPage.NewHireOnBoarding())))
        
        print "Clicked on Triggered"
        
        #Checking New hire on boarding is selected 
        if driver.find_element_by_xpath("//input[@id='trigger-new-hire']").is_selected():
            print "Radio button New Hire Onboarding is selected"
        else:
            print "Radio button New Hire Onboarding is not selected"
            raise Exception
        
        

        
        print "Selecting New to group"
        driver.find_element_by_xpath(planAssignPage.NewtoGroup()).click()
        wait.until(EC.visibility_of_element_located((By.XPATH,planAssignPage.AddGroupTextField()))) 
        
        searchGroups=driver.find_element_by_xpath(planAssignPage.AddGroupTextField())
        
        webdriver.ActionChains(driver).move_to_element(searchGroups).click().send_keys(groupName).perform()
        groupdisplayed=wait.until(EC.visibility_of_element_located((By.XPATH,"(//div[@role='option']/span)[1]")))
        
        webdriver.ActionChains(driver).move_to_element(groupdisplayed).click().perform()
        
        
        print "Checking Group is displayed in Grid"
        
        if driver.find_element_by_xpath("//table/tbody/tr/td[1]/a").text==groupName:
            print "Group '"+groupName+"' is displayed in grid"
        else:
            print "Group is not displayed in Grid"
            raise Exception
        
        
        print "Clicking on checkbox 'Also assign to learners who currently match this criteria' "
        
        driver.find_element_by_xpath(planAssignPage.CheckBoxCurrentUserMatching()).click()
        wait.until(EC.visibility_of_element_located((By.XPATH,planAssignPage.YourCriteriaMatchesBox())))
        
        
        
        print "Saving Trigger"
        planAssignPage.saveTrigger()
        
        
        wait.until(EC.visibility_of_element_located((By.XPATH,planAssignPage.confirmPopupYesSaveButton())))
        driver.find_element_by_xpath(planAssignPage.confirmPopupYesSaveButton()).click()
        
        print "Clicked on Yes,Save button from pop up"
        
        print "Checking for In Campaign details page Trigger is displayed"
        try:
            elements.TriggerDisplayedInGridForGroup(groupName)
            print "Trigger with group name '"+groupName+"' is displayed in Campaign details page"
        except Exception:
            print "Trigger is not displayed"
            raise Exception
        
        
        #Verifying Campaign Detail page is displayed
        print "\nVerifying campaign detail page is displayed"
        
        if elements.campaignDetailPageHeaderText()==campaignTitle:
            print "Campaign detail page is displayed"
        else:
            print "Campaign detail page is not displayed"
            raise Exception
        
        #verifying in Campaigns displayed in Campaigns grid
        elements.searchingForlesson(campaignTitle)
        
        if elements.actualCampTitleINGrid()==campaignTitle:
            print "Campaign '"+campaignTitle+"' displayed in Grid"
        
        else:
            print "Campaign is not displayed in Grid"
            raise Exception
        
         
         
        print "\nVerifying 'Y' is displayed for created triggred campaign"
        hasTrigger=driver.find_element_by_xpath("(//tr/td[1]/a[.='"+campaignTitle+"']/../../td[2])[1]").text
        
        if hasTrigger=="Y":
            print "'Y' displayed in Has trigger column"
        else:
            print "Invalid data in Hastrigger column"
            
            
            
    def groupCreateForCampaign(self,groupName,FirstName):
        
        driver.refresh()
        wait=WebDriverWait(driver, 60)
        
        group=GroupsPageElements()
        
        group.adminSideMenuUnexpanded()
        print "Clicked on admin icon"
         
        group.groupSideMenuExpanded()  
        print "Clicked on Group icon"
        
        
        print "Checking Group page is displayed"
        if driver.find_element_by_xpath(group.groupPageHeader()).is_displayed():
            print "Group page is displayed successfully"
        else:
            print "Group page is not displayed"
            raise Exception
        
        
        group.createGroupButton()
        print "Clicked on Create Group button"
        
        group.enteringGroupname(groupName)
        
        print "Group Name entered....."
        
        group.nextButton()
        print "Clicked on Next button"
        
        wait.until(EC.visibility_of_element_located((By.XPATH,group.groupDetailPageHeader())))
        print "Checking group is created"
        
        time.sleep(4)
        headerText=driver.find_element_by_xpath(group.groupDetailPageHeader()).text
        
        if headerText==groupName:
            print "Group created successfully"
            
        else:
            print "Group not created"
            raise Exception
        
        
        print "Adding user to Group by name"
        group.addByNameButton()
        group.addByName(FirstName)
        
        print "Verifying user is added to group"
        group.groupAddedInList(FirstName)
        
        group.saveButton()
        print "Clicked on Save button"

        group.saveButtonPopup()
        print "Clicked on Save from popup"
        
        
        print "Checking Created Group is displayed in Grid"
        group.groupsLinkFromBreadCrumb()
        
        print "Searching Group"
        group.createdGroupDisplayedInGrid(groupName)
         
        
        driver.refresh()
    
    
    def assignCampWithOneLessonForGroupTriggered(self):
        
        book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        first_sheet = book.sheet_by_name('CampAssign')
        
        
        #Campaign
        cell1 = first_sheet.cell(208,1)
        campaignTitle = cell1.value
        
        cell1 = first_sheet.cell(209,1)
        campDescription = cell1.value
        
        cell1 = first_sheet.cell(225,1)
        groupName = cell1.value
        
        groupNameId = groupName.split("_")
        emp = groupNameId[0]+"_"
        ids = groupNameId[1]
        empId = int(ids)+1
        GroupNameUpdated= emp+str(empId)
        
        cell1 = first_sheet.cell(210,1)
        minPassingScore = cell1.value
        
        cell1 = first_sheet.cell(211,1)
        numberOfAttempts = cell1.value
        
        cell1 = first_sheet.cell(212,1)
        lessonName = cell1.value
        
        cell2 = first_sheet.cell(213,1)
        questionCard= cell2.value
        
        cell2 = first_sheet.cell(214,1)
        ans1= cell2.value
        
        cell2 = first_sheet.cell(215,1)
        ans2= cell2.value
        
        
        
        
        
        #Learner
        cell = first_sheet.cell(217,1)
        FirstName = cell.value
        
        FirstNameId = FirstName.split("_")
        emp = FirstNameId[0]+"_"
        ids = FirstNameId[1]
        empId = int(ids)+1
        FirstNameUpdated= emp+str(empId)
        
        cell = first_sheet.cell(218,1)
        LastName = cell.value
        
        LastNameID = LastName.split("_")
        emp = LastNameID[0]+"_"
        ids = LastNameID[1]
        empId = int(ids)+1
        LastNameUpdated= emp+str(empId)
        
        
        
        cell = first_sheet.cell(219,1)
        Email = cell.value
        
        EmailId = Email.split("_")
        emp = EmailId[0]+"_"
        ids = EmailId[1]
        remaining="_"+EmailId[2]
        empId = int(ids)+1
        EmailIdUpdated= emp+str(empId)+remaining
        
        cell = first_sheet.cell(220,1)
        EmployeeId = cell.value
        
        Employee = EmployeeId.split("_")
        emp = Employee[0]+"_"
        ids = Employee[1]
        empId = int(ids)+1
        EmployeeIdUpdated= emp+str(empId)
        
        cell = first_sheet.cell(221,1)
        Password = cell.value
        
        cell = first_sheet.cell(222,1)
        NewPassword = cell.value
        
        
        #For Original User
        book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        s_sheet = book.sheet_by_name('Login_Credentials')
        
        cell = s_sheet.cell(1,1)
        url = cell.value
        
        cell = s_sheet.cell(3,1)
        username = cell.value
        
        cell = s_sheet.cell(3,2)
        password = cell.value
        
        #updating user values
        wb = load_workbook(os.path.join('Test_Data/TestData.xlsx'))
            #print (wb.sheetnames)
        
        sheet = wb['CampAssign']
        
        sheet.cell(row=218, column=2).value = FirstNameUpdated
        sheet.cell(row=219, column=2).value = LastNameUpdated
        sheet.cell(row=220, column=2).value = EmailIdUpdated
        sheet.cell(row=221, column=2).value = EmployeeIdUpdated
        sheet.cell(row=226, column=2).value = GroupNameUpdated
        
        
        wb.save(os.path.join('Test_Data/TestData.xlsx'))
            
        print "All User Data Updated in Excel"
        
        
        
        
        try:
            print "\nCreating a New Learner\n"
            lt=CreateLearner()
            lt.createNewLearnerMain(FirstName, LastName, Email, EmployeeId, Password, NewPassword, url, username, password)
            
            
            print "\nCreating Group\n"
            lk=AssignCampLessonToGroupTriggered()
            lk.groupCreateForCampaign(groupName, FirstName)
            
            print "\nCreating a lesson\n"
            ot=CreateLessonDifferentLessons()
            ot.lessonWithQuestion(lessonName, questionCard, ans1, ans2)
            
            
            print "\nCreating Campaign\n"
            
            lk.campaignToGroupTriggered(campaignTitle, campDescription, lessonName, groupName, minPassingScore, numberOfAttempts)
            
            
            
            print "\n----Test Execution Completed----\n"
            
            
        except Exception as e:
            traceback.print_exc()
            print (e)
            raise Exception 
         
        finally:
            driver.save_screenshot("ScreenShots/AssignCampLessonToGroupTriggered.png") 
            second_sheet = book.sheet_by_name('Login_Credentials')
            cell = second_sheet.cell(1,1)
            url = cell.value
            driver.get(url)
            
            try:
                WebDriverWait(driver, 5).until(EC.alert_is_present())

                alert = driver.switch_to.alert
                alert.accept()
                print("alert accepted")
            except Exception:
                print("no alert")
                
            
            
            wait=WebDriverWait(driver, 60)
            unDropDown=wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")))
            driver.execute_script('arguments[0].click()',unDropDown)
            
            signOut=driver.find_element_by_xpath("html/body/div/div/div[1]/div[2]/div[2]/a")
            driver.execute_script('arguments[0].click()',signOut)
            
            
            
            element = wait.until(EC.presence_of_element_located((By.ID, "password")))
            
            if driver.title == "Grovo":
                print("Grovo Application URL Opened")
            else:
                raise Exception.message
    
            print "Grovo Sign-In page is displayed"
            
            print "Entering User name"
            driver.find_element_by_xpath(".//*[@id='username']").send_keys(username)
           
            print "Entering Password"
            element.send_keys(password)
            
            element.send_keys(Keys.TAB)
            print "Clicking on Sign_In button"
            driver.find_element_by_xpath("//*[@id='submitButton']").click()
            
            print "Successfully Logged Into Users account"
            time.sleep(3)
            
            
            try:
                d=DeleteLesson()
                d.lessonDelete(lessonName) 
            except Exception:
                driver.get(url)

        
if __name__ == '__main__':
    
    btc=BaseTestClass()
    btc.UserLogin()
    
    hk=AssignCampLessonToGroupTriggered()
    hk.assignCampWithOneLessonForGroupTriggered()
    
        