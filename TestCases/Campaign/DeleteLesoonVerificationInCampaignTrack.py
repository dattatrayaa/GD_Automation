'''
Created on 25-Jul-2018

@author: Sheethu C
'''
import os.path
import time
import traceback
from BaseTestClass import BaseTestClass
from BaseTestClass import driver
from openpyxl.reader.excel import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import xlrd
from BaseTestClass import projectPath
from CampaignPageElements import CampPage
from CreateLessonDifferentLessons import CreateLessonDifferentLessons
from DeleteLesson import DeleteLesson
from GroupsPageElements import GroupsPageElements
from AddTriggerForPageElements import AddTriggerPage
from BaseTestClass import projectPath
from CreateTrackComman import CreateTrackComman
class DeleteLesoonVerificationInCampaignTrack:
    def campWithTrackTextlesson(self,campaignTitle,campDescription,trackName):
        elements=CampPage()
        
        wait=WebDriverWait(driver, 60)
        
        print "\n\nCreating Campaign"
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.campaignButtonFromSideMenuXpath())))
        elements.campaignButtonFromSideMenu()
        
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.createCampaignButtonXpath())))

        if elements.campaignsPageHeaderText()=="Campaigns":
            print "Campaigns page displayed"
        else:
            print "Campaigns page is not displayed"
            raise Exception
        
        
        print "Clicking on Create Campaign button"
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.createCampaignButtonXpath())))
        elements.createCampaignButton()
        
        
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.Camp_titleXpath())))
        print "Create Campaign page is displayed"
        
                  
        elements.titleTextField(campaignTitle)
        print "Title entered ::campTitle"
        
        elements.descriptionField(campDescription)
        print "Description entered ::campDescription"
        
        print "Adding Track"
        #Add Track 
        
        elements.addTrackButton()
        
        #Searching track and adding
        elements.searchTracksAndSelect(trackName)
        
        #checking lesson count is 0
        time.sleep(6)
        ct=driver.find_element_by_xpath("//li/div[2]/h4[.='"+trackName+"']/../../div[2]/small").text
        count =ct.split(' ')
        print count[0]
        if count[0] == str(0):
            print "lesson deleted"
        else:
            raise Exception
        #Adding to Campaign
        elements.addToCampaignTrack()
        
        
        #Verifying Added Track is displayed in Grid
        print "\nVerifying Added Track is displayed in Grid"
        if trackName in elements.firstTrackInGrid():
            print "Track displayed in grid"
        else:
            print "Track not displayed in grid"
        
        wait.until(EC.element_to_be_clickable((By.XPATH,elements.SaveAndExit_ButtonXpath())))
        #Clicking on save & exit button
        print "Clicking on save & exit button"
        elements.saveAndExitButton()
        
        #Verifying campaign detail page is displayed
        print "\nVerifying campaign detail page is displayed"
        
        if elements.campaignDetailPageHeaderText()==campaignTitle:
            print "Campaign detail page is displayed"
        else:
            print "Campaign detail page is not displayed"
            raise Exception
        
        #verifying in Campaigns displayed in Campaigns grid
        elements.searchingForlesson(campaignTitle)
        
        if elements.actualCampTitleINGrid()==campaignTitle:
            print "Campaign '"+campaignTitle+"' displayed in Grid"
        
        else:
            print "Campaign is not displayed in Grid"
            raise Exception
        
        print "\n----Text Execution Completed----\n"
    def mainDeleteLesoonVerificationInCampaignTrack(self):
        try:
            book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
            first_sheet = book.sheet_by_name("CampAssign")
            
            #lesson Creation
            
            cell = first_sheet.cell(861,1)
            lessonName = cell.value
            
            cell = first_sheet.cell(862,1)
            questionCard = cell.value
            
            cell = first_sheet.cell(863,1)
            ans1 = cell.value
            
            cell = first_sheet.cell(864,1)
            ans2 = cell.value
            
            #track Creation
            
            cell = first_sheet.cell(865,1)
            titleOfTrack = cell.value
            
            cell = first_sheet.cell(866,1)
            Imagefilepath = cell.value
            
            cell = first_sheet.cell(867,1)
            description = cell.value
            
            cell = first_sheet.cell(868,1)
            tagName = cell.value
            
            cell = first_sheet.cell(869,1)
            expectedSuccessText = cell.value
            
            
            #campaign Creation
            
            cell = first_sheet.cell(870,1)
            campaignTitle = cell.value
            
            cell = first_sheet.cell(871,1)
            campDescription = cell.value
            
            cell = first_sheet.cell(872,1)
            numberOfAttempts = cell.value
            
            cell = first_sheet.cell(873,1)
            minPassingScore = cell.value
            
            
            
        
            book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
            second_sheet = book.sheet_by_name('Login_Credentials')
            cell = second_sheet.cell(1,1)
            url = cell.value
        
            cell = second_sheet.cell(3,1)
            username = cell.value
        
            cell = second_sheet.cell(3,2)
            password = cell.value
            
            print "\nCreating a lesson\n"
            #ot=CreateLessonDifferentLessons()
            #ot.lessonWithQuestion(lessonName, questionCard, ans1, ans2)
            
            print "Create Track"
            #obj1=CreateTrackComman()
            #obj1.createTrack(titleOfTrack, Imagefilepath, description, tagName, lessonName, expectedSuccessText) 
            
            print "Delete lesson"
            #d=DeleteLesson()
            #d.lessonDelete(lessonName)  
            
            print "Creating Campaign"
            
            obj2=DeleteLesoonVerificationInCampaignTrack()
            obj2.campWithTrackTextlesson(campaignTitle, campDescription, titleOfTrack)
            
        except Exception as e:
            traceback.print_exc()
            print (e)
            raise Exception   
          
        finally:
            
            second_sheet = book.sheet_by_name('Login_Credentials')
            cell = second_sheet.cell(1,1)
            url = cell.value
            driver.get(url)
            try:
                WebDriverWait(driver, 5).until(EC.alert_is_present())

                alert = driver.switch_to.alert
                alert.accept()
                print("alert accepted")
            except Exception:
                print("no alert")
    
        
        
        
if __name__=='__main__':
    
    btc=BaseTestClass()
    btc.userLogin() 

    gr=DeleteLesoonVerificationInCampaignTrack()
    gr.mainDeleteLesoonVerificationInCampaignTrack() 
    
    
    
    
    
    
    
    
    
    
    
    
    