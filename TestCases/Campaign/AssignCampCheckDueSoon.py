'''
Created on 11-Apr-2018

@author: dattatraya
'''
import os.path
import time
import traceback

from BaseTestClass import BaseTestClass
from BaseTestClass import driver
from openpyxl.reader.excel import load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import xlrd

from CampaignPageElements import CampPage
from CreateLearnerNew import CreateLearner
from CreateLessonDifferentLessons import CreateLessonDifferentLessons
from DeleteLesson import DeleteLesson
from PlanAssignmentForPageElements import PlanAssignmentForPageElements
from selenium import webdriver
from BaseTestClass import projectPath
class AssignCampCheckDueSoon:
    
    
    def campaignToCheckDueSoon(self,campaignTitle,campDescription,lessonName,FirstName,minPassingScore,numberOfAttempts):
        
        elements=CampPage()
        
        wait=WebDriverWait(driver, 60)
        
        driver.refresh()
        print "\n\nCreating Campaign"
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.campaignButtonFromSideMenuXpath())))
        elements.campaignButtonFromSideMenu()
        
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.createCampaignButtonXpath())))

        if elements.campaignsPageHeaderText()=="Campaigns":
            print "Campaigns page displayed"
        else:
            print "Campaigns page is not displayed"
            raise Exception
        
        
        print "Clicking on Create Campaign button"
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.createCampaignButtonXpath())))
        elements.createCampaignButton()
        
        
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.Camp_titleXpath())))
        print "Create Campaign page is displayed"
        
                  
        elements.titleTextField(campaignTitle)
        print "Title entered"
        
        elements.descriptionField(campDescription)
        print "Description entered"
        
        print "Adding Lesson"
        #Add lesson button
        elements.addlessonButton()
        
        #Waiting until first lesson in pop is displayed
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.FirstLessonWaitXpath())))
        
        #Searching lesson by its name
        elements.searchLesson(lessonName)
        
        #Waiting until lesson displayed
        elements.waitUntilSearchedLessonDisplayed(lessonName)
        
        #selecting searched lesson
        elements.selectSearchedLesson(lessonName)
        
        
        
        #waiting until add to campaign button is click able
        wait.until(EC.element_to_be_clickable((By.XPATH,elements.AddToCampaign_ButtonXpath())))
        
        #Clicking on Add to Campaign button
        elements.addToCampaignButton()
        
        
        time.sleep(2)
        #Verifying Added lesson is displayed in Grid
        print "\nVerifying Added lesson is displayed in Grid"
        if elements.firstLessonInGrid()==lessonName:
            print "Lesson displayed in grid"
        else:
            print "Lesson not displayed in grid"
            raise Exception
        
        
        print "Setting minimum passing score"
        elements.makeThisAsAGradedCampaign()
        elements.setMinimumPassingScore(minPassingScore)
        print "Setting maximum number of attempts"
        elements.setAMaxNoOfAttempts(numberOfAttempts)
        
        
        
        wait.until(EC.element_to_be_clickable((By.XPATH,elements.SaveAndExit_ButtonXpath())))
        print "Clicking on Save & Exit button"
        elements.saveAndPlanAssignmentbutton()
        
        planAssign=PlanAssignmentForPageElements()
        if campaignTitle in planAssign.planAssignmentHeader():
            print "'"+planAssign.planAssignmentHeader()+"' is displayed"
        else:
            print "Plan Assignment page is not displayed"
            raise Exception
        
        
        print "Assigning assignment to a user"
        selectUser=wait.until(EC.visibility_of_element_located((By.XPATH,planAssign.AddGroupTextField())))
        webdriver.ActionChains(driver).move_to_element(selectUser).click(selectUser).send_keys(FirstName).perform()
        print "'"+FirstName+"' is enetered"
        groupdisplayed=wait.until(EC.visibility_of_element_located((By.XPATH,"(//div[@role='option']/span)[1]")))
        webdriver.ActionChains(driver).move_to_element(groupdisplayed).click().perform()
        print "'"+FirstName+" is selected"
        
        
        print "Checking User is displayed in Grid"
        userInGrid=wait.until(EC.visibility_of_element_located((By.XPATH,"//table/tbody/tr/td[1]")))
        if FirstName in userInGrid.text:
            print "'"+userInGrid.text+"' is displayed"
        else:
            print "User in grid is not displayed"
            raise Exception


        
        
        print "Click on Send Assignment"
        wait.until(EC.visibility_of_element_located((By.XPATH,planAssign.assignButton())))
        sendAssign=wait.until(EC.element_to_be_clickable((By.XPATH,planAssign.assignButton())))
        sendAssign.click()
        print "Clicked on Send Assignment button"
        
        
        planAssign.sendAssignmentButtonPopup()
        print "Clicked on Yes,Send assignment button from pop up"
    
        
        
        
        #Verifying Campaign Detail page is displayed
        print "\nVerifying campaign detail page is displayed"
        
        if elements.campaignDetailPageHeaderText()==campaignTitle:
            print "Campaign detail page is displayed"
        else:
            print "Campaign detail page is not displayed"
            raise Exception
        
        #verifying in Campaigns displayed in Campaigns grid
        elements.searchingForlesson(campaignTitle)
        
        if elements.actualCampTitleINGrid()==campaignTitle:
            print "Campaign '"+campaignTitle+"' displayed in Grid"
        
        else:
            print "Campaign is not displayed in Grid"
            raise Exception
        
        
        
    def checkingForUserDueSoon(self,EmailId,Password,campaignTitle,lessonName,status,username,password):
        
        
        print "\n\nChecking sent assignment is displayed for User"
        wait=WebDriverWait(driver, 60)
        driver.refresh()
        wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")))
        
        
        print "Logging out"
        print "Clicking on Username Dropdown"
        ele =driver.find_element_by_xpath(".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")
        driver.execute_script('arguments[0].click()',ele)
        elem=driver.find_element_by_xpath("html/body/div/div/div[1]/div[2]/div[2]/a")
        driver.execute_script('arguments[0].click()',elem)
        
        print "Clicked on Sign Out option"
        
        element = wait.until(EC.presence_of_element_located((By.ID, "password")))
        
        if driver.title == "Grovo":
            print("Grovo Application URL Opened")
        else:
            raise Exception.message

        print "Grovo Sign-In page is displayed"
        
        print "Entering User name"
        driver.find_element_by_xpath(".//*[@id='username']").send_keys(EmailId)
       
        print "Entering Password"
        element.send_keys(Password)
        
        element.send_keys(Keys.TAB)
        print "Clicking on Sign_In button"
        driver.find_element_by_xpath("//*[@id='submitButton']").click()
        
        print "Successfully Logged Into Users account"
        time.sleep(30)
        
        driver.refresh()
        time.sleep(10)
        driver.refresh()
        driver.refresh()
        
        print "Checking in my Learning page Assignment is displayed"
        wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[3]/div[2]/div/div[2]/div/div/div/div[1]/div/div/div[2]/div[1]/div/div/a[1]/span")))
        
        unDropDown1=wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")))
        driver.execute_script('arguments[0].click()',unDropDown1)
        
        mylearning=wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[1]/div[2]/div[1]/a[1]")))
        driver.execute_script('arguments[0].click()',mylearning)
        
        assignmentText=wait.until(EC.visibility_of_element_located((By.XPATH,"(//div[2]/div[1]/div/div/a[1]/span[contains(.,'"+campaignTitle+"')])[1]")))
        
        
        
        if campaignTitle in assignmentText.text:
            print "Campaign displayed for User in My learning page"
            
        else:
            print "Campaign not displayed for User"
            raise Exception
        
       
        print "Checking Assignment is displayed with Due Soon status"
        
        try:
            wait.until(EC.visibility_of_element_located((By.XPATH,"(//a/span[contains(.,'"+campaignTitle+"')]/span[.='"+status+"'])[1]")))
            print "Status '"+status+"' is displayed for '"+campaignTitle+"'"
        except Exception:
            print "Status displayed is invalid"
            raise Exception
        
        
        print "Starting assignment"
        startButton=wait.until(EC.visibility_of_element_located((By.XPATH,"(//a/span[contains(.,'"+campaignTitle+"')]/../../a[2])[1]")))
        startButton.click()
        
        print "Waiting for lesson name to be displayed"
        
        lessonNameForUser=wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div/div/div[4]/div/div/div/div[1]/h1/span")))
        lessonNameForUserText=lessonNameForUser.text
        if lessonName==lessonNameForUserText:
            print "Verified Lesson '"+lessonNameForUserText+"' is displayed for User"
        else:
            print "Lesson displayed is not valid"
            raise Exception
        
        
        backButtoon=wait.until(EC.element_to_be_clickable((By.XPATH,".//*[@id='content']/div/div/div/div[1]/span/button")))
        backButtoon.click()
        
        unDropDown=wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")))
        driver.execute_script('arguments[0].click()',unDropDown)
        
        signOut=driver.find_element_by_xpath("html/body/div/div/div[1]/div[2]/div[2]/a")
        driver.execute_script('arguments[0].click()',signOut)
        
        
        
        element = wait.until(EC.presence_of_element_located((By.ID, "password")))
        
        if driver.title == "Grovo":
            print("Grovo Application URL Opened")
        else:
            raise Exception.message

        print "Grovo Sign-In page is displayed"
        
        print "Entering User name"
        driver.find_element_by_xpath(".//*[@id='username']").send_keys(username)
       
        print "Entering Password"
        element.send_keys(password)
        
        element.send_keys(Keys.TAB)
        print "Clicking on Sign_In button"
        driver.find_element_by_xpath("//*[@id='submitButton']").click()
        
        print "Successfully Logged Into Users account"
        
        time.sleep(5)    
        
        
        
    def assignCampToCheckDueSoonStatusForUser(self):
        
        book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        first_sheet = book.sheet_by_name('CampAssign')
        
        
        #Campaign
        cell1 = first_sheet.cell(658,1)
        campaignTitle = cell1.value
        
        cell1 = first_sheet.cell(659,1)
        campDescription = cell1.value
        
        
        
        cell1 = first_sheet.cell(660,1)
        minPassingScore = cell1.value
        
        cell1 = first_sheet.cell(661,1)
        numberOfAttempts = cell1.value
        
        cell1 = first_sheet.cell(662,1)
        lessonName = cell1.value
        
        cell2 = first_sheet.cell(663,1)
        questionCard= cell2.value
        
        cell2 = first_sheet.cell(664,1)
        ans1= cell2.value
        
        cell2 = first_sheet.cell(665,1)
        ans2= cell2.value
        
        
        
        
        
        #Learner 1
        cell = first_sheet.cell(667,1)
        FirstName = cell.value
        
        FirstNameId = FirstName.split("_")
        emp = FirstNameId[0]+"_"
        ids = FirstNameId[1]
        empId = int(ids)+1
        FirstNameUpdated= emp+str(empId)
        
        cell = first_sheet.cell(668,1)
        LastName = cell.value
        
        LastNameID = LastName.split("_")
        emp = LastNameID[0]+"_"
        ids = LastNameID[1]
        empId = int(ids)+1
        LastNameUpdated= emp+str(empId)
        
        
        
        cell = first_sheet.cell(669,1)
        Email = cell.value
        
        EmailId = Email.split("_")
        emp = EmailId[0]+"_"
        ids = EmailId[1]
        remaining="_"+EmailId[2]
        empId = int(ids)+1
        EmailIdUpdated= emp+str(empId)+remaining
        
        cell = first_sheet.cell(670,1)
        EmployeeId = cell.value
        
        Employee = EmployeeId.split("_")
        emp = Employee[0]+"_"
        ids = Employee[1]
        empId = int(ids)+1
        EmployeeIdUpdated= emp+str(empId)
        
        cell = first_sheet.cell(671,1)
        Password = cell.value
        
        cell = first_sheet.cell(672,1)
        NewPassword = cell.value
        
        cell = first_sheet.cell(673,1)
        role = cell.value
        
        cell = first_sheet.cell(674,1)
        status = cell.value
        
        #For Original User
        book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        s_sheet = book.sheet_by_name('Login_Credentials')
        
        cell = s_sheet.cell(1,1)
        url = cell.value
        
        cell = s_sheet.cell(3,1)
        username = cell.value
        
        cell = s_sheet.cell(3,2)
        password = cell.value
        
        
        
        
        #updating user values
        wb = load_workbook(os.path.join('Test_Data/TestData.xlsx'))
            #print (wb.sheetnames)
        
        sheet = wb['CampAssign']
        
        sheet.cell(row=668, column=2).value = FirstNameUpdated
        sheet.cell(row=669, column=2).value = LastNameUpdated
        sheet.cell(row=670, column=2).value = EmailIdUpdated
        sheet.cell(row=671, column=2).value = EmployeeIdUpdated
        
        
        
        wb.save(os.path.join('Test_Data/TestData.xlsx'))
            
        print "All User Data Updated in Excel"
        
        
        
        
        try:
            print "\nCreating a New Learner\n"
            lt=CreateLearner()
            lt.createNewLearnerMain(FirstName, LastName, Email, EmployeeId, Password, NewPassword, url, username, password)
            
            
            
            print "\nCreating a lesson\n"
            ot=CreateLessonDifferentLessons()
            ot.lessonWithQuestion(lessonName, questionCard, ans1, ans2)
            
            
            print "\nCreating Campaign\n"
            lk=AssignCampCheckDueSoon()
            lk.campaignToCheckDueSoon(campaignTitle, campDescription, lessonName, FirstName, minPassingScore, numberOfAttempts)
            
            lk.checkingForUserDueSoon(Email, Password, campaignTitle, lessonName, status, username, password)
            
            d=DeleteLesson()
            d.lessonDelete(lessonName)
            print "\n----Test Execution Completed----\n"
            
            
        except Exception as e:
            traceback.print_exc()
            print (e)
            raise Exception 
         
        finally:  
            second_sheet = book.sheet_by_name('Login_Credentials')
            cell = second_sheet.cell(1,1)
            url = cell.value
            driver.get(url)
            try:
                WebDriverWait(driver, 5).until(EC.alert_is_present())

                alert = driver.switch_to.alert
                alert.accept()
                print("alert accepted")
            except Exception:
                print("no alert")
            
            wait=WebDriverWait(driver, 60)
            unDropDown=wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")))
            driver.execute_script('arguments[0].click()',unDropDown)
            
            signOut=driver.find_element_by_xpath("html/body/div/div/div[1]/div[2]/div[2]/a")
            driver.execute_script('arguments[0].click()',signOut)
            
            
            
            element = wait.until(EC.presence_of_element_located((By.ID, "password")))
            
            
            time.sleep(3)
            if driver.title == "Grovo":
                print("Grovo Application URL Opened")
            else:
                raise Exception.message
    
            print "Grovo Sign-In page is displayed"
            
            print "Entering User name"
            driver.find_element_by_xpath(".//*[@id='username']").send_keys(username)
           
            print "Entering Password"
            element.send_keys(password)
            
            element.send_keys(Keys.TAB)
            print "Clicking on Sign_In button"
            driver.find_element_by_xpath("//*[@id='submitButton']").click()
            
            print "Successfully Logged Into Users account"
            time.sleep(3)
            
            
            
            
if __name__ == '__main__':
    
    btc=BaseTestClass()
    btc.UserLogin()
    
    ne=AssignCampCheckDueSoon()
    ne.assignCampToCheckDueSoonStatusForUser()

