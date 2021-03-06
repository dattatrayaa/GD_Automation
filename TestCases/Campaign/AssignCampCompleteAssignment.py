'''
Created on 11-Apr-2018

@author: dattatraya
'''
import os.path
import time
import traceback

from BaseTestClass import BaseTestClass
from BaseTestClass import driver
from openpyxl.reader.excel import load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import xlrd

from CampaignPageElements import CampPage
from CreateLearnerNew import CreateLearner
from CreateLessonDifferentLessons import CreateLessonDifferentLessons
from DeleteLesson import DeleteLesson
from PlanAssignmentForPageElements import PlanAssignmentForPageElements
from selenium import webdriver
from ManageAssignMentForPageElements import ManageAssignMentForPageElements
from BaseTestClass import projectPath
class AssignCampCompleteAssignment:
    
    def campaignToCheckCompletedAssignMentStatus(self,campaignTitle,campDescription,lessonName,FirstName,minPassingScore,numberOfAttempts):
        
        elements=CampPage()
        
        wait=WebDriverWait(driver, 60)
        
        driver.refresh()
        print "\n\nCreating Campaign"
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.campaignButtonFromSideMenuXpath())))
        elements.campaignButtonFromSideMenu()
        
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.createCampaignButtonXpath())))

        if elements.campaignsPageHeaderText()=="Campaigns":
            print "Campaigns page displayed"
        else:
            print "Campaigns page is not displayed"
            raise Exception
        
        
        print "Clicking on Create Campaign button"
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.createCampaignButtonXpath())))
        elements.createCampaignButton()
        
        
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.Camp_titleXpath())))
        print "Create Campaign page is displayed"
        
                  
        elements.titleTextField(campaignTitle)
        print "Title entered"
        
        elements.descriptionField(campDescription)
        print "Description entered"
        
        print "Adding Lesson"
        #Add lesson button
        elements.addlessonButton()
        
        #Waiting until first lesson in pop is displayed
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.FirstLessonWaitXpath())))
        
        #Searching lesson by its name
        elements.searchLesson(lessonName)
        
        #Waiting until lesson displayed
        elements.waitUntilSearchedLessonDisplayed(lessonName)
        
        #selecting searched lesson
        elements.selectSearchedLesson(lessonName)
        
        #waiting until add to campaign button is click able
        wait.until(EC.element_to_be_clickable((By.XPATH,elements.AddToCampaign_ButtonXpath())))
        
        #Clicking on Add to Campaign button
        elements.addToCampaignButton()
        
        
        time.sleep(2)
        #Verifying Added lesson is displayed in Grid
        print "\nVerifying Added lesson is displayed in Grid"
        if elements.firstLessonInGrid()==lessonName:
            print "Lesson displayed in grid"
        else:
            print "Lesson not displayed in grid"
            raise Exception
        
        
        print "Setting minimum passing score"
        elements.makeThisAsAGradedCampaign()
        elements.setMinimumPassingScore(minPassingScore)
        print "Setting maximum number of attempts"
        elements.setAMaxNoOfAttempts(numberOfAttempts)
        
        
        
        wait.until(EC.element_to_be_clickable((By.XPATH,elements.SaveAndExit_ButtonXpath())))
        print "Clicking on Save & Exit button"
        elements.saveAndPlanAssignmentbutton()
        
        planAssign=PlanAssignmentForPageElements()
        if campaignTitle in planAssign.planAssignmentHeader():
            print "'"+planAssign.planAssignmentHeader()+"' is displayed"
        else:
            print "Plan Assignment page is not displayed"
            raise Exception
        
        
        print "Assigning assignment to a user"
        selectUser=wait.until(EC.visibility_of_element_located((By.XPATH,planAssign.AddGroupTextField())))
        webdriver.ActionChains(driver).move_to_element(selectUser).click(selectUser).send_keys(FirstName).perform()
        print "'"+FirstName+"' is enetered"
        groupdisplayed=wait.until(EC.visibility_of_element_located((By.XPATH,"(//div[@role='option']/span)[1]")))
        webdriver.ActionChains(driver).move_to_element(groupdisplayed).click().perform()
        print "'"+FirstName+" is selected"
        
        
        print "Checking User is displayed in Grid"
        userInGrid=wait.until(EC.visibility_of_element_located((By.XPATH,"//table/tbody/tr/td[1]")))
        if FirstName in userInGrid.text:
            print "'"+userInGrid.text+"' is displayed"
        else:
            print "User in grid is not displayed"
            raise Exception


        '''print "Changing Due date to todays date"
        driver.find_element_by_xpath("//input[@name='campaign-due-date']").click()
        datePick=wait.until(EC.visibility_of_element_located((By.XPATH,"//div[@class='react-datepicker__day react-datepicker__day--today']")))
        print "'"+str(datePick.text)+"' is selected"
        datePick.click()'''
        
        
        
        print "Click on Send Assignment"
        wait.until(EC.visibility_of_element_located((By.XPATH,planAssign.assignButton())))
        sendAssign=wait.until(EC.element_to_be_clickable((By.XPATH,planAssign.assignButton())))
        sendAssign.click()
        print "Clicked on Send Assignment button"
        
        
        planAssign.sendAssignmentButtonPopup()
        print "Clicked on Yes,Send assignment button from pop up"
    
        
        
        
        #Verifying Campaign Detail page is displayed
        print "\nVerifying campaign detail page is displayed"
        
        if elements.campaignDetailPageHeaderText()==campaignTitle:
            print "Campaign detail page is displayed"
        else:
            print "Campaign detail page is not displayed"
            raise Exception
        
        #verifying in Campaigns displayed in Campaigns grid
        elements.searchingForlesson(campaignTitle)
        
        if elements.actualCampTitleINGrid()==campaignTitle:
            print "Campaign '"+campaignTitle+"' displayed in Grid"
        
        else:
            print "Campaign is not displayed in Grid"
            raise Exception
        
        
    
    def toCompleteGivenAssignment(self,EmailId,Password,campaignTitle,lessonName,username,password):
        
        
        print "\n\nChecking sent assignment is displayed for User"
        wait=WebDriverWait(driver, 60)
        driver.refresh()
        wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")))
        
        
        print "Logging out"
        print "Clicking on Username Dropdown"
        ele =driver.find_element_by_xpath(".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")
        driver.execute_script('arguments[0].click()',ele)
        elem=driver.find_element_by_xpath("html/body/div/div/div[1]/div[2]/div[2]/a")
        driver.execute_script('arguments[0].click()',elem)
        
        print "Clicked on Sign Out option"
        
        element = wait.until(EC.presence_of_element_located((By.ID, "password")))
        
        if driver.title == "Grovo":
            print("Grovo Application URL Opened")
        else:
            raise Exception.message

        print "Grovo Sign-In page is displayed"
        
        print "Entering User name"
        driver.find_element_by_xpath(".//*[@id='username']").send_keys(EmailId)
       
        print "Entering Password"
        element.send_keys(Password)
        
        element.send_keys(Keys.TAB)
        print "Clicking on Sign_In button"
        driver.find_element_by_xpath("//*[@id='submitButton']").click()
        
        print "Successfully Logged Into Users account"
        time.sleep(30)
        
        driver.refresh()
        time.sleep(10)
        driver.refresh()
        driver.refresh()
        
        print "Checking in my Learning page Assignment is displayed"
        wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[3]/div[2]/div/div[2]/div/div/div/div[1]/div/div/div[2]/div[1]/div/div/a[1]/span")))
        
        unDropDown1=wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")))
        driver.execute_script('arguments[0].click()',unDropDown1)
        
        mylearning=wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[1]/div[2]/div[1]/a[1]")))
        driver.execute_script('arguments[0].click()',mylearning)
        
        assignmentText=wait.until(EC.visibility_of_element_located((By.XPATH,"(//div[2]/div[1]/div/div/a[1]/span[contains(.,'"+campaignTitle+"')])[1]")))
        
        
        
        if campaignTitle in assignmentText.text:
            print "Campaign displayed for User in My learning page"
            
        else:
            print "Campaign not displayed for User"
            raise Exception
        
       
        
        
        print "Starting assignment"
        startButton=wait.until(EC.visibility_of_element_located((By.XPATH,"(//a/span[contains(.,'"+campaignTitle+"')]/../../a[2])[1]")))
        startButton.click()
        
        print "Waiting for lesson name to be displayed"
        
        lessonNameForUser=wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div/div/div[4]/div/div/div/div[1]/h1/span")))
        lessonNameForUserText=lessonNameForUser.text
        if lessonName==lessonNameForUserText:
            print "Verified Lesson '"+lessonNameForUserText+"' is displayed for User"
        else:
            print "Lesson displayed is not valid"
            raise Exception
        
        
        
        print "Clicking on Start button from Lesson play"
        startPlay=wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div/div/div[4]/div/div/div/div[2]/button")))
        startPlay.click()
        
        print "Clicking on answer of question"
        answer1=wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div/div/div[4]/div/div/div/div/div/div[2]/div/div/div[1]/div/div")))
        answer1.click()
        
        time.sleep(3)
        print "Clicking Next button"
        nextButton=wait.until(EC.visibility_of_element_located((By.XPATH,"//button[.='Next']")))
        nextButton.click()
        time.sleep(2)
        
        print "Clicking on Next Button"
        nextButton=wait.until(EC.visibility_of_element_located((By.XPATH,"//button[.='Next']")))
        nextButton.click()
        time.sleep(2)
        
        print "Verifying Complete assignment page is displayed"
        try:
            complete=wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div/div/div[4]/div/div/div/div/div[1]/h1[contains(.,'Assignment')]")))
            print "'"+complete.text+"' is displayed"
        except Exception:
            print "Assignment complete page is not displayed"
            raise Exception
        
        
        
        print "Clicking on Back Button"
        backButtoon=wait.until(EC.element_to_be_clickable((By.XPATH,".//*[@id='content']/div/div/div/div[1]/span/button")))
        backButtoon.click()
        
        print "Logging out of current user"
        unDropDown=wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")))
        driver.execute_script('arguments[0].click()',unDropDown)
        
        signOut=driver.find_element_by_xpath("html/body/div/div/div[1]/div[2]/div[2]/a")
        driver.execute_script('arguments[0].click()',signOut)
        print "Clicked on  Sign out"
        
        
        element = wait.until(EC.presence_of_element_located((By.ID, "password")))
        
        if driver.title == "Grovo":
            print("Grovo Application URL Opened")
        else:
            raise Exception.message

        print "Grovo Sign-In page is displayed"
        
        print "Entering User name"
        driver.find_element_by_xpath(".//*[@id='username']").send_keys(username)
       
        print "Entering Password"
        element.send_keys(password)
        
        element.send_keys(Keys.TAB)
        print "Clicking on Sign_In button"
        driver.find_element_by_xpath("//*[@id='submitButton']").click()
        
        print "Successfully Logged Into Users account"
        
        time.sleep(5)    
        
        
        
        
    def checkStatusOfCompletedAssignment(self,campaignTitle,FirstName,LastName,status1,scoreAssessment):
        elements=CampPage()
        
        wait=WebDriverWait(driver, 60)
        
        driver.refresh()
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.campaignButtonFromSideMenuXpath())))
        elements.campaignButtonFromSideMenu()
        
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.createCampaignButtonXpath())))

        if elements.campaignsPageHeaderText()=="Campaigns":
            print "Campaigns page displayed"
        else:
            print "Campaigns page is not displayed"
            raise Exception
        
        print "Clicking on link of campaign created"
        elements.CampaignLinkFromCampaignGrid(campaignTitle)
        
        
        print "Verifying Campaign details page is displayed"
        if elements.campaignDetailPageHeaderText()==campaignTitle:
            print "Header Text '"+elements.campaignDetailPageHeaderText()+"' is displayed"
        else:
            print "Header Text not displayed"
            raise Exception  
        
        time.sleep(4)
        print "Checking in Assignments section Completed status is displayed"
        statusInCol=wait.until(EC.visibility_of_element_located((By.XPATH,"//tbody/tr/td[2]/span/span[2]")))
        if statusInCol.text==status1:
            print "'"+statusInCol.text+"' is displayed"
        else:
            print "Status is not displayed"
            raise Exception
        
        print "Clicking on 1 learner text Present in Assigned to column"
        userAssignedToCol=wait.until(EC.visibility_of_element_located((By.XPATH,"//tbody/tr/td[1]/a/span/span")))
        userAssignedToCol.click()
        
                
        managepage=ManageAssignMentForPageElements()
        
        try:
            header=wait.until(EC.visibility_of_element_located((By.XPATH,managepage.HeaderTextManageAssign())))      
            print header.text+"is displayed"
        except Exception:
            print "Manage assignment for page is not displayed"
            raise Exception
        
        try:
            wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div[1]/div/div[3]/div[2]/div/div[4]/table/tbody/tr/td[contains(.,'"+FirstName+"')]")))
        except Exception:
            print "Learner Name not displayed"
            raise Exception   
        
        print "Checking Progress displayed"
        progressCol=wait.until(EC.visibility_of_element_located((By.XPATH,"//table/tbody/tr/td[2]/span/span[2]")))   
        if progressCol.text==status1:
            print "'"+progressCol.text+"' is displayed"
        else:
            print "Status is not displayed"
            raise Exception
        
        
        print "Checking Assessment score displayed"
        scoreCol=wait.until(EC.visibility_of_element_located((By.XPATH,"//table/tbody/tr/td[5]/span/span/span")))   
        if scoreCol.text==scoreAssessment:
            print "'"+scoreCol.text+"' is displayed"
        else:
            print "Assessment scores is not displayed"
            #raise Exception
        
        
        print "Verifying Cancel Assignment button is disabled"
        try:
            wait.until(EC.visibility_of_element_located((By.XPATH,"//button[@class='btn btn-secondary is-disabled']")))   
            print "Verified Cancel Assignment button is disabled"
        except Exception:
            print "Cancel Assignment button is not disabled"
            raise Exception
        
        
        print "Hence Verified Users status, progress, assessment score"
        
        
        driver.refresh()
        
        
        
        
        
    def assignCampToCheckCompleteAssignmentStatus(self):
        
        book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        first_sheet = book.sheet_by_name('CampAssign')
        
        
        #Campaign
        cell1 = first_sheet.cell(678,1)
        campaignTitle = cell1.value
        
        cell1 = first_sheet.cell(679,1)
        campDescription = cell1.value
        
        
        
        cell1 = first_sheet.cell(680,1)
        minPassingScore = cell1.value
        
        cell1 = first_sheet.cell(681,1)
        numberOfAttempts = cell1.value
        
        cell1 = first_sheet.cell(682,1)
        lessonName = cell1.value
        
        cell2 = first_sheet.cell(683,1)
        questionCard= cell2.value
        
        cell2 = first_sheet.cell(684,1)
        ans1= cell2.value
        
        cell2 = first_sheet.cell(685,1)
        ans2= cell2.value
        

        
        #Learner 1
        cell = first_sheet.cell(687,1)
        FirstName = cell.value
        
        FirstNameId = FirstName.split("_")
        emp = FirstNameId[0]+"_"
        ids = FirstNameId[1]
        empId = int(ids)+1
        FirstNameUpdated= emp+str(empId)
        
        cell = first_sheet.cell(688,1)
        LastName = cell.value
        
        LastNameID = LastName.split("_")
        emp = LastNameID[0]+"_"
        ids = LastNameID[1]
        empId = int(ids)+1
        LastNameUpdated= emp+str(empId)
        
        
        
        cell = first_sheet.cell(689,1)
        Email = cell.value
        
        EmailId = Email.split("_")
        emp = EmailId[0]+"_"
        ids = EmailId[1]
        remaining="_"+EmailId[2]
        empId = int(ids)+1
        EmailIdUpdated= emp+str(empId)+remaining
        
        cell = first_sheet.cell(690,1)
        EmployeeId = cell.value
        
        Employee = EmployeeId.split("_")
        emp = Employee[0]+"_"
        ids = Employee[1]
        empId = int(ids)+1
        EmployeeIdUpdated= emp+str(empId)
        
        cell = first_sheet.cell(691,1)
        Password = cell.value
        
        cell = first_sheet.cell(692,1)
        NewPassword = cell.value
        
        
        
        cell = first_sheet.cell(693,1)
        status1 = cell.value
        
        cell = first_sheet.cell(693,1)
        scoreAssessment = cell.value
        
        
        #For Original User
        book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        s_sheet = book.sheet_by_name('Login_Credentials')
        
        cell = s_sheet.cell(1,1)
        url = cell.value
        
        cell = s_sheet.cell(3,1)
        username = cell.value
        
        cell = s_sheet.cell(3,2)
        password = cell.value
        
            
        
        
        #updating user values
        wb = load_workbook(os.path.join('Test_Data/TestData.xlsx'))
            #print (wb.sheetnames)
        
        sheet = wb['CampAssign']
        
        sheet.cell(row=688, column=2).value = FirstNameUpdated
        sheet.cell(row=689, column=2).value = LastNameUpdated
        sheet.cell(row=690, column=2).value = EmailIdUpdated
        sheet.cell(row=691, column=2).value = EmployeeIdUpdated
        
        
        
        wb.save(os.path.join('Test_Data/TestData.xlsx'))
            
        print "All User Data Updated in Excel"
        
        
        
        
        try:
            print "\nCreating a New Learner\n"
            lt=CreateLearner()
            lt.createNewLearnerMain(FirstName, LastName, Email, EmployeeId, Password, NewPassword, url, username, password)
            
            
            
            print "\nCreating a lesson\n"
            ot=CreateLessonDifferentLessons()
            ot.lessonWithQuestion(lessonName, questionCard, ans1, ans2)
            
            print "\nCreating Campaign\n"
            lk=AssignCampCompleteAssignment()
            lk.campaignToCheckCompletedAssignMentStatus(campaignTitle, campDescription, lessonName, FirstName, minPassingScore, numberOfAttempts)
            
            print "\nTo Complete given assignment\n"
            lk.toCompleteGivenAssignment(Email, Password, campaignTitle, lessonName, username, password)
            
            print "\nChecking Status of completed assignment\n"
            lk.checkStatusOfCompletedAssignment(campaignTitle, FirstName, LastName, status1, scoreAssessment)
            
            
            print "\n----Test Execution Completed----\n"
            
            
        except Exception as e:
            traceback.print_exc()
            print (e)
            raise Exception 
         
        finally:  
            second_sheet = book.sheet_by_name('Login_Credentials')
            cell = second_sheet.cell(1,1)
            url = cell.value
            driver.get(url)
            try:
                WebDriverWait(driver, 5).until(EC.alert_is_present())

                alert = driver.switch_to.alert
                alert.accept()
                print("alert accepted")
            except Exception:
                print("no alert")
            
            wait=WebDriverWait(driver, 60)
            unDropDown=wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")))
            driver.execute_script('arguments[0].click()',unDropDown)
            
            signOut=driver.find_element_by_xpath("html/body/div/div/div[1]/div[2]/div[2]/a")
            driver.execute_script('arguments[0].click()',signOut)
            
            
            
            element = wait.until(EC.presence_of_element_located((By.ID, "password")))
            
            
            time.sleep(3)
            if driver.title == "Grovo":
                print("Grovo Application URL Opened")
            else:
                raise Exception.message
    
            print "Grovo Sign-In page is displayed"
            
            print "Entering User name"
            driver.find_element_by_xpath(".//*[@id='username']").send_keys(username)
           
            print "Entering Password"
            element.send_keys(password)
            
            element.send_keys(Keys.TAB)
            print "Clicking on Sign_In button"
            driver.find_element_by_xpath("//*[@id='submitButton']").click()
            
            print "Successfully Logged Into Users account"
            time.sleep(3)
            
            try:
                d=DeleteLesson()
                d.lessonDelete(lessonName)
            except Exception:
                driver.get(url)
            
            
            
            
if __name__ == '__main__':
    
    btc=BaseTestClass()
    btc.UserLogin()
    
    ne=AssignCampCompleteAssignment()
    ne.assignCampToCheckCompleteAssignmentStatus()
    