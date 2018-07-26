'''
Created on 24-Jul-2018

@author: Sheethu C
'''
import os.path
import time
import traceback
from BaseTestClass import BaseTestClass
from BaseTestClass import driver
from openpyxl.reader.excel import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import xlrd
from BaseTestClass import projectPath
from CampaignPageElements import CampPage
from CreateLearnerNew import CreateLearner
from CreateLessonDifferentLessons import CreateLessonDifferentLessons
from DeleteLesson import DeleteLesson
from PlanAssignmentForPageElements import PlanAssignmentForPageElements
from GroupsPageElements import GroupsPageElements
from AddTriggerForPageElements import AddTriggerPage
from BaseTestClass import projectPath
from GroupDelete import GroupDelete
from CreateUserXpath import CreateUserXpath
class LearnerCountForAssignOptionForTriggering:
    def groupCreation(self,groupName,FirstName1,FirstName2):
        driver.refresh()
        wait=WebDriverWait(driver, 60)
        
        group=GroupsPageElements()
        
        group.adminSideMenuUnexpanded()
        print "Clicked on admin icon"
         
        group.groupSideMenuExpanded()  
        print "Clicked on Group icon"
        
        
        print "Checking Group page is displayed"
        if driver.find_element_by_xpath(group.groupPageHeader()).is_displayed():
            print "Group page is displayed successfully"
        else:
            print "Group page is not displayed"
            raise Exception
        
        
        group.createGroupButton()
        print "Clicked on Create Group button"
        
        group.enteringGroupname(groupName)
        
        print "Group Name entered....."
        
        group.nextButton()
        print "Clicked on Next button"
        
        wait.until(EC.visibility_of_element_located((By.XPATH,group.groupDetailPageHeader())))
        print "Checking group is created"
        
        time.sleep(4)
        headerText=driver.find_element_by_xpath(group.groupDetailPageHeader()).text
        
        if headerText==groupName:
            print "Group created successfully"
            
        else:
            print "Group not created"
            raise Exception
        
        
        print "Adding user to Group by name"
        group.addByNameButton()
        group.addByName(FirstName1)
        group.addByName(FirstName2)
        
        
        print "Verifying user is added to group"
        group.groupAddedInList(FirstName1)
        group.groupAddedInList(FirstName2)
        
    
        group.saveButton()
        print "Clicked on Save button"

        group.saveButtonPopup()
        print "Clicked on Save from popup"
        
        
        print "Checking Created Group is displayed in Grid"
        group.groupsLinkFromBreadCrumb()
        
        print "Searching Group"
        group.createdGroupDisplayedInGrid(groupName)
        driver.refresh()
    def campaignCreationWithTrigger(self,campaignTitle,lessonName,campDescription,numberOfAttempts,minPassingScore,groupName1,groupName2):
        elements=CampPage()
        
        wait=WebDriverWait(driver, 60)
        driver.refresh()
        
        print "\n\nCreating Campaign"
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.campaignButtonFromSideMenuXpath())))
        elements.campaignButtonFromSideMenu()
        
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.createCampaignButtonXpath())))

        if elements.campaignsPageHeaderText()=="Campaigns":
            print "Campaigns page displayed"
        else:
            print "Campaigns page is not displayed"
            raise Exception
        
        
        print "Clicking on Create Campaign button"
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.createCampaignButtonXpath())))
        elements.createCampaignButton()
        
        
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.Camp_titleXpath())))
        print "Create Campaign page is displayed"
        
                  
        elements.titleTextField(campaignTitle)
        print "Title entered"
        
        elements.descriptionField(campDescription)
        print "Description entered"
        
        print "Adding Lesson"
        #Add lesson button
        elements.addlessonButton()
        
        #Waiting until first lesson in pop is displayed
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.FirstLessonWaitXpath())))
        
        #Searching lesson by its name
        elements.searchLesson(lessonName)
        
        #Waiting until lesson displayed
        elements.waitUntilSearchedLessonDisplayed(lessonName)
        
        #selecting searched lesson
        elements.selectSearchedLesson(lessonName)
        
        #waiting until add to campaign button is click able
        wait.until(EC.element_to_be_clickable((By.XPATH,elements.AddToCampaign_ButtonXpath())))
        
        #Clicking on Add to Campaign button
        elements.addToCampaignButton()
        
        
        time.sleep(2)
        #Verifying Added lesson is displayed in Grid
        print "\nVerifying Added lesson is displayed in Grid"
        if elements.firstLessonInGrid()==lessonName:
            print "Lesson displayed in grid"
        else:
            print "Lesson not displayed in grid"
            raise Exception
        
        
        print "Making this as a Graded campaign"
        elements.makeThisAsAGradedCampaign()
        print "Setting minimum passing score"
        elements.setMinimumPassingScore(minPassingScore)
        print "Setting maximum no of attempts"
        elements.setAMaxNoOfAttempts(numberOfAttempts)
        
        
        wait.until(EC.element_to_be_clickable((By.XPATH,elements.SaveAndExit_ButtonXpath())))
        print "Clicking on Save and Exit button"
        elements.saveAndExitButton()
        
        
        
        
        #Verifying Campaign Detail page is displayed
        print "\nVerifying campaign detail page is displayed"
        
        if elements.campaignDetailPageHeaderText()==campaignTitle:
            print "Campaign detail page is displayed"
        else:
            print "Campaign detail page is not displayed"
            raise Exception
        
        
        print "\nCreating trigger\n"
        elements.addTriggerButton()
        
        
        addTriggerPage=AddTriggerPage()
        
        
        header=wait.until(EC.visibility_of_element_located((By.XPATH,addTriggerPage.HeaderTextAddTrigger()))) 
        
        #Verifying Plan assignment for page is displayed
        print "Verifying Add trigger for page is displayed"
        if campaignTitle in header.text :
            print "Page '"+header.text+"' is displayed"
        else:
            print "Add trigger page is not displayed"
            raise Exception
        
        
        #Checking New hire on boarding is selected 
        if driver.find_element_by_xpath("//input[@id='trigger-new-hire']").is_selected():
            print "Radio button New Hire Onboarding is selected"
        else:
            print "Radio button New Hire Onboarding is not selected"
            raise Exception
        
        

        
        print "Selecting New to group"
        addTriggerPage.newToGroupRadio()
        searchGroups=wait.until(EC.visibility_of_element_located((By.XPATH,addTriggerPage.SelectGroupTextBoxTriggerPage()))) 
        
        
        webdriver.ActionChains(driver).move_to_element(searchGroups).click().send_keys(groupName1).perform()
        groupdisplayed=wait.until(EC.visibility_of_element_located((By.XPATH,"(//div[@role='option']/span)[1]")))
        
        webdriver.ActionChains(driver).move_to_element(groupdisplayed).click().perform()
        
        
        print "selecting second group"
        
        searchGroups=wait.until(EC.visibility_of_element_located((By.XPATH,addTriggerPage.SelectGroupTextBoxTriggerPage()))) 
        
        
        webdriver.ActionChains(driver).move_to_element(searchGroups).click().send_keys(groupName1).perform()
        groupdisplayed=wait.until(EC.visibility_of_element_located((By.XPATH,"(//div[@role='option']/span)[1]")))
        
        webdriver.ActionChains(driver).move_to_element(groupdisplayed).click().perform()
        
        print "Checking Group is displayed in Grid"
        
        if driver.find_element_by_xpath("//table/tbody/tr/td[1]/a").text==groupName1:
            print "Group '"+groupName1+"' is displayed in grid"
        else:
            print "Group is not displayed in Grid"
            raise Exception
        
        time.sleep(2)
        
        #Selecting second group
        
        searchGroups1=driver.find_element_by_xpath(addTriggerPage.SelectGroupTextBoxTriggerPage())
        webdriver.ActionChains(driver).move_to_element(searchGroups1).click().send_keys(groupName2).perform()
        groupdisplayed=wait.until(EC.visibility_of_element_located((By.XPATH,"(//div[@role='option']/span)[1]")))
        
        webdriver.ActionChains(driver).move_to_element(groupdisplayed).click().perform()
        
        
        print "Checking Group is displayed in Grid"
        
        if driver.find_element_by_xpath("//table/tbody/tr[1]/td[1]/a").text==groupName2:
            print "Group '"+groupName2+"' is displayed in grid"
        else:
            print "Group is not displayed in Grid"
            raise Exception
        
        print "Two groups triggered"
        
        
        
        print "Clicking on checkbox 'Also assign to learners who currently match this criteria' "
        
        
        check=wait.until(EC.visibility_of_element_located((By.XPATH,addTriggerPage.AlsoAssignToCurrentUserCheckbox())))
        check.click()
        wait.until(EC.visibility_of_element_located((By.XPATH,addTriggerPage.YourCriteriaMatchesBox())))
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div[1]/div/div[3]/div[2]/div/div[2]/section[1]/div/div[3]/div[3]")))
        print "Checking the criteria for duplicate laerner in the group"
        cnt= driver.find_element_by_xpath("html/body/div[1]/div/div[3]/div[2]/div/div[2]/section[1]/div/div[3]/div[3]").text
        learnercount=cnt.rsplit(' ',4)[3]
        print learnercount
        if learnercount == str(3):
            print "Duplicate count of the user is not displaying"
        else:
            print "Dupliacte count of the user is displaying"
            raise Exception
        print "Saving this trigger"
        addTriggerPage.saveTriggerButton()
        
        
        print "Checking for In Campaign details page Trigger is displayed"
        
        try:
            elements.TriggerDisplayedInGridForGroup(groupName1)
            print "Trigger with group name '"+groupName1+"' is displayed in Campaign details page"
        except Exception:
            print "Trigger is not displayed"
            raise Exception
       
       
       
        try:
            elements.TriggerDisplayedInGridForGroup(groupName1)
            print "Trigger with group name '"+groupName1+"' is displayed in Campaign details page"
        except Exception:
            print "Trigger is not displayed"
            raise Exception
        
        
        
        
        #verifying in Campaigns displayed in Campaigns grid
        elements.searchingForlesson(campaignTitle)
        
        if elements.actualCampTitleINGrid()==campaignTitle:
            print "Campaign '"+campaignTitle+"' displayed in Grid"
        
        else:
            print "Campaign is not displayed in Grid"
            raise Exception
        
         
         
        print "\nVerifying 'Y' is displayed for created triggred campaign"
        hasTrigger=driver.find_element_by_xpath("(//tr/td[1]/a[.='"+campaignTitle+"']/../../td[2])[1]").text
        
        if hasTrigger=="Y":
            print "'Y' displayed in Has trigger column"
        else:
            print "Invalid data in Hastrigger column"
        
        
    def updateUser(self):
        print "Updating Details"
        book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        second_sheet = book.sheet_by_name('CampAssign')
        #First User
        cell = second_sheet.cell(797,1)
        FirstName = cell.value
        FirstNameId = FirstName.split("_")
        emp = FirstNameId[0]+"_"
        ids = FirstNameId[1]
        empId1 = int(ids)+1
        FirstNameUpdated1= emp+str(empId1)
        
        
        cell2 = second_sheet.cell(800,1)
        LearnId = cell2.value
        LearnsId = LearnId.split("#")
        emp = LearnsId[0]+"#"
        ids = LearnsId[1]
        empId = int(ids)+1
        LearnerEmpId1 = emp+str(empId)
        
        
        cell1 = second_sheet.cell(799,1)
        LearnEmailId = cell1.value
        LearnerEmail = LearnEmailId.split("@")
        spitValue = LearnerEmail[0][:4]
        email = spitValue+str(empId)
        id = "@"+LearnerEmail[1]
        LearnerEmailId1 = email+id
        
        
        
        #Second User
        cell = second_sheet.cell(803,1)
        FirstName = cell.value
        FirstNameId = FirstName.split("_")
        emp = FirstNameId[0]+"_"
        ids = FirstNameId[1]
        empId1 = int(ids)+1
        FirstNameUpdated2= emp+str(empId1)
        
        
        cell2 = second_sheet.cell(806,1)
        LearnId = cell2.value
        LearnsId = LearnId.split("#")
        emp = LearnsId[0]+"#"
        ids = LearnsId[1]
        empId = int(ids)+1
        LearnerEmpId2 = emp+str(empId)
        
        
        cell1 = second_sheet.cell(805,1)
        LearnEmailId = cell1.value
        LearnerEmail = LearnEmailId.split("@")
        spitValue = LearnerEmail[0][:4]
        email = spitValue+str(empId)
        id = "@"+LearnerEmail[1]
        LearnerEmailId2 = email+id
        
        
        #Third User
        cell = second_sheet.cell(809,1)
        FirstName = cell.value
        FirstNameId = FirstName.split("_")
        emp = FirstNameId[0]+"_"
        ids = FirstNameId[1]
        empId1 = int(ids)+1
        FirstNameUpdated3= emp+str(empId1)
        
        cell2 = second_sheet.cell(812,1)
        LearnId = cell2.value
        LearnsId = LearnId.split("#")
        emp = LearnsId[0]+"#"
        ids = LearnsId[1]
        empId = int(ids)+1
        LearnerEmpId3 = emp+str(empId)
        
        
        cell1 = second_sheet.cell(811,1)
        LearnEmailId = cell1.value
        LearnerEmail = LearnEmailId.split("@")
        spitValue = LearnerEmail[0][:4]
        email = spitValue+str(empId)
        id = "@"+LearnerEmail[1]
        LearnerEmailId3 = email+id
        
        wb = load_workbook(os.path.join('Test_Data/TestData.xlsx'))
        
        #print (wb.sheetnames)
    
        sheet = wb['CampAssign']
        
        sheet.cell(row=798, column=2).value = FirstNameUpdated1    
        sheet.cell(row=800, column=2).value = LearnerEmailId1
        sheet.cell(row=801, column=2).value = LearnerEmpId1
        
        sheet.cell(row=804, column=2).value = FirstNameUpdated2   
        sheet.cell(row=806, column=2).value = LearnerEmailId2
        sheet.cell(row=807, column=2).value = LearnerEmpId2
        
        sheet.cell(row=810, column=2).value = FirstNameUpdated3    
        sheet.cell(row=812, column=2).value = LearnerEmailId3
        sheet.cell(row=813, column=2).value = LearnerEmpId3
        
        wb.save(os.path.join('Test_Data/TestData.xlsx'))
        
        
    def mainLearnerCountForAssignOptionForTriggering(self):
        try:
            book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
            first_sheet = book.sheet_by_name("CampAssign")
            
            cell = first_sheet.cell(795,1)
            groupName1 = cell.value
            
            cell = first_sheet.cell(796,1)
            groupName2 = cell.value
            
            cell = first_sheet.cell(797,1)
            FirstName1= cell.value
        
            cell = first_sheet.cell(798,1)
            LastName1 = cell.value
        
            cell = first_sheet.cell(799,1)
            Email1 = cell.value
        
            cell = first_sheet.cell(800,1)
            EmployeeId1 = cell.value
        
            cell = first_sheet.cell(801,1)
            Password1 = cell.value
        
            cell = first_sheet.cell(802,1)
            NewPassword1 = cell.value
            
            
            #Second User
            
            cell = first_sheet.cell(803,1)
            FirstName2= cell.value
        
            cell = first_sheet.cell(804,1)
            LastName2 = cell.value
        
            cell = first_sheet.cell(805,1)
            Email2 = cell.value
        
            cell = first_sheet.cell(806,1)
            EmployeeId2 = cell.value
        
            cell = first_sheet.cell(807,1)
            Password2 = cell.value
        
            cell = first_sheet.cell(808,1)
            NewPassword2 = cell.value
            
            #Third User
            
            cell = first_sheet.cell(809,1)
            FirstName3= cell.value
        
            cell = first_sheet.cell(810,1)
            LastName3= cell.value
        
            cell = first_sheet.cell(811,1)
            Email3 = cell.value
        
            cell = first_sheet.cell(812,1)
            EmployeeId3 = cell.value
        
            cell = first_sheet.cell(813,1)
            Password3 = cell.value
        
            cell = first_sheet.cell(814,1)
            NewPassword3 = cell.value
            
            #lesson Creation
            cell = first_sheet.cell(815,1)
            lessonName = cell.value
           
            cell = first_sheet.cell(816,1)
            questionCard = cell.value
            
            cell = first_sheet.cell(817,1)
            ans1 = cell.value
            
            cell = first_sheet.cell(818,1)
            ans2 = cell.value
            
            #campaign Creation
            
            cell = first_sheet.cell(819,1)
            campaignTitle = cell.value
            
            cell = first_sheet.cell(820,1)
            campDescription = cell.value
            
            cell = first_sheet.cell(821,1)
            numberOfAttempts = cell.value
            
            cell = first_sheet.cell(822,1)
            minPassingScore = cell.value
            
        
            book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
            second_sheet = book.sheet_by_name('Login_Credentials')
            cell = second_sheet.cell(1,1)
            url = cell.value
        
            cell = second_sheet.cell(3,1)
            username = cell.value
        
            cell = second_sheet.cell(3,2)
            password = cell.value
            
            print "Creating Users"
            ob=CreateLearner()
            #ob.createNewLearnerMain(FirstName1,LastName1,Email1,EmployeeId1,Password1,NewPassword1,url,username,password)
            #ob.createNewLearnerMain(FirstName2,LastName2,Email2,EmployeeId2,Password2,NewPassword2,url,username,password)
            #ob.createNewLearnerMain(FirstName3,LastName3,Email3,EmployeeId3,Password3,NewPassword3,url,username,password)
            print "Creatimng Groups"
            grp=LearnerCountForAssignOptionForTriggering()
            grp.groupCreation(groupName1,FirstName1,FirstName2)
            grp.groupCreation(groupName2,FirstName1,FirstName3)
            print "\nCreating a lesson\n"
            #ot=CreateLessonDifferentLessons()
            #ot.lessonWithQuestion(lessonName, questionCard, ans1, ans2)
            
            print "Creating Campaign with trigger"
            cam=LearnerCountForAssignOptionForTriggering() 
            cam.campaignCreationWithTrigger(campaignTitle,lessonName,campDescription,numberOfAttempts,minPassingScore,groupName1,groupName2)
            
        except Exception as e:
            traceback.print_exc()
            print (e)
            raise Exception   
          
        finally:
            ob=LearnerCountForAssignOptionForTriggering()
            ob.updateUser()
            grpd= GroupDelete()
            grpd.deleteGroup(groupName1)
            grpd.deleteGroup(groupName2)
            d=DeleteLesson()
            d.lessonDelete(lessonName)  
            second_sheet = book.sheet_by_name('Login_Credentials')
            cell = second_sheet.cell(1,1)
            url = cell.value
            driver.get(url)
            try:
                WebDriverWait(driver, 5).until(EC.alert_is_present())

                alert = driver.switch_to.alert
                alert.accept()
                print("alert accepted")
            except Exception:
                print("no alert")
    
        
        
        
if __name__=='__main__':
    
    btc=BaseTestClass()
    btc.userLogin() 

    gr=LearnerCountForAssignOptionForTriggering()
    gr.mainLearnerCountForAssignOptionForTriggering() 