'''
Created on 28-Mar-2018

@author: dattatraya
'''

import os.path
import time
import traceback

from BaseTestClass import driver
from openpyxl.reader.excel import load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import xlrd
from BaseTestClass import projectPath
from DeleteLesson import DeleteLesson
from CampaignPageElements import CampPage
from CreateLearnerNew import CreateLearner
from CreateLessonDifferentLessons import CreateLessonDifferentLessons
from ManageAssignMentForPageElements import ManageAssignMentForPageElements
from BaseTestClass import projectPath

class AssignCampToCancelSentAssignment:
    
    def campaignTolearnerToCancel(self,campaignTitle,campDescription,lessonName,nameOFuser):
        
        elements=CampPage()
        driver.refresh()
        wait=WebDriverWait(driver, 60)
        
        driver.refresh()
        print "\n\nCreating Campaign"
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.campaignButtonFromSideMenuXpath())))
        elements.campaignButtonFromSideMenu()
        
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.createCampaignButtonXpath())))

        if elements.campaignsPageHeaderText()=="Campaigns":
            print "Campaigns page displayed"
        else:
            print "Campaigns page is not displayed"
            raise Exception
        
        
        print "Clicking on Create Campaign button"
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.createCampaignButtonXpath())))
        elements.createCampaignButton()
        
        
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.Camp_titleXpath())))
        print "Create Campaign page is displayed"
        
                  
        elements.titleTextField(campaignTitle)
        print "Title entered ::campTitle"
        
        elements.descriptionField(campDescription)
        print "Description entered ::campDescription"
        
        print "Adding Lesson"
        #Add lesson button
        elements.addlessonButton()
        
        #Waiting until first lesson in pop is displayed
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.FirstLessonWaitXpath())))
        
        #Searching lesson by its name
        elements.searchLesson(lessonName)
        
        #Waiting until lesson displayed
        elements.waitUntilSearchedLessonDisplayed(lessonName)
        
        #selecting searched lesson
        elements.selectSearchedLesson(lessonName)
        
        #waiting until add to campaign button is click able
        wait.until(EC.element_to_be_clickable((By.XPATH,elements.AddToCampaign_ButtonXpath())))
        
        #Clicking on Add to Campaign button
        elements.addToCampaignButton()
        
        
        
        #Verifying Added lesson is displayed in Grid
        print "\nVerifying Added lesson is displayed in Grid"
        if elements.firstLessonInGrid()==lessonName:
            print "Lesson displayed in grid"
        else:
            print "Lesson not displayed in grid"
        
        wait.until(EC.element_to_be_clickable((By.XPATH,elements.SaveAndExit_ButtonXpath())))
        #Clicking on save & exit button
        print "Clicking on Save and plan assignmet button"
        elements.saveAndPlanAssignmentbutton()
        
        
        #Verifying Plan assignment for page is displayed
        print "Verifying Plan assignment for page is displayed"
        if campaignTitle in elements.planAssignementForHeaderText():
            print "Page '"+elements.planAssignementForHeaderText()+"' is displayed"
        else:
            print "Plan assignment page is not displayed"
            raise Exception
        
        
        #Checking One time triggered radio button is selected by Default
        if driver.find_element_by_xpath("//input[@id='assignment-one-time']").is_selected():
            print "Radio button One Time triggered is selected"
        else:
            print "Radio button is not selected"
            raise Exception
        
        
        print "Difference between two dates"
        from datetime import datetime
        date_format = "%m/%d/%Y"
        a = datetime.strptime(elements.assignDateValue(), date_format)
        b = datetime.strptime(elements.dueDateValue(), date_format)
        delta = b - a
        print "The difference between two dates is ::"
        print delta.days
        
        
        #Adding user to this campaign
        elements.addingUser(nameOFuser)
        
        
        print "Verifying User is displayed in Grid"
        if nameOFuser in elements.userInGrid():
            print "User is displayed in Grid"
        else:
            print "user is not displayed in grid"
            raise Exception        
        
        
        
        elements.sendAssignment()
        print "Clicked on Send assignment button"
        
        
        print "Verifying Confirm pop up is displayed"
        if elements.confirmPopupsendingAssignement().is_displayed():
            print "Confirm assignment pop up is displayed"
        else:
            print "Confirm assignment pop up is not displayed"
            raise Exception
        
        
        elements.sendAssignmentFromPopup()
        print "Clicked on Send Assignment button from pop up"
        
        
        
        #Verifying Campaign Detail page is displayed
        print "\nVerifying campaign detail page is displayed"
        
        if elements.campaignDetailPageHeaderText()==campaignTitle:
            print "Campaign detail page is displayed"
        else:
            print "Campaign detail page is not displayed"
            raise Exception
        
        #verifying in Campaigns displayed in Campaigns grid
        elements.searchingForlesson(campaignTitle)
        
        if elements.actualCampTitleINGrid()==campaignTitle:
            print "Campaign '"+campaignTitle+"' displayed in Grid"
        
        else:
            print "Campaign is not displayed in Grid"
            raise Exception
        
        
        


    def assignmentCheck(self,EmailId,Password,campaignTitle,lessonName,username,password):
        
        print "\n\nChecking sent assignment is displayed for User"
        wait=WebDriverWait(driver, 60)
        driver.refresh()
        wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")))
        
        
        print "Logging out"
        print "Clicking on Username Dropdown"
        ele =driver.find_element_by_xpath(".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")
        driver.execute_script('arguments[0].click()',ele)
        elem=driver.find_element_by_xpath("html/body/div/div/div[1]/div[2]/div[2]/a")
        driver.execute_script('arguments[0].click()',elem)
        
        print "Clicked on Sign Out option"
        
        element = wait.until(EC.presence_of_element_located((By.ID, "password")))
        
        if driver.title == "Grovo":
            print("Grovo Application URL Opened")
        else:
            raise Exception.message

        print "Grovo Sign-In page is displayed"
        
        print "Entering User name"
        driver.find_element_by_xpath(".//*[@id='username']").send_keys(EmailId)
       
        print "Entering Password"
        element.send_keys(Password)
        
        element.send_keys(Keys.TAB)
        print "Clicking on Sign_In button"
        driver.find_element_by_xpath("//*[@id='submitButton']").click()
        
        print "Successfully Logged Into Users account"
        time.sleep(4)
        
        
        ele3=wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")))
        driver.execute_script('arguments[0].click()',ele3)
        
        ele4 =driver.find_element_by_xpath(".//*[@id='content']/div/div[1]/div[2]/div[1]/a[1]")
        driver.execute_script('arguments[0].click()',ele4)  
        
        
        print "Checking Assignment is displayed for user"
        wait.until(EC.visibility_of_all_elements_located((By.XPATH,"//div/div[1]/div/div/a[1]/span")))
        driver.find_element_by_xpath("//div/div[1]/div/div/a[1]/span[contains(.,'"+campaignTitle+"')]")
        
        if driver.find_element_by_xpath("//div/div[1]/div/div/a[1]/span[contains(.,'"+campaignTitle+"')]").is_displayed():
            print "Campaign displayed for User in Current assignment section"
            
        else:
            print "Campaign not displayed for User"
            raise Exception
        
        print "Starting assignment"
        
        
        startButton=wait.until(EC.element_to_be_clickable((By.XPATH,"(//div/div[1]/div/div/a[1]/span[contains(.,'"+campaignTitle+"')]/../../a[2])[1]")))
        startButton.click()
        
        print "Waiting for lesson name to be displayed"
        
        lessonNameForUser=wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div/div/div[4]/div/div/div/div[1]/h1/span")))
        lessonNameForUserText=lessonNameForUser.text
        if lessonName==lessonNameForUserText:
            print "Verified Lesson '"+lessonNameForUserText+"' is displayed for User"
        else:
            print "Lesson displayed is not valid"
            raise Exception
        
        
        backButtoon=wait.until(EC.element_to_be_clickable((By.XPATH,".//*[@id='content']/div/div/div/div[1]/span/button")))
        backButtoon.click()
        
        unDropDown=wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")))
        driver.execute_script('arguments[0].click()',unDropDown)
        
        signOut=driver.find_element_by_xpath("html/body/div/div/div[1]/div[2]/div[2]/a")
        driver.execute_script('arguments[0].click()',signOut)
        
        
        
        element = wait.until(EC.presence_of_element_located((By.ID, "password")))
        
        if driver.title == "Grovo":
            print("Grovo Application URL Opened")
        else:
            raise Exception.message

        print "Grovo Sign-In page is displayed"
        
        print "Entering User name"
        driver.find_element_by_xpath(".//*[@id='username']").send_keys(username)
       
        print "Entering Password"
        element.send_keys(password)
        
        element.send_keys(Keys.TAB)
        print "Clicking on Sign_In button"
        driver.find_element_by_xpath("//*[@id='submitButton']").click()
        
        print "Successfully Logged Into Users account"
        
        time.sleep(5)
    
    
    def cancelAssignment(self,campaignTitle,nameOFuser):
        
        elements=CampPage()
        
        wait=WebDriverWait(driver, 120)
        driver.refresh()
        
        print "\n\nCreating Campaign"
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.campaignButtonFromSideMenuXpath())))
        elements.campaignButtonFromSideMenu()
        
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.createCampaignButtonXpath())))

        if elements.campaignsPageHeaderText()=="Campaigns":
            print "Campaigns page displayed"
        else:
            print "Campaigns page is not displayed"
            raise Exception
        
        print "Clicking on Created Campaign from Grid"
        
        elements.CampaignLinkFromCampaignGrid(campaignTitle)
        
       
        user=wait.until(EC.visibility_of_element_located((By.XPATH,"//tbody/tr/td[1]/a/span/span")))
        user.click()
        print "Clicked on User"
        
        print "Cancelling the Assignment"
        
        
        manage=ManageAssignMentForPageElements()
        manage.cancelAssignMentClick()
        
        
        print "Clicked on Cancel Assignment"
        
        print "Clicked on Cancel Assignment from pop up"
        
        manage.cancelAssignmentButtonFromPopup()
        
        
        driver.refresh()
        
        
    def checkCancelledAssigment(self,EmailId,Password,username,password,campaignTitle):
        wait=WebDriverWait(driver, 60)
        print "Logging out as current user"
        print "Clicking on Username Drop down"
        driver.refresh()
        ele =driver.find_element_by_xpath(".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")
        driver.execute_script('arguments[0].click()',ele)
        elem=driver.find_element_by_xpath("html/body/div/div/div[1]/div[2]/div[2]/a")
        driver.execute_script('arguments[0].click()',elem)
        
        print "Clicked on Sign Out option"
        
        element = wait.until(EC.presence_of_element_located((By.ID, "password")))
        
        if driver.title == "Grovo":
            print("Grovo Application URL Opened")
        else:
            raise Exception.message

        print "Grovo Sign-In page is displayed"
        
        print "Entering User name"
        driver.find_element_by_xpath(".//*[@id='username']").send_keys(EmailId)
       
        print "Entering Password"
        element.send_keys(Password)
        
        element.send_keys(Keys.TAB)
        print "Clicking on Sign_In button"
        driver.find_element_by_xpath("//*[@id='submitButton']").click()
        
        print "Successfully Logged Into Users account"
        time.sleep(30)
        
        driver.refresh()
        time.sleep(10)
        driver.refresh()
        driver.refresh()
        
        ele3=wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")))
        driver.execute_script('arguments[0].click()',ele3)
        
        ele4 =driver.find_element_by_xpath(".//*[@id='content']/div/div[1]/div[2]/div[1]/a[1]")
        driver.execute_script('arguments[0].click()',ele4)
        
        listsCamps=wait.until(EC.visibility_of_all_elements_located((By.XPATH,"//div/div[1]/div/div/a[1]/span")))
        
        count = len(listsCamps)
        for count in range (0,count):
            if listsCamps[count].text!=campaignTitle:
                print "Campaign is not displayed"
            else:
                print "Campaign is displayed"
                raise Exception
                
        
        
        print "Logging in as a Original user"
        
        unDropDown=wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")))
        driver.execute_script('arguments[0].click()',unDropDown)
        
        signOut=driver.find_element_by_xpath("html/body/div/div/div[1]/div[2]/div[2]/a")
        driver.execute_script('arguments[0].click()',signOut)
        
        
        
        element = wait.until(EC.presence_of_element_located((By.ID, "password")))
        
        if driver.title == "Grovo":
            print("Grovo Application URL Opened")
        else:
            raise Exception.message

        print "Grovo Sign-In page is displayed"
        
        print "Entering User name"
        driver.find_element_by_xpath(".//*[@id='username']").send_keys(username)
       
        print "Entering Password"
        element.send_keys(password)
        
        element.send_keys(Keys.TAB)
        print "Clicking on Sign_In button"
        driver.find_element_by_xpath("//*[@id='submitButton']").click()
        
        print "Successfully Logged Into Users account"
        
        time.sleep(5)
        
    
    def assignCampToCancelAssignment(self):
        
        book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        first_sheet = book.sheet_by_name('CampAssign')
        
        cell1 = first_sheet.cell(325,1)
        campaignTitle = cell1.value
        
        cell1 = first_sheet.cell(326,1)
        campDescription = cell1.value
        
        cell1 = first_sheet.cell(332,1)
        nameOFuser = cell1.value
        
        cell1 = first_sheet.cell(327,1)
        lessonName = cell1.value
        
        
        
        cell2 = first_sheet.cell(328,1)
        questionCard= cell2.value
        
        cell2 = first_sheet.cell(329,1)
        ans1= cell2.value
        
        cell2 = first_sheet.cell(330,1)
        ans2= cell2.value
        
        
        #Learner
        cell = first_sheet.cell(332,1)
        FirstName = cell.value
        
        FirstNameId = FirstName.split("_")
        emp = FirstNameId[0]+"_"
        ids = FirstNameId[1]
        empId = int(ids)+1
        FirstNameUpdated= emp+str(empId)
        
        cell = first_sheet.cell(333,1)
        LastName = cell.value
        
        LastNameID = LastName.split("_")
        emp = LastNameID[0]+"_"
        ids = LastNameID[1]
        empId = int(ids)+1
        LastNameUpdated= emp+str(empId)
        
        
        
        cell = first_sheet.cell(334,1)
        Email = cell.value
        
        EmailId = Email.split("_")
        emp = EmailId[0]+"_"
        ids = EmailId[1]
        remaining="_"+EmailId[2]
        empId = int(ids)+1
        EmailIdUpdated= emp+str(empId)+remaining
        
        cell = first_sheet.cell(335,1)
        EmployeeId = cell.value
        
        Employee = EmployeeId.split("_")
        emp = Employee[0]+"_"
        ids = Employee[1]
        empId = int(ids)+1
        EmployeeIdUpdated= emp+str(empId)
        
        cell = first_sheet.cell(336,1)
        Password = cell.value
        
        cell = first_sheet.cell(337,1)
        NewPassword = cell.value
        
        cell = first_sheet.cell(338,1)
        role = cell.value
        
        #For Original User
        book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        s_sheet = book.sheet_by_name('Login_Credentials')
        
        cell = s_sheet.cell(1,1)
        url = cell.value
        
        cell = s_sheet.cell(3,1)
        username = cell.value
        
        cell = s_sheet.cell(3,2)
        password = cell.value
        
        
        wb = load_workbook(os.path.join('Test_Data/TestData.xlsx'))
            #print (wb.sheetnames)
        ws = wb.active
        sheet = wb['CampAssign']
        
        
        sheet.cell(row=333, column=2).value = FirstNameUpdated
        sheet.cell(row=334, column=2).value = LastNameUpdated
        sheet.cell(row=335, column=2).value = EmailIdUpdated
        sheet.cell(row=336, column=2).value = EmployeeIdUpdated
        
        
        wb.save(os.path.join('Test_Data/TestData.xlsx'))
            
        print "All User Data Updated in Excel"
        
        try:
            print "\nCreating a New Learner\n"
            ot=CreateLearner()
            ot.createNewLearnerMain(FirstName, LastName, Email, EmployeeId, Password, NewPassword, url, username, password)
            
            print "\nCreating a lesson\n"
            obj=CreateLessonDifferentLessons()
            obj.lessonWithQuestion(lessonName, questionCard, ans1, ans2)
            print "\nCreating Campaign\n"
            
            mb=AssignCampToCancelSentAssignment()
            mb.campaignTolearnerToCancel(campaignTitle, campDescription, lessonName, nameOFuser)
            
            print "\nChecking Assignment is displayed for user\n"
            mb.assignmentCheck(Email, Password, campaignTitle, lessonName, username, password)
            
            print "\nCancelling the Current assignment\n"
            mb.cancelAssignment(campaignTitle, nameOFuser)
            
            print "\nChecking for assignment is not displaying for User\n"
            mb.checkCancelledAssigment(Email, Password, username, password, campaignTitle)
            
            print "\Deleting lesson"
            
            
            print "\n----Text Execution Completed----\n"
            
            
        except Exception as e:
            traceback.print_exc()
            print (e)
            raise Exception 
         
        finally:  
            driver.save_screenshot("ScreenShots/AssignCampToCancelSentAssignment.png")
            
            
            second_sheet = book.sheet_by_name('Login_Credentials')
            cell = second_sheet.cell(1,1)
            url = cell.value
            driver.get(url)
            try:
                WebDriverWait(driver, 5).until(EC.alert_is_present())

                alert = driver.switch_to.alert
                alert.accept()
                print("alert accepted")
            except Exception:
                print("no alert")
                
            
            wait=WebDriverWait(driver, 60)
            unDropDown=wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")))
            driver.execute_script('arguments[0].click()',unDropDown)
            
            signOut=driver.find_element_by_xpath("html/body/div/div/div[1]/div[2]/div[2]/a")
            driver.execute_script('arguments[0].click()',signOut)
            
            
            
            element = wait.until(EC.presence_of_element_located((By.ID, "password")))
            
            if driver.title == "Grovo":
                print("Grovo Application URL Opened")
            else:
                raise Exception.message
    
            print "Grovo Sign-In page is displayed"
            
            print "Entering User name"
            driver.find_element_by_xpath(".//*[@id='username']").send_keys(username)
           
            print "Entering Password"
            element.send_keys(password)
            
            element.send_keys(Keys.TAB)
            print "Clicking on Sign_In button"
            driver.find_element_by_xpath("//*[@id='submitButton']").click()
            
            print "Successfully Logged Into Users account"
            time.sleep(3)
            
            try:
                oj=DeleteLesson()
                oj.lessonDelete(lessonName)
            except Exception:
                driver.get(url)
            
