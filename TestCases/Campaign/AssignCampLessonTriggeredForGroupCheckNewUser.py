'''
Created on 23-Mar-2018

@author: dattatraya
'''

import os.path
import time
import traceback

from BaseTestClass import BaseTestClass
from BaseTestClass import driver
from openpyxl.reader.excel import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import xlrd

from CampaignPageElements import CampPage
from CreateLearnerNew import CreateLearner
from CreateLessonDifferentLessons import CreateLessonDifferentLessons
from DeleteLesson import DeleteLesson
from GroupsPageElements import GroupsPageElements
from BaseTestClass import projectPath

class AssignCampLessonTriggeredForGroupCheckNewUser:
    
    def campaignToGroupTriggeredCheckNewUser(self,campaignTitle,campDescription,lessonName,groupName,minPassingScore,numberOfAttempts,criteriaActualText,FirstName,LastName):
        
        elements=CampPage()
        driver.refresh()
        wait=WebDriverWait(driver, 60)
        
        
        print "\n\nCreating Campaign"
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.campaignButtonFromSideMenuXpath())))
        elements.campaignButtonFromSideMenu()
        
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.createCampaignButtonXpath())))

        if elements.campaignsPageHeaderText()=="Campaigns":
            print "Campaigns page displayed"
        else:
            print "Campaigns page is not displayed"
            raise Exception
        
        
        print "Clicking on Create Campaign button"
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.createCampaignButtonXpath())))
        elements.createCampaignButton()
        
        
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.Camp_titleXpath())))
        print "Create Campaign page is displayed"
        
                  
        elements.titleTextField(campaignTitle)
        print "Title entered"
        
        elements.descriptionField(campDescription)
        print "Description entered"
        
        print "Adding Lesson"
        #Add lesson button
        elements.addlessonButton()
        
        #Waiting until first lesson in pop is displayed
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.FirstLessonWaitXpath())))
        
        #Searching lesson by its name
        elements.searchLesson(lessonName)
        
        #Waiting until lesson displayed
        elements.waitUntilSearchedLessonDisplayed(lessonName)
        
        #selecting searched lesson
        elements.selectSearchedLesson(lessonName)
        
        #waiting until add to campaign button is click able
        wait.until(EC.element_to_be_clickable((By.XPATH,elements.AddToCampaign_ButtonXpath())))
        
        #Clicking on Add to Campaign button
        elements.addToCampaignButton()
        
        
        time.sleep(2)
        #Verifying Added lesson is displayed in Grid
        print "\nVerifying Added lesson is displayed in Grid"
        if elements.firstLessonInGrid()==lessonName:
            print "Lesson displayed in grid"
        else:
            print "Lesson not displayed in grid"
            raise Exception
        
        
        print "Making this as a Graded campaign"
        elements.makeThisAsAGradedCampaign()
        print "Settng minimum passing score"
        elements.setMinimumPassingScore(minPassingScore)
        print "Setting maximum no of attempts"
        elements.setAMaxNoOfAttempts(numberOfAttempts)
        
        
        wait.until(EC.element_to_be_clickable((By.XPATH,elements.SaveAndExit_ButtonXpath())))
        print "Clicking on Save and plan assignmet button"
        elements.saveAndPlanAssignmentbutton()
        
        
        
        #Verifying Plan assignment for page is displayed
        print "Verifying Plan assignment for page is displayed"
        if campaignTitle in elements.planAssignementForHeaderText():
            print "Page '"+elements.planAssignementForHeaderText()+"' is displayed"
        else:
            print "Plan assignment page is not displayed"
            raise Exception
        
        
        #Checking One time triggered radio button is selected by Default
        if driver.find_element_by_xpath("//input[@id='assignment-one-time']").is_selected():
            print "Radio button One Time triggered is selected"
        else:
            print "Radio button is not selected"
            raise Exception
        
        
        
        print "Clicking on Triggered radio button"
        
        driver.find_element_by_xpath(".//*[@id='content']/div/div[3]/div[2]/div/section/div/div[2]/label/span[2]").click()
        wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[3]/div[2]/div/div[2]/section[1]/div/div[1]/label/span[2]")))
        
        print "Clicked on Triggered"
        
        #Checking New hire on boarding is selected 
        if driver.find_element_by_xpath("//input[@id='trigger-new-hire']").is_selected():
            print "Radio button New Hire Onboarding is selected"
        else:
            print "Radio button New Hire Onboarding is not selected"
            raise Exception
        
        

        
        print "Selecting New to group"
        driver.find_element_by_xpath(".//*[@id='content']/div/div[3]/div[2]/div/div[2]/section[1]/div/div[2]/label/span[2]").click()
        wait.until(EC.visibility_of_element_located((By.XPATH,"//div[@class='Select-placeholder']"))) 
        
        searchGroups=driver.find_element_by_xpath("//div[@class='Select-placeholder']")
        
        webdriver.ActionChains(driver).move_to_element(searchGroups).click().send_keys(groupName).perform()
        groupdisplayed=wait.until(EC.visibility_of_element_located((By.XPATH,"(//div[@role='option']/span)[1]")))
        
        webdriver.ActionChains(driver).move_to_element(groupdisplayed).click().perform()
        
        
        print "Checking Group is displayed in Grid"
        
        if driver.find_element_by_xpath("//table/tbody/tr/td[1]/a").text==groupName:
            print "Group '"+groupName+"' is displayed in grid"
        else:
            print "Group is not displayed in Grid"
            raise Exception
        
        
        print "Clicking on checkbox 'Also assign to learners who currently match this criteria' "
        
        driver.find_element_by_xpath(".//*[@id='content']/div/div[3]/div[2]/div/div[2]/section[1]/div/div[3]/div[2]/label/span[2]").click()
        wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[3]/div[2]/div/div[2]/section[1]/div/div[3]/div[3]")))
        
        
        print "Checking Learner count is displayed after checking checkbox"
        criteria=driver.find_element_by_xpath(".//*[@id='content']/div/div[3]/div[2]/div/div[2]/section[1]/div/div[3]/div[3]").text
        
        
        if criteriaActualText in criteria:
            print "Text '"+criteria+"' is displayed after checking checkboox"
        else:
            print "Text not displayed"
            raise Exception
        
        
        #Learner Count 
        print "Checking Learner count with button is displayed"
        learnerCount=driver.find_element_by_xpath(".//*[@id='content']/div/div[3]/div[2]/div/div[2]/section[1]/div/div[3]/div[3]/button")
        
        if learnerCount.is_displayed():
            print "Learner count '"+learnerCount.text+"' is displayed"
            
        else:
            print "Learner Count is not displayed"
            raise Exception
               
        learnerCount.click()
        
        try:
            wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div[4]/div/div/div[2]/div[1]/div[2]/table/tbody/tr/td[1]")))
        except Exception:
            print "View Learners Pop up is not displayed"
            raise Exception
        
        
        
        #View Learner
        print "Checking User name is displayed in View Learners grid"
        learnerName=driver.find_element_by_xpath("html/body/div[4]/div/div/div[2]/div[1]/div[2]/table/tbody/tr/td[1]").text
        print learnerName
        if FirstName in learnerName:
            print "Learner Name displayed as ::"+FirstName+" "+LastName
        else:
            print "Learner name is displayed is not valid"
            raise Exception
        
        
        driver.find_element_by_xpath("html/body/div[4]/div/div/div[2]/div[2]/button").click()
        wait.until(EC.invisibility_of_element_located((By.XPATH,"html/body/div[4]/div/div/div[1]")))
        
        
        print "Saving Trigger"
        wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[3]/div[2]/div/div[2]/section[2]/div/div[2]/button")))
        saveTrigger=wait.until(EC.element_to_be_clickable((By.XPATH,".//*[@id='content']/div/div[3]/div[2]/div/div[2]/section[2]/div/div[2]/button")))
        saveTrigger.click()
        
        
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div[4]/div/div/div[2]/div[2]/button[1]")))
        driver.find_element_by_xpath("html/body/div[4]/div/div/div[2]/div[2]/button[1]").click()
        
        print "Clicked on Yes,Save button from pop up"
        newtrigger=wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[3]/div[2]/div/div[3]/table/tbody/tr[1]/td[1]/div[2]/span")))

        print "Checking for In Campaign details page Trigger is displayed"
        if newtrigger.text==groupName:
            print "Trigger with group name is displayed in Campaign details page"
        else:
            print "Trigger is not displayed"
            raise Exception
        
        
        #Verifying Campaign Detail page is displayed
        print "\nVerifying campaign detail page is displayed"
        
        if elements.campaignDetailPageHeaderText()==campaignTitle:
            print "Campaign detail page is displayed"
        else:
            print "Campaign detail page is not displayed"
            raise Exception
        
        #verifying in Campaigns displayed in Campaigns grid
        elements.searchingForlesson(campaignTitle)
        
        if elements.actualCampTitleINGrid()==campaignTitle:
            print "Campaign '"+campaignTitle+"' displayed in Grid"
        
        else:
            print "Campaign is not displayed in Grid"
            raise Exception
        
         
         
        print "\nVerifying 'Y' is displayed for created triggred campaign"
        hasTrigger=driver.find_element_by_xpath("(//tr/td[1]/a[.='"+campaignTitle+"']/../../td[2])[1]").text
        
        if hasTrigger=="Y":
            print "'Y' displayed in Has trigger column"
        else:
            print "Invalid data in Hastrigger column"
            
            
            
    def groupCreateForCampaign(self,groupName,FirstName):
        
        driver.refresh()
        wait=WebDriverWait(driver, 60)
        
        group=GroupsPageElements()
        
        group.adminSideMenuUnexpanded()
        print "Clicked on admin icon"
         
        group.groupSideMenuExpanded()  
        print "Clicked on Group icon"
        
        
        print "Checking Group page is displayed"
        if driver.find_element_by_xpath(group.groupPageHeader()).is_displayed():
            print "Group page is displayed successfully"
        else:
            print "Group page is not displayed"
            raise Exception
        
        
        group.createGroupButton()
        print "Clicked on Create Group button"
        
        group.enteringGroupname(groupName)
        
        print "Group Name entered....."
        
        group.nextButton()
        print "Clicked on Next button"
        
        wait.until(EC.visibility_of_element_located((By.XPATH,group.groupDetailPageHeader())))
        print "Checking group is created"
        
        time.sleep(4)
        headerText=driver.find_element_by_xpath(group.groupDetailPageHeader()).text
        
        if headerText==groupName:
            print "Group created successfully"
            
        else:
            print "Group not created"
            raise Exception
        
        
        print "Adding user to Group by name"
        group.addByNameButton()
        group.addByName(FirstName)
        
        print "Verifying user is added to group"
        group.groupAddedInList(FirstName)
        
        group.saveButton()
        print "Clicked on Save button"

        group.saveButtonPopup()
        print "Clicked on Save from popup"
        
        
        print "Checking Created Group is displayed in Grid"
        group.groupsLinkFromBreadCrumb()
        
        print "Searching Group"
        group.createdGroupDisplayedInGrid(groupName)
         
        
        driver.refresh()
        
    def newUserAddingToGroup(self,groupName,FirstName1):
        driver.refresh()
        
        wait=WebDriverWait(driver, 60)
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div/div/div[3]/div[1]/div/nav/div/div[2]/div[6]/a")))
        driver.find_element_by_xpath("html/body/div/div/div[3]/div[1]/div/nav/div/div[2]/div[6]/a").click()
        print "Clicked on admin icon"
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div/div/div[3]/div[1]/div/nav/div/div[2]/div[6]")))
        driver.find_element_by_xpath("html/body/div/div/div[3]/div[1]/div/nav/div/div[2]/div[6]").click()
        print "Clicked on Admin"
         
         
        driver.find_element_by_xpath(".//*[@id='content']/div/div[3]/div[1]/div/nav/div/div[2]/div[6]/div/ul/li[2]/a").click()   
        print "Clicked on Group icon"
        
        
        
        print "Checking Group page is displayed"
        if driver.find_element_by_xpath(".//*[@id='content']/div/div[3]/div[2]/div/header/h1").is_displayed():
            print "Group page is displayed successfully"
        else:
            print "Group page is not displayed"
            raise Exception
        
        print "Searching for Group"
        
        wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='search-groups']")))
        driver.find_element_by_xpath(".//*[@id='search-groups']").send_keys(groupName)
        ele=wait.until(EC.visibility_of_element_located((By.XPATH,"//table/tbody/tr/td[2]/a[.='"+groupName+"']")))
        print "Group found"
        ele.click()
        
        
        print "Adding user to Group by name"
        addByName=wait.until(EC.element_to_be_clickable((By.XPATH,".//*[@id='content']/div/div[3]/div[2]/div/div[2]/div/div/div[1]/div/div/div[2]/div/button")))
        addByName.click()
        
        
        wait.until(EC.visibility_of_element_located((By.XPATH,"//div[@class='Select-placeholder']")))
        
        names=driver.find_element_by_xpath("//div[@class='Select-placeholder']")
        
        webdriver.ActionChains(driver).move_to_element(names).click().send_keys(FirstName1).perform()
        userDisplayed=wait.until(EC.visibility_of_element_located((By.XPATH,"//div[@role='option']/span[contains(.,'"+FirstName1+"')]")))
        
        webdriver.ActionChains(driver).move_to_element(userDisplayed).click().perform()
        print "User selected and added to group"
        
        tableData=wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[3]/div[2]/div/div[4]/table/tbody/tr/td[.='"+FirstName1+"']")))
        
        print "Verifying user is added to group"
        if tableData.text==FirstName1:
            print "User is displayed in grid"
        else:
            print "User is not displayed in grid"
        
        
        saveButton=wait.until(EC.element_to_be_clickable((By.XPATH,".//*[@id='content']/div/div[3]/div[2]/div/header/div[2]/button[2]")))
        saveButton.click()

        saveFrompopup=wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div[4]/div/div/div[2]/div/button[2]")))
        saveFrompopup.click()
        print "Clicked on Save from popup"
        
        wait.until(EC.invisibility_of_element_located((By.XPATH,"html/body/div[4]/div/div/div/div")))
        wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[3]/div[2]/div/header/div[2]/button")))
        
        print "New User is successfully added to group"
        
        driver.refresh()
            
        
    def assignmentCheck(self,EmailId,Password,campaignTitle,lessonName,username,password):
        driver.refresh()
        print "\n\nChecking sent assignment is displayed for User"
        wait=WebDriverWait(driver, 60)
        driver.refresh()
        wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")))
        
        
        print "Logging out"
        print "Clicking on Username Dropdown"
        ele =driver.find_element_by_xpath(".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")
        driver.execute_script('arguments[0].click()',ele)
        elem=driver.find_element_by_xpath("html/body/div/div/div[1]/div[2]/div[2]/a")
        driver.execute_script('arguments[0].click()',elem)
        
        print "Clicked on Sign Out option"
        
        element = wait.until(EC.presence_of_element_located((By.ID, "password")))
        
        if driver.title == "Grovo":
            print("Grovo Application URL Opened")
        else:
            raise Exception.message

        print "Grovo Sign-In page is displayed"
        
        print "Entering User name"
        driver.find_element_by_xpath(".//*[@id='username']").send_keys(EmailId)
       
        print "Entering Password"
        element.send_keys(Password)
        
        element.send_keys(Keys.TAB)
        print "Clicking on Sign_In button"
        driver.find_element_by_xpath("//*[@id='submitButton']").click()
        
        print "Successfully Logged Into Users account"
        time.sleep(30)
        
        driver.refresh()
        time.sleep(10)
        driver.refresh()
        driver.refresh()
        
        print "Checking in my Learning page Assignment is displayed"
        wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[3]/div[2]/div/div[2]/div/div/div/div[1]/div/div/div[2]/div[1]/div/div/a[1]/span")))
        
        unDropDown1=wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")))
        driver.execute_script('arguments[0].click()',unDropDown1)
        
        mylearning=wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[1]/div[2]/div[1]/a[1]")))
        driver.execute_script('arguments[0].click()',mylearning)
        
        assignmentText=wait.until(EC.visibility_of_element_located((By.XPATH,"//div[2]/div[1]/div/div/a[1]/span[contains(.,'"+campaignTitle+"')]")))
        
        
        
        if campaignTitle in assignmentText.text:
            print "Campaign displayed for User in My learning page"
            
        else:
            print "Campaign not displayed for User"
            raise Exception
        
        print "Starting assignment"
        
        
        startButton=wait.until(EC.visibility_of_element_located((By.XPATH,"//a/span[contains(.,'"+campaignTitle+"')]/../../a[2]")))
        startButton.click()
        
        print "Waiting for lesson name to be displayed"
        
        lessonNameForUser=wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div/div/div[4]/div/div/div/div[1]/h1/span")))
        lessonNameForUserText=lessonNameForUser.text
        if lessonName==lessonNameForUserText:
            print "Verified Lesson '"+lessonNameForUserText+"' is displayed for User"
        else:
            print "Lesson displayed is not valid"
            raise Exception
        
        
        backButtoon=wait.until(EC.element_to_be_clickable((By.XPATH,".//*[@id='content']/div/div/div/div[1]/span/button")))
        backButtoon.click()
        
        unDropDown=wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")))
        driver.execute_script('arguments[0].click()',unDropDown)
        
        signOut=driver.find_element_by_xpath("html/body/div/div/div[1]/div[2]/div[2]/a")
        driver.execute_script('arguments[0].click()',signOut)
        
        
        
        element = wait.until(EC.presence_of_element_located((By.ID, "password")))
        
        if driver.title == "Grovo":
            print("Grovo Application URL Opened")
        else:
            raise Exception.message

        print "Grovo Sign-In page is displayed"
        
        print "Entering User name"
        driver.find_element_by_xpath(".//*[@id='username']").send_keys(username)
       
        print "Entering Password"
        element.send_keys(password)
        
        element.send_keys(Keys.TAB)
        print "Clicking on Sign_In button"
        driver.find_element_by_xpath("//*[@id='submitButton']").click()
        
        print "Successfully Logged Into Users account"
        
        time.sleep(5)
    
    
    def assignCampWithOneLessonForGroupTriggeredCheckAssignmentForNewUser(self):
        
        book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        first_sheet = book.sheet_by_name('CampAssign')
        
        
        #Campaign
        cell1 = first_sheet.cell(230,1)
        campaignTitle = cell1.value
        
        cell1 = first_sheet.cell(231,1)
        campDescription = cell1.value
        
        cell1 = first_sheet.cell(252,1)
        groupName = cell1.value
        
        groupNameId = groupName.split("_")
        emp = groupNameId[0]+"_"
        ids = groupNameId[1]
        empId = int(ids)+1
        GroupNameUpdated= emp+str(empId)
        
        cell1 = first_sheet.cell(232,1)
        minPassingScore = cell1.value
        
        cell1 = first_sheet.cell(233,1)
        numberOfAttempts = cell1.value
        
        cell1 = first_sheet.cell(234,1)
        lessonName = cell1.value
        
        cell2 = first_sheet.cell(235,1)
        questionCard= cell2.value
        
        cell2 = first_sheet.cell(236,1)
        ans1= cell2.value
        
        cell2 = first_sheet.cell(237,1)
        ans2= cell2.value
        
        cell2 = first_sheet.cell(250,1)
        criteriaActualText= cell2.value
        
        
        
        
        #Learner 1
        cell = first_sheet.cell(239,1)
        FirstName = cell.value
        
        FirstNameId = FirstName.split("_")
        emp = FirstNameId[0]+"_"
        ids = FirstNameId[1]
        empId = int(ids)+1
        FirstNameUpdated= emp+str(empId)
        
        cell = first_sheet.cell(240,1)
        LastName = cell.value
        
        LastNameID = LastName.split("_")
        emp = LastNameID[0]+"_"
        ids = LastNameID[1]
        empId = int(ids)+1
        LastNameUpdated= emp+str(empId)
        
        
        
        cell = first_sheet.cell(241,1)
        Email = cell.value
        
        EmailId = Email.split("_")
        emp = EmailId[0]+"_"
        ids = EmailId[1]
        remaining="_"+EmailId[2]
        empId = int(ids)+1
        EmailIdUpdated= emp+str(empId)+remaining
        
        cell = first_sheet.cell(242,1)
        EmployeeId = cell.value
        
        Employee = EmployeeId.split("_")
        emp = Employee[0]+"_"
        ids = Employee[1]
        empId = int(ids)+1
        EmployeeIdUpdated= emp+str(empId)
        
        cell = first_sheet.cell(248,1)
        Password = cell.value
        
        cell = first_sheet.cell(249,1)
        NewPassword = cell.value
        
        cell = first_sheet.cell(250,1)
        role = cell.value
        
        #For Original User
        book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        s_sheet = book.sheet_by_name('Login_Credentials')
        
        cell = s_sheet.cell(1,1)
        url = cell.value
        
        cell = s_sheet.cell(3,1)
        username = cell.value
        
        cell = s_sheet.cell(3,2)
        password = cell.value
        
        
        #Learner 2
        
        cell = first_sheet.cell(244,1)
        FirstName1 = cell.value
        
        FirstNameId1 = FirstName1.split("_")
        emp = FirstNameId1[0]+"_"
        ids = FirstNameId1[1]
        empId = int(ids)+1
        FirstNameUpdated1= emp+str(empId)
        
        cell = first_sheet.cell(245,1)
        LastName1 = cell.value
        
        LastNameID1 = LastName1.split("_")
        emp = LastNameID1[0]+"_"
        ids = LastNameID1[1]
        empId = int(ids)+1
        LastNameUpdated1= emp+str(empId)
        
        
        
        cell = first_sheet.cell(246,1)
        Email1 = cell.value
        
        EmailId1 = Email1.split("_")
        emp = EmailId1[0]+"_"
        ids = EmailId1[1]
        remaining="_"+EmailId[2]
        empId = int(ids)+1
        EmailIdUpdated1= emp+str(empId)+remaining
        
        cell = first_sheet.cell(247,1)
        EmployeeId1 = cell.value
        
        Employee1 = EmployeeId1.split("_")
        emp = Employee1[0]+"_"
        ids = Employee1[1]
        empId = int(ids)+1
        EmployeeIdUpdated1= emp+str(empId)
        
        
        #updating user values
        wb = load_workbook(os.path.join('Test_Data/TestData.xlsx'))
            #print (wb.sheetnames)
        
        sheet = wb['CampAssign']
        
        sheet.cell(row=240, column=2).value = FirstNameUpdated
        sheet.cell(row=241, column=2).value = LastNameUpdated
        sheet.cell(row=242, column=2).value = EmailIdUpdated
        sheet.cell(row=243, column=2).value = EmployeeIdUpdated
        
        sheet.cell(row=253, column=2).value = GroupNameUpdated
        
        
        sheet.cell(row=245, column=2).value = FirstNameUpdated1
        sheet.cell(row=246, column=2).value = LastNameUpdated1
        sheet.cell(row=247, column=2).value = EmailIdUpdated1
        sheet.cell(row=248, column=2).value = EmployeeIdUpdated1
        
        wb.save(os.path.join('Test_Data/TestData.xlsx'))
            
        print "All User Data Updated in Excel"
        
        
        
        
        try:
            print "\nCreating a New Learner\n"
            lt=CreateLearner()
            lt.createNewLearnerMain(FirstName, LastName, Email, EmployeeId, Password, NewPassword, url, username, password)
            
            print "\nCreating Group\n"
            lk=AssignCampLessonTriggeredForGroupCheckNewUser()
            lk.groupCreateForCampaign(groupName, FirstName)
            
            print "\nCreating a lesson\n"
            ot=CreateLessonDifferentLessons()
            ot.lessonWithQuestion(lessonName, questionCard, ans1, ans2)
            
            
            print "\nCreating Campaign\n"
            lk.campaignToGroupTriggeredCheckNewUser(campaignTitle, campDescription, lessonName, groupName, minPassingScore, numberOfAttempts, criteriaActualText, FirstName, LastName)
            
            print "\nCreating second user\n"
            driver.refresh()
            lt.createNewLearnerMain(FirstName1, LastName1, Email1, EmployeeId1, Password, NewPassword, url, username, password)
            
            lk.newUserAddingToGroup(groupName, FirstName1)
            
            print "\nChecking for new user Assignment is sent and displayed in Home page"
            lk.assignmentCheck(Email1, Password, campaignTitle, lessonName, username, password)
            
            
            print "\n----Test Execution Completed----\n"
            
            
        except Exception as e:
            traceback.print_exc()
            print (e)
            raise Exception 
         
        finally:
            driver.save_screenshot("ScreenShots/AssignCampLessonTriggeredForGroupCheckNewUser.png")
              
            second_sheet = book.sheet_by_name('Login_Credentials')
            cell = second_sheet.cell(1,1)
            url = cell.value
            driver.get(url)
            
            try:
                WebDriverWait(driver, 5).until(EC.alert_is_present())

                alert = driver.switch_to.alert
                alert.accept()
                print("alert accepted")
            except Exception:
                print("no alert")
                
            
            
            wait=WebDriverWait(driver, 60)
            unDropDown=wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")))
            driver.execute_script('arguments[0].click()',unDropDown)
            
            signOut=driver.find_element_by_xpath("html/body/div/div/div[1]/div[2]/div[2]/a")
            driver.execute_script('arguments[0].click()',signOut)
            
            
            
            element = wait.until(EC.presence_of_element_located((By.ID, "password")))
            
            
            time.sleep(3)
            if driver.title == "Grovo":
                print("Grovo Application URL Opened")
            else:
                raise Exception.message
    
            print "Grovo Sign-In page is displayed"
            
            print "Entering User name"
            driver.find_element_by_xpath(".//*[@id='username']").send_keys(username)
           
            print "Entering Password"
            element.send_keys(password)
            
            element.send_keys(Keys.TAB)
            print "Clicking on Sign_In button"
            driver.find_element_by_xpath("//*[@id='submitButton']").click()
            
            print "Successfully Logged Into Users account"
            time.sleep(3)
            
            
            try:
                d=DeleteLesson()
                d.lessonDelete(lessonName) 
            except Exception:
                driver.get(url)


if __name__ == '__main__':
    
    btc=BaseTestClass()
    btc.UserLogin()
    
    ne=AssignCampLessonTriggeredForGroupCheckNewUser()
    ne.assignCampWithOneLessonForGroupTriggeredCheckAssignmentForNewUser()



