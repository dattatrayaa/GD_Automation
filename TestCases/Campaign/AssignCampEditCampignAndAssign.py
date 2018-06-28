'''
Created on 02-Apr-2018

@author: dattatraya
'''

import os.path
import time
import traceback

from BaseTestClass import BaseTestClass
from BaseTestClass import driver
from openpyxl.reader.excel import load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import xlrd

from CampaignPageElements import CampPage
from CreateLearnerNew import CreateLearner
from CreateLessonDifferentLessons import CreateLessonDifferentLessons
from DeleteLesson import DeleteLesson
from BaseTestClass import projectPath
class AssignCampEditCampignAndAssign:
    
    def campaignEditAndAssign(self,campaignTitle,campDescription,lessonName,minPassingScore,numberOfAttempts):
        
        elements=CampPage()
        
        wait=WebDriverWait(driver, 60)
        
        driver.refresh()
        print "\n\nCreating Campaign"
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.campaignButtonFromSideMenuXpath())))
        elements.campaignButtonFromSideMenu()
        
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.createCampaignButtonXpath())))

        if elements.campaignsPageHeaderText()=="Campaigns":
            print "Campaigns page displayed"
        else:
            print "Campaigns page is not displayed"
            raise Exception
        
        
        print "Clicking on Create Campaign button"
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.createCampaignButtonXpath())))
        elements.createCampaignButton()
        
        
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.Camp_titleXpath())))
        print "Create Campaign page is displayed"
        
                  
        elements.titleTextField(campaignTitle)
        print "Title entered"
        
        elements.descriptionField(campDescription)
        print "Description entered"
        
        print "Adding Lesson"
        #Add lesson button
        elements.addlessonButton()
        
        #Waiting until first lesson in pop is displayed
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.FirstLessonWaitXpath())))
        
        #Searching lesson by its name
        elements.searchLesson(lessonName)
        
        #Waiting until lesson displayed
        elements.waitUntilSearchedLessonDisplayed(lessonName)
        
        #selecting searched lesson
        elements.selectSearchedLesson(lessonName)
        
        #waiting until add to campaign button is click able
        wait.until(EC.element_to_be_clickable((By.XPATH,elements.AddToCampaign_ButtonXpath())))
        
        #Clicking on Add to Campaign button
        elements.addToCampaignButton()
        
        
        time.sleep(2)
        #Verifying Added lesson is displayed in Grid
        print "\nVerifying Added lesson is displayed in Grid"
        if elements.firstLessonInGrid()==lessonName:
            print "Lesson displayed in grid"
        else:
            print "Lesson not displayed in grid"
            raise Exception
        
        
        print "Making this as a Graded campaign"
        elements.makeThisAsAGradedCampaign()
        print "Setting minimum passing score"
        elements.setMinimumPassingScore(minPassingScore)
        print "Setting maximum no of attempts"
        elements.setAMaxNoOfAttempts(numberOfAttempts)
        
        
        wait.until(EC.element_to_be_clickable((By.XPATH,elements.SaveAndExit_ButtonXpath())))
        print "Clicking on Save and Exit button"
        elements.saveAndExitButton()
        
        
        
        
        #Verifying Campaign Detail page is displayed
        print "\nVerifying campaign detail page is displayed"
        
        if elements.campaignDetailPageHeaderText()==campaignTitle:
            print "Campaign detail page is displayed"
        else:
            print "Campaign detail page is not displayed"
            raise Exception
        
        
        elements.searchingForlesson(campaignTitle)
        
        if elements.actualCampTitleINGrid()==campaignTitle:
            print "Campaign '"+campaignTitle+"' displayed in Grid"
        
        else:
            print "Campaign is not displayed in Grid"
            raise Exception
        
        
        
    def editCampaignAndAssignToLearner(self,campaignTitle,newCampaignTitle,campDescription,newDescription,lessonName2,newDuration,newMinPassingScore,newNumberOfAttempts,nameOFuser):
        
        elements=CampPage()
        driver.refresh()
        wait=WebDriverWait(driver, 60)
        
        
        print "\n\nCreating Campaign"
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.campaignButtonFromSideMenuXpath())))
        elements.campaignButtonFromSideMenu()
        
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.createCampaignButtonXpath())))

        if elements.campaignsPageHeaderText()=="Campaigns":
            print "Campaigns page displayed"
        else:
            print "Campaigns page is not displayed"
            raise Exception
        
        
        print "Clicking on Edit Button of Campaign ::"+campaignTitle
        editButton=wait.until(EC.visibility_of_element_located((By.XPATH,"(//table/tbody/tr/td[1]/a[.='"+campaignTitle+"']/../../td[4]/a[1])[1]")))
        editButton.click()
        okayGotIt=wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div[4]/div/div/div[2]/div[2]/button[1]")))
        
        print "Confirm popup is displayed"
        okayGotIt.click()
        
        wait.until(EC.visibility_of_element_located((By.XPATH,"//header/h1[contains(.,'Edit')]")))
        
        print "Verifying Edit campaign page is displayed"
        header=driver.find_element_by_xpath("//header/h1[contains(.,'Edit')]").text
        if campaignTitle in header:
            print "'"+header+"' page is displayed"
        else:
            print "Edit Campaign page is not displayed"
            raise Exception
        
        
        
        print "\nEditing existing data"
        
        existingTitle=driver.find_element_by_xpath(".//*[@id='campaign-title']").get_attribute("value")
        
        if existingTitle==campaignTitle:
            print "Existing title of campaign ::"+existingTitle
        else:
            print "Title is not displayed"
            raise Exception
        
        time.sleep(2)
        driver.find_element_by_xpath(".//*[@id='campaign-title']").clear()
        time.sleep(1)
        elements.titleTextField(newCampaignTitle)
        
        
        newCampaignTitleEntered=driver.find_element_by_xpath(".//*[@id='campaign-title']").get_attribute("value")
        
        if newCampaignTitleEntered!=campaignTitle:
            print "New title entered ::"+newCampaignTitle
        else:
            print "New title not entered"
            raise Exception
        
        
        time.sleep(2)
        driver.find_element_by_xpath(".//*[@id='campaign-description']").clear()
        time.sleep(1)
        elements.descriptionField(newDescription)
        print "New Discription entered ::"+newDescription
        
        newDescriptionEntered=driver.find_element_by_xpath(".//*[@id='campaign-title']").get_attribute("value")
        
        if newDescriptionEntered!=campDescription:
            print "New Description entered ::"+newDescription
        else:
            print "New Description not entered"
            raise Exception
        
        
        
        
        
        print "Removing old lesson"
        driver.find_element_by_xpath("//ul/li/div[2]/div/div/button[2]").click()
        try:
            wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[3]/div[2]/div/section[2]/div[2]/div")))
        except Exception:
            print "Lesson not deleted"
            raise Exception
        
        
        
        
        print "Adding new Lesson"
        
        elements.addlessonButton()
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.FirstLessonWaitXpath())))
        #Searching lesson by its name
        elements.searchLesson(lessonName2)
        #Waiting until lesson displayed
        elements.waitUntilSearchedLessonDisplayed(lessonName2)
        #selecting searched lesson
        elements.selectSearchedLesson(lessonName2)
        #waiting until add to campaign button is click able
        wait.until(EC.element_to_be_clickable((By.XPATH,elements.AddToCampaign_ButtonXpath())))
        #Clicking on Add to Campaign button
        elements.addToCampaignButton()
        time.sleep(2)
        
        
        #Verifying Added lesson is displayed in Grid
        print "\nVerifying newly Added lesson is displayed in Grid"
        if elements.firstLessonInGrid()==lessonName2:
            print "Lesson displayed in grid"
        else:
            print "Lesson not displayed in grid"
            raise Exception
        
        time.sleep(2)
        duration=driver.find_element_by_xpath(".//*[@id='input-campaign-duration']")
        duration.clear()
        time.sleep(2)
        duration.send_keys(str(newDuration))
        
        
        print "Setting minimum passing score"
        elements.setMinimumPassingScore(newMinPassingScore)
        print "Setting maximum no of attempts"
        elements.setAMaxNoOfAttempts(newNumberOfAttempts)
        
        elements.saveButton()
        
        
        #Verifying Campaign Detail page is displayed
        print "\nVerifying campaign detail page is displayed"
        
        if elements.campaignDetailPageHeaderText()==newCampaignTitle:
            print "Campaign detail page is displayed"
        else:
            print "Campaign detail page is not displayed"
            raise Exception
        
        
        
        print "Assigning Edited Campaign to learner"
        elements.assignButton()
        
        headerText=wait.until(EC.element_to_be_clickable((By.XPATH,"//header/h1"))).text
        
        if newCampaignTitle in headerText:
            print "Plan assignment for page is displayed"
        else:
            print "Plan assignment for page is not displayed"
            
            
            
        #Checking One time triggered radio button is selected by Default
        if driver.find_element_by_xpath("//input[@id='assignment-one-time']").is_selected():
            print "Radio button One Time triggered is selected"
        else:
            print "Radio button is not selected"
            raise Exception
        
        time.sleep(3)
        print "Difference between two dates"
        from datetime import datetime
        date_format = "%m/%d/%Y"
        a = datetime.strptime(elements.assignDateValue(), date_format)
        b = datetime.strptime(elements.dueDateValue(), date_format)
        delta = b - a
        print "The difference between two dates is ::"
        print delta.days
        
        
        
        #Adding user to this campaign
        elements.addingUser(nameOFuser)
        
        
        print "Verifying User is displayed in Grid"
        if nameOFuser in elements.userInGrid():
            print "User is displayed in Grid"
        else:
            print "user is not displayed in grid"
            raise Exception        
        
        
        
        
        elements.sendAssignment()
        print "Clicked on Send assignment button"
        
        
        print "Verifying Confirm pop up is displayed"
        if elements.confirmPopupsendingAssignement().is_displayed():
            print "Confirm assignment pop up is displayed"
        else:
            print "Confirm assignment pop up is not displayed"
            raise Exception
        
        
        elements.sendAssignmentFromPopup()
        print "Clicked on Send Assignment button from pop up"
        
        
        
        #Verifying Campaign Detail page is displayed
        print "\nVerifying campaign detail page is displayed"
        
        if elements.campaignDetailPageHeaderText()==newCampaignTitle:
            print "Campaign detail page is displayed"
        else:
            print "Campaign detail page is not displayed"
            raise Exception
        
        
        elements.searchingForlesson(newCampaignTitle)
        #verifying in Campaigns displayed in Campaigns grid
        
        if elements.actualCampTitleINGrid()==newCampaignTitle:
            print "Campaign '"+newCampaignTitle+"' displayed in Grid"
        
        else:
            print "Campaign is not displayed in Grid"
            raise Exception
        
        
        
        
    def assignEditedCampaign(self):
        
        book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        first_sheet = book.sheet_by_name('CampAssign')
        
        
        #Campaign
        cell1 = first_sheet.cell(405,1)
        campaignTitle = cell1.value
        
        cell1 = first_sheet.cell(406,1)
        campDescription = cell1.value
        
        
        cell1 = first_sheet.cell(407,1)
        minPassingScore = cell1.value
        
        cell1 = first_sheet.cell(408,1)
        numberOfAttempts = cell1.value
        
        cell1 = first_sheet.cell(409,1)
        lessonName1 = cell1.value
        
        cell1 = first_sheet.cell(410,1)
        lessonName2 = cell1.value
        
        cell2 = first_sheet.cell(411,1)
        questionCard= cell2.value
        
        cell2 = first_sheet.cell(412,1)
        ans1= cell2.value
        
        cell2 = first_sheet.cell(413,1)
        ans2= cell2.value
        
        
        
        #Edit Campaign
        cell1 = first_sheet.cell(415,1)
        newCampaignTitle = cell1.value
        
        cell1 = first_sheet.cell(416,1)
        newDescription = cell1.value
        
        
        cell1 = first_sheet.cell(417,1)
        newMinPassingScore = cell1.value
        
        cell1 = first_sheet.cell(418,1)
        newNumberOfAttempts = cell1.value
        
        cell1 = first_sheet.cell(419,1)
        newDuration = cell1.value
    
        
        
        #Learner
        cell = first_sheet.cell(421,1)
        FirstName = cell.value
        
        FirstNameId = FirstName.split("_")
        emp = FirstNameId[0]+"_"
        ids = FirstNameId[1]
        empId = int(ids)+1
        FirstNameUpdated= emp+str(empId)
        
        cell = first_sheet.cell(422,1)
        LastName = cell.value
        
        LastNameID = LastName.split("_")
        emp = LastNameID[0]+"_"
        ids = LastNameID[1]
        empId = int(ids)+1
        LastNameUpdated= emp+str(empId)
        
        
        
        cell = first_sheet.cell(423,1)
        Email = cell.value
        
        EmailId = Email.split("_")
        emp = EmailId[0]+"_"
        ids = EmailId[1]
        remaining="_"+EmailId[2]
        empId = int(ids)+1
        EmailIdUpdated= emp+str(empId)+remaining
        
        cell = first_sheet.cell(424,1)
        EmployeeId = cell.value
        
        Employee = EmployeeId.split("_")
        emp = Employee[0]+"_"
        ids = Employee[1]
        empId = int(ids)+1
        EmployeeIdUpdated= emp+str(empId)
        
        cell = first_sheet.cell(425,1)
        Password = cell.value
        
        cell = first_sheet.cell(426,1)
        NewPassword = cell.value
        
        cell = first_sheet.cell(427,1)
        role = cell.value
        
        #For Original User
        book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        s_sheet = book.sheet_by_name('Login_Credentials')
        
        cell = s_sheet.cell(1,1)
        url = cell.value
        
        cell = s_sheet.cell(3,1)
        username = cell.value
        
        cell = s_sheet.cell(3,2)
        password = cell.value
        
        #updating user values
        wb = load_workbook(os.path.join('Test_Data/TestData.xlsx'))
            #print (wb.sheetnames)
        
        sheet = wb['CampAssign']
        
        sheet.cell(row=422, column=2).value = FirstNameUpdated
        sheet.cell(row=423, column=2).value = LastNameUpdated
        sheet.cell(row=424, column=2).value = EmailIdUpdated
        sheet.cell(row=425, column=2).value = EmployeeIdUpdated
        
        
        
        wb.save(os.path.join('Test_Data/TestData.xlsx'))
            
        print "All User Data Updated in Excel"
        
        
        
        
        try:
            print "\nCreating a New Learner\n"
            lt=CreateLearner()
            lt.createNewLearnerMain(FirstName, LastName, Email, EmployeeId, Password, NewPassword, url, username, password)
            
            
            print "\nCreating a lesson\n"
            ot=CreateLessonDifferentLessons()
            ot.lessonWithQuestion(lessonName1, questionCard, ans1, ans2)
            ot.lessonWithQuestion(lessonName2, questionCard, ans1, ans2)
            
            print "\nCreating Campaign\n"
            lk=AssignCampEditCampignAndAssign()
            lk.campaignEditAndAssign(campaignTitle, campDescription, lessonName1, minPassingScore, numberOfAttempts)
            
            lk.editCampaignAndAssignToLearner(campaignTitle, newCampaignTitle, campDescription, newDescription, lessonName2, newDuration, newMinPassingScore, newNumberOfAttempts, FirstName)
            
            
            
            print "\n----Text Execution Completed----\n"
            
            
        except Exception as e:
            traceback.print_exc()
            print (e)
            raise Exception 
         
        finally:  
            driver.save_screenshot("ScreenShots/AssignCampEditCampignAndAssign.png") 
            
            second_sheet = book.sheet_by_name('Login_Credentials')
            cell = second_sheet.cell(1,1)
            url = cell.value
            driver.get(url)
            
            try:
                WebDriverWait(driver, 5).until(EC.alert_is_present())

                alert = driver.switch_to.alert
                alert.accept()
                print("alert accepted")
            except Exception:
                print("no alert")
                
            
            
            wait=WebDriverWait(driver, 60)
            unDropDown=wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")))
            driver.execute_script('arguments[0].click()',unDropDown)
            
            signOut=driver.find_element_by_xpath("html/body/div/div/div[1]/div[2]/div[2]/a")
            driver.execute_script('arguments[0].click()',signOut)
            
            
            
            element = wait.until(EC.presence_of_element_located((By.ID, "password")))
            
            if driver.title == "Grovo":
                print("Grovo Application URL Opened")
            else:
                raise Exception.message
    
            print "Grovo Sign-In page is displayed"
            
            print "Entering User name"
            driver.find_element_by_xpath(".//*[@id='username']").send_keys(username)
           
            print "Entering Password"
            element.send_keys(password)
            
            element.send_keys(Keys.TAB)
            print "Clicking on Sign_In button"
            driver.find_element_by_xpath("//*[@id='submitButton']").click()
            
            print "Successfully Logged Into Users account"
            time.sleep(3)
            
            try:
                d=DeleteLesson()
                d.lessonDelete(lessonName1)
                d.lessonDelete(lessonName2)
            except Exception:
                driver.get(url)
            
            
if __name__ == '__main__':
    
    btc=BaseTestClass()
    btc.UserLogin()
    
    ne=AssignCampEditCampignAndAssign()
    ne.assignEditedCampaign()
            
            
        
    