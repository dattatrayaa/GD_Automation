'''
Created on 15-Mar-2018

@author: dattatraya
'''

import os.path
import time
import traceback

from BaseTestClass import BaseTestClass
from BaseTestClass import driver
from openpyxl.reader.excel import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import xlrd

from CampaignPageElements import CampPage
from CreateLearnerNew import CreateLearner
from CreateLessonDifferentLessons import CreateLessonDifferentLessons
from DeleteLesson import DeleteLesson
from TracksPageElements import TracksPageElements
from BaseTestClass import projectPath

class AssignCampOneTrackContainsFourLessons:
    def campaignTolearnerOneTrackWithFourLessons(self,campaignTitle,campDescription,nameOFuser,trackName):
        
        elements=CampPage()
        
        wait=WebDriverWait(driver, 60)
        
        driver.refresh()
        print "\n\nCreating Campaign"
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.campaignButtonFromSideMenuXpath())))
        elements.campaignButtonFromSideMenu()
        
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.createCampaignButtonXpath())))

        if elements.campaignsPageHeaderText()=="Campaigns":
            print "Campaigns page displayed"
        else:
            print "Campaigns page is not displayed"
            raise Exception
        
        
        print "Clicking on Create Campaign button"
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.createCampaignButtonXpath())))
        elements.createCampaignButton()
        
        
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.Camp_titleXpath())))
        print "Create Campaign page is displayed"
        
                  
        elements.titleTextField(campaignTitle)
        print "Title entered"
        
        elements.descriptionField(campDescription)
        print "Description entered"
        
       
        print "Adding Track to Campaign"
        elements.addTrackButton()
        elements.searchTracksAndSelect(trackName)
        elements.addToCampaignTrack()
        print "Track Added to track"
        
        
        
        time.sleep(2)
        print "\nVerifying Added Track is displayed in Grid"
        if trackName in elements.firstTrackInGrid():
            print "Track displayed in grid"
        else:
            print "Track not displayed in Grid"
            raise Exception
        
        
        wait.until(EC.element_to_be_clickable((By.XPATH,elements.SaveAndExit_ButtonXpath())))
        print "Clicking on Save and plan assignmet button"
        elements.saveAndPlanAssignmentbutton()
        
        
        
        #Verifying Plan assignment for page is displayed
        print "Verifying Plan assignment for page is displayed"
        if campaignTitle in elements.planAssignementForHeaderText():
            print "Page '"+elements.planAssignementForHeaderText()+"' is displayed"
        else:
            print "Plan assignment page is not displayed"
            raise Exception
        
        
        #Checking One time triggered radio button is selected by Default
        if driver.find_element_by_xpath("//input[@id='assignment-one-time']").is_selected():
            print "Radio button One Time triggered is selected"
        else:
            print "Radio button is not selected"
            raise Exception
        
        time.sleep(3)
        print "Difference between two dates"
        from datetime import datetime
        date_format = "%m/%d/%Y"
        a = datetime.strptime(elements.assignDateValue(), date_format)
        b = datetime.strptime(elements.dueDateValue(), date_format)
        delta = b - a
        print "The difference between two dates is ::"
        print delta.days
        
        
        #Adding user to this campaign
        elements.addingUser(nameOFuser)
        
        
        print "Verifying User is displayed in Grid"
        if nameOFuser in elements.userInGrid():
            print "User is displayed in Grid"
        else:
            print "user is not displayed in grid"
            raise Exception        
        
        
        
        elements.sendAssignment()
        print "Clicked on Send assignment button"
        
        
        print "Verifying Confirm pop up is displayed"
        if elements.confirmPopupsendingAssignement().is_displayed():
            print "Confirm assignment pop up is displayed"
        else:
            print "Confirm assignment pop up is not displayed"
            raise Exception
        
        
        elements.sendAssignmentFromPopup()
        print "Clicked on Send Assignment button from pop up"
        
        
        
        #Verifying Campaign Detail page is displayed
        print "\nVerifying campaign detail page is displayed"
        
        if elements.campaignDetailPageHeaderText()==campaignTitle:
            print "Campaign detail page is displayed"
        else:
            print "Campaign detail page is not displayed"
            raise Exception
        
        #verifying in Campaigns displayed in Campaigns grid
        elements.searchingForlesson(campaignTitle)
        
        if elements.actualCampTitleINGrid()==campaignTitle:
            print "Campaign '"+campaignTitle+"' displayed in Grid"
        
        else:
            print "Campaign is not displayed in Grid"
            raise Exception
        
        
    def createTrackwithFourLessons(self,titleOfTrack,Imagefilepath,description,tagName,lessonName1,lessonName2,lessonName3,lessonName4,expectedSuccessText):
        
        print "\nCreating track with one lesson contains Text Card"
        driver.refresh()
        wait=WebDriverWait(driver, 120)
        track=TracksPageElements()
        
        print "Clicking on Lessons button from side menu"
        track.lessonsButton()
        
        print "Clicking on Track button from side menu"
        track.tracksButton()
        
        track.createTracksButton()
        
        print "Entering title"
        titlefield=wait.until(EC.visibility_of_element_located((By.XPATH,track.trackTitle())))
        titlefield.send_keys(titleOfTrack)
        print "Title entered ::"+titleOfTrack
        
        driver.find_element_by_css_selector('input[type="file"]').send_keys(Imagefilepath)
        print "waiting to upload image"
        wait.until(EC.visibility_of_element_located((By.XPATH,track.imageDisplayed())))
        print "Image uploaded"
        
        print "Entering Description"
        driver.find_element_by_xpath(track.trackDescription()).send_keys(description)
        print "Description entered ::"+description 
        
        
        print "Adding tag"
        addTags=driver.find_element_by_xpath(track.addingTag())
        webdriver.ActionChains(driver).move_to_element(addTags).click().send_keys(tagName).perform()
        
        option=wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='react-select-2--option-0']")))
        webdriver.ActionChains(driver).move_to_element(option).click(option).perform()
        
        driver.find_element_by_xpath(track.trackDescription()).send_keys(" ")
        
        
        print "\nAdding created two lessons"
        
        print "Clicking on Add lessons button"
       
        driver.execute_script("window.scrollTo(0, 0);")
        
        track.addLessonButton()
                
        wait.until(EC.visibility_of_element_located((By.XPATH,track.serchLesson())))
        
        searchLesson=driver.find_element_by_xpath(track.serchLesson())
        print "Searching for first lesson in Add lessons pop up"
        searchLesson.send_keys(lessonName1)
        track.searchedLessonAdd(lessonName1)
        
        time.sleep(2)
        print "Searching for Second lesson in Add lessons pop up"
        searchLesson.clear()
        searchLesson.send_keys(lessonName2)
        track.searchedLessonAdd(lessonName2)
        
        time.sleep(2)
        print "Searching for Third lesson in Add lessons pop up"
        searchLesson.clear()
        searchLesson.send_keys(lessonName3)
        track.searchedLessonAdd(lessonName3)
        
        
        
        print "Searching for Fourth lesson in Add lessons pop up"
        searchLesson.clear()
        time.sleep(1)
        searchLesson.send_keys(lessonName4)
        track.searchedLessonAdd(lessonName4)
        
     
        print "Adding to Track"
        track.addingToTrack()
        
        print "\nChecking added lesson is selected lesson from Pop up"
        
        wait.until(EC.visibility_of_element_located((By.XPATH,"//li[1]/div[2]/div/h4/div")))
        lessonTextAddedToGrid=driver.find_element_by_xpath("//li[1]/div[2]/div/h4/div").text
        if lessonTextAddedToGrid==lessonName1:
            print "Selected Lesson '"+lessonName1+"' is displayed in grid of tracks page"
        
        else:
            print "Lesson is not displayed"
            raise Exception
        
        lessonTextAddedToGrid1=driver.find_element_by_xpath("//li[2]/div[2]/div/h4/div").text
        if lessonTextAddedToGrid1==lessonName2:
            print "Selected Lesson '"+lessonName2+"' is displayed in grid of tracks page"
        
        else:
            print "Lesson is not displayed"
            raise Exception
        
        lessonTextAddedToGrid2=driver.find_element_by_xpath("//li[3]/div[2]/div/h4/div").text
        if lessonTextAddedToGrid2==lessonName3:
            print "Selected Lesson '"+lessonName3+"' is displayed in grid of tracks page"
        
        else:
            print "Lesson is not displayed"
            raise Exception
        
        lessonTextAddedToGrid2=driver.find_element_by_xpath("//li[4]/div[2]/div/h4/div").text
        if lessonTextAddedToGrid2==lessonName4:
            print "Selected Lesson '"+lessonName4+"' is displayed in grid of tracks page"
        
        else:
            print "Lesson is not displayed"
            raise Exception
        
        
        
        
        #Publishing Track
        print "Clicking on Publish Track button"
        
        track.publishButton()        
        print "\nVerifying Success message is displaying"
        
        
        track.successmessage(expectedSuccessText, titleOfTrack)
        
        
        track.sideMenuTracksExpanded()
        
        wait.until(EC.visibility_of_element_located((By.XPATH,"//tbody/tr/td[2]/a[.='"+titleOfTrack+"']")))
        
        trackInGrid=driver.find_element_by_xpath("//tbody/tr/td[2]/a[.='"+titleOfTrack+"']").text
        
        if trackInGrid==titleOfTrack:
            print "Track '"+trackInGrid+"' is displayed in grid"
        else:
            print "Track is not displayed in grid"
            raise Exception
        
        driver.find_element_by_xpath(".//*[@id='content']/div/div[3]/div[1]/div/nav/div/div[4]").click()
             
        
        
    
        
    
        
    def assignCampOneTrackContainsFourLessons(self):
        
        book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        first_sheet = book.sheet_by_name('CampAssign')
        
        
        #Campaign
        cell1 = first_sheet.cell(48,1)
        campaignTitle = cell1.value
        
        cell1 = first_sheet.cell(49,1)
        campDescription = cell1.value
        
        cell1 = first_sheet.cell(68,1)
        nameOFuser = cell1.value
        
        
        
        #lesson
        cell2 = first_sheet.cell(50,1)
        lessonName1= cell2.value
        
        
        cell2 = first_sheet.cell(51,1)
        lessonName2= cell2.value
        
        cell2 = first_sheet.cell(52,1)
        lessonName3= cell2.value
        
        cell2 = first_sheet.cell(53,1)
        lessonName4= cell2.value
        
        cell2 = first_sheet.cell(54,1)
        textCard= cell2.value
        
        cell2 = first_sheet.cell(55,1)
        Imagefilepath1= cell2.value
        
        
        cell2 = first_sheet.cell(56,1)
        videoPath= cell2.value
        
        cell2 = first_sheet.cell(57,1)
        timeToUploadVideo= cell2.value
        
        cell2 = first_sheet.cell(58,1)
        questionCard= cell2.value
        
        cell2 = first_sheet.cell(59,1)
        ans1= cell2.value
        
        cell2 = first_sheet.cell(60,1)
        ans2= cell2.value
        
        
        
        
        #track
        cell2 = first_sheet.cell(62,1)
        titleOfTrack= cell2.value
        
        cell2 = first_sheet.cell(63,1)
        Imagefilepath= cell2.value
        
        cell2 = first_sheet.cell(64,1)
        description= cell2.value
        
        cell2 = first_sheet.cell(65,1)
        tagName= cell2.value
        
        cell2 = first_sheet.cell(66,1)
        expectedSuccessText= cell2.value
        
        
        #Learner
        cell = first_sheet.cell(68,1)
        FirstName = cell.value
        
        FirstNameId = FirstName.split("_")
        emp = FirstNameId[0]+"_"
        ids = FirstNameId[1]
        empId = int(ids)+1
        FirstNameUpdated= emp+str(empId)
        
        cell = first_sheet.cell(69,1)
        LastName = cell.value
        
        LastNameID = LastName.split("_")
        emp = LastNameID[0]+"_"
        ids = LastNameID[1]
        empId = int(ids)+1
        LastNameUpdated= emp+str(empId)
        
        
        
        cell = first_sheet.cell(70,1)
        Email = cell.value
        
        EmailId = Email.split("_")
        emp = EmailId[0]+"_"
        ids = EmailId[1]
        remaining="_"+EmailId[2]
        empId = int(ids)+1
        EmailIdUpdated= emp+str(empId)+remaining
        
        cell = first_sheet.cell(71,1)
        EmployeeId = cell.value
        
        Employee = EmployeeId.split("_")
        emp = Employee[0]+"_"
        ids = Employee[1]
        empId = int(ids)+1
        EmployeeIdUpdated= emp+str(empId)
        
        cell = first_sheet.cell(72,1)
        Password = cell.value
        
        cell = first_sheet.cell(73,1)
        NewPassword = cell.value
        
        cell = first_sheet.cell(74,1)
        role = cell.value
        
        #For Original User
        book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        s_sheet = book.sheet_by_name('Login_Credentials')
        
        cell = s_sheet.cell(1,1)
        url = cell.value
        
        cell = s_sheet.cell(3,1)
        username = cell.value
        
        cell = s_sheet.cell(3,2)
        password = cell.value
        
        #updating user values
        wb = load_workbook(os.path.join('Test_Data/TestData.xlsx'))
            #print (wb.sheetnames)
        
        sheet = wb['CampAssign']
        
        sheet.cell(row=69, column=2).value = FirstNameUpdated
        sheet.cell(row=70, column=2).value = LastNameUpdated
        sheet.cell(row=71, column=2).value = EmailIdUpdated
        sheet.cell(row=72, column=2).value = EmployeeIdUpdated
        
        
        
        wb.save(os.path.join('Test_Data/TestData.xlsx'))
            
        print "All User Data Updated in Excel"
        
        
        
        
        try:
            print "\nCreating a New Learner\n"
            lt=CreateLearner()
            lt.createNewLearnerMain(FirstName, LastName, Email, EmployeeId, Password, NewPassword, url, username, password)
            
            
            print "\nCreating a lesson\n"
            ot=CreateLessonDifferentLessons()
            ot.lessonWithText(lessonName1, textCard)
            ot.lessonWithImage(lessonName2, Imagefilepath1)
            ot.lessonWithVideo(lessonName3, videoPath, timeToUploadVideo)
            ot.lessonWithQuestion(lessonName4, questionCard, ans1, ans2)
            
            
            print "\nCreating Track\n"
            jk=AssignCampOneTrackContainsFourLessons()
            jk.createTrackwithFourLessons(titleOfTrack, Imagefilepath, description, tagName, lessonName1, lessonName2, lessonName3, lessonName4, expectedSuccessText)
            
            print "\nCreating Campaign\n"
            jk.campaignTolearnerOneTrackWithFourLessons(campaignTitle, campDescription, nameOFuser, titleOfTrack)
            
            
            print "\n----Text Execution Completed----\n"
            
            
        except Exception as e:
            traceback.print_exc()
            print (e)
            raise Exception 
         
        finally: 
            driver.save_screenshot("ScreenShots/AssignCampOneTrackContainsFourLessons.png") 
            
            second_sheet = book.sheet_by_name('Login_Credentials')
            cell = second_sheet.cell(1,1)
            url = cell.value
            driver.get(url)
            try:
                WebDriverWait(driver, 5).until(EC.alert_is_present())

                alert = driver.switch_to.alert
                alert.accept()
                print("alert accepted")
            except Exception:
                print("no alert")
                
            
            wait=WebDriverWait(driver, 60)
            unDropDown=wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")))
            driver.execute_script('arguments[0].click()',unDropDown)
            
            signOut=driver.find_element_by_xpath("html/body/div/div/div[1]/div[2]/div[2]/a")
            driver.execute_script('arguments[0].click()',signOut)
            
            
            
            element = wait.until(EC.presence_of_element_located((By.ID, "password")))
            
            if driver.title == "Grovo":
                print("Grovo Application URL Opened")
            else:
                raise Exception.message
    
            print "Grovo Sign-In page is displayed"
            
            print "Entering User name"
            driver.find_element_by_xpath(".//*[@id='username']").send_keys(username)
           
            print "Entering Password"
            element.send_keys(password)
            
            element.send_keys(Keys.TAB)
            print "Clicking on Sign_In button"
            driver.find_element_by_xpath("//*[@id='submitButton']").click()
            
            print "Successfully Logged Into Users account"
            time.sleep(3)
            
            
            try:
                d=DeleteLesson()
                d.lessonDelete(lessonName1)
                d.lessonDelete(lessonName2)
                d.lessonDelete(lessonName3)
                d.lessonDelete(lessonName4)
            except Exception:
                driver.get(url)
            
if __name__ == '__main__':
    
    btc=BaseTestClass()
    btc.UserLogin()
    
    ne=AssignCampOneTrackContainsFourLessons()
    ne.assignCampOneTrackContainsFourLessons()
            
                       
