'''
Created on 25-Jul-2018

@author: Sheethu C
'''

import os.path
import time
import traceback
from BaseTestClass import BaseTestClass
from BaseTestClass import driver
from openpyxl.reader.excel import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import xlrd
from BaseTestClass import projectPath
from CampaignPageElements import CampPage
from CreateLearnerNew import CreateLearner
from CreateLessonDifferentLessons import CreateLessonDifferentLessons
from DeleteLesson import DeleteLesson
from CreateTrackComman import CreateTrackComman
from PlanAssignmentForPageElements import PlanAssignmentForPageElements
from GroupsPageElements import GroupsPageElements
from AddTriggerForPageElements import AddTriggerPage
from BaseTestClass import projectPath
from GroupDelete import GroupDelete
from CreateUserXpath import CreateUserXpath
class VerificationOfEditedCampaignWithoutLesson:
    def createcreator(self,FirstName,LastName,Email,EmployeeId,Password):
        createuser =CreateUserXpath()
        wait=WebDriverWait(driver, 10)
        driver.refresh()
        wait.until(EC.visibility_of_element_located((By.XPATH,createuser.adminSideMenu())))
        driver.find_element_by_xpath(createuser.adminSideMenu()).click()
        print "Clicked on admin icon"
        driver.find_element_by_xpath(createuser.usersideMenu()).click()
        print "clicked on Users"
        wait.until(EC.visibility_of_element_located((By.XPATH,createuser.addEditUserButton())))
        wait.until(EC.element_to_be_clickable((By.XPATH,createuser.addEditUserButton())))
        driver.find_element_by_xpath(createuser.addEditUserButton()).click()
        print "Clicked on Add or editUser"
        wait.until(EC.visibility_of_element_located((By.XPATH,createuser.addAnIndividualUser())))
        wait.until(EC.element_to_be_clickable((By.XPATH,createuser.addAnIndividualUser())))
        driver.find_element_by_xpath(createuser.addAnIndividualUser()).click()
        print "Clicked on Add An individual User"
        
        wait.until(EC.visibility_of_element_located((By.XPATH,createuser.addUserPageWait())))
        print "Verifying Add user Page"

        if driver.find_element_by_xpath(createuser.addUserPageWait()).is_displayed():
            print("Add user Page is displayed")
        else:
            print ""
            raise Exception
        
        print "Verifying First Name field"
        
        wait.until(EC.visibility_of_element_located((By.ID,createuser.firstName())))
        if driver.find_element_by_id("create-edit-user-search-firstName").is_displayed():
            print(" First Name field displayed")
        else:
            print ""
            raise Exception
        driver.find_element_by_id(createuser.firstName()).send_keys(FirstName)
        print "FirstName is Entered ::"+FirstName
        print "Last NAme verifying"
        if driver.find_element_by_id(createuser.lastName()).is_displayed():
            print(" Last Name field displayed")
        else:
            print ""
            raise Exception
        driver.find_element_by_id(createuser.lastName()).send_keys(LastName)
        print "Last Name is Entered ::"+LastName
        print "Email verifying"
        if driver.find_element_by_id(createuser.email()).is_displayed():
            print("Email field displayed")
        else:
            print ""
            raise Exception
        driver.find_element_by_id(createuser.email()).send_keys(Email)
        print "Email is Entered ::"+Email
        
        print "Employee ID verifying"
        if driver.find_element_by_id(createuser.employeId()).is_displayed():
            print("Employee ID field displayed")
        else:
            print ""
            raise Exception
        driver.find_element_by_id(createuser.employeId()).send_keys(EmployeeId)
        print "Employee ID  is Entered ::"+EmployeeId
        print "Inherited Role Verifying"
        if driver.find_element_by_xpath(createuser.inheritedRole()).is_displayed():
            print("Inherited Role field displayed")
        else:
            print ""
            raise Exception
        print "Selecting Creator in Direct Roles"

        dd1=driver.find_element_by_xpath(createuser.directRoleDisplay())
        webdriver.ActionChains(driver).move_to_element(dd1).click().send_keys("creator").perform()
        wait.until(EC.visibility_of_element_located((By.XPATH,createuser.directCreatorRole())))
        ele =driver.find_element_by_xpath(createuser.directCreatorRole())
        
        webdriver.ActionChains(driver).move_to_element(ele).click().perform()
        
        print "Password Field is Verifying"
        if driver.find_element_by_id(createuser.password()).is_displayed():
            print("Password field displayed")
        else:
            print ""
            raise Exception
        driver.find_element_by_id(createuser.password()).send_keys(Password)
        print "Password is Entered ::"+Password
        
        wait.until(EC.visibility_of_element_located((By.XPATH,createuser.addButton())))
        wait.until(EC.element_to_be_clickable((By.XPATH,createuser.addButton())))
        driver.find_element_by_xpath(createuser.addButton()).click()
        print "Clicked on add button"
        wait.until(EC.visibility_of_element_located((By.XPATH,createuser.saveButton())))
        driver.find_element_by_xpath(createuser.saveButton()).click()
        print "Clicked on Save"
        print "Searching for the Created User"
        wait.until(EC.visibility_of_element_located((By.XPATH,createuser.userPageWait())))
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div/div/div[3]/div[2]/div/header/h1")))
        wait.until(EC.visibility_of_element_located((By.XPATH,"//table/tbody/tr[1]/td[2]")))
        driver.find_element_by_id("search-users").send_keys(FirstName)
        time.sleep(4)
        wait.until(EC.visibility_of_element_located((By.XPATH,"//table/tbody/tr/td[.='"+Email+"']/../td[2]/a")))
        ele =driver.find_element_by_xpath("//table/tbody/tr/td[.='"+Email+"']/../td[2]/a").text
        if(ele==FirstName):
            print("Created User Verified")
        else:
            print ""
            raise Exception  
        driver.find_element_by_xpath(createuser.profileClick()).click()
        print "Clicked on Account"
        driver.find_element_by_xpath(createuser.signOut()).click()
        print "Clicked on signOut Button"
        wait.until(EC.visibility_of_element_located((By.ID,createuser.loginUserName())))
        time.sleep(4)
        
        
    def updateUser(self):
        print "Updating Details"
        book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        second_sheet = book.sheet_by_name('CampAssign')
        #First User
        cell = second_sheet.cell(848,1)
        FirstName = cell.value
        FirstNameId = FirstName.split("_")
        emp = FirstNameId[0]+"_"
        ids = FirstNameId[1]
        empId1 = int(ids)+1
        FirstNameUpdated1= emp+str(empId1)
        
        
        cell2 = second_sheet.cell(851,1)
        LearnId = cell2.value
        LearnsId = LearnId.split("#")
        emp = LearnsId[0]+"#"
        ids = LearnsId[1]
        empId = int(ids)+1
        LearnerEmpId1 = emp+str(empId)
        
        
        cell1 = second_sheet.cell(850,1)
        LearnEmailId = cell1.value
        LearnerEmail = LearnEmailId.split("@")
        spitValue = LearnerEmail[0][:4]
        email = spitValue+str(empId)
        id = "@"+LearnerEmail[1]
        LearnerEmailId1 = email+id
        
        wb = load_workbook(os.path.join('Test_Data/TestData.xlsx'))
        
        #print (wb.sheetnames)
    
        sheet = wb['CampAssign']
        
        sheet.cell(row=849, column=2).value = FirstNameUpdated1    
        sheet.cell(row=851, column=2).value = LearnerEmailId1
        sheet.cell(row=852, column=2).value = LearnerEmpId1
        
        
        wb.save(os.path.join('Test_Data/TestData.xlsx'))
    def createCreatorUserLogin(self,Email,Password,NewPassword):
        createuser =CreateUserXpath()
        print "Reading data from excel sheet"
        wait=WebDriverWait(driver, 80)
        wait.until(EC.visibility_of_element_located((By.ID,createuser.loginUserName())))
        driver.find_element_by_id(createuser.loginUserName()).send_keys(Email)
       
        print "Entering Password"
        element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID,createuser.loginPassword())))
        element.send_keys(Password)
        
        element.send_keys(Keys.TAB)
        print "Clicking on Sign_In button"
        driver.find_element_by_id(createuser.submitButton()).click()
        wait.until(EC.visibility_of_element_located((By.ID,createuser.currentPassword())))
        driver.find_element_by_id(createuser.currentPassword()).send_keys(Password)
        print "Current Password is entered :"+Password
        driver.find_element_by_id(createuser.newPassword()).send_keys(NewPassword)
        print "New Password is entered :"+NewPassword
        wait.until(EC.visibility_of_element_located((By.XPATH,createuser.newSubmit())))
        driver.find_element_by_xpath(createuser.newSubmit()).click()
        time.sleep(4)
        wait.until(EC.visibility_of_element_located((By.ID,createuser.homepageSearchWait())))
        print "Home Page is Loaded"
        print "Creator Page Verification"
        driver.find_element_by_xpath(createuser.createicon()).click()
        #Home =driver.find_element_by_xpath(createuser.homeicon())
        #Library = driver.find_element_by_xpath(createuser.libraryicon())
        #Create = driver.find_element_by_xpath(createuser.createicon())
        #campaign = driver.find_element_by_xpath(createuser.campaignicon())
        #if (Home.is_displayed() and Library.is_displayed() and (Create.is_displayed()) and campaign.is_displayed()):
            #print"Home page for Creator displayed"             
        #else:
            #print"Home page for Creator not displayed"
            #raise Exception 
        
    def campWithTrackTextlesson(self,campaignTitle,campDescription,lessonName):
        elements=CampPage()
        
        wait=WebDriverWait(driver, 60)
        driver.refresh()
        print "\n\nCreating Campaign"
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.campaignButtonFromSideMenuXpath())))
        elements.campaignButtonFromSideMenu()
        
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.createCampaignButtonXpath())))

        if elements.campaignsPageHeaderText()=="Campaigns":
            print "Campaigns page displayed"
        else:
            print "Campaigns page is not displayed"
            raise Exception
        
        
        print "Clicking on Create Campaign button"
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.createCampaignButtonXpath())))
        elements.createCampaignButton()
        
        
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.Camp_titleXpath())))
        print "Create Campaign page is displayed"
        
                  
        elements.titleTextField(campaignTitle)
        print "Title entered ::campTitle"
        
        elements.descriptionField(campDescription)
        print "Description entered ::campDescription"
        
        print "Adding Lesson"
        print "Adding Lesson"
        #Add lesson button
        elements.addlessonButton()
        
        #Waiting untill first lesson in pop is displayed
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.FirstLessonWaitXpath())))
        
        #Searching lesson by its name
        elements.searchLesson(lessonName)
        
        #Waiting until lesson displayed
        elements.waitUntilSearchedLessonDisplayed(lessonName)
        
        #selecting searched lesson
        elements.selectSearchedLesson(lessonName)
        
        #waiting until add to campaign button is clickable
        wait.until(EC.element_to_be_clickable((By.XPATH,elements.AddToCampaign_ButtonXpath())))
        
        #Clicking on Add to Campaign button
        elements.addToCampaignButton()
        
        #Verifying Added lesson is displayed in Grid
        print "\nVerifying Added lesson is displayed in Grid"
        if elements.firstLessonInGrid()==lessonName:
            print "Lesson displayed in grid"
        else:
            print "Lesson not displayed in grid"
        
        wait.until(EC.element_to_be_clickable((By.XPATH,elements.SaveAndExit_ButtonXpath())))
        #Clicking on save & exit button
        print "Clicking on save & exit button"
        elements.saveAndExitButton()
        
        '''#verifying success message
        print "\nVerifying success message"
        
        if elements.successMessage()==actualSuccessMessage:
            print "Message '"+actualSuccessMessage+"' is displayed"
        else:
            print "Success message is not displayed properly"
            raise Exception'''
        
        #Verifying campaign detail page is displayed
        print "\nVerifying campaign detail page is displayed"
        
        if elements.campaignDetailPageHeaderText()==campaignTitle:
            print "Campaign detail page is displayed"
        else:
            print "Campaign detail page is not displayed"
            raise Exception
        
       
        elements.searchingForlesson(campaignTitle)
        
        if elements.actualCampTitleINGrid()==campaignTitle:
            print "Campaign '"+campaignTitle+"' displayed in Grid"
        
        else:
            print "Campaign is not displayed in Grid"
            raise Exception
         
        print "\n----Text Execution Completed----\n"
    def editCampaign(self,campaignTitle):
        elements=CampPage()
        driver.refresh()
        wait=WebDriverWait(driver, 60)
        
        
        print "\n\nCreating Campaign"
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.campaignButtonFromSideMenuXpath())))
        elements.campaignButtonFromSideMenu()
        
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.createCampaignButtonXpath())))

        if elements.campaignsPageHeaderText()=="Campaigns":
            print "Campaigns page displayed"
        else:
            print "Campaigns page is not displayed"
            raise Exception
        
        
        print "Clicking on Edit Button of Campaign ::"+campaignTitle
        editButton=wait.until(EC.visibility_of_element_located((By.XPATH,"(//table/tbody/tr/td[1]/a[.='"+campaignTitle+"']/../../td[4]/a[1])[1]")))
        editButton.click()
        okayGotIt=wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div[4]/div/div/div[2]/div[2]/button[1]")))
        
        print "Confirm popup is displayed"
        okayGotIt.click()
        
        wait.until(EC.visibility_of_element_located((By.XPATH,"//header/h1[contains(.,'Edit')]")))
        
        print "Verifying Edit campaign page is displayed"
        header=driver.find_element_by_xpath("//header/h1[contains(.,'Edit')]").text
        if campaignTitle in header:
            print "'"+header+"' page is displayed"
        else:
            print "Edit Campaign page is not displayed"
            raise Exception
        #checking lesson is added
        if driver.find_element_by_xpath("html/body/div[1]/div/div[3]/div[2]/div/section[2]/div[2]/div/div/h4").is_displayed():
            print "without lesson You cant save the campaign"
            raise Exception
        elements.saveButton()   
    def mainVerificationOfEditedCampaignWithoutLesson(self):
        try:
            
            book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
            first_sheet = book.sheet_by_name("CampAssign")
            
            #user
            
            cell = first_sheet.cell(848,1)
            FirstName = cell.value
            
            cell = first_sheet.cell(849,1)
            LastName = cell.value
            
            cell = first_sheet.cell(850,1)
            Email = cell.value
            
            cell = first_sheet.cell(851,1)
            EmployeeId = cell.value
            
            cell = first_sheet.cell(852,1)
            Password = cell.value
            
            cell = first_sheet.cell(853,1)
            NewPassword = cell.value
            
            #lesson Creation
            
            cell = first_sheet.cell(854,1)
            lessonname = cell.value
            
            cell = first_sheet.cell(855,1)
            textCard = cell.value
            
        
            
            #campaign
            
            cell = first_sheet.cell(856,1)
            campaignTitle = cell.value
            
            cell = first_sheet.cell(857,1)
            campDescription = cell.value
            
        
            
            #User Creation
            obj2 =VerificationOfEditedCampaignWithoutLesson()
            obj2.createcreator(FirstName,LastName,Email,EmployeeId,Password)
            obj2.createCreatorUserLogin(Email,Password,NewPassword)
            #lesson Creation
            obj=CreateLessonDifferentLessons()
            obj.lessonWithText(lessonname, textCard)
            
            
            
            #campaign Creation
            
            obj2.campWithTrackTextlesson(campaignTitle, campDescription, lessonname)
            
            #deleting lesson
            
            d=DeleteLesson()
            d.lessonDelete(lessonname) 
            
            #editing the campaign
            obj2.editCampaign(campaignTitle)
            
             
            
            
            
            
            
        except Exception as e:
            traceback.print_exc()
            print (e)
            raise Exception   
          
        finally:
            ob=VerificationOfEditedCampaignWithoutLesson()
            ob.updateUser()
            
            second_sheet = book.sheet_by_name('Login_Credentials')
            cell = second_sheet.cell(1,1)
            url = cell.value
            driver.get(url)
            try:
                WebDriverWait(driver, 5).until(EC.alert_is_present())

                alert = driver.switch_to.alert
                alert.accept()
                print("alert accepted")
            except Exception:
                print("no alert")
    
        
        
        
if __name__=='__main__':
    
    btc=BaseTestClass()
    btc.userLogin() 

    gr=VerificationOfEditedCampaignWithoutLesson()
    gr.mainVerificationOfEditedCampaignWithoutLesson()          
        
        
        
        
        
        