'''
Created on 10-Apr-2018

@author: dattatraya
'''
import os.path
import time
import traceback

from BaseTestClass import BaseTestClass
from BaseTestClass import driver
from openpyxl.reader.excel import load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import xlrd
from BaseTestClass import projectPath
from CampaignPageElements import CampPage
from CreateLearnerNew import CreateLearner
from CreateLessonDifferentLessons import CreateLessonDifferentLessons
from DeleteLesson import DeleteLesson
from PlanAssignmentForPageElements import PlanAssignmentForPageElements
from BaseTestClass import projectPath
class AssignCampVerifyPlanAssignmentPage:
    
    def campaignToVerifyManageAssignPage(self,campaignTitle,campDescription,lessonName,nameOFuser,minPassingScore,numberOfAttempts):
        
        elements=CampPage()
        
        wait=WebDriverWait(driver, 60)
        
        driver.refresh()
        print "\n\nCreating Campaign"
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.campaignButtonFromSideMenuXpath())))
        elements.campaignButtonFromSideMenu()
        
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.createCampaignButtonXpath())))

        if elements.campaignsPageHeaderText()=="Campaigns":
            print "Campaigns page displayed"
        else:
            print "Campaigns page is not displayed"
            raise Exception
        
        
        print "Clicking on Create Campaign button"
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.createCampaignButtonXpath())))
        elements.createCampaignButton()
        
        
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.Camp_titleXpath())))
        print "Create Campaign page is displayed"
        
                  
        elements.titleTextField(campaignTitle)
        print "Title entered"
        
        elements.descriptionField(campDescription)
        print "Description entered"
        
        print "Adding Lesson"
        #Add lesson button
        elements.addlessonButton()
        
        #Waiting until first lesson in pop is displayed
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.FirstLessonWaitXpath())))
        
        #Searching lesson by its name
        elements.searchLesson(lessonName)
        
        #Waiting until lesson displayed
        elements.waitUntilSearchedLessonDisplayed(lessonName)
        
        #selecting searched lesson
        elements.selectSearchedLesson(lessonName)
        
        #waiting until add to campaign button is click able
        wait.until(EC.element_to_be_clickable((By.XPATH,elements.AddToCampaign_ButtonXpath())))
        
        #Clicking on Add to Campaign button
        elements.addToCampaignButton()
        
        
        time.sleep(2)
        #Verifying Added lesson is displayed in Grid
        print "\nVerifying Added lesson is displayed in Grid"
        if elements.firstLessonInGrid()==lessonName:
            print "Lesson displayed in grid"
        else:
            print "Lesson not displayed in grid"
            raise Exception
        
        
        print "Setting minimum passing score"
        elements.makeThisAsAGradedCampaign()
        elements.setMinimumPassingScore(minPassingScore)
        print "Setting maximum number of attempts"
        elements.setAMaxNoOfAttempts(numberOfAttempts)
        
        
        
        wait.until(EC.element_to_be_clickable((By.XPATH,elements.SaveAndExit_ButtonXpath())))
        print "Clicking on Save & Exit button"
        elements.saveAndExitButton()
        
        
        
        
        #Verifying Campaign Detail page is displayed
        print "\nVerifying campaign detail page is displayed"
        
        if elements.campaignDetailPageHeaderText()==campaignTitle:
            print "Campaign detail page is displayed"
        else:
            print "Campaign detail page is not displayed"
            raise Exception
        
        #verifying in Campaigns displayed in Campaigns grid
        elements.searchingForlesson(campaignTitle)
        
        if elements.actualCampTitleINGrid()==campaignTitle:
            print "Campaign '"+campaignTitle+"' displayed in Grid"
        
        else:
            print "Campaign is not displayed in Grid"
            raise Exception
        
        
    def verifyPlanAssignmentPage(self,campaignTitle,planAssignActualText,actualAssignmentTypeText,actualOneTimeTriggeredText,actualTriggeredText,
                                 actualSetUpSectionText,actualAssignDateText,actualDueDateText,actualSearchUserGrpText,actualplaceholderText,
                                 actualGridHeaderText,actualEmptyGridmsg,actualNewHireLabel,actualNewTogrpLabel,actualSelectGroup,actualsearchGrptriggerText,
                                 actualGrpGridHeaderText,actualGrpEmptyMessage,actualLabelCheecboxText,actualMessageAfterClickonChckbox,
                                 actualDueHeaderText,actualSendAssignButtonText,actualCancelButtonText):
        
        elements=CampPage()
        
        wait=WebDriverWait(driver, 60)
        
        driver.refresh()
        print "\n\nCreating Campaign"
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.campaignButtonFromSideMenuXpath())))
        elements.campaignButtonFromSideMenu()
        
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.createCampaignButtonXpath())))

        if elements.campaignsPageHeaderText()=="Campaigns":
            print "Campaigns page displayed"
        else:
            print "Campaigns page is not displayed"
            raise Exception
        
        print "Clicking on link of campaign created"
        elements.CampaignLinkFromCampaignGrid(campaignTitle)
        
        
        print "Verifying Campaign details page is displayed"
        if elements.campaignDetailPageHeaderText()==campaignTitle:
            print "Header Text '"+elements.campaignDetailPageHeaderText()+"' is displayed"
        else:
            print "Header Text not displayed"
            raise Exception  
        
        print "Clicking on Assign button from Campaign detail page"
        assignButton=wait.until(EC.visibility_of_element_located((By.XPATH,elements.assignButtonCampaignDetailsPage())))
        assignButton.click()
        print "Clicked on Assign button"
        planAssign=PlanAssignmentForPageElements()
        
        wait.until(EC.visibility_of_element_located((By.XPATH,planAssign.headerTextPlanAssignment())))
        print "Verifying breadcrumb"
        
        
        try:
            wait.until(EC.visibility_of_element_located((By.XPATH,"//div/div/a[.='Campaigns']/../a[2 and contains(.,'"+campaignTitle+"')]/../span[3 and .='"+planAssignActualText+"']")))
            print "'Campaigns>"+campaignTitle+">"+planAssignActualText+"' is displayed in breadcrumb"
        except Exception:
            print "BreadCrumb is not displayed"
            raise Exception
        
        
        print "Verifying Header"
        driver.find_element_by_xpath(planAssign.headerTextPlanAssignment())
        if "Plan assignment for: "+campaignTitle:
            print "'Plan assignment for: "+campaignTitle+"' is displayed"
        else:
            print "Header is not displayed properly"
            raise Exception
        
        
        print "Verifying Section Assignment type header"
        assignmentTypeHeader=wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[3]/div[2]/div/section[1]/h2")))
        print assignmentTypeHeader.text
        if actualAssignmentTypeText in assignmentTypeHeader.text:
            print "'"+assignmentTypeHeader.text+"' is displayed"
        else:
            print "Assignment type section header is not displayed properly"
            raise Exception
        
        
        #Checking One time triggered radio button is selected by Default
        if driver.find_element_by_xpath("//input[@id='assignment-one-time']").is_selected():
            print "Radio button One Time triggered is selected"
        else:
            print "Radio button is not selected"
            raise Exception
        
        
        print "Verifying one time triggered text"
        oneTimeText=driver.find_element_by_xpath(".//*[@id='content']/div/div[3]/div[2]/div/section[1]/div/div[1]/label/span[1]")
        if oneTimeText.text==actualOneTimeTriggeredText:
            print "'"+oneTimeText.text+"' is displayed"
        else:
            print "One time triggered text is not displayed"
            raise Exception
        
        
        print "Verifying Triggered text is displayed"
        triggeredText=driver.find_element_by_xpath(".//*[@id='content']/div/div[3]/div[2]/div/section[1]/div/div[2]/label/span[1]")
        if triggeredText.text==actualTriggeredText:
            print "'"+triggeredText.text+"' is displayed"
        else:
            print "Triggered text is not displayed"
            raise Exception
        
        
        print "Verifying Set up assignment section header"
        setUp=driver.find_element_by_xpath(".//*[@id='content']/div/div[3]/div[2]/div/section[2]/h2")
        if setUp.text==actualSetUpSectionText:
            print "'"+setUp.text+"' is displayed"
        else:
            print "Set up assignment text is not displayed"
            raise Exception
        
        
        
        
        time.sleep(3)
        print "Difference between two dates"
        from datetime import datetime
        date_format = "%m/%d/%Y"
        a = datetime.strptime(elements.assignDateValue(), date_format)
        b = datetime.strptime(elements.dueDateValue(), date_format)
        delta = b - a
        print "Assign Date ::"+elements.assignDateValue()
        print "Due Date ::"+elements.dueDateValue()
        
        print "The difference between two dates is ::"
        print delta.days
        
        print "Verifying label displayed"
        wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[3]/div[2]/div/section[2]/div/div[1]/div/div[1]/label")))
        assignDate=driver.find_element_by_xpath(".//*[@id='content']/div/div[3]/div[2]/div/section[2]/div/div[1]/div/div[1]/label")
        print assignDate.text
        print actualAssignDateText
        if actualAssignDateText in assignDate.text:
            print "'"+assignDate.text+"' is displayed"
        else:
            print "Assign date label is not  displayed"
            raise Exception
        
        dueDate=driver.find_element_by_xpath(".//*[@id='content']/div/div[3]/div[2]/div/section[2]/div/div[1]/div/div[2]/label")
        if actualDueDateText in dueDate.text:
            print "'"+dueDate.text+"' is displayed"
        else:
            print "Due date label is not  displayed"
            raise Exception
        
        print "Verifying label displayed for searchUserGroup field"
        searchUserGrp=driver.find_element_by_xpath(".//*[@id='content']/div/div[3]/div[2]/div/section[2]/div/div[2]/div/div[1]/label")
        if searchUserGrp.text==actualSearchUserGrpText:
            print "'"+searchUserGrp.text+"' is displayed"
        else:
            print "Search user group label is not  displayed"
            raise Exception
        
        
        print "Verifying Placeholder displayed in Select user field"
        placeholder=driver.find_element_by_xpath("//div[@class='Select-placeholder']")
        if actualplaceholderText in placeholder.text:
            print "'"+placeholder.text+"' is displayed"
        else:
            print "Placeholder is not displayed properly"
            raise Exception
        
        
        print "Verifying Grid header text"
        gridHeadergr=driver.find_element_by_xpath(".//*[@id='content']/div/div[3]/div[2]/div/section[2]/div/div[3]/h3")
        if gridHeadergr.text==actualGridHeaderText:
            print "'"+gridHeadergr.text+"' is displayed"
        else:
            print "Grid header text is not displayed properly"
            raise Exception
        
        print "Verifying Empty grid message"
        emptyGrid=driver.find_element_by_xpath(".//*[@id='content']/div/div[3]/div[2]/div/section[2]/div/div[4]")
        if emptyGrid.text==actualEmptyGridmsg:
            print "'"+emptyGrid.text+"' is displayed"
        else:
            print "Grid empty message is not displayed"
            raise Exception
        
        
        time.sleep(3)
        print "Clicking on Triggered"
        planAssign.triggredRadioButton()
    
        wait.until(EC.visibility_of_element_located((By.XPATH,planAssign.NewHireOnBoarding())))
        
        print "Verifying New Hire onboarding is selected"
        if driver.find_element_by_xpath("//input[@id='trigger-new-hire']").is_selected():
            print "New hire onboarding is selected"
        else:
            print "New hire on-boarding is not selected"
            raise Exception
        
        
        print "Verifying label"
        newHireLabel=driver.find_element_by_xpath(".//*[@id='content']/div/div[3]/div[2]/div/div[2]/section[1]/div/div[1]/label/span[1]")
        if newHireLabel.text==actualNewHireLabel:
            print "'"+newHireLabel.text+"' is displayed"
        else:
            print "New hire onboarding label is not valid"
            raise Exception
        
        newToGroup=driver.find_element_by_xpath(".//*[@id='content']/div/div[3]/div[2]/div/div[2]/section[1]/div/div[2]/label/span[1]")
        if newToGroup.text==actualNewTogrpLabel:
            print "'"+newToGroup.text+"' is displayed"
        else:
            print "New to group label is not valid"
            raise Exception
        
        
        print "Clicking on New to Group radio button"
        driver.find_element_by_xpath(planAssign.NewtoGroup()).click()
        
        wait.until(EC.visibility_of_element_located((By.XPATH,planAssign.AddGroupTextField())))
        
        print "Verifying Search groups label"
        wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[3]/div[2]/div/div[2]/section[1]/div/div[3]/div[1]/div[1]/div/div[1]/label")))
        selectGrp=driver.find_element_by_xpath(".//*[@id='content']/div/div[3]/div[2]/div/div[2]/section[1]/div/div[3]/div[1]/div[1]/div/div[1]/label")
        print selectGrp.text
        print actualSelectGroup
        if selectGrp.text in actualSelectGroup:
            print "'"+selectGrp.text+"' is displayed"
        else:
            print "Label for select group is not displayed properly"
            #raise Exception
        
        print "Verifying Placeholder is displayed"
        searchGrpTriggered=driver.find_element_by_xpath(planAssign.AddGroupTextField())
        if actualsearchGrptriggerText in searchGrpTriggered.text:
            print "'"+searchGrpTriggered.text+"' is displayed"
        else:
            print "Select group text field is not displayed"
            raise Exception 
        
        
        print "Verifying Group grid Header"
        grpGridHeader=driver.find_element_by_xpath(".//*[@id='content']/div/div[3]/div[2]/div/div[2]/section[1]/div/div[3]/div[1]/div[2]/h3")
        if grpGridHeader.text==actualGrpGridHeaderText:
            print "'"+grpGridHeader.text+"' is displayed"
        else:
            print "Group grid header is not displayed properly"
            raise Exception
        
        
        print "Verifying group empty message"
        grpmsg=driver.find_element_by_xpath(".//*[@id='content']/div/div[3]/div[2]/div/div[2]/section[1]/div/div[3]/div[1]/div[3]")
        if grpmsg.text==actualGrpEmptyMessage:
            print "'"+grpmsg.text+"' is displayed"
        else:
            print "Group empty message is not displayed"
            raise Exception
        
        print "Verifying checkbox label"
        labelCheckbox=driver.find_element_by_xpath(".//*[@id='content']/div/div[3]/div[2]/div/div[2]/section[1]/div/div[3]/div[2]/label/span[1]")
        if labelCheckbox.text==actualLabelCheecboxText:
            print "'"+labelCheckbox.text+"' is displayed"  
        else:
            print "Checkbox label is not displayed properly"
            raise Exception
        
        print "Clicking on Checkbox"
        driver.find_element_by_xpath(".//*[@id='content']/div/div[3]/div[2]/div/div[2]/section[1]/div/div[3]/div[2]/label/span[2]").click()
        print "Verifying message is displayed"
        wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[3]/div[2]/div/div[2]/section[1]/div/div[3]/div[3]")))
        msgAfterCheckbox=driver.find_element_by_xpath(".//*[@id='content']/div/div[3]/div[2]/div/div[2]/section[1]/div/div[3]/div[3]")
        if msgAfterCheckbox.text==actualMessageAfterClickonChckbox:
            print "'"+msgAfterCheckbox.text+"' is displayed"
        else:
            print "Message displayed after click on Checkbox is not valid"
            raise Exception
        
        print "Verifying Due Section header"
        dueHeader=driver.find_element_by_xpath(".//*[@id='content']/div/div[3]/div[2]/div/div[2]/section[2]/h2")  
        if dueHeader.text==actualDueHeaderText:
            print "'"+dueHeader.text+"' is displayed"
        else:
            print "Due header is not displayed properly"
            raise Exception
        
        
        print "Verify Send assignment button text"
        sendAssignButton=driver.find_element_by_xpath(".//*[@id='content']/div/div[3]/div[2]/div/div[2]/section[2]/div/div[2]/span/button")
        if sendAssignButton.text==actualSendAssignButtonText:
            print "'"+sendAssignButton.text+"' is displayed"
        else:
            print "Save and send assignment button text is not valid"
            raise Exception
        
        print "Verify Cancel button text"
        cancelButton=driver.find_element_by_xpath(".//*[@id='content']/div/div[3]/div[2]/div/div[2]/section[2]/div/div[2]/a")
        if cancelButton.text==actualCancelButtonText:
            print "'"+cancelButton.text+"' is displayed"
        else:
            print "Cancel button text is not valid"
            raise Exception
        
        
        driver.refresh()
        
    def assignCampToVerifyPlanAssignmentForPage(self):
        
        book = xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        first_sheet = book.sheet_by_name('CampAssign')
        
        
        #Campaign
        cell1 = first_sheet.cell(596,1)
        campaignTitle = cell1.value
        
        cell1 = first_sheet.cell(597,1)
        campDescription = cell1.value
        
        cell1 = first_sheet.cell(598,1)
        minPassingScore = cell1.value
        
        cell1 = first_sheet.cell(599,1)
        numberOfAttempts = cell1.value
        
        cell1 = first_sheet.cell(600,1)
        lessonName = cell1.value
        
        cell2 = first_sheet.cell(601,1)
        questionCard= cell2.value
        
        cell2 = first_sheet.cell(602,1)
        ans1= cell2.value
        
        cell2 = first_sheet.cell(603,1)
        ans2= cell2.value
        
        
        
        #Learner 1
        cell = first_sheet.cell(605,1)
        FirstName = cell.value
        
        FirstNameId = FirstName.split("_")
        emp = FirstNameId[0]+"_"
        ids = FirstNameId[1]
        empId = int(ids)+1
        FirstNameUpdated= emp+str(empId)
        
        cell = first_sheet.cell(606,1)
        LastName = cell.value
        
        LastNameID = LastName.split("_")
        emp = LastNameID[0]+"_"
        ids = LastNameID[1]
        empId = int(ids)+1
        LastNameUpdated= emp+str(empId)
        
        
        
        cell = first_sheet.cell(607,1)
        Email = cell.value
        
        EmailId = Email.split("_")
        emp = EmailId[0]+"_"
        ids = EmailId[1]
        remaining="_"+EmailId[2]
        empId = int(ids)+1
        EmailIdUpdated= emp+str(empId)+remaining
        
        cell = first_sheet.cell(608,1)
        EmployeeId = cell.value
        
        Employee = EmployeeId.split("_")
        emp = Employee[0]+"_"
        ids = Employee[1]
        empId = int(ids)+1
        EmployeeIdUpdated= emp+str(empId)
        
        
        cell = first_sheet.cell(609,1)
        Password = cell.value
        
        cell = first_sheet.cell(610,1)
        NewPassword = cell.value
        
        
        
        #Ui elements for Manage Assignment for page
        cell = first_sheet.cell(612,1)
        planAssignActualText = cell.value
        cell = first_sheet.cell(613,1)
        actualAssignmentTypeText = cell.value
        cell = first_sheet.cell(614,1)
        actualOneTimeTriggeredText = cell.value
        cell = first_sheet.cell(615,1)
        actualTriggeredText = cell.value
        cell = first_sheet.cell(616,1)
        actualSetUpSectionText = cell.value
        cell = first_sheet.cell(617,1)
        actualAssignDateText = cell.value
        cell = first_sheet.cell(618,1)
        actualDueDateText = cell.value
        cell = first_sheet.cell(619,1)
        actualSearchUserGrpText = cell.value
        cell = first_sheet.cell(620,1)
        actualplaceholderText = cell.value
        cell = first_sheet.cell(621,1)
        actualGridHeaderText = cell.value
        cell = first_sheet.cell(622,1)
        actualEmptyGridmsg = cell.value
        cell = first_sheet.cell(623,1)
        actualNewHireLabel = cell.value
        cell = first_sheet.cell(624,1)
        actualNewTogrpLabel = cell.value
        cell = first_sheet.cell(625,1)
        actualSelectGroup = cell.value
        cell = first_sheet.cell(626,1)
        actualsearchGrptriggerText = cell.value
        
        cell = first_sheet.cell(627,1)
        actualGrpGridHeaderText = cell.value
        cell = first_sheet.cell(628,1)
        actualGrpEmptyMessage = cell.value
        cell = first_sheet.cell(629,1)
        actualLabelCheecboxText = cell.value
        cell = first_sheet.cell(630,1)
        actualMessageAfterClickonChckbox = cell.value
        cell = first_sheet.cell(631,1)
        actualDueHeaderText = cell.value
        cell = first_sheet.cell(632,1)
        actualSendAssignButtonText = cell.value
        cell = first_sheet.cell(633,1)
        actualCancelButtonText = cell.value
        
        
        
        
        
        
        
        
        #For Original User
        book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        s_sheet = book.sheet_by_name('Login_Credentials')
        
        cell = s_sheet.cell(1,1)
        url = cell.value
        
        cell = s_sheet.cell(3,1)
        username = cell.value
        
        cell = s_sheet.cell(3,2)
        password = cell.value
        
        
        
        
        #updating user values
        wb = load_workbook(os.path.join('Test_Data/TestData.xlsx'))
            #print (wb.sheetnames)
        
        sheet = wb['CampAssign']
        
        sheet.cell(row=606, column=2).value = FirstNameUpdated
        sheet.cell(row=607, column=2).value = LastNameUpdated
        sheet.cell(row=608, column=2).value = EmailIdUpdated
        sheet.cell(row=609, column=2).value = EmployeeIdUpdated
        
        
        
        wb.save(os.path.join('Test_Data/TestData.xlsx'))
            
        print "All User Data Updated in Excel"
        
        
        
        
        try:
            print "\nCreating a New Learner\n"
            lt=CreateLearner()
            lt.createNewLearnerMain(FirstName, LastName, Email, EmployeeId, Password, NewPassword, url, username, password)
            
            
            print "\nCreating Group\n"
            lk=AssignCampVerifyPlanAssignmentPage()
            
            print "\nCreating a lesson\n"
            ot=CreateLessonDifferentLessons()
            ot.lessonWithQuestion(lessonName, questionCard, ans1, ans2)
            
            
            print "\nCreating Campaign\n"
            lk.campaignToVerifyManageAssignPage(campaignTitle, campDescription, lessonName, FirstName, minPassingScore, numberOfAttempts)            
            
            print "\nVerifying Manage Assignment for page\n"
            lk.verifyPlanAssignmentPage(campaignTitle, planAssignActualText, actualAssignmentTypeText, actualOneTimeTriggeredText, actualTriggeredText,
                                         actualSetUpSectionText, actualAssignDateText, actualDueDateText, actualSearchUserGrpText, actualplaceholderText,
                                          actualGridHeaderText, actualEmptyGridmsg, actualNewHireLabel, actualNewTogrpLabel, actualSelectGroup, 
                                          actualsearchGrptriggerText, actualGrpGridHeaderText, actualGrpEmptyMessage, actualLabelCheecboxText, 
                                          actualMessageAfterClickonChckbox, actualDueHeaderText, actualSendAssignButtonText, actualCancelButtonText)
            
            
            
            
            
            print "\n----Test Execution Completed----\n"
            
            
        except Exception as e:
            traceback.print_exc()
            print (e)
            raise Exception 
         
        finally: 
            driver.save_screenshot("ScreenShots/AssignCampVerifyPlanAssignmentPage.png") 
            
            second_sheet = book.sheet_by_name('Login_Credentials')
            cell = second_sheet.cell(1,1)
            url = cell.value
            driver.get(url)
            try:
                WebDriverWait(driver, 5).until(EC.alert_is_present())

                alert = driver.switch_to.alert
                alert.accept()
                print("alert accepted")
            except Exception:
                print("no alert")
                
            
            
            wait=WebDriverWait(driver, 60)
            unDropDown=wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")))
            driver.execute_script('arguments[0].click()',unDropDown)
            
            signOut=driver.find_element_by_xpath("html/body/div/div/div[1]/div[2]/div[2]/a")
            driver.execute_script('arguments[0].click()',signOut)
            
            
            
            element = wait.until(EC.presence_of_element_located((By.ID, "password")))
            
            if driver.title == "Grovo":
                print("Grovo Application URL Opened")
            else:
                raise Exception.message
    
            print "Grovo Sign-In page is displayed"
            
            print "Entering User name"
            driver.find_element_by_xpath(".//*[@id='username']").send_keys(username)
           
            print "Entering Password"
            element.send_keys(password)
            
            element.send_keys(Keys.TAB)
            print "Clicking on Sign_In button"
            driver.find_element_by_xpath("//*[@id='submitButton']").click()
            
            print "Successfully Logged Into Users account"
            time.sleep(3)
            
            try:
                d=DeleteLesson()
                d.lessonDelete(lessonName)
            except Exception:
                driver.get(url) 
                
if __name__ == '__main__':
    
    btc=BaseTestClass()
    btc.UserLogin()
    
    ne=AssignCampVerifyPlanAssignmentPage()
    ne.assignCampToVerifyPlanAssignmentForPage()
            
            
