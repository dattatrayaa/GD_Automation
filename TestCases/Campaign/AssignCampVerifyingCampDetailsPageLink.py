'''
Created on 09-Apr-2018

@author: dattatraya
'''
import os.path
import time
import traceback

from BaseTestClass import BaseTestClass
from BaseTestClass import driver
from openpyxl.reader.excel import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import xlrd
from BaseTestClass import projectPath
from AddTriggerForPageElements import AddTriggerPage
from CampaignPageElements import CampPage
from CreateLearnerNew import CreateLearner
from CreateLessonDifferentLessons import CreateLessonDifferentLessons
from DeleteLesson import DeleteLesson
from GroupsPageElements import GroupsPageElements
from BaseTestClass import projectPath

class AssignCampVerifyingCampDetailsPageLink:
    
    def campaignWithTwoTriggerForCampDetail(self,campaignTitle,campDescription,lessonName,groupName,minPassingScore,numberOfAttempts):
        
        elements=CampPage()
        driver.refresh()
        wait=WebDriverWait(driver, 60)
        
        
        print "\n\nCreating Campaign"
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.campaignButtonFromSideMenuXpath())))
        elements.campaignButtonFromSideMenu()
        
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.createCampaignButtonXpath())))

        if elements.campaignsPageHeaderText()=="Campaigns":
            print "Campaigns page displayed"
        else:
            print "Campaigns page is not displayed"
            raise Exception
        
        
        print "Clicking on Create Campaign button"
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.createCampaignButtonXpath())))
        elements.createCampaignButton()
        
        
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.Camp_titleXpath())))
        print "Create Campaign page is displayed"
        
                  
        elements.titleTextField(campaignTitle)
        print "Title entered"
        
        elements.descriptionField(campDescription)
        print "Description entered"
        
        print "Adding Lesson"
        #Add lesson button
        elements.addlessonButton()
        
        #Waiting until first lesson in pop is displayed
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.FirstLessonWaitXpath())))
        
        #Searching lesson by its name
        elements.searchLesson(lessonName)
        
        #Waiting until lesson displayed
        elements.waitUntilSearchedLessonDisplayed(lessonName)
        
        #selecting searched lesson
        elements.selectSearchedLesson(lessonName)
        
        #waiting until add to campaign button is click able
        wait.until(EC.element_to_be_clickable((By.XPATH,elements.AddToCampaign_ButtonXpath())))
        
        #Clicking on Add to Campaign button
        elements.addToCampaignButton()
        
        
        time.sleep(2)
        #Verifying Added lesson is displayed in Grid
        print "\nVerifying Added lesson is displayed in Grid"
        if elements.firstLessonInGrid()==lessonName:
            print "Lesson displayed in grid"
        else:
            print "Lesson not displayed in grid"
            raise Exception
        
        
        print "Making this as a Graded campaign"
        elements.makeThisAsAGradedCampaign()
        print "Setting minimum passing score"
        elements.setMinimumPassingScore(minPassingScore)
        print "Setting maximum no of attempts"
        elements.setAMaxNoOfAttempts(numberOfAttempts)
        
        
        wait.until(EC.element_to_be_clickable((By.XPATH,elements.SaveAndExit_ButtonXpath())))
        print "Clicking on Save and Exit button"
        elements.saveAndExitButton()
        
        
        
        
        #Verifying Campaign Detail page is displayed
        print "\nVerifying campaign detail page is displayed"
        
        if elements.campaignDetailPageHeaderText()==campaignTitle:
            print "Campaign detail page is displayed"
        else:
            print "Campaign detail page is not displayed"
            raise Exception
        
        
        addTriggerPage=AddTriggerPage()
        
        
        print "\nCreating trigger\n"
        
        print "\nAdd Trigger"
        elements.addTriggerButton()
        
        
        header=wait.until(EC.visibility_of_element_located((By.XPATH,addTriggerPage.HeaderTextAddTrigger()))) 
        
        #Verifying Plan assignment for page is displayed
        print "Verifying Add trigger for page is displayed"
        if campaignTitle in header.text :
            print "Page '"+header.text+"' is displayed"
        else:
            print "Plan assignment page is not displayed"
            raise Exception
        
        
        #Checking New hire on boarding is selected 
        if driver.find_element_by_xpath("//input[@id='trigger-new-hire']").is_selected():
            print "Radio button New Hire Onboarding is selected"
        else:
            print "Radio button New Hire Onboarding is not selected"
            raise Exception
        
        

        
        print "Selecting New to group"
        addTriggerPage.newToGroupRadio()
        
        
        searchGroups=wait.until(EC.visibility_of_element_located((By.XPATH,addTriggerPage.SelectGroupTextBoxTriggerPage()))) 
        
        print "Adding Group"
        webdriver.ActionChains(driver).move_to_element(searchGroups).click().send_keys(groupName).perform()
        groupdisplayed=wait.until(EC.visibility_of_element_located((By.XPATH,"(//div[@role='option']/span)[1]")))
        
        webdriver.ActionChains(driver).move_to_element(groupdisplayed).click().perform()
        
        
        print "Checking Group is displayed in Grid"
        
        if driver.find_element_by_xpath("//table/tbody/tr/td[1]/a").text==groupName:
            print "Group '"+groupName+"' is displayed in grid"
        else:
            print "Group is not displayed in Grid"
            raise Exception
        
        
        print "Clicking on checkbox 'Also assign to learners who currently match this criteria' "
        
        
        check=wait.until(EC.visibility_of_element_located((By.XPATH,addTriggerPage.AlsoAssignToCurrentUserCheckbox())))
        check.click()
        wait.until(EC.visibility_of_element_located((By.XPATH,addTriggerPage.YourCriteriaMatchesBox())))
        
        print "Saving this trigger"
        addTriggerPage.saveTriggerButton()
        
        print "Checking for In Campaign details page Trigger is displayed"
        try:
            elements.TriggerDisplayedInGridForGroup(groupName)
            print "Trigger with group name '"+groupName+"' is displayed in Campaign details page"
        except Exception:
            print "Trigger is not displayed"
            raise Exception
       
       
        
        
        #verifying in Campaigns displayed in Campaigns grid
        elements.searchingForlesson(campaignTitle)
        
        if elements.actualCampTitleINGrid()==campaignTitle:
            print "Campaign '"+campaignTitle+"' displayed in Grid"
        
        else:
            print "Campaign is not displayed in Grid"
            raise Exception
        
         
         
        print "\nVerifying 'Y' is displayed for created triggred campaign"
        hasTrigger=driver.find_element_by_xpath("(//tr/td[1]/a[.='"+campaignTitle+"']/../../td[2])[1]").text
        
        if hasTrigger=="Y":
            print "'Y' displayed in Has trigger column"
        else:
            print "Invalid data in Hastrigger column"
            
    



    def campDetailPageVerify(self,campaignTitle,actualEditText,actualDuplicateText,actualAssignText,actualLastupdatedGText,
                             actualUpdatedByText,actualDurationDays,actualMinPassScore,actualAllowedAttemptsText,actualReportingText,
                             actualTriggersText, actualAssignmentHeaderText,actualAssignedto,actualStatusText,actualAssignedByText,
                             actualAssignedDateText,actualDueDateText,lessonName,questioncount,timeDisplayed):
        
        elements=CampPage()
        driver.refresh()
        wait=WebDriverWait(driver, 60)
        
        
        print "\n\nCreating Campaign"
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.campaignButtonFromSideMenuXpath())))
        elements.campaignButtonFromSideMenu()
        
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.createCampaignButtonXpath())))

        if elements.campaignsPageHeaderText()=="Campaigns":
            print "Campaigns page displayed"
        else:
            print "Campaigns page is not displayed"
            raise Exception
        
        
        print "Clicking on Created Campaign link from Campaigns Grid"
        elements.CampaignLinkFromCampaignGrid(campaignTitle)
        
        
        print "Verifying Campaign details page is displayed"
        if elements.campaignDetailPageHeaderText()==campaignTitle:
            print "Header Text '"+elements.campaignDetailPageHeaderText()+"' is displayed"
        else:
            print "Header Text not displayed"
            raise Exception
        
        
        print "Verifying Edit, Duplicate, Assign links are displayed in Header"
        try:
            edit=wait.until(EC.visibility_of_element_located((By.XPATH,elements.editButtonCampaignDetailsPage())))
            if edit.text==actualEditText:
                print "'"+edit.text+"' button is displayed in Campaign detail page"
        except Exception:
            print "Edit button is not displayed"
            raise Exception
        
        try:
            duplicate=wait.until(EC.visibility_of_element_located((By.XPATH,elements.duplicateButtonCampaignDetailsPage())))
            if duplicate.text==actualDuplicateText:
                print "'"+duplicate.text+"' button is displayed in Campaign detail page"
        except Exception:
            print "Duplicate button is not displayed"
            raise Exception
        
        try:
            assign=wait.until(EC.visibility_of_element_located((By.XPATH,elements.assignButtonCampaignDetailsPage())))
            if assign.text==actualAssignText:
                print "'"+assign.text+"' button is displayed in Campaign detail page"
        except Exception:
            print "Assign button is not displayed"
            raise Exception
        
        
        print "Verifying Last Updated Text is displayed"
        try:    
            lastupdated=wait.until(EC.visibility_of_element_located((By.XPATH,"//span[contains(.,'"+actualLastupdatedGText+"')]")))     
            print "Text '"+lastupdated.text+"' is displayed"
        except Exception:
            print "Text '"+actualLastupdatedGText+"' is not displayed"
            raise Exception
            
        
        print "Verifying Updated by Text is displayed"
        try:    
            updatedby=wait.until(EC.visibility_of_element_located((By.XPATH,"//span[contains(.,'"+actualUpdatedByText+"')]")))     
            print "Text '"+updatedby.text+"' is displayed"
        except Exception:
            print "Text '"+actualUpdatedByText+"' is not displayed"
        
        
        print "Verifying Duration (Days) Text is displayed"
        try:    
            durationdays=wait.until(EC.visibility_of_element_located((By.XPATH,"//ul/li/small[contains(.,'"+actualDurationDays+"')]")))     
            print "Text '"+durationdays.text+"' is displayed"
        except Exception:
            print "Text '"+actualDurationDays+"' is not displayed"
            
            
        print "Verifying Min passing score Text is displayed"
        try:    
            minPassScore=wait.until(EC.visibility_of_element_located((By.XPATH,"//ul/li/small[contains(.,'"+actualMinPassScore+"')]")))     
            print "Text '"+minPassScore.text+"' is displayed"
        except Exception:
            print "Text '"+actualMinPassScore+"' is not displayed"  
            
        
        print "Verifying Allowed Attempts Text is displayed"
        try:    
            attempts=wait.until(EC.visibility_of_element_located((By.XPATH,"//ul/li/small[contains(.,'"+actualAllowedAttemptsText+"')]")))     
            print "Text '"+attempts.text+"' is displayed"
        except Exception:
            print "Text '"+actualAllowedAttemptsText+"' is not displayed"   
        
        
        print "Verifying Campaign Reporting Text is displayed"
        try:    
            report=wait.until(EC.visibility_of_element_located((By.XPATH,"//ul/li/small[contains(.,'"+actualReportingText+"')]")))     
            print "Text '"+report.text+"' is displayed"
        except Exception:
            print "Text '"+actualReportingText+"' is not displayed"
            
            
        
        
        print "Verifying Triggers Header Text"
        try:
            triggerText=wait.until(EC.visibility_of_element_located((By.XPATH,"//h3[contains(.,'"+actualTriggersText+"')]"))) 
            driver.find_element_by_xpath("//h3[contains(.,'"+actualTriggersText+"')]")    
            print "Text '"+triggerText.text+"' is displayed"
        except Exception:
            print "Text '"+actualTriggersText+"' is not displayed"
        
        
        print "Verifying add trigger button is displayed"
        try:
            if driver.find_element_by_xpath(".//*[@id='content']/div/div[3]/div[2]/div/div[3]/div/div/div/a").is_displayed():
                print "Add Trigger button is displayed"
        except Exception:
            print "Add trigger button is not displayed"
            raise Exception
        
        
        
        print "Verifying New to group text is displayed"
        try:
            wait.until(EC.visibility_of_element_located((By.XPATH,"//tbody/tr/td/div/span[contains(.,'New to group')]")))
            print "New to group text is displayed"
        except Exception:
            print "New to group text is not displayed"
            raise Exception
        
        time.sleep(5)
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        
        
        
        wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[3]/div[2]/div/div[5]/div[2]/table/thead/tr/th[1]")))
        '''print "Verifying Assignment header Text"
        wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[3]/div[2]/div/div[5]/div[1]/div/div/h3")))
        assHeaderText=driver.find_element_by_xpath(".//*[@id='content']/div/div[3]/div[2]/div/div[5]/div[1]/div/div/h3")
        if assHeaderText.text in actualAssignmentHeaderText:
            print "'"+assHeaderText.text+"' is displayed in Header"
        else: 
            print "Header Text is not displayed"
            raise Exception'''
        
        
        print "Verifying Column texts of assignment table is displayed"
        col1=driver.find_element_by_xpath(".//*[@id='content']/div/div[3]/div[2]/div/div[5]/div[2]/table/thead/tr/th[1]")
        if col1.text==actualAssignedto:
            print "'"+col1.text+"' text is displayed"
        else:
            print ""+actualAssignedto+" text is not displayed"
            raise Exception
        
        col2=driver.find_element_by_xpath(".//*[@id='content']/div/div[3]/div[2]/div/div[5]/div[2]/table/thead/tr/th[2]")
        if col2.text==actualStatusText:
            print "'"+col2.text+"' text is displayed"
        else:
            print ""+actualStatusText+" text is not displayed"
            raise Exception
        
        col3=driver.find_element_by_xpath(".//*[@id='content']/div/div[3]/div[2]/div/div[5]/div[2]/table/thead/tr/th[3]")
        if col3.text==actualAssignedByText:
            print "'"+col3.text+"' text is displayed"
        else:
            print ""+actualAssignedByText+" text is not displayed"
            raise Exception
        
        col4=driver.find_element_by_xpath(".//*[@id='content']/div/div[3]/div[2]/div/div[5]/div[2]/table/thead/tr/th[4]")
        if col4.text==actualAssignedDateText:
            print "'"+col4.text+"' text is displayed"
        else:
            print ""+actualAssignedDateText+" text is not displayed"
            raise Exception
        
        col5=driver.find_element_by_xpath(".//*[@id='content']/div/div[3]/div[2]/div/div[5]/div[2]/table/thead/tr/th[5]")
        if col5.text==actualDueDateText:
            print "'"+col5.text+"' text is displayed"
        else:
            print ""+actualAssignedDateText+" text is not displayed"
            raise Exception
        
        
        
        
        
        print "Clicking on Campaign Content tab"
        campaignContentTab=wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[3]/div[2]/div/div[4]/ul/li[2]")))
        campaignContentTab.click()
        
        try:
            print "Verifying lesson displayed is correct"
            lessText=wait.until(EC.visibility_of_element_located((By.XPATH,"//li/div[2]/div/h4/div/span[contains(.,'"+lessonName+"')]")))
            print "Lesson '"+lessText.text+"' is displayed in grid in  Campaign content"
        except Exception:
            print "Lesson not displayed in Campaign content tab"
            raise Exception
        
        
        
        print "Verifying count of question displayed is correct"
        qcount=driver.find_element_by_xpath(".//*[@id='content']/div/div[3]/div[2]/div/div[5]/div/ul/li/div[2]/div/h4/div/small")
        if questioncount in qcount.text:
            print "count '"+qcount.text+"' is displayed in grid in  Campaign content"
        else:
            print "Question count is not displayed in Campaign content tab"
            raise Exception
        
        
        print "Verifying time displayed is correct"
        r=driver.find_element_by_xpath(".//*[@id='content']/div/div[3]/div[2]/div/div[5]/div/ul/li/div[2]/div/div/small")
        if timeDisplayed in r.text:
            print "Time '"+r.text+"' is displayed in grid in  Campaign content"
        else:
            print "Time is not displayed in Campaign content tab"
            raise Exception
        
        
        
        
        print "Verifying Preview link is displayed"
        wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[3]/div[2]/div/div[5]/div/ul/li/div[2]/div/div/button")))
        pre=driver.find_element_by_xpath(".//*[@id='content']/div/div[3]/div[2]/div/div[5]/div/ul/li/div[2]/div/div/button")
        if pre.is_displayed():
            print "Preview link is displayed"
            print "Clicking on Preview link"
            pre.click()
        else:
            print "Time is not displayed in Campaign content tab"
            raise Exception
        
        try:
            print "Verifying preview is displayed for lesson desktop view"
            wait.until(EC.visibility_of_element_located((By.XPATH,"//div[1]/h1/span[contains(.,'"+lessonName+"')]")))
            print "Preview is displayed for lesson :'"+lessonName+"' in desktop view"
        except Exception: 
            print "Preview is not displayed for lesson in desktop view"
            raise Exception
        try:
            print "Verifying preview is displayed for lesson in mobile view"
            mobileButton=wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[4]/div[1]/div[2]/button[2]")))
            mobileButton.click()
            wait.until(EC.visibility_of_element_located((By.XPATH,"//div[1]/h1/span[contains(.,'"+lessonName+"')]")))
            print "Preview is displayed for lesson :'"+lessonName+"' in Mobile view"
            print "Clicking on Close icon"
            driver.find_element_by_xpath(".//*[@id='content']/div/div[4]/div[1]/div[1]/span").click()
            wait.until(EC.visibility_of_element_located((By.XPATH,"//ul/li/div[2]/div/h4/div/span")))
        except Exception: 
            print "Preview is not displayed for lesson in Mobile view"
            raise Exception
        
        
        
        print "All Elements verified in Campaign details page"
        
        driver.refresh()  
            
    def groupCreateForCampaign(self,groupName,FirstName):
        
        driver.refresh()
        wait=WebDriverWait(driver, 60)
        
        group=GroupsPageElements()
        
        group.adminSideMenuUnexpanded()
        print "Clicked on admin icon"
         
        group.groupSideMenuExpanded()  
        print "Clicked on Group icon"
        
        
        print "Checking Group page is displayed"
        if driver.find_element_by_xpath(group.groupPageHeader()).is_displayed():
            print "Group page is displayed successfully"
        else:
            print "Group page is not displayed"
            raise Exception
        
        
        group.createGroupButton()
        print "Clicked on Create Group button"
        
        group.enteringGroupname(groupName)
        
        print "Group Name entered....."
        
        group.nextButton()
        print "Clicked on Next button"
        
        wait.until(EC.visibility_of_element_located((By.XPATH,group.groupDetailPageHeader())))
        print "Checking group is created"
        
        time.sleep(4)
        headerText=driver.find_element_by_xpath(group.groupDetailPageHeader()).text
        
        if headerText==groupName:
            print "Group created successfully"
            
        else:
            print "Group not created"
            raise Exception
        
        
        print "Adding user to Group by name"
        group.addByNameButton()
        group.addByName(FirstName)
        
        print "Verifying user is added to group"
        group.groupAddedInList(FirstName)
        
        group.saveButton()
        print "Clicked on Save button"

        group.saveButtonPopup()
        print "Clicked on Save from popup"
        
        
        print "Checking Created Group is displayed in Grid"
        group.groupsLinkFromBreadCrumb()
        
        print "Searching Group"
        group.createdGroupDisplayedInGrid(groupName)
         
        
        driver.refresh()
        
        
    def assignCampForVerifyCampaignDetailsPage(self):
        
        book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        first_sheet = book.sheet_by_name('CampAssign')
        
        
        #Campaign
        cell1 = first_sheet.cell(504,1)
        campaignTitle = cell1.value
        
        cell1 = first_sheet.cell(505,1)
        campDescription = cell1.value
        
        cell1 = first_sheet.cell(520,1)
        groupName = cell1.value
        
        groupNameId = groupName.split("_")
        emp = groupNameId[0]+"_"
        ids = groupNameId[1]
        empId = int(ids)+1
        GroupNameUpdated= emp+str(empId)
        
        cell1 = first_sheet.cell(506,1)
        minPassingScore = cell1.value
        
        cell1 = first_sheet.cell(507,1)
        numberOfAttempts = cell1.value
        
        cell1 = first_sheet.cell(508,1)
        lessonName = cell1.value
        
        cell2 = first_sheet.cell(509,1)
        questionCard= cell2.value
        
        cell2 = first_sheet.cell(510,1)
        ans1= cell2.value
        
        cell2 = first_sheet.cell(511,1)
        ans2= cell2.value
        
        
        
        
        
        #Learner 1
        cell = first_sheet.cell(513,1)
        FirstName = cell.value
        
        FirstNameId = FirstName.split("_")
        emp = FirstNameId[0]+"_"
        ids = FirstNameId[1]
        empId = int(ids)+1
        FirstNameUpdated= emp+str(empId)
        
        cell = first_sheet.cell(514,1)
        LastName = cell.value
        
        LastNameID = LastName.split("_")
        emp = LastNameID[0]+"_"
        ids = LastNameID[1]
        empId = int(ids)+1
        LastNameUpdated= emp+str(empId)
        
        
        
        cell = first_sheet.cell(515,1)
        Email = cell.value
        
        EmailId = Email.split("_")
        emp = EmailId[0]+"_"
        ids = EmailId[1]
        remaining="_"+EmailId[2]
        empId = int(ids)+1
        EmailIdUpdated= emp+str(empId)+remaining
        
        cell = first_sheet.cell(516,1)
        EmployeeId = cell.value
        
        Employee = EmployeeId.split("_")
        emp = Employee[0]+"_"
        ids = Employee[1]
        empId = int(ids)+1
        EmployeeIdUpdated= emp+str(empId)
        
        cell = first_sheet.cell(517,1)
        Password = cell.value
        
        cell = first_sheet.cell(518,1)
        NewPassword = cell.value
        
        
        
        #UI elements of Campaign Detail page
        cell = first_sheet.cell(522,1)
        actualEditText = cell.value
        cell = first_sheet.cell(523,1)
        actualDuplicateText = cell.value
        cell = first_sheet.cell(524,1)
        actualAssignText = cell.value
        cell = first_sheet.cell(525,1)
        actualLastupdatedGText = cell.value
        cell = first_sheet.cell(526,1)
        actualUpdatedByText = cell.value
        cell = first_sheet.cell(527,1)
        actualDurationDays = cell.value
        cell = first_sheet.cell(528,1)
        actualMinPassScore = cell.value
        cell = first_sheet.cell(529,1)
        actualAllowedAttemptsText = cell.value
        cell = first_sheet.cell(530,1)
        actualReportingText = cell.value
        cell = first_sheet.cell(531,1)
        actualTriggersText = cell.value
        cell = first_sheet.cell(532,1)
        actualAssignmentHeaderText = cell.value
        cell = first_sheet.cell(533,1)
        actualAssignedto = cell.value
        cell = first_sheet.cell(534,1)
        actualStatusText = cell.value
        cell = first_sheet.cell(535,1)
        actualAssignedByText = cell.value
        cell = first_sheet.cell(536,1)
        actualAssignedDateText = cell.value
        cell = first_sheet.cell(537,1)
        actualDueDateText = cell.value
        cell = first_sheet.cell(538,1)
        questioncount = cell.value
        cell = first_sheet.cell(539,1)
        timeDisplayed = cell.value
        
        #For Original User
        book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        s_sheet = book.sheet_by_name('Login_Credentials')
        
        cell = s_sheet.cell(1,1)
        url = cell.value
        
        cell = s_sheet.cell(3,1)
        username = cell.value
        
        cell = s_sheet.cell(3,2)
        password = cell.value
        
        
        
        
        #updating user values
        wb = load_workbook(os.path.join('Test_Data/TestData.xlsx'))
            #print (wb.sheetnames)
        
        sheet = wb['CampAssign']
        
        sheet.cell(row=514, column=2).value = FirstNameUpdated
        sheet.cell(row=515, column=2).value = LastNameUpdated
        sheet.cell(row=516, column=2).value = EmailIdUpdated
        sheet.cell(row=517, column=2).value = EmployeeIdUpdated
        
        sheet.cell(row=521, column=2).value = GroupNameUpdated
        
        
        wb.save(os.path.join('Test_Data/TestData.xlsx'))
            
        print "All User Data Updated in Excel"
        
        
        
        
        try:
            print "\nCreating a New Learner\n"
            lt=CreateLearner()
            lt.createNewLearnerMain(FirstName, LastName, Email, EmployeeId, Password, NewPassword, url, username, password)
            
            
            print "\nCreating Group\n"
            lk=AssignCampVerifyingCampDetailsPageLink()
            lk.groupCreateForCampaign(groupName, FirstName)
            
            print "\nCreating a lesson\n"
            ot=CreateLessonDifferentLessons()
            ot.lessonWithQuestion(lessonName, questionCard, ans1, ans2)
            
            
            print "\nCreating Campaign\n"
            lk.campaignWithTwoTriggerForCampDetail(campaignTitle, campDescription, lessonName, groupName, minPassingScore, numberOfAttempts)
            
            print "\nChecking Campaign detail page is displayed with  all Required elements"
            lk.campDetailPageVerify(campaignTitle, actualEditText, actualDuplicateText, actualAssignText, actualLastupdatedGText, 
                                    actualUpdatedByText, actualDurationDays, actualMinPassScore, actualAllowedAttemptsText, 
                                    actualReportingText, actualTriggersText, actualAssignmentHeaderText, actualAssignedto, actualStatusText, 
                                    actualAssignedByText, actualAssignedDateText, actualDueDateText, lessonName, questioncount, timeDisplayed)
            
            
            
            
            print "\n----Test Execution Completed----\n"
            
            
        except Exception as e:
            traceback.print_exc()
            print (e)
            raise Exception 
         
        finally:  
            driver.save_screenshot("ScreenShots/AssignCampVerifyingCampDetailsPageLink.png")
            
            second_sheet = book.sheet_by_name('Login_Credentials')
            cell = second_sheet.cell(1,1)
            url = cell.value
            driver.get(url)
            try:
                WebDriverWait(driver, 5).until(EC.alert_is_present())

                alert = driver.switch_to.alert
                alert.accept()
                print("alert accepted")
            except Exception:
                print("no alert")
                
            
            
            wait=WebDriverWait(driver, 60)
            unDropDown=wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")))
            driver.execute_script('arguments[0].click()',unDropDown)
            
            signOut=driver.find_element_by_xpath("html/body/div/div/div[1]/div[2]/div[2]/a")
            driver.execute_script('arguments[0].click()',signOut)
            
            
            
            element = wait.until(EC.presence_of_element_located((By.ID, "password")))
            
            if driver.title == "Grovo":
                print("Grovo Application URL Opened")
            else:
                raise Exception.message
    
            print "Grovo Sign-In page is displayed"
            
            print "Entering User name"
            driver.find_element_by_xpath(".//*[@id='username']").send_keys(username)
           
            print "Entering Password"
            element.send_keys(password)
            
            element.send_keys(Keys.TAB)
            print "Clicking on Sign_In button"
            driver.find_element_by_xpath("//*[@id='submitButton']").click()
            
            print "Successfully Logged Into Users account"
            time.sleep(3)
            
            
            try:
                d=DeleteLesson()
                d.lessonDelete(lessonName)
            except Exception:
                driver.get(url) 
                
if __name__ == '__main__':
    
    btc=BaseTestClass()
    btc.UserLogin()
    
    ne=AssignCampVerifyingCampDetailsPageLink()
    ne.assignCampForVerifyCampaignDetailsPage()
                
                
            
