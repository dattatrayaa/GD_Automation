'''
Created on 02-Apr-2018

@author: dattatraya
'''
import os.path
import time
import traceback

from BaseTestClass import BaseTestClass
from BaseTestClass import driver
from openpyxl.reader.excel import load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import xlrd

from CampaignPageElements import CampPage
from CreateLearnerNew import CreateLearner
from CreateLessonDifferentLessons import CreateLessonDifferentLessons
from DeleteLesson import DeleteLesson
from BaseTestClass import projectPath

class AssignCampSelfEnrollNoDueDate:
    
    def campaignSelfEnroll(self,campaignTitle,campDescription,lessonName,minPassingScore,numberOfAttempts):
        
        elements=CampPage()
        
        wait=WebDriverWait(driver, 60)
        
        
        print "\n\nCreating Campaign"
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.campaignButtonFromSideMenuXpath())))
        elements.campaignButtonFromSideMenu()
        
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.createCampaignButtonXpath())))

        if elements.campaignsPageHeaderText()=="Campaigns":
            print "Campaigns page displayed"
        else:
            print "Campaigns page is not displayed"
            raise Exception
        
        
        print "Clicking on Create Campaign button"
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.createCampaignButtonXpath())))
        elements.createCampaignButton()
        
        
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.Camp_titleXpath())))
        print "Create Campaign page is displayed"
        
                  
        elements.titleTextField(campaignTitle)
        print "Title entered"
        
        elements.descriptionField(campDescription)
        print "Description entered"
        
        print "Adding Lesson"
        #Add lesson button
        elements.addlessonButton()
        
        #Waiting until first lesson in pop is displayed
        wait.until(EC.visibility_of_element_located((By.XPATH,elements.FirstLessonWaitXpath())))
        
        #Searching lesson by its name
        elements.searchLesson(lessonName)
        
        #Waiting until lesson displayed
        elements.waitUntilSearchedLessonDisplayed(lessonName)
        
        #selecting searched lesson
        elements.selectSearchedLesson(lessonName)
        
        #waiting until add to campaign button is click able
        wait.until(EC.element_to_be_clickable((By.XPATH,elements.AddToCampaign_ButtonXpath())))
        
        #Clicking on Add to Campaign button
        elements.addToCampaignButton()
        
        
        time.sleep(2)
        #Verifying Added lesson is displayed in Grid
        print "\nVerifying Added lesson is displayed in Grid"
        if elements.firstLessonInGrid()==lessonName:
            print "Lesson displayed in grid"
        else:
            print "Lesson not displayed in grid"
            raise Exception
        
        
        print "Making this as a Graded campaign"
        elements.makeThisAsAGradedCampaign()
        print "Setting minimum passing score"
        elements.setMinimumPassingScore(minPassingScore)
        print "Setting maximum no of attempts"
        elements.setAMaxNoOfAttempts(numberOfAttempts)
        
        
        wait.until(EC.element_to_be_clickable((By.XPATH,elements.SaveAndExit_ButtonXpath())))
        print "Clicking on Save and Exit button"
        elements.saveAndExitButton()
        
        
        
        
        #Verifying Campaign Detail page is displayed
        print "\nVerifying campaign detail page is displayed"
        
        if elements.campaignDetailPageHeaderText()==campaignTitle:
            print "Campaign detail page is displayed"
        else:
            print "Campaign detail page is not displayed"
            raise Exception
        
        
        
        
        print "Clicking on  Get Self Enroll link"
        wait.until(EC.visibility_of_element_located((By.XPATH,"//header/div[2]/div/button")))
        getEnroll=wait.until(EC.element_to_be_clickable((By.XPATH,"//header/div[2]/div/button")))
        getEnroll.click()
        wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[3]/div[2]/div/header/div[2]/div/div/div[1]")))
        print "Checking No Due Date checkbox is selected"
        
        noDueDate=driver.find_element_by_xpath("//input[@id='self-enroll-no-due-date']")
        if noDueDate.is_selected():
            print "No Due Date checkbox is selected"
            
        else:
            print "No Due Date checckbox is not selected"
            raise Exception
        
        
        
        print "Copying link"
        
        
        copyLink=wait.until(EC.element_to_be_clickable((By.XPATH,".//*[@id='content']/div/div[3]/div[2]/div/header/div[2]/div/div/div[1]/div[3]/button")))
        copyLink.click()
        time.sleep(2)
        print "Checking Link is copied"
        self.linkname=driver.find_element_by_xpath(".//*[@id='content']/div/div[3]/div[2]/div/header/div[2]/div/div/div[1]/div[2]/input").get_attribute("value")
            
        print "Link '"+self.linkname+"' is copied to clipboard"
        
        
        #verifying in Campaigns displayed in Campaigns grid
        elements.searchingForlesson(campaignTitle)
        
        if elements.actualCampTitleINGrid()==campaignTitle:
            print "Campaign '"+campaignTitle+"' displayed in Grid"
        
        else:
            print "Campaign is not displayed in Grid"
            raise Exception
        
         
         
        
        
    def checkLinkCampaignEnrolled(self,campaignTitle,EmailId,Password,lessonName,username,password):
        
        print "\nLogin as Learner with self enroll link"
        wait=WebDriverWait(driver, 60)
        driver.refresh()
        wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")))
        
        
        print "Logging out"
        print "Clicking on Username Dropdown"
        ele =driver.find_element_by_xpath(".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")
        driver.execute_script('arguments[0].click()',ele)
        elem=driver.find_element_by_xpath("html/body/div/div/div[1]/div[2]/div[2]/a")
        driver.execute_script('arguments[0].click()',elem)
        
        print "Clicked on Sign Out option"
        
        wait.until(EC.presence_of_element_located((By.ID, "password")))
        
        print "Pasting enroll link in browser"
        driver.get(self.linkname)
        
        
        
        
        element = wait.until(EC.presence_of_element_located((By.ID, "password")))
        if driver.title == "Grovo":
            print("Grovo Application URL Opened")
        else:
            raise Exception.message

        print "Grovo Sign-In page is displayed"
        
        print "Entering User name"
        driver.find_element_by_xpath(".//*[@id='username']").send_keys(EmailId)
       
        print "Entering Password"
        element.send_keys(Password)
        
        element.send_keys(Keys.TAB)
        print "Clicking on Sign_In button"
        driver.find_element_by_xpath("//*[@id='submitButton']").click()
        
        print "Successfully Logged Into Users account"
        time.sleep(30)
        
        driver.refresh()
        time.sleep(10)
        driver.refresh()
        driver.refresh()
        
        try:
            wait.until(EC.visibility_of_element_located((By.XPATH,"//header/h1/em[.='"+campaignTitle+"']")))
            print "Enrolled campaign is displayed for learner"
        except: raise Exception
        
        
        
        print "Starting assignment"
        
        
        startButton=wait.until(EC.visibility_of_element_located((By.XPATH,"//li/a/div[2]/div[1]/div/button")))
        startButton.click()
        
        print "Waiting for lesson name to be displayed"
        
        lessonNameForUser=wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div/div/div[4]/div/div/div/div[1]/h1/span")))
        lessonNameForUserText=lessonNameForUser.text
        if lessonName==lessonNameForUserText:
            print "Verified Lesson '"+lessonNameForUserText+"' is displayed for User"
        else:
            print "Lesson displayed is not valid"
            raise Exception
        
        
        backButtoon=wait.until(EC.element_to_be_clickable((By.XPATH,".//*[@id='content']/div/div/div/div[1]/span/button")))
        backButtoon.click()
        
        unDropDown=wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")))
        driver.execute_script('arguments[0].click()',unDropDown)
        
        signOut=driver.find_element_by_xpath("html/body/div/div/div[1]/div[2]/div[2]/a")
        driver.execute_script('arguments[0].click()',signOut)
        
        
        
        element = wait.until(EC.presence_of_element_located((By.ID, "password")))
        
        if driver.title == "Grovo":
            print("Grovo Application URL Opened")
        else:
            raise Exception.message

        print "Grovo Sign-In page is displayed"
        
        print "Entering User name"
        driver.find_element_by_xpath(".//*[@id='username']").send_keys(username)
       
        print "Entering Password"
        element.send_keys(password)
        
        element.send_keys(Keys.TAB)
        print "Clicking on Sign_In button"
        driver.find_element_by_xpath("//*[@id='submitButton']").click()
        
        print "Successfully Logged Into Users account"
        
        time.sleep(5)
    
    
    
    
    def assignCampWithEnrollLinkNoDueDate(self):
        
        book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        first_sheet = book.sheet_by_name('CampAssign')
        
        
        #Campaign
        cell1 = first_sheet.cell(368,1)
        campaignTitle = cell1.value
        
        cell1 = first_sheet.cell(369,1)
        campDescription = cell1.value
        
        
        cell1 = first_sheet.cell(370,1)
        minPassingScore = cell1.value
        
        cell1 = first_sheet.cell(371,1)
        numberOfAttempts = cell1.value
        
        cell1 = first_sheet.cell(372,1)
        lessonName = cell1.value
        
        cell2 = first_sheet.cell(373,1)
        questionCard= cell2.value
        
        cell2 = first_sheet.cell(374,1)
        ans1= cell2.value
        
        cell2 = first_sheet.cell(375,1)
        ans2= cell2.value
        
        
        
        
        
        #Learner 1
        cell = first_sheet.cell(377,1)
        FirstName = cell.value
        
        FirstNameId = FirstName.split("_")
        emp = FirstNameId[0]+"_"
        ids = FirstNameId[1]
        empId = int(ids)+1
        FirstNameUpdated= emp+str(empId)
        
        cell = first_sheet.cell(378,1)
        LastName = cell.value
        
        LastNameID = LastName.split("_")
        emp = LastNameID[0]+"_"
        ids = LastNameID[1]
        empId = int(ids)+1
        LastNameUpdated= emp+str(empId)
        
        
        
        cell = first_sheet.cell(379,1)
        Email = cell.value
        
        EmailId = Email.split("_")
        emp = EmailId[0]+"_"
        ids = EmailId[1]
        remaining="_"+EmailId[2]
        empId = int(ids)+1
        EmailIdUpdated= emp+str(empId)+remaining
        
        cell = first_sheet.cell(380,1)
        EmployeeId = cell.value
        
        Employee = EmployeeId.split("_")
        emp = Employee[0]+"_"
        ids = Employee[1]
        empId = int(ids)+1
        EmployeeIdUpdated= emp+str(empId)
        
        cell = first_sheet.cell(381,1)
        Password = cell.value
        
        cell = first_sheet.cell(382,1)
        NewPassword = cell.value
        
        cell = first_sheet.cell(383,1)
        role = cell.value
        
        #For Original User
        book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        s_sheet = book.sheet_by_name('Login_Credentials')
        
        cell = s_sheet.cell(1,1)
        url = cell.value
        
        cell = s_sheet.cell(3,1)
        username = cell.value
        
        cell = s_sheet.cell(3,2)
        password = cell.value
        
        
        
        #updating user values
        wb = load_workbook(os.path.join('Test_Data/TestData.xlsx'))
            #print (wb.sheetnames)
        
        sheet = wb['CampAssign']
        
        sheet.cell(row=378, column=2).value = FirstNameUpdated
        sheet.cell(row=379, column=2).value = LastNameUpdated
        sheet.cell(row=380, column=2).value = EmailIdUpdated
        sheet.cell(row=381, column=2).value = EmployeeIdUpdated
        
        
        wb.save(os.path.join('Test_Data/TestData.xlsx'))
            
        print "All User Data Updated in Excel"
        
        
        
        
        try:
            print "\nCreating a New Learner\n"
            lt=CreateLearner()
            lt.createNewLearnerMain(FirstName, LastName, Email, EmployeeId, Password, NewPassword, url, username, password)
            
            
            print "\nCreating a lesson\n"
            ot=CreateLessonDifferentLessons()
            ot.lessonWithQuestion(lessonName, questionCard, ans1, ans2)
            
            
            print "\nCreating Campaign\n"
            lk=AssignCampSelfEnrollNoDueDate()
            lk.campaignSelfEnroll(campaignTitle, campDescription, lessonName, minPassingScore, numberOfAttempts)
            
            print "\nOpening enroll link for new user and assigning campaign"
            lk.checkLinkCampaignEnrolled(campaignTitle, Email, Password, lessonName, username, password)
            
            d=DeleteLesson()
            d.lessonDelete(lessonName)
            
            print "\n----Test Execution Completed----\n"
            
            
        except Exception as e:
            traceback.print_exc()
            print (e)
            raise Exception 
         
        finally:  
            second_sheet = book.sheet_by_name('Login_Credentials')
            cell = second_sheet.cell(1,1)
            url = cell.value
            driver.get(url)
            try:
                WebDriverWait(driver, 5).until(EC.alert_is_present())

                alert = driver.switch_to.alert
                alert.accept()
                print("alert accepted")
            except Exception:
                print("no alert")
            
            wait=WebDriverWait(driver, 60)
            unDropDown=wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")))
            driver.execute_script('arguments[0].click()',unDropDown)
            
            signOut=driver.find_element_by_xpath("html/body/div/div/div[1]/div[2]/div[2]/a")
            driver.execute_script('arguments[0].click()',signOut)
            
            
            
            element = wait.until(EC.presence_of_element_located((By.ID, "password")))
            
            
            time.sleep(3)
            if driver.title == "Grovo":
                print("Grovo Application URL Opened")
            else:
                raise Exception.message
    
            print "Grovo Sign-In page is displayed"
            
            print "Entering User name"
            driver.find_element_by_xpath(".//*[@id='username']").send_keys(username)
           
            print "Entering Password"
            element.send_keys(password)
            
            element.send_keys(Keys.TAB)
            print "Clicking on Sign_In button"
            driver.find_element_by_xpath("//*[@id='submitButton']").click()
            
            print "Successfully Logged Into Users account"
            time.sleep(3)
            
            
            
            
if __name__ == '__main__':
    
    btc=BaseTestClass()
    btc.UserLogin()
    
    hk=AssignCampSelfEnrollNoDueDate()
    hk.assignCampWithEnrollLinkNoDueDate()
        


    
    
    
    