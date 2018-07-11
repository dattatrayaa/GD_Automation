'''
Created on 02-Jul-2018

@author: geethukn
'''
from Tkinter import Tk
from operator import contains
import os.path
import time
import traceback

from openpyxl.reader.excel import load_workbook
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import wait, expected_conditions as EC
from selenium.webdriver.support.select import Select
import xlrd

from ADPElements import ADPXpath
from BaseTestClass import BaseTestClass
from BaseTestClass import WebDriverWait
from BaseTestClass import driver
from BaseTestClass import projectPath

class ADPIntegration:
    def adpIntegration(self):
        ADPIntegrations = ADPXpath()
        
        wait=WebDriverWait(driver,90)
        wait.until(EC.visibility_of_element_located((By.XPATH,ADPIntegrations.adminMenu())))
        # Click on Admin from side menu
        driver.find_element_by_xpath(ADPIntegrations.adminMenu()).click()
        print "Click on Admin from side menu"
        
        
        wait=WebDriverWait(driver,90)
        wait.until(EC.visibility_of_element_located((By.XPATH,ADPIntegrations.integrationMenu())))
        review = driver.find_element_by_xpath(ADPIntegrations.integrationMenu())
        webdriver.ActionChains(driver).move_to_element(review).click(review).perform()
        
        
        wait=WebDriverWait(driver,90)
        wait.until(EC.visibility_of_element_located((By.XPATH,ADPIntegrations.adpConfigure_Button())))
        # Click on Configure button
        driver.find_element_by_xpath(ADPIntegrations.adpConfigure_Button()).click()
        print "Click on Configure button"
        
        wait=WebDriverWait(driver,90)
        wait.until(EC.visibility_of_element_located((By.XPATH,ADPIntegrations.goToADP())))
        # Click on Go To ADP Button
        driver.find_element_by_xpath(ADPIntegrations.goToADP()).click()
        print "Click on Go To ADP button"
        
        time.sleep(6)
        window_after = driver.window_handles[1]
        driver.switch_to.window(window_after)
        wait.until(EC.visibility_of_element_located((By.XPATH,ADPIntegrations.adpBuyNow())))
        wait.until(EC.element_to_be_clickable((By.XPATH,ADPIntegrations.adpBuyNow())))
        # Click on Go To ADP Button
        review = driver.find_element_by_xpath(ADPIntegrations.adpBuyNow())
        webdriver.ActionChains(driver).move_to_element(review).click(review).perform()
        print "Click on Buy now button button"
        
        #login to ADPMarket place
        wait.until(EC.visibility_of_element_located((By.XPATH,ADPIntegrations.adpLoginusename())))
        print "Reading data from excel sheet"
        book=xlrd.open_workbook(os.path.join('TestData.xlsx'))
        
        first_sheet = book.sheet_by_name('ADP User details')

        # read a cell
        cell = first_sheet.cell(1,0)
        ADPUserID = cell.value
        print "UserId : "+ADPUserID
        cell = first_sheet.cell(1,1)
        Password = cell.value
        print "Password : "+Password
       
        userid =driver.find_element_by_xpath(ADPIntegrations.adpLoginusename())        
        webdriver.ActionChains(driver).move_to_element(userid).click(userid).perform()
        userid.send_keys(ADPUserID)
        password =driver.find_element_by_xpath(ADPIntegrations.adpLoginpassword())        
        webdriver.ActionChains(driver).move_to_element(password).click(password).perform()
        password.send_keys(Password)
        driver.find_element_by_xpath(ADPIntegrations.adpsigninbutton()).click()
        print "Click on Sign_In button"
        
        #adding Discount code
        print "Reading data from excel sheet"
        book=xlrd.open_workbook(os.path.join('TestData.xlsx'))
        first_sheet = book.sheet_by_name('ADP User details')
        # read a cell
        cell = first_sheet.cell(1,2)
        Discountcodevalue = cell.value
        print "Discount Code : "+Discountcodevalue
        wait.until(EC.visibility_of_element_located((By.XPATH,ADPIntegrations.discountCode())))
        Discountcode = driver.find_element_by_xpath(ADPIntegrations.discountCode())
        webdriver.ActionChains(driver).move_to_element(Discountcode).click(Discountcode).perform()
        Discountcode.send_keys(Discountcodevalue)
        driver.find_element_by_xpath(ADPIntegrations.apply()).click()
        print "Click on Apply button"
        
        driver.find_element_by_xpath(ADPIntegrations.continuebutton()).click()
        print "Click on Continue button"
        
        window_after = driver.window_handles[0]
        driver.switch_to.window(window_after)
        
        wait.until(EC.visibility_of_element_located((By.XPATH,ADPIntegrations.grovoidcopy())))
        
        #Fetching - Your Grovo ID and saving in Excel
        driver.find_element_by_xpath(ADPIntegrations.grovoidcopy()).click()
        print "Copy on Your Grovo ID"         
        # Writing the copied API key in to Excel 
        data = Tk().clipboard_get()
        #Pasting the copied API key in to excel
        wb = load_workbook(os.path.join('D:/_WorkSpace/EclipseWS/PythonAutomation/src/TestData.xlsx'))
        #print (wb.sheetnames)
        sheet = wb['ADP User details']
        sheet.cell(row=2, column=4).value=data
        wb.save(os.path.join('D:/_WorkSpace/EclipseWS/PythonAutomation/src/TestData.xlsx'))
        print "Pasting the copied Grovo ID in to Test Data Excel"
        
        #Fetching - Your Grovo Organization ID and saving in Excel
        driver.find_element_by_xpath(ADPIntegrations.grovoorganizationidcopy()).click()
        print "Copy on Your Grovo Organization ID"        
        # Writing the copied API key in to Excel 
        data = Tk().clipboard_get()
        #Pasting the copied API key in to excel
        wb = load_workbook(os.path.join('D:/_WorkSpace/EclipseWS/PythonAutomation/src/TestData.xlsx'))
        #print (wb.sheetnames)
        sheet = wb['ADP User details']
        sheet.cell(row=2, column=5).value=data
        wb.save(os.path.join('D:/_WorkSpace/EclipseWS/PythonAutomation/src/TestData.xlsx'))
        print "Pasting the copied  Grovo Organization ID in to Test Data Excel"
        
        window_after = driver.window_handles[1]
        driver.switch_to.window(window_after)
        
        #Clearing user_id from ADP
        wait.until(EC.visibility_of_element_located((By.XPATH,ADPIntegrations.adpuserid())))
        driver.find_element_by_xpath(ADPIntegrations.adpuserid()).click()
        driver.find_element_by_xpath(ADPIntegrations.adpuserid()).clear()
        #adding user_id
        print "Reading data from excel sheet"
        book=xlrd.open_workbook(os.path.join('TestData.xlsx'))
        first_sheet = book.sheet_by_name('ADP User details')
        # read a cell
        cell = first_sheet.cell(1,3)
        adpuser_id = cell.value
        print "user_id : "+adpuser_id
        driver.find_element_by_xpath(ADPIntegrations.adpuserid()).send_keys(adpuser_id)
        
        #Clearing enterprise_id from ADP
        wait.until(EC.visibility_of_element_located((By.XPATH,ADPIntegrations.enterpriseid())))
        driver.find_element_by_xpath(ADPIntegrations.enterpriseid()).click()
        driver.find_element_by_xpath(ADPIntegrations.enterpriseid()).clear()
        #adding enterprise_id
        print "Reading data from excel sheet"
        book=xlrd.open_workbook(os.path.join('TestData.xlsx'))
        first_sheet = book.sheet_by_name('ADP User details')
        # read a cell
        cell = first_sheet.cell(1,4)
        enterpriseid = cell.value
        print "enterprise_id : "+enterpriseid
        driver.find_element_by_xpath(ADPIntegrations.enterpriseid()).send_keys(enterpriseid)
        
        driver.find_element_by_xpath(ADPIntegrations.continuebuttonafteradding()).click()
        print"Click  on Continue button"
        
        #Click on Terms and condition
        wait.until(EC.visibility_of_element_located((By.XPATH,ADPIntegrations.agreeToTermsCheckbox())))
        TermsCheckbox = driver.find_element_by_xpath(ADPIntegrations.agreeToTermsCheckbox())
        webdriver.ActionChains(driver).move_to_element(TermsCheckbox).click(TermsCheckbox).perform()
        driver.find_element_by_xpath(ADPIntegrations.placeOrder()).click()
        print "Click on Place Order button "
        
        #Click on Go to myapps
        wait.until(EC.visibility_of_element_located((By.XPATH,ADPIntegrations.gotomyapps())))
        gotomyappsbutton = driver.find_element_by_xpath(ADPIntegrations.gotomyapps())
        webdriver.ActionChains(driver).move_to_element(gotomyappsbutton).click(gotomyappsbutton).perform()
        print "Click on GO TO MYAPPS button"
        print"You successfully subscribed to Grovo Microlearning Solution and can start using the application from your MyApps page. We have sent you an email with your receipt."
        
        #driver.find_element_by_tag_name('body').send_keys(Keys.COMMAND + 'w')
        
        window_after = driver.window_handles[0]
        driver.switch_to.window(window_after)
        
        driver.find_element_by_xpath(ADPIntegrations.connectbutton()).click()
        print"Click  on Connect button"
        
        wait.until(EC.visibility_of_element_located((By.XPATH,ADPIntegrations.consenturl())))
        driver.find_element_by_xpath(ADPIntegrations.consenturl()).click()
        print"Click on Consent URL from Grovo"
        
        time.sleep(6)
        window_after = driver.window_handles[2]
        driver.switch_to.window(window_after)
        
        #Click on Manage button from Pending consents
        wait.until(EC.visibility_of_element_located((By.XPATH,ADPIntegrations.managebutton())))
        driver.find_element_by_xpath(ADPIntegrations.managebutton()).click()
        print"Click on Manage button from ADP Pending Consents"
        
        wait.until(EC.visibility_of_element_located((By.XPATH,ADPIntegrations.allowbutton())))
        allow = driver.find_element_by_xpath(ADPIntegrations.allowbutton())
        webdriver.ActionChains(driver).move_to_element(allow).click(allow).perform()
        print "Click on Allow button"
        
        
        wait.until(EC.visibility_of_element_located((By.XPATH,ADPIntegrations.continuefromconsent())))
        driver.find_element_by_xpath(ADPIntegrations.continuefromconsent()).click()
        print"Click on Continue button from ADP Pending Consents"
        
        window_after = driver.window_handles[0]
        driver.switch_to.window(window_after)
        
        wait.until(EC.visibility_of_element_located((By.XPATH,ADPIntegrations.connectbutton())))
        driver.find_element_by_xpath(ADPIntegrations.connectbutton()).click()
        print"Click on Connect from Grovo"
        
        #Select Screen
        wait.until(EC.visibility_of_element_located((By.XPATH,ADPIntegrations.search())))
        
        
        
        
        
        
        # select check box for user field name
        
        driver.find_element_by_xpath(ADPIntegrations.userfield_checkbox()).click()
        
        time.sleep(2)
        driver.find_element_by_xpath(ADPIntegrations.userfield_checkbox()).click()
        print "Select only required user fields by double selecting the check box next to USER FIELD NAME"
        
        time.sleep(2)
        
        # Clicking on Next Button
        
        driver.find_element_by_xpath(ADPIntegrations.nextbutton_selectscreen()).click()
        print "Clicking on Next button from select screen"
        
        time.sleep(2)
        
        # Clicking on Sync button from Confirm screen
        driver.find_element_by_xpath(ADPIntegrations.syncbutton_confirmscreen()).click()
        print "Clicking on Sync button from select screen"
        
        time.sleep(2)
        
        # Clicking on Check now link
        
        while True:
            
            try:
                
                wait.until(EC.visibility_of_element_located((By.XPATH, ADPIntegrations.connected_status())))
                element = wait.until(EC.element_to_be_clickable((By.XPATH, ADPIntegrations.checknow_link())))
                time.sleep(9)
                element.click()
                
            except:
                print "Failed to find check now link"
                break
                              
        #driver.find_element_by_xpath(bamboohr.checknow_link()).click()
        print "Clicking on Check now link"
        
        wait.until(EC.visibility_of_element_located((By.XPATH,ADPIntegrations.connected_status())))
        print "Wait until the status Connected is displayed"
        
        time.sleep(5)
        # Clicking on Users link in side menu tab
        
        print "Going to view the users managed by ADPHRIS"
        
        driver.find_element_by_xpath(ADPIntegrations.users_tab()).click()
        print "Clicking on Users menu"
        
        # wait for the search user field
        wait.until(EC.visibility_of_element_located((By.XPATH,ADPIntegrations.search_users_field())))
        
        print "Enter the First name of synced user in search users field"
        
        cell3= first_sheet.cell(1,6)
        user_firstname = cell3.value 
        
        webdriver.ActionChains(driver).move_to_element(driver.find_element_by_xpath(ADPIntegrations.search_users_field())).click( driver.find_element_by_xpath(ADPIntegrations.search_users_field())).perform()
        
        driver.find_element_by_xpath(ADPIntegrations.search_users_field()).send_keys(user_firstname)
        print "The First name of synced user is "+" "+user_firstname
        
        # Going to Click and open the created users
        time.sleep(10)
        
        wait.until(EC.visibility_of_element_located((By.XPATH,"//table/tbody/tr[1]/td[2]/a[.='"+user_firstname+"']")))
        driver.find_element_by_xpath("//table/tbody/tr[1]/td[2]/a[.='"+user_firstname+"']").click()
        print "Click on the synced user"
        
        # Going to get the First name
        
        wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='create-edit-user-search-firstName']")))
        actual_firstname= driver.find_element_by_xpath(".//*[@id='create-edit-user-search-firstName']").get_attribute('value')
        print "The displayed first name is "+actual_firstname
        
        actual_lastname= driver.find_element_by_xpath(".//*[@id='create-edit-user-search-lastName']").get_attribute('value')
        print "The displayed last name is "+actual_lastname
        
        
        actual_emailidd= driver.find_element_by_xpath(".//*[@id='create-edit-user-search-username']").get_attribute('value')
        print "The displayed email id is "+actual_emailidd
         
                
        actual_employeeid= driver.find_element_by_xpath(".//*[@id='create-edit-user-search-employeeId']").get_attribute('value')
        print "The displayed employee id is "+actual_employeeid
        
        
        cell3= first_sheet.cell(1,6)
        user_firstname = cell3.value 
        
        cell4= first_sheet.cell(1,7)
        expected_lastname = cell4.value 
        
        cell5= first_sheet.cell(1,8)
        expected_emailid = cell5.value 
        
        cell6= first_sheet.cell(1,9)
        expected_employeeid = cell6.value 
        
        #print expected_lastname
        # print expected_emailid
        # print expected_employeeid
        
        if(actual_firstname == user_firstname):
            
            if(actual_lastname == expected_lastname):
                
                if(actual_emailidd == expected_emailid):
                    
                    if(actual_employeeid == expected_employeeid):
                    
                        print "The synced user details are matching as in ADP Market place"
           
        else:
            print "Failed to validate the synced user details"
            raise Exception
        
        # Going to Click on User Attributes tab from side menu
        
        driver.find_element_by_xpath(ADPIntegrations.user_attributes_tab()).click()
        print "Clicking on User attributes tab"
        
        # Going to Click on Grovo attributes tab
        driver.find_element_by_xpath(ADPIntegrations.grovo_attributres_subtab()).click()
        print "Clicking on Grovo Attributes tab"
        
        # Going to fetch the HSRIS Integrations text for the synced attributes
        
        #get the text for First Name.Last name, Primary email, Employee id
        
        wait.until(EC.visibility_of_element_located((By.XPATH,".//table/tbody/tr/td[1]/a[.='First Name']/../../td[2]")))
        
        First_name_source= driver.find_element_by_xpath(".//table/tbody/tr/td[1]/a[.='First Name']/../../td[2]").text
        Last_name_source= driver.find_element_by_xpath(".//table/tbody/tr/td[1]/a[.='Last Name']/../../td[2]").text
        Employeeid_source= driver.find_element_by_xpath(".//table/tbody/tr/td[1]/a[.='Employee ID']/../../td[2]").text
        Primary_email= driver.find_element_by_xpath(".//table/tbody/tr/td[1]/a[.='Primary Email']/../../td[2]").text
        
        
        
        if (First_name_source == "HRIS Integration"):
            
            if(Last_name_source == "HRIS Integration"):
                
                if(Employeeid_source == "HRIS Integration"):
                    
                    if(Primary_email == "HRIS Integration"):
                        
                        print "The source for synced attributes are displaying as HRIS Integration"
                        
        else:
                         
            print "Failed to validate the attribute source details"
            raise Exception
        
        
    def createuser_in_adp(self,Employee_FirstName,Employee_LastName,Work_Email):
        
          
        ADPIntegrations = ADPXpath()
        wait=WebDriverWait(driver, 60)
        print "Going to create a user in ADP account"
        
        
        book=xlrd.open_workbook(os.path.join('TestData.xlsx'))
        first_sheet = book.sheet_by_name('ADP User details')
        
        cell4= first_sheet.cell(1,5)
        adpmarketplaceurl = cell4.value 
        
        cell5= first_sheet.cell(1,0)
        adpusername = cell5.value 
        
        cell6= first_sheet.cell(1,1)
        adppassword = cell6.value 
        
        cell7= first_sheet.cell(1,14)
        hire_date = cell7.value 
        
        cell8= first_sheet.cell(1,13)
        reason_forhire = cell8.value
        
        cell9= first_sheet.cell(1,15)
        company_code = cell9.value
        
        cell10= first_sheet.cell(1,16)
        taxid_type = cell10.value
        
        cell11= first_sheet.cell(1,17)
        gender = cell11.value
        
        cell12= first_sheet.cell(1,18)
        dob = cell12.value
        
        cell13= first_sheet.cell(1,19)
        address_line = cell13.value
        
        cell14= first_sheet.cell(1,20)
        adp_city = cell14.value
        
        cell15= first_sheet.cell(1,21)
        state = cell15.value
        
        cell16= first_sheet.cell(1,22)
        zipcodes = cell16.value
        
        cell17= first_sheet.cell(1,23)
        file_number = cell17.value
        
        cell18= first_sheet.cell(1,24)
        regularpayrate = cell18.value
        
        cell19= first_sheet.cell(1,25)
        federalsq = cell19.value
        
        print "Going to access ADP account url"+" "+adpmarketplaceurl
        driver.get(adpmarketplaceurl)
        driver.maximize_window()
        
        wait.until(EC.visibility_of_element_located((By.XPATH,ADPIntegrations.loginurladp())))
        driver.find_element_by_xpath(ADPIntegrations.loginurladp()).click()
        
        wait.until(EC.visibility_of_element_located((By.XPATH,ADPIntegrations.adpusername())))
        driver.find_element_by_xpath(ADPIntegrations.adpusername()).send_keys(adpusername)
        print "Enter ADP email for login"
        
        driver.find_element_by_xpath(ADPIntegrations.adpsubmit()).click()
        
        wait.until(EC.visibility_of_element_located((By.XPATH,ADPIntegrations.adppassword())))
        driver.find_element_by_xpath(ADPIntegrations.adppassword()).send_keys(adppassword)
        print "Enter ADP password for login"
        
        driver.find_element_by_xpath(ADPIntegrations.passwordsubmit()).click()
        
        # Waiting until the Dash board is displaying
        
        wait.until(EC.visibility_of_element_located((By.XPATH,ADPIntegrations.adpprocess())))
        driver.find_element_by_xpath(ADPIntegrations.adpprocess()).click()
        print "Click on Process from ADP Market place"
        
        driver.find_element_by_xpath(ADPIntegrations.adphr()).click()
        print "Click on HR menu from ADP Market place"
        
        driver.find_element_by_xpath(ADPIntegrations.hireandrehire()).click()
        print "Click on Hire and Rehire link from ADP Market place"
        
        wait.until(EC.visibility_of_element_located((By.XPATH,ADPIntegrations.quickhire())))
        driver.find_element_by_xpath(ADPIntegrations.quickhire()).click()
        print "Click on Quick hire button from ADP Market place"
        
        #Enter First Name
        wait.until(EC.visibility_of_element_located((By.XPATH,ADPIntegrations.firstnameadp())))
        driver.find_element_by_xpath(ADPIntegrations.firstnameadp()).send_keys(Employee_FirstName)
        print "Enter employee first name as "+Employee_FirstName
        
        #Enter Last Name
        driver.find_element_by_xpath(ADPIntegrations.lastnameadp()).send_keys(Employee_LastName)
        print "Enter employee Last name as "+Employee_LastName
        
        # Hire Date
        driver.find_element_by_xpath(ADPIntegrations.hiredateadp()).send_keys(hire_date)
        print "Enter Hire date as "+hire_date
       
        # Reason for Hire 
        driver.find_element_by_xpath(ADPIntegrations.reasonforhire()).send_keys(reason_forhire)
        print "Enter Reason for Hire as "+reason_forhire
        
        # Company code
        driver.find_element_by_xpath(ADPIntegrations.companycode()).send_keys(company_code)
        print "Enter company code as "+company_code
        
        # Tax id type
        driver.find_element_by_xpath(ADPIntegrations.taxidtype()).send_keys(taxid_type)
        driver.find_element_by_xpath(ADPIntegrations.taxidtypearrow()).click()
        print "Enter Tax Id type as "+taxid_type
        
        driver.find_element_by_xpath(ADPIntegrations.taxidappliedfor()).click()
        
        # Gender
        driver.find_element_by_xpath(ADPIntegrations.gender()).send_keys(gender)
        print "Enter Gender as "+gender
        
        #DOB
        driver.find_element_by_xpath(ADPIntegrations.dob()).send_keys(dob)
        print "Enter DOB as "+dob
        
        #Address Line1
        driver.find_element_by_xpath(ADPIntegrations.addressline()).send_keys(address_line)
        print "Enter Address as "+address_line
        
        #City
        driver.find_element_by_xpath(ADPIntegrations.city()).send_keys(adp_city)
        print "Enter City as "+adp_city
        
        #State
        driver.find_element_by_xpath(ADPIntegrations.state()).send_keys(state)
        print "Enter State as "+state
        
        #Zip code
        zcode = driver.find_element_by_xpath(ADPIntegrations.zipcode())
        webdriver.ActionChains(driver).move_to_element(zcode).click(zcode).perform()
        driver.find_element_by_xpath(ADPIntegrations.zipcode()).send_keys(int(zipcodes))
        #print "Enter Zip code as "+int(zipcodes)
        
        #File Number
        filenumber = driver.find_element_by_xpath(ADPIntegrations.filenumber())
        webdriver.ActionChains(driver).move_to_element(filenumber).click(filenumber).perform()
        driver.find_element_by_xpath(ADPIntegrations.filenumber()).send_keys(int(file_number))
        #print "Enter File Number as "+file_number
        
        #regularpayrate
        
        driver.find_element_by_xpath(ADPIntegrations.regularpayrate()).send_keys(int(regularpayrate))
        #print "Enter Regular pay rate as "+regularpayrate
        
        #Federal
        federals = driver.find_element_by_xpath(ADPIntegrations.federal())
        webdriver.ActionChains(driver).move_to_element(federals).click(federals).perform()
        driver.find_element_by_xpath(ADPIntegrations.federal()).send_keys(int(federalsq))
        #print "Enter Federals as "+federalsq
        
        done_adp = driver.find_element_by_xpath(ADPIntegrations.doneadp())
        webdriver.ActionChains(driver).move_to_element(done_adp).click(done_adp).perform()
        print "Click on Done button"
        
        driver.find_element_by_xpath(ADPIntegrations.donepopup()).click()
        print "Click on Done from Popup"
        
        wait.until(EC.visibility_of_element_located((By.XPATH,ADPIntegrations.searchadp())))
        driver.find_element_by_xpath(ADPIntegrations.searchadp()).send_keys(Employee_FirstName)
        print"Enter Employee First name: "+Employee_FirstName
        
         
        wait.until(EC.visibility_of_element_located((By.XPATH,ADPIntegrations.clicksearckfn())))
        driver.find_element_by_xpath(ADPIntegrations.clicksearckfn()).click()
        print"Click on First Name"
        
        wait.until(EC.visibility_of_element_located((By.XPATH,ADPIntegrations.clickonname())))
        driver.find_element_by_xpath(ADPIntegrations.clickonname()).click()
        print"Click on Name arrow"
        
        wait.until(EC.visibility_of_element_located((By.XPATH,ADPIntegrations.personaldetails())))
        driver.find_element_by_xpath(ADPIntegrations.personaldetails()).click()
        print"Click on Personal profile link"
        
        wait.until(EC.visibility_of_element_located((By.XPATH,ADPIntegrations.clickoncontacts())))
        driver.find_element_by_xpath(ADPIntegrations.clickoncontacts()).click()
        print"click on Edit contacts"
        
        
        wait.until(EC.visibility_of_element_located((By.XPATH,ADPIntegrations.enteremail())))
        driver.find_element_by_xpath(ADPIntegrations.enteremail()).send_keys(Work_Email)
        print "Enter Email as "+Work_Email
        
        driver.find_element_by_xpath(ADPIntegrations.doneemail()).click()
        print"click on Done button"   
        
    def updating_the_employee_values_and_startmain(self):
            
            book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
            first_sheet = book.sheet_by_name('ADP User details')
            
            #Firstname 
            cell = first_sheet.cell(1,10)
            Employee_FirstName = cell.value
            FirstNameId = Employee_FirstName.split("_")
            emp = FirstNameId[0]+"_"
            ids = FirstNameId[1]
            empId = int(ids)+1
            FirstNameUpdated= emp+str(empId)
            
            #Lastname
            cell = first_sheet.cell(1,11)
            Employee_LastName = cell.value
            LastNameID = Employee_LastName.split("_")
            emp = LastNameID[0]+"_"
            ids = LastNameID[1]
            empId = int(ids)+1
            LastNameUpdated= emp+str(empId)
            
            #Work Email
            cell = first_sheet.cell(1,12)
            Work_Email = cell.value
            EmailId = Work_Email.split("_")
            emp = EmailId[0]+"_"
            ids = EmailId[1]
            remaining="_"+EmailId[2]
            empId = int(ids)+1
            EmailIdUpdated= emp+str(empId)+remaining
            
            
            #updating user values
            wb = load_workbook(os.path.join('Test_Data/TestData.xlsx'))
            #print (wb.sheetnames)
            
            sheet = wb['ADP User details']
            
            sheet.cell(row=2, column=9).value = FirstNameUpdated
            sheet.cell(row=2, column=10).value = LastNameUpdated
            sheet.cell(row=2, column=11).value =EmailIdUpdated
            
           wb.save(os.path.join('Test_Data/TestData.xlsx'))
            
            print "All User Data Updated in Excel"
            
            obj1= ADPIntegration()
            obj1.createuser_in_adp(FirstNameUpdated, LastNameUpdated, EmailIdUpdated)
                  
if __name__ == '__main__':
    
    
    ob= BaseTestClass()
    ob.UserLogin()
    obj1= ADPIntegration()
    #obj1.adpIntegration()
    obj1.updating_the_employee_values_and_startmain()
  
   
    print "Test executed successfully"
    
    
         
        
        
