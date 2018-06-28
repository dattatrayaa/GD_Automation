'''
Created on 10-Apr-2018

@author: Sheethu C
'''

from operator import contains
import os.path
import time
import traceback

from openpyxl.reader.excel import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import xlrd

from BaseTestClass import BaseTestClass
from BaseTestClass import driver
from BaseTestClass import projectPath

class RoleXpathElements():
    
    def checkboxUser(self):
        if(driver.find_element_by_xpath("html/body/div[1]/div/div[3]/div[2]/div/div[2]/div[2]/div[6]/div[2]/div[2]/div/div[2]/div[1]/div/label/input")).is_selected:
                                        
            print "User is selected"
        else:
            raise Exception
    def checkboxContent(self):
        if(driver.find_element_by_xpath("html/body/div[1]/div/div[3]/div[2]/div/div[2]/div[2]/div[6]/div[2]/div[2]/div/div[2]/div[2]/div/label/input")).is_selected:
            print "content is selected"
        else:
            raise Exception
    def checkboxRoles(self):
        if(driver.find_element_by_xpath("html/body/div[1]/div/div[3]/div[2]/div/div[2]/div[2]/div[6]/div[2]/div[2]/div/div[2]/div[3]/div/label/input")).is_selected:
            print "content is selected"
        else:
            raise Exception
    def checkboxTag(self):
        if(driver.find_element_by_xpath("html/body/div[1]/div/div[3]/div[2]/div/div[2]/div[2]/div[6]/div[2]/div[2]/div/div[2]/div[4]/div/label/input")).is_selected:
            print "content is selected"
        else:
            raise Exception
        
    def checkboxBranding(self): 
        if(driver.find_element_by_xpath("html/body/div[1]/div/div[3]/div[2]/div/div[2]/div[2]/div[6]/div[2]/div[2]/div/div[2]/div[5]/div/label/input")).is_selected:
            print "content is selected"
        else:
            raise Exception  
        
    def checkboxIntegration(self):
        if(driver.find_element_by_xpath("html/body/div[1]/div/div[3]/div[2]/div/div[2]/div[2]/div[6]/div[2]/div[2]/div/div[2]/div[6]/div/label/input")).is_selected:
            print "content is selected"
        else:
            raise Exception
    def checkCreate(self):
        Create=driver.find_element_by_xpath("html/body/div[1]/div/div[3]/div[2]/div/div[2]/div[2]/div[3]/div[1]/div").get_attribute("class")
                                           
        if(Create == "slide-toggle is-on"):
            print "Create Is selected"
        else:
            raise Exception
        
    def checkAssign(self):
        assign=driver.find_element_by_xpath("html/body/div[1]/div/div[3]/div[2]/div/div[2]/div[2]/div[4]/div[1]/div").get_attribute("class")
        if(assign == "slide-toggle is-on"):
            print "Create Is selected"
        else:
            raise Exception
        
    def checkReport(self):
        report=driver.find_element_by_xpath("html/body/div[1]/div/div[3]/div[2]/div/div[2]/div[2]/div[5]/div[1]/div").get_attribute("class")
        if(report == "slide-toggle is-on"):
            print "Create Is selected"
        else:
            raise Exception   
        
    def checkAdministrator(self):
        admini=driver.find_element_by_xpath("html/body/div[1]/div/div[3]/div[2]/div/div[2]/div[2]/div[6]/div[1]/div").get_attribute("class")
        if(admini == "slide-toggle is-on"):
            print "Create Is selected"
        else:
            raise Exception
    def searchRole(self,RoleName,UserName):   
        wait=WebDriverWait(driver, 80)
        driver.refresh()
        wait.until(EC.visibility_of_element_located((By.XPATH,"(//a[@href='/admin/users'])[1]")))
        driver.find_element_by_xpath("(//a[@href='/admin/users'])[1]").click()
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div/div/div[3]/div[1]/div/nav/div/div[2]/div[6]/div/ul/li[3]")))
        driver.find_element_by_xpath("html/body/div/div/div[3]/div[1]/div/nav/div/div[2]/div[6]/div/ul/li[3]").click()
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div/div/div[3]/div[2]/div/header/h1")))
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div/div/div[3]/div[2]/div/div/ul/li[2]/a")))
        driver.find_element_by_xpath("html/body/div/div/div[3]/div[2]/div/div/ul/li[2]/a").click()
        time.sleep(4)
        wait.until(EC.invisibility_of_element_located((By.XPATH,"//tbody/tr/td[1]/span/a[.='Creator']")))
        wait.until(EC.visibility_of_element_located((By.XPATH,"//tbody/tr[1]/td[4]/button")))
        time.sleep(4)
        wait.until(EC.visibility_of_element_located((By.ID,"search-roles")))
        driver.find_element_by_id("search-roles").send_keys(RoleName)
        wait.until(EC.visibility_of_element_located((By.XPATH,"//tbody/tr/td[.='"+RoleName+"']/span/a/../../../td[4]/button")))
        time.sleep(4)
        driver.find_element_by_xpath("//tbody/tr/td[.='"+RoleName+"']/span/a").click()
        time.sleep(4)
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div/div/div[3]/div[2]/div/header/h1/em/div/span/div[.='"+RoleName+"']")))
        time.sleep(4)
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div[1]/div/div[3]/div[2]/div/div[2]/ul/li[2]/a")))
        time.sleep(4)
        driver.find_element_by_xpath("html/body/div[1]/div/div[3]/div[2]/div/div[2]/ul/li[2]/a").click()
        time.sleep(4)
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div[1]/div/div[3]/div[2]/div/div[2]/div[2]/div[1]/div/div[2]/div/span[1]/div[1]")))
        if(driver.find_element_by_xpath("html/body/div[1]/div/div[3]/div[2]/div/div[2]/div[2]/div[1]/div/div[2]/div/span[1]/div[1]")).is_displayed():
            print "Grant this role to users or groups below is displaying"
        else:
            raise  Exception
        dd1=driver.find_element_by_xpath("html/body/div[1]/div/div[3]/div[2]/div/div[2]/div[2]/div[1]/div/div[2]/div/span[1]/div[1]")
        webdriver.ActionChains(driver).move_to_element(dd1).click().send_keys(UserName).perform()
        time.sleep(4)
        wait.until(EC.visibility_of_element_located((By.XPATH,"//div[@role='option']/span[contains(.,'"+UserName+"')]")))
        dd1=driver.find_element_by_xpath("//div[@role='option']/span[contains(.,'"+UserName+"')]")
        webdriver.ActionChains(driver).move_to_element(dd1).click().perform()
        time.sleep(4)
        wait.until(EC.visibility_of_element_located((By.XPATH,"//table/tbody/tr/td[.='"+UserName+"']")))
        time.sleep(4)
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div/div/div[3]/div[2]/div/header/div/button[.='Save']")))
        wait.until(EC.element_to_be_clickable((By.XPATH,"html/body/div/div/div[3]/div[2]/div/header/div/button[.='Save']")))
        elemt=driver.find_element_by_xpath("html/body/div/div/div[3]/div[2]/div/header/div/button[.='Save']")
        driver.execute_script('arguments[0].click()',elemt)
        time.sleep(4)
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div[4]/div/div/div[2]/div[2]/button[1]")))
        wait.until(EC.element_to_be_clickable((By.XPATH,"html/body/div[4]/div/div/div[2]/div[2]/button[1]")))
        elemm=driver.find_element_by_xpath("html/body/div[4]/div/div/div[2]/div[2]/button[1]")
        driver.execute_script("arguments[0].click()", elemm)
        time.sleep(4)
        wait.until(EC.element_to_be_clickable((By.XPATH,"html/body/div[1]/div/div[3]/div[2]/div/header/h1")))
    def CreateUserAssign(self):
        book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        first_sheet = book.sheet_by_name('UserAssignToRole')
        print("Fetching the First Name, LastName,Email,EmployeeId and password from Excel Sheet\n")
        # read a cell
        cell = first_sheet.cell(1,1)
        FirstName = cell.value
        print FirstName
        
        cell = first_sheet.cell(1,2)
        LastName = cell.value
        print LastName
        
        
        cell = first_sheet.cell(1,3)
        Email = cell.value
       
        
        cell = first_sheet.cell(1,4)
        EmployeeId = cell.value
        
        cell = first_sheet.cell(1,5)
        Password = cell.value
        
        cell = first_sheet.cell(1,6)
        NewPassword = cell.value
        
        wait=WebDriverWait(driver, 60)
        driver.refresh()
        wait.until(EC.visibility_of_element_located((By.XPATH,"(//a[@href='/admin/users'])[1]")))
        driver.find_element_by_xpath("(//a[@href='/admin/users'])[1]").click()
        time.sleep(4)
        print "Clicked on admin icon"
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div/div/div[3]/div[1]/div/nav/div/div[2]/div[6]/div/ul/li[1]")))
        driver.find_element_by_xpath("html/body/div/div/div[3]/div[1]/div/nav/div/div[2]/div[6]/div/ul/li[1]").click()
        time.sleep(4)
        print "clicked on Users"
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div/div/div[3]/div[2]/div/header/div/div/a")))
        wait.until(EC.element_to_be_clickable((By.XPATH,"html/body/div/div/div[3]/div[2]/div/header/div/div/a")))
        driver.find_element_by_xpath("html/body/div/div/div[3]/div[2]/div/header/div/div").click()
        time.sleep(4)
        print "Clicked on Add or editUser"
        driver.find_element_by_xpath("html/body/div/div/div[3]/div[2]/div/header/div").click()
        time.sleep(4)
        print "Clicked on Add An individual User"
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div/div/div[3]/div[2]/div/header")))
        
        print "Verifying Add user Page"
        if driver.find_element_by_xpath("html/body/div/div/div[3]/div[2]/div/header").is_displayed():
            print("Add user Page is displayed")
        else:
            print ""
            raise Exception
        print "Verifying First Name field"
        wait.until(EC.visibility_of_element_located((By.ID,"create-edit-user-search-firstName")))
        if driver.find_element_by_id("create-edit-user-search-firstName").is_displayed():
            print(" First Name field displayed")
        else:
            print ""
            raise Exception
        driver.find_element_by_id("create-edit-user-search-firstName").send_keys(FirstName)
        time.sleep(4)
        print "FirstName is Entered ::"+FirstName
        print "Last NAme verifying"
        if driver.find_element_by_id("create-edit-user-search-lastName").is_displayed():
            print(" Last Name field displayed")
        else:
            print ""
            raise Exception
        driver.find_element_by_id("create-edit-user-search-lastName").send_keys(LastName)
        time.sleep(4)
        print "Last Name is Entered ::"+LastName
        print "Email verifying"
        if driver.find_element_by_id("create-edit-user-search-username").is_displayed():
            print("Email field displayed")
        else:
            print ""
            raise Exception
        driver.find_element_by_id("create-edit-user-search-username").send_keys(Email)
        time.sleep(4)
        print "Email is Entered ::"+Email
        
        print "Employee ID verifying"
        if driver.find_element_by_id("create-edit-user-search-employeeId").is_displayed():
            print("Employee ID field displayed")
        else:
            print ""
            raise Exception
        driver.find_element_by_id("create-edit-user-search-employeeId").send_keys(EmployeeId)
        time.sleep(4)
        print "Employee ID  is Entered ::"+EmployeeId
        print "Inherited Role Verifying"
        if driver.find_element_by_xpath("html/body/div/div/div[3]/div[2]/div/div[2]/div/div/div[1]/div[6]/div").is_displayed():
            print("Inherited Role field displayed")
        else:
            print ""
            raise Exception
        print "Password Field is Verifying"
        if driver.find_element_by_id("create-edit-user-search-new-password").is_displayed():
            print("Password field displayed")
        else:
            print ""
            raise Exception
        driver.find_element_by_id("create-edit-user-search-new-password").send_keys(Password)
        time.sleep(4)
        print "Password is Entered ::"+Password
        
        print "Clicking on add button"
        wait.until(EC.visibility_of_element_located((By.XPATH,"//button[.='Add']")))
        wait.until(EC.element_to_be_clickable((By.XPATH,"//button[.='Add']")))
        driver.find_element_by_xpath("//button[.='Add']").click()
        time.sleep(4)
        print "Clicked on add button"
        wait.until(EC.visibility_of_element_located((By.XPATH,"//button[.='Save']")))
        driver.find_element_by_xpath("//button[.='Save']").click()
        time.sleep(4)
        print "Clicked on Save"
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div/div/div[3]/div[2]/div/header/h1")))
        print "Searching for the Created User"
        driver.find_element_by_id("search-users").send_keys(FirstName)
        
        wait.until(EC.visibility_of_element_located((By.XPATH,"//table/tbody/tr/td[.='"+Email+"']/../td[2]/a")))
        ele =driver.find_element_by_xpath("//table/tbody/tr/td[.='"+Email+"']/../td[2]/a").text
        print ele
        if(ele==FirstName):
            print("Created User Verified")
        else:
            print ""
            raise Exception  
        driver.find_element_by_xpath(".//*[@id='content']/div/div[1]/div/nav/div[2]/a/span[3]").click()
        time.sleep(4)
        print "Clicked on Account"
        driver.find_element_by_xpath(".//*[@id='content']/div/div[1]/div[2]/div[2]/a").click()
        time.sleep(4)
        print "Clicked on signOut Button"
        wait.until(EC.visibility_of_element_located((By.ID,"username")))
        
        print "Entering with the created User"
        driver.find_element_by_xpath(".//*[@id='username']").send_keys(Email)
        print "Entering Password"
        element = WebDriverWait(driver,80).until(EC.presence_of_element_located((By.ID, "password")))
        element.send_keys(Password)
        element.send_keys(Keys.TAB)
        print "Clicking on Sign_In button"
        driver.find_element_by_xpath("//*[@id='submitButton']").click()
        wait.until(EC.visibility_of_element_located((By.ID,"currentPassword")))
        driver.find_element_by_id("currentPassword").send_keys(Password)
        print "Current Password is entered :"+Password
        driver.find_element_by_id("newPassword").send_keys(NewPassword)
        print "New Password is entered :"+NewPassword
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div[2]/div/div/div[2]/div[1]/div[2]/button")))
        wait.until(EC.element_to_be_clickable((By.XPATH,"html/body/div[2]/div/div/div[2]/div[1]/div[2]/button")))
        driver.find_element_by_xpath("html/body/div[2]/div/div/div[2]/div[1]/div[2]/button").click()
        time.sleep(4)
        wait.until(EC.visibility_of_element_located((By.ID,"global-header-search")))
        print "Home Page is Loaded"
        expectedresult = "Home"
        expectedresult1="Library"
        actualresult = driver.find_element_by_xpath(".//*[@id='content']/div/div[1]/div/nav/div[1]/a[2]/span").text
        actualresult1 = driver.find_element_by_xpath(".//*[@id='content']/div/div[1]/div/nav/div[1]/a[3]/span").text
        if(expectedresult == actualresult):
            if(expectedresult1==actualresult1):
                print"User is able to login and Dashboard is displayed.."
        else:
            print"User not able to login.."
            raise Exception    
        print "Sign out "
        ele =driver.find_element_by_xpath(".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")
        driver.execute_script('arguments[0].click()',ele)
        elem=driver.find_element_by_xpath("html/body/div/div/div[1]/div[2]/div[2]/a")
        driver.execute_script('arguments[0].click()',elem)
        
        print "Again Entering The User"
        book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        first_sheet = book.sheet_by_name('Login_Credentials')
        cell = first_sheet.cell(1,1)
        url = cell.value
        cell = first_sheet.cell(3,1)
        username = cell.value
        print username
        cell = first_sheet.cell(3,2)
        password = cell.value
        print password  
        driver.get(url)
        element = WebDriverWait(driver,80).until(EC.presence_of_element_located((By.ID, "password")))
        if driver.title == "Grovo":
            print("Grovo Application URL Opened")
        else:
            raise Exception.message

        print "Grovo Sign-In page is displayed"
        
        print "Entering User name"
        driver.find_element_by_xpath(".//*[@id='username']").send_keys(username)
       
        print "Entering Password"
        element.send_keys(password)
        
        element.send_keys(Keys.TAB)
        print "Clicking on Sign_In button"
        driver.find_element_by_xpath("//*[@id='submitButton']").click()
        
        print "Successfully Loged Into Grovo Application"
        time.sleep(5) 
    def CreateUserReport(self):
        book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        first_sheet = book.sheet_by_name('UserAssignToRole')
        print("Fetching the First Name, LastName,Email,EmployeeId and password from Excel Sheet\n")
        # read a cell
        cell = first_sheet.cell(2,1)
        FirstName = cell.value
        print FirstName
        
        cell = first_sheet.cell(2,2)
        LastName = cell.value
        print LastName
        
        
        cell = first_sheet.cell(2,3)
        Email = cell.value
       
        
        cell = first_sheet.cell(2,4)
        EmployeeId = cell.value
        
        cell = first_sheet.cell(2,5)
        Password = cell.value
        
        cell = first_sheet.cell(2,6)
        NewPassword = cell.value
        
        wait=WebDriverWait(driver, 60)
        driver.refresh()
        wait.until(EC.visibility_of_element_located((By.XPATH,"(//a[@href='/admin/users'])[1]")))
        driver.find_element_by_xpath("(//a[@href='/admin/users'])[1]").click()
        time.sleep(4)
        print "Clicked on admin icon"
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div/div/div[3]/div[1]/div/nav/div/div[2]/div[6]/div/ul/li[1]")))
        driver.find_element_by_xpath("html/body/div/div/div[3]/div[1]/div/nav/div/div[2]/div[6]/div/ul/li[1]").click()
        time.sleep(4)
        print "clicked on Users"
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div/div/div[3]/div[2]/div/header/div/div/a")))
        wait.until(EC.element_to_be_clickable((By.XPATH,"html/body/div/div/div[3]/div[2]/div/header/div/div/a")))
        driver.find_element_by_xpath("html/body/div/div/div[3]/div[2]/div/header/div/div").click()
        time.sleep(4)
        print "Clicked on Add or editUser"
        driver.find_element_by_xpath("html/body/div/div/div[3]/div[2]/div/header/div").click()
        time.sleep(4)
        print "Clicked on Add An individual User"
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div/div/div[3]/div[2]/div/header")))
        
        print "Verifying Add user Page"
        if driver.find_element_by_xpath("html/body/div/div/div[3]/div[2]/div/header").is_displayed():
            print("Add user Page is displayed")
        else:
            print ""
            raise Exception
        print "Verifying First Name field"
        wait.until(EC.visibility_of_element_located((By.ID,"create-edit-user-search-firstName")))
        if driver.find_element_by_id("create-edit-user-search-firstName").is_displayed():
            print(" First Name field displayed")
        else:
            print ""
            raise Exception
        driver.find_element_by_id("create-edit-user-search-firstName").send_keys(FirstName)
        time.sleep(4)
        print "FirstName is Entered ::"+FirstName
        print "Last NAme verifying"
        if driver.find_element_by_id("create-edit-user-search-lastName").is_displayed():
            print(" Last Name field displayed")
        else:
            print ""
            raise Exception
        driver.find_element_by_id("create-edit-user-search-lastName").send_keys(LastName)
        time.sleep(4)
        print "Last Name is Entered ::"+LastName
        print "Email verifying"
        if driver.find_element_by_id("create-edit-user-search-username").is_displayed():
            print("Email field displayed")
        else:
            print ""
            raise Exception
        driver.find_element_by_id("create-edit-user-search-username").send_keys(Email)
        time.sleep(4)
        print "Email is Entered ::"+Email
        
        print "Employee ID verifying"
        if driver.find_element_by_id("create-edit-user-search-employeeId").is_displayed():
            print("Employee ID field displayed")
        else:
            print ""
            raise Exception
        driver.find_element_by_id("create-edit-user-search-employeeId").send_keys(EmployeeId)
        time.sleep(4)
        print "Employee ID  is Entered ::"+EmployeeId
        print "Inherited Role Verifying"
        if driver.find_element_by_xpath("html/body/div/div/div[3]/div[2]/div/div[2]/div/div/div[1]/div[6]/div").is_displayed():
            print("Inherited Role field displayed")
        else:
            print ""
            raise Exception
        print "Password Field is Verifying"
        if driver.find_element_by_id("create-edit-user-search-new-password").is_displayed():
            print("Password field displayed")
        else:
            print ""
            raise Exception
        driver.find_element_by_id("create-edit-user-search-new-password").send_keys(Password)
        time.sleep(4)
        print "Password is Entered ::"+Password
        
        print "Clicking on add button"
        wait.until(EC.visibility_of_element_located((By.XPATH,"//button[.='Add']")))
        wait.until(EC.element_to_be_clickable((By.XPATH,"//button[.='Add']")))
        driver.find_element_by_xpath("//button[.='Add']").click()
        time.sleep(4)
        print "Clicked on add button"
        wait.until(EC.visibility_of_element_located((By.XPATH,"//button[.='Save']")))
        driver.find_element_by_xpath("//button[.='Save']").click()
        time.sleep(4)
        print "Clicked on Save"
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div/div/div[3]/div[2]/div/header/h1")))
        print "Searching for the Created User"
        driver.find_element_by_id("search-users").send_keys(FirstName)
        
        wait.until(EC.visibility_of_element_located((By.XPATH,"//table/tbody/tr/td[.='"+Email+"']/../td[2]/a")))
        ele =driver.find_element_by_xpath("//table/tbody/tr/td[.='"+Email+"']/../td[2]/a").text
        print ele
        if(ele==FirstName):
            print("Created User Verified")
        else:
            print ""
            raise Exception  
        driver.find_element_by_xpath(".//*[@id='content']/div/div[1]/div/nav/div[2]/a/span[3]").click()
        time.sleep(4)
        print "Clicked on Account"
        driver.find_element_by_xpath(".//*[@id='content']/div/div[1]/div[2]/div[2]/a").click()
        time.sleep(4)
        print "Clicked on signOut Button"
        wait.until(EC.visibility_of_element_located((By.ID,"username")))
        
        print "Entering with the created User"
        driver.find_element_by_xpath(".//*[@id='username']").send_keys(Email)
        print "Entering Password"
        element = WebDriverWait(driver,80).until(EC.presence_of_element_located((By.ID, "password")))
        element.send_keys(Password)
        element.send_keys(Keys.TAB)
        print "Clicking on Sign_In button"
        driver.find_element_by_xpath("//*[@id='submitButton']").click()
        wait.until(EC.visibility_of_element_located((By.ID,"currentPassword")))
        driver.find_element_by_id("currentPassword").send_keys(Password)
        print "Current Password is entered :"+Password
        driver.find_element_by_id("newPassword").send_keys(NewPassword)
        print "New Password is entered :"+NewPassword
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div[2]/div/div/div[2]/div[1]/div[2]/button")))
        wait.until(EC.element_to_be_clickable((By.XPATH,"html/body/div[2]/div/div/div[2]/div[1]/div[2]/button")))
        driver.find_element_by_xpath("html/body/div[2]/div/div/div[2]/div[1]/div[2]/button").click()
        time.sleep(4)
        wait.until(EC.visibility_of_element_located((By.ID,"global-header-search")))
        print "Home Page is Loaded"
        expectedresult = "Home"
        expectedresult1="Library"
        actualresult = driver.find_element_by_xpath(".//*[@id='content']/div/div[1]/div/nav/div[1]/a[2]/span").text
        actualresult1 = driver.find_element_by_xpath(".//*[@id='content']/div/div[1]/div/nav/div[1]/a[3]/span").text
        if(expectedresult == actualresult):
            if(expectedresult1==actualresult1):
                print"User is able to login and Dashboard is displayed.."
        else:
            print"User not able to login.."
            raise Exception    
        print "Sign out "
        ele =driver.find_element_by_xpath(".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")
        driver.execute_script('arguments[0].click()',ele)
        elem=driver.find_element_by_xpath("html/body/div/div/div[1]/div[2]/div[2]/a")
        driver.execute_script('arguments[0].click()',elem)
        
        print "Again Entering The User"
        book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        first_sheet = book.sheet_by_name('Login_Credentials')
        cell = first_sheet.cell(1,1)
        url = cell.value
        cell = first_sheet.cell(3,1)
        username = cell.value
        print username
        cell = first_sheet.cell(3,2)
        password = cell.value
        print password  
        driver.get(url)
        element = WebDriverWait(driver,80).until(EC.presence_of_element_located((By.ID, "password")))
        if driver.title == "Grovo":
            print("Grovo Application URL Opened")
        else:
            raise Exception.message

        print "Grovo Sign-In page is displayed"
        
        print "Entering User name"
        driver.find_element_by_xpath(".//*[@id='username']").send_keys(username)
       
        print "Entering Password"
        element.send_keys(password)
        
        element.send_keys(Keys.TAB)
        print "Clicking on Sign_In button"
        driver.find_element_by_xpath("//*[@id='submitButton']").click()
        
        print "Successfully Loged Into Grovo Application"
        time.sleep(5)     
    def CreateUserLearn(self):
        book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        first_sheet = book.sheet_by_name('UserAssignToRole')
        print("Fetching the First Name, LastName,Email,EmployeeId and password from Excel Sheet\n")
        # read a cell
        cell = first_sheet.cell(3,1)
        FirstName = cell.value
        print FirstName
        
        cell = first_sheet.cell(3,2)
        LastName = cell.value
        print LastName
        
        
        cell = first_sheet.cell(3,3)
        Email = cell.value
       
        
        cell = first_sheet.cell(3,4)
        EmployeeId = cell.value
        
        cell = first_sheet.cell(3,5)
        Password = cell.value
        
        cell = first_sheet.cell(3,6)
        NewPassword = cell.value
        
        wait=WebDriverWait(driver, 60)
        driver.refresh()
        wait.until(EC.visibility_of_element_located((By.XPATH,"(//a[@href='/admin/users'])[1]")))
        driver.find_element_by_xpath("(//a[@href='/admin/users'])[1]").click()
        time.sleep(4)
        print "Clicked on admin icon"
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div/div/div[3]/div[1]/div/nav/div/div[2]/div[6]/div/ul/li[1]")))
        driver.find_element_by_xpath("html/body/div/div/div[3]/div[1]/div/nav/div/div[2]/div[6]/div/ul/li[1]").click()
        time.sleep(4)
        print "clicked on Users"
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div/div/div[3]/div[2]/div/header/div/div/a")))
        wait.until(EC.element_to_be_clickable((By.XPATH,"html/body/div/div/div[3]/div[2]/div/header/div/div/a")))
        driver.find_element_by_xpath("html/body/div/div/div[3]/div[2]/div/header/div/div").click()
        time.sleep(4)
        print "Clicked on Add or editUser"
        driver.find_element_by_xpath("html/body/div/div/div[3]/div[2]/div/header/div").click()
        time.sleep(4)
        print "Clicked on Add An individual User"
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div/div/div[3]/div[2]/div/header")))
        
        print "Verifying Add user Page"
        if driver.find_element_by_xpath("html/body/div/div/div[3]/div[2]/div/header").is_displayed():
            print("Add user Page is displayed")
        else:
            print ""
            raise Exception
        print "Verifying First Name field"
        wait.until(EC.visibility_of_element_located((By.ID,"create-edit-user-search-firstName")))
        if driver.find_element_by_id("create-edit-user-search-firstName").is_displayed():
            print(" First Name field displayed")
        else:
            print ""
            raise Exception
        driver.find_element_by_id("create-edit-user-search-firstName").send_keys(FirstName)
        time.sleep(4)
        print "FirstName is Entered ::"+FirstName
        print "Last NAme verifying"
        if driver.find_element_by_id("create-edit-user-search-lastName").is_displayed():
            print(" Last Name field displayed")
        else:
            print ""
            raise Exception
        driver.find_element_by_id("create-edit-user-search-lastName").send_keys(LastName)
        time.sleep(4)
        print "Last Name is Entered ::"+LastName
        print "Email verifying"
        if driver.find_element_by_id("create-edit-user-search-username").is_displayed():
            print("Email field displayed")
        else:
            print ""
            raise Exception
        driver.find_element_by_id("create-edit-user-search-username").send_keys(Email)
        time.sleep(4)
        print "Email is Entered ::"+Email
        
        print "Employee ID verifying"
        if driver.find_element_by_id("create-edit-user-search-employeeId").is_displayed():
            print("Employee ID field displayed")
        else:
            print ""
            raise Exception
        driver.find_element_by_id("create-edit-user-search-employeeId").send_keys(EmployeeId)
        time.sleep(4)
        print "Employee ID  is Entered ::"+EmployeeId
        print "Inherited Role Verifying"
        if driver.find_element_by_xpath("html/body/div/div/div[3]/div[2]/div/div[2]/div/div/div[1]/div[6]/div").is_displayed():
            print("Inherited Role field displayed")
        else:
            print ""
            raise Exception
        print "Password Field is Verifying"
        if driver.find_element_by_id("create-edit-user-search-new-password").is_displayed():
            print("Password field displayed")
        else:
            print ""
            raise Exception
        driver.find_element_by_id("create-edit-user-search-new-password").send_keys(Password)
        time.sleep(4)
        print "Password is Entered ::"+Password
        
        print "Clicking on add button"
        wait.until(EC.visibility_of_element_located((By.XPATH,"//button[.='Add']")))
        wait.until(EC.element_to_be_clickable((By.XPATH,"//button[.='Add']")))
        driver.find_element_by_xpath("//button[.='Add']").click()
        time.sleep(4)
        print "Clicked on add button"
        wait.until(EC.visibility_of_element_located((By.XPATH,"//button[.='Save']")))
        driver.find_element_by_xpath("//button[.='Save']").click()
        time.sleep(4)
        print "Clicked on Save"
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div/div/div[3]/div[2]/div/header/h1")))
        print "Searching for the Created User"
        driver.find_element_by_id("search-users").send_keys(FirstName)
        
        wait.until(EC.visibility_of_element_located((By.XPATH,"//table/tbody/tr/td[.='"+Email+"']/../td[2]/a")))
        ele =driver.find_element_by_xpath("//table/tbody/tr/td[.='"+Email+"']/../td[2]/a").text
        print ele
        if(ele==FirstName):
            print("Created User Verified")
        else:
            print ""
            raise Exception  
        driver.find_element_by_xpath(".//*[@id='content']/div/div[1]/div/nav/div[2]/a/span[3]").click()
        time.sleep(4)
        print "Clicked on Account"
        driver.find_element_by_xpath(".//*[@id='content']/div/div[1]/div[2]/div[2]/a").click()
        time.sleep(4)
        print "Clicked on signOut Button"
        wait.until(EC.visibility_of_element_located((By.ID,"username")))
        
        print "Entering with the created User"
        driver.find_element_by_xpath(".//*[@id='username']").send_keys(Email)
        print "Entering Password"
        element = WebDriverWait(driver,80).until(EC.presence_of_element_located((By.ID, "password")))
        element.send_keys(Password)
        element.send_keys(Keys.TAB)
        print "Clicking on Sign_In button"
        driver.find_element_by_xpath("//*[@id='submitButton']").click()
        wait.until(EC.visibility_of_element_located((By.ID,"currentPassword")))
        driver.find_element_by_id("currentPassword").send_keys(Password)
        print "Current Password is entered :"+Password
        driver.find_element_by_id("newPassword").send_keys(NewPassword)
        print "New Password is entered :"+NewPassword
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div[2]/div/div/div[2]/div[1]/div[2]/button")))
        wait.until(EC.element_to_be_clickable((By.XPATH,"html/body/div[2]/div/div/div[2]/div[1]/div[2]/button")))
        driver.find_element_by_xpath("html/body/div[2]/div/div/div[2]/div[1]/div[2]/button").click()
        time.sleep(4)
        wait.until(EC.visibility_of_element_located((By.ID,"global-header-search")))
        print "Home Page is Loaded"
        expectedresult = "Home"
        expectedresult1="Library"
        actualresult = driver.find_element_by_xpath(".//*[@id='content']/div/div[1]/div/nav/div[1]/a[2]/span").text
        actualresult1 = driver.find_element_by_xpath(".//*[@id='content']/div/div[1]/div/nav/div[1]/a[3]/span").text
        if(expectedresult == actualresult):
            if(expectedresult1==actualresult1):
                print"User is able to login and Dashboard is displayed.."
        else:
            print"User not able to login.."
            raise Exception    
        print "Sign out "
        ele =driver.find_element_by_xpath(".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")
        driver.execute_script('arguments[0].click()',ele)
        elem=driver.find_element_by_xpath("html/body/div/div/div[1]/div[2]/div[2]/a")
        driver.execute_script('arguments[0].click()',elem)
        
        print "Again Entering The User"
        book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        first_sheet = book.sheet_by_name('Login_Credentials')
        cell = first_sheet.cell(1,1)
        url = cell.value
        cell = first_sheet.cell(3,1)
        username = cell.value
        print username
        cell = first_sheet.cell(3,2)
        password = cell.value
        print password  
        driver.get(url)
        element = WebDriverWait(driver,80).until(EC.presence_of_element_located((By.ID, "password")))
        if driver.title == "Grovo":
            print("Grovo Application URL Opened")
        else:
            raise Exception.message

        print "Grovo Sign-In page is displayed"
        
        print "Entering User name"
        driver.find_element_by_xpath(".//*[@id='username']").send_keys(username)
       
        print "Entering Password"
        element.send_keys(password)
        
        element.send_keys(Keys.TAB)
        print "Clicking on Sign_In button"
        driver.find_element_by_xpath("//*[@id='submitButton']").click()
        
        print "Successfully Loged Into Grovo Application"
        time.sleep(5)   
    def CreateUsercreator(self):
        book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        first_sheet = book.sheet_by_name('UserAssignToRole')
        print("Fetching the First Name, LastName,Email,EmployeeId and password from Excel Sheet\n")
        # read a cell
        cell = first_sheet.cell(4,1)
        FirstName = cell.value
        print FirstName
        
        cell = first_sheet.cell(4,2)
        LastName = cell.value
        print LastName
        
        
        cell = first_sheet.cell(4,3)
        Email = cell.value
       
        
        cell = first_sheet.cell(4,4)
        EmployeeId = cell.value
        
        cell = first_sheet.cell(4,5)
        Password = cell.value
        
        cell = first_sheet.cell(4,6)
        NewPassword = cell.value
        
        wait=WebDriverWait(driver, 60)
        driver.refresh()
        wait.until(EC.visibility_of_element_located((By.XPATH,"(//a[@href='/admin/users'])[1]")))
        driver.find_element_by_xpath("(//a[@href='/admin/users'])[1]").click()
        time.sleep(4)
        print "Clicked on admin icon"
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div/div/div[3]/div[1]/div/nav/div/div[2]/div[6]/div/ul/li[1]")))
        driver.find_element_by_xpath("html/body/div/div/div[3]/div[1]/div/nav/div/div[2]/div[6]/div/ul/li[1]").click()
        time.sleep(4)
        print "clicked on Users"
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div/div/div[3]/div[2]/div/header/div/div/a")))
        wait.until(EC.element_to_be_clickable((By.XPATH,"html/body/div/div/div[3]/div[2]/div/header/div/div/a")))
        driver.find_element_by_xpath("html/body/div/div/div[3]/div[2]/div/header/div/div").click()
        time.sleep(4)
        print "Clicked on Add or editUser"
        driver.find_element_by_xpath("html/body/div/div/div[3]/div[2]/div/header/div").click()
        time.sleep(4)
        print "Clicked on Add An individual User"
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div/div/div[3]/div[2]/div/header")))
        
        print "Verifying Add user Page"
        if driver.find_element_by_xpath("html/body/div/div/div[3]/div[2]/div/header").is_displayed():
            print("Add user Page is displayed")
        else:
            print ""
            raise Exception
        print "Verifying First Name field"
        wait.until(EC.visibility_of_element_located((By.ID,"create-edit-user-search-firstName")))
        if driver.find_element_by_id("create-edit-user-search-firstName").is_displayed():
            print(" First Name field displayed")
        else:
            print ""
            raise Exception
        driver.find_element_by_id("create-edit-user-search-firstName").send_keys(FirstName)
        time.sleep(4)
        print "FirstName is Entered ::"+FirstName
        print "Last NAme verifying"
        if driver.find_element_by_id("create-edit-user-search-lastName").is_displayed():
            print(" Last Name field displayed")
        else:
            print ""
            raise Exception
        driver.find_element_by_id("create-edit-user-search-lastName").send_keys(LastName)
        time.sleep(4)
        print "Last Name is Entered ::"+LastName
        print "Email verifying"
        if driver.find_element_by_id("create-edit-user-search-username").is_displayed():
            print("Email field displayed")
        else:
            print ""
            raise Exception
        driver.find_element_by_id("create-edit-user-search-username").send_keys(Email)
        time.sleep(4)
        print "Email is Entered ::"+Email
        
        print "Employee ID verifying"
        if driver.find_element_by_id("create-edit-user-search-employeeId").is_displayed():
            print("Employee ID field displayed")
        else:
            print ""
            raise Exception
        driver.find_element_by_id("create-edit-user-search-employeeId").send_keys(EmployeeId)
        time.sleep(4)
        print "Employee ID  is Entered ::"+EmployeeId
        print "Inherited Role Verifying"
        if driver.find_element_by_xpath("html/body/div/div/div[3]/div[2]/div/div[2]/div/div/div[1]/div[6]/div").is_displayed():
            print("Inherited Role field displayed")
        else:
            print ""
            raise Exception
        print "Password Field is Verifying"
        if driver.find_element_by_id("create-edit-user-search-new-password").is_displayed():
            print("Password field displayed")
        else:
            print ""
            raise Exception
        driver.find_element_by_id("create-edit-user-search-new-password").send_keys(Password)
        time.sleep(4)
        print "Password is Entered ::"+Password
        
        print "Clicking on add button"
        wait.until(EC.visibility_of_element_located((By.XPATH,"//button[.='Add']")))
        wait.until(EC.element_to_be_clickable((By.XPATH,"//button[.='Add']")))
        driver.find_element_by_xpath("//button[.='Add']").click()
        time.sleep(4)
        print "Clicked on add button"
        wait.until(EC.visibility_of_element_located((By.XPATH,"//button[.='Save']")))
        driver.find_element_by_xpath("//button[.='Save']").click()
        time.sleep(4)
        print "Clicked on Save"
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div/div/div[3]/div[2]/div/header/h1")))
        print "Searching for the Created User"
        driver.find_element_by_id("search-users").send_keys(FirstName)
        
        wait.until(EC.visibility_of_element_located((By.XPATH,"//table/tbody/tr/td[.='"+Email+"']/../td[2]/a")))
        ele =driver.find_element_by_xpath("//table/tbody/tr/td[.='"+Email+"']/../td[2]/a").text
        print ele
        if(ele==FirstName):
            print("Created User Verified")
        else:
            print ""
            raise Exception  
        driver.find_element_by_xpath(".//*[@id='content']/div/div[1]/div/nav/div[2]/a/span[3]").click()
        time.sleep(4)
        print "Clicked on Account"
        driver.find_element_by_xpath(".//*[@id='content']/div/div[1]/div[2]/div[2]/a").click()
        time.sleep(4)
        print "Clicked on signOut Button"
        wait.until(EC.visibility_of_element_located((By.ID,"username")))
        
        print "Entering with the created User"
        driver.find_element_by_xpath(".//*[@id='username']").send_keys(Email)
        print "Entering Password"
        element = WebDriverWait(driver,80).until(EC.presence_of_element_located((By.ID, "password")))
        element.send_keys(Password)
        element.send_keys(Keys.TAB)
        print "Clicking on Sign_In button"
        driver.find_element_by_xpath("//*[@id='submitButton']").click()
        wait.until(EC.visibility_of_element_located((By.ID,"currentPassword")))
        driver.find_element_by_id("currentPassword").send_keys(Password)
        print "Current Password is entered :"+Password
        driver.find_element_by_id("newPassword").send_keys(NewPassword)
        print "New Password is entered :"+NewPassword
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div[2]/div/div/div[2]/div[1]/div[2]/button")))
        wait.until(EC.element_to_be_clickable((By.XPATH,"html/body/div[2]/div/div/div[2]/div[1]/div[2]/button")))
        driver.find_element_by_xpath("html/body/div[2]/div/div/div[2]/div[1]/div[2]/button").click()
        time.sleep(4)
        wait.until(EC.visibility_of_element_located((By.ID,"global-header-search")))
        print "Home Page is Loaded"
        expectedresult = "Home"
        expectedresult1="Library"
        actualresult = driver.find_element_by_xpath(".//*[@id='content']/div/div[1]/div/nav/div[1]/a[2]/span").text
        actualresult1 = driver.find_element_by_xpath(".//*[@id='content']/div/div[1]/div/nav/div[1]/a[3]/span").text
        if(expectedresult == actualresult):
            if(expectedresult1==actualresult1):
                print"User is able to login and Dashboard is displayed.."
        else:
            print"User not able to login.."
            raise Exception    
        print "Sign out "
        ele =driver.find_element_by_xpath(".//*[@id='content']/div/div[1]/div[1]/nav/div[2]/a/span[3]")
        driver.execute_script('arguments[0].click()',ele)
        elem=driver.find_element_by_xpath("html/body/div/div/div[1]/div[2]/div[2]/a")
        driver.execute_script('arguments[0].click()',elem)
        
        print "Again Entering The User"
        book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        first_sheet = book.sheet_by_name('Login_Credentials')
        cell = first_sheet.cell(1,1)
        url = cell.value
        cell = first_sheet.cell(3,1)
        username = cell.value
        print username
        cell = first_sheet.cell(3,2)
        password = cell.value
        print password  
        driver.get(url)
        element = WebDriverWait(driver,80).until(EC.presence_of_element_located((By.ID, "password")))
        if driver.title == "Grovo":
            print("Grovo Application URL Opened")
        else:
            raise Exception.message

        print "Grovo Sign-In page is displayed"
        
        print "Entering User name"
        driver.find_element_by_xpath(".//*[@id='username']").send_keys(username)
       
        print "Entering Password"
        element.send_keys(password)
        
        element.send_keys(Keys.TAB)
        print "Clicking on Sign_In button"
        driver.find_element_by_xpath("//*[@id='submitButton']").click()
        
        print "Successfully Loged Into Grovo Application"
        time.sleep(5)
    def updateUserAssign(self):
        
        print "Updating Details"
        book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        second_sheet = book.sheet_by_name('UserAssignToRole')
        cellf =second_sheet.cell(1,1)
        fn = cellf.value
        cell2 = second_sheet.cell(1,4)
        LId = cell2.value
        cell1 = second_sheet.cell(1,3)
        LEmail = cell1.value
        
        
        cell = second_sheet.cell(1,1)
        FirstName = cell.value
        FirstNameId = FirstName.split("_")
        emp = FirstNameId[0]+"_"
        ids = FirstNameId[1]
        empId1 = int(ids)+1
        FirstNameUpdated= emp+str(empId1)
        
        
        cell2 = second_sheet.cell(1,4)
        LearnId = cell2.value
        LearnsId = LearnId.split("#")
        emp = LearnsId[0]+"#"
        ids = LearnsId[1]
        empId = int(ids)+1
        LearnerEmpId = emp+str(empId)
        
        
        cell1 = second_sheet.cell(1,3)
        LearnEmailId = cell1.value
        LearnerEmail = LearnEmailId.split("@")
        spitValue = LearnerEmail[0][:4]
        email = spitValue+str(empId)
        id = "@"+LearnerEmail[1]
        LearnerEmailId = email+id
       
       
        
       
        
        wb = load_workbook(os.path.join('Test_Data/TestData.xlsx'))
        
        #print (wb.sheetnames)
    
        sheet = wb['UserAssignToRole']
        
        sheet.cell(row=2, column=2).value = FirstNameUpdated    
        sheet.cell(row=2, column=4).value = LearnerEmailId
        sheet.cell(row=2, column=5).value = LearnerEmpId
        
       
        wb.save(os.path.join('Test_Data/TestData.xlsx'))
        
       
    def updateUserLearn(self):
        
        print "Updating Details"
        book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        second_sheet = book.sheet_by_name('UserAssignToRole')
       
        cellf =second_sheet.cell(3,1)
        fn = cellf.value
        cell6 = second_sheet.cell(3,4)
        MId = cell6.value
        cell5 = second_sheet.cell(3,3)
        MEmail = cell5.value
       
        
        cell = second_sheet.cell(3,1)
        FirstName = cell.value
        FirstNameId = FirstName.split("_")
        emp = FirstNameId[0]+"_"
        ids = FirstNameId[1]
        empId1 = int(ids)+1
        FirstNameUser3= emp+str(empId1)
        
        cell6 = second_sheet.cell(3,4)
        MasterAdminId = cell6.value
        MasterAdminsId = MasterAdminId.split("#")
        emp2 = MasterAdminsId[0]+"#"
        ids2 = MasterAdminsId[1]
        empId2 = int(ids2)+1
        MasterAdminsEmpId = emp2+str(empId2)
        
        
        cell5 = second_sheet.cell(3,3)
        MasterEmail = cell5.value
        MasterAdminEmail = MasterEmail.split("@")
        spitValue2 = MasterAdminEmail[0][:4]
        email2 = spitValue2+str(empId2)
        id2 = "@"+MasterAdminEmail[1]
        MasterAdminEmailId = email2+id2
        
        
        
        
        
       
        
        wb = load_workbook(os.path.abspath(os.path.join(os.path.dirname(__file__),'TestData.xlsx')))
        
        #print (wb.sheetnames)
    
        sheet = wb['UserAssignToRole']
        
        
        
        sheet.cell(row=4, column=2).value = FirstNameUser3    
        sheet.cell(row=4, column=4).value = MasterAdminEmailId
        sheet.cell(row=4, column=5).value = MasterAdminsEmpId
        
        
        
        wb.save(os.path.abspath(os.path.join(os.path.dirname(__file__),'TestData.xlsx')))
       
        
        
        
    def updateUserReport(self):
        
       
        print "Updating Details"
        book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
        second_sheet = book.sheet_by_name('UserAssignToRole')
        
        
        cellf =second_sheet.cell(2,1)
        fn = cellf.value
        cell8 = second_sheet.cell(2,4)
        LAId = cell8.value
        cell7 = second_sheet.cell(2,3)
        LAEmail = cell7.value
        
        
        cell = second_sheet.cell(2,1)
        FirstName = cell.value
        FirstNameId = FirstName.split("_")
        emp = FirstNameId[0]+"_"
        ids = FirstNameId[1]
        empId1 = int(ids)+1
        FirstNameUser4= emp+str(empId1)
        
        
        cell8 = second_sheet.cell(2,4)
        LearningAdminId = cell8.value
        LearningAdminsId = LearningAdminId.split("#")
        emp3 = LearningAdminsId[0]+"#"
        ids3 = LearningAdminsId[1]
        empId3 = int(ids3)+1
        LearningAdminsEmpId = emp3+str(empId3)
        print LearningAdminsEmpId
        
        cell7 = second_sheet.cell(2,3)
        LearningEmail = cell7.value
        LearningAdminEmail = LearningEmail.split("@")
        spitValue3 = LearningAdminEmail[0][:4]
        email3 = spitValue3+str(empId3)
        id3 = "@"+LearningAdminEmail[1]
        LearningAdminEmailId = email3+id3
        print LearningAdminEmailId
       
        
        wb = load_workbook(os.path.join('Test_Data/TestData.xlsx'))
        
        #print (wb.sheetnames)
    
        sheet = wb['UserAssignToRole']
       
        sheet.cell(row=3, column=2).value = FirstNameUser4    
        sheet.cell(row=3, column=4).value = LearningAdminEmailId
        sheet.cell(row=3, column=5).value = LearningAdminsEmpId
        
        
        wb.save(os.path.join('Test_Data/TestData.xlsx'))
       
    def updateUserCreator(self):
        
            print "Updating Details"
            book=xlrd.open_workbook(os.path.join('Test_Data/TestData.xlsx'))
            second_sheet = book.sheet_by_name('UserAssignToRole')
            
            cell4 = second_sheet.cell(4,4)
            CId = cell4.value
            cell3 = second_sheet.cell(4,3)
            CEmail = cell3.value
            
            
            
            cell = second_sheet.cell(4,1)
            FirstName = cell.value
            FirstNameId = FirstName.split("_")
            emp = FirstNameId[0]+"_"
            ids = FirstNameId[1]
            empId1 = int(ids)+1
            updatedUser2= emp+str(empId1)
           
            
            cell4 = second_sheet.cell(4,4)
            CreatorId = cell4.value
            CreatorsId = CreatorId.split("#")
            emp1 = CreatorsId[0]+"#"
            ids1 = CreatorsId[1]
            empId1 = int(ids1)+1
            CreatorsEmpId = emp1+str(empId1)
            print CreatorsEmpId
            
            
            cell3 = second_sheet.cell(4,3)
            CreateEmailId = cell3.value
            CreatorEmail = CreateEmailId.split("@")
            spitValue1 = CreatorEmail[0][:4]
            email1 = spitValue1+str(empId1)
            id1 = "@"+CreatorEmail[1]
            CreatorEmailId = email1+id1
            print CreatorEmailId
            
           
            
            wb = load_workbook(os.path.join('Test_Data/TestData.xlsx'))
            
            #print (wb.sheetnames)
        
            sheet = wb['UserAssignToRole']
            sheet.cell(row=5, column=2).value = updatedUser2    
            sheet.cell(row=5, column=4).value = CreatorEmailId
            sheet.cell(row=5, column=5).value = CreatorsEmpId
            
            
         
            
           
            wb.save(os.path.join('Test_Data/TestData.xlsx'))
            
       
    def deactivateUser(self,UserName):  
        wait= WebDriverWait(driver,80)
        driver.refresh()
        wait.until(EC.visibility_of_element_located((By.XPATH,"(//a[@href='/admin/users'])[1]")))
        driver.find_element_by_xpath("(//a[@href='/admin/users'])[1]").click()
        time.sleep(4)
        print "Clicked on admin icon"
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div/div/div[3]/div[1]/div/nav/div/div[2]/div[6]/div/ul/li[1]")))
        driver.find_element_by_xpath("html/body/div/div/div[3]/div[1]/div/nav/div/div[2]/div[6]/div/ul/li[1]").click()
        time.sleep(4)
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div/div/div[3]/div[2]/div/header/h1")))
        wait.until(EC.visibility_of_element_located((By.ID,"search-users")))
        driver.find_element_by_id("search-users").send_keys(UserName)
        time.sleep(4)
        wait.until(EC.visibility_of_element_located((By.XPATH,"//table/tbody/tr/td[.='"+UserName+"']/../td[.='Deactivate']/button")))
        driver.find_element_by_xpath("//table/tbody/tr/td[.='"+UserName+"']/../td[.='Deactivate']/button").click()
        time.sleep(4)
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div[4]/div/div/div[2]/div[2]/button[1]")))
        wait.until(EC.element_to_be_clickable((By.XPATH,"html/body/div[4]/div/div/div[2]/div[2]/button[1]")))
        driver.find_element_by_xpath("html/body/div[4]/div/div/div[2]/div[2]/button[1]").click()     
        time.sleep(4)    
          
            
    def createClick(self):    
        wait =WebDriverWait(driver,80) 
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div[1]/div/div[3]/div[2]/div/div[2]/div[2]/div[3]/div[1]/div")))
        driver.find_element_by_xpath("html/body/div[1]/div/div[3]/div[2]/div/div[2]/div[2]/div[3]/div[1]/div").click()
        time.sleep(3)  
    def assignClick(self):  
        wait =WebDriverWait(driver,80)   
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div[1]/div/div[3]/div[2]/div/div[2]/div[2]/div[4]/div[1]/div")))
        elemm=driver.find_element_by_xpath("html/body/div[1]/div/div[3]/div[2]/div/div[2]/div[2]/div[4]/div[1]/div")
        webdriver.ActionChains(driver).move_to_element(elemm).perform()
        driver.execute_script("arguments[0].click()", elemm)    
    def reportClick(self):     
        wait =WebDriverWait(driver,80)   
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div[1]/div/div[3]/div[2]/div/div[2]/div[2]/div[5]/div[1]/div")))
        wait.until(EC.element_to_be_clickable((By.XPATH,"html/body/div[1]/div/div[3]/div[2]/div/div[2]/div[2]/div[5]/div[1]/div")))
        elemm=driver.find_element_by_xpath("html/body/div[1]/div/div[3]/div[2]/div/div[2]/div[2]/div[5]/div[1]/div")
        webdriver.ActionChains(driver).move_to_element(elemm).perform()
        driver.execute_script("arguments[0].click()", elemm)
        time.sleep(3)
    def adminClick(self):   
        wait =WebDriverWait(driver,80)        
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div[1]/div/div[3]/div[2]/div/div[2]/div[2]/div[6]/div[1]/div")))
        elemm=driver.find_element_by_xpath("html/body/div[1]/div/div[3]/div[2]/div/div[2]/div[2]/div[6]/div[1]/div")
        webdriver.ActionChains(driver).move_to_element(elemm).perform()
        driver.execute_script("arguments[0].click()", elemm)   
    def userClick(self):    
        wait =WebDriverWait(driver,80) 
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div[1]/div/div[3]/div[2]/div/div[2]/div[2]/div[6]/div[2]/div[2]/div/div[2]/div[1]/div/label/span['Users & groups']")))
        elemm=driver.find_element_by_xpath("html/body/div[1]/div/div[3]/div[2]/div/div[2]/div[2]/div[6]/div[2]/div[2]/div/div[2]/div[1]/div/label/span[2]")
        webdriver.ActionChains(driver).move_to_element(elemm).perform()
        driver.execute_script("arguments[0].click()", elemm)
        time.sleep(3)   
    def contentClick(self):     
        wait =WebDriverWait(driver,80)
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div[1]/div/div[3]/div[2]/div/div[2]/div[2]/div[6]/div[2]/div[2]/div/div[2]/div[2]/div/label/span['Content manager']")))
        wait.until(EC.element_to_be_clickable((By.XPATH,"html/body/div[1]/div/div[3]/div[2]/div/div[2]/div[2]/div[6]/div[2]/div[2]/div/div[2]/div[2]/div/label/span[2]")))
        elemm=driver.find_element_by_xpath("html/body/div[1]/div/div[3]/div[2]/div/div[2]/div[2]/div[6]/div[2]/div[2]/div/div[2]/div[2]/div/label/span[2]")
        webdriver.ActionChains(driver).move_to_element(elemm).perform()
        driver.execute_script("arguments[0].click()", elemm)
        time.sleep(3)    
    def roleClick(self):  
        wait =WebDriverWait(driver,80)       
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div[1]/div/div[3]/div[2]/div/div[2]/div[2]/div[6]/div[2]/div[2]/div/div[2]/div[3]/div/label/span['Roles & access']")))
        wait.until(EC.element_to_be_clickable((By.XPATH,"html/body/div[1]/div/div[3]/div[2]/div/div[2]/div[2]/div[6]/div[2]/div[2]/div/div[2]/div[3]/div/label/span[2]")))
        elemm=driver.find_element_by_xpath("html/body/div[1]/div/div[3]/div[2]/div/div[2]/div[2]/div[6]/div[2]/div[2]/div/div[2]/div[3]/div/label/span[2]")
        webdriver.ActionChains(driver).move_to_element(elemm).perform()
        driver.execute_script("arguments[0].click()", elemm)
        time.sleep(3)   
    def tagClick(self):  
        wait =WebDriverWait(driver,80)       
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div[1]/div/div[3]/div[2]/div/div[2]/div[2]/div[6]/div[2]/div[2]/div/div[2]/div[4]/div/label/span['Tag manager']")))
        wait.until(EC.element_to_be_clickable((By.XPATH,"html/body/div[1]/div/div[3]/div[2]/div/div[2]/div[2]/div[6]/div[2]/div[2]/div/div[2]/div[4]/div/label/span[2]")))
        elemm=driver.find_element_by_xpath("html/body/div[1]/div/div[3]/div[2]/div/div[2]/div[2]/div[6]/div[2]/div[2]/div/div[2]/div[4]/div/label/span[2]")
        webdriver.ActionChains(driver).move_to_element(elemm).perform()
        driver.execute_script("arguments[0].click()", elemm)
        time.sleep(3) 
    def brandingClick(self):   
        wait =WebDriverWait(driver,80) 
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div[1]/div/div[3]/div[2]/div/div[2]/div[2]/div[6]/div[2]/div[2]/div/div[2]/div[5]/div/label/span['Branding']")))
        wait.until(EC.element_to_be_clickable((By.XPATH,"html/body/div[1]/div/div[3]/div[2]/div/div[2]/div[2]/div[6]/div[2]/div[2]/div/div[2]/div[5]/div/label/span[2]")))
        elemm=driver.find_element_by_xpath("html/body/div[1]/div/div[3]/div[2]/div/div[2]/div[2]/div[6]/div[2]/div[2]/div/div[2]/div[5]/div/label/span[2]")
        webdriver.ActionChains(driver).move_to_element(elemm).perform()
        driver.execute_script("arguments[0].click()", elemm)
        time.sleep(3)   
    def integrateClick(self):   
        wait =WebDriverWait(driver,80)
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div[1]/div/div[3]/div[2]/div/div[2]/div[2]/div[6]/div[2]/div[2]/div/div[2]/div[6]/div/label/span[1]")))
        wait.until(EC.element_to_be_clickable((By.XPATH,"html/body/div[1]/div/div[3]/div[2]/div/div[2]/div[2]/div[6]/div[2]/div[2]/div/div[2]/div[6]/div/label/span[2]")))
        elemm=driver.find_element_by_xpath("html/body/div[1]/div/div[3]/div[2]/div/div[2]/div[2]/div[6]/div[2]/div[2]/div/div[2]/div[6]/div/label/span[2]")
        webdriver.ActionChains(driver).move_to_element(elemm).perform()
        driver.execute_script("arguments[0].click()", elemm)
        time.sleep(3)    
       
    def adminSideMenu(self):
        time.sleep(2)
        return "(//a[@href='/admin/users'])[1]"   
        
    def roleSideMenu(self):
        time.sleep(2)
        return "html/body/div/div/div[3]/div[1]/div/nav/div/div[2]/div[6]/div/ul/li[3]"     
        
    def createRole(self):
        time.sleep(2)
        return "html/body/div[1]/div/div[3]/div[2]/div/header/div/div/button"   
    
    def roleCreation(self,RoleName,Description):
        wait =WebDriverWait(driver,80)
        wait.until(EC.visibility_of_element_located((By.ID,"create-group")))
        driver.find_element_by_id("create-group").clear()
        driver.find_element_by_id("create-group").send_keys(RoleName)
        print "Entering Role Name"+RoleName
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div[4]/div/div/div[2]/div/div[2]/button[1]")))
        wait.until(EC.element_to_be_clickable((By.XPATH,"html/body/div[4]/div/div/div[2]/div/div[2]/button[1]")))
        elemt=driver.find_element_by_xpath("html/body/div[4]/div/div/div[2]/div/div[2]/button[1]")
        driver.execute_script('arguments[0].click()',elemt)
        time.sleep(3)
        print "clicked on Next Button"
        wait.until(EC.visibility_of_element_located((By.XPATH,"//h1/em/div/span/div[.='"+RoleName+"']")))
        print "Created Role Page Dispalyed"
        time.sleep(5)
        wait.until(EC.visibility_of_element_located((By.ID,"role-description")))
        driver.find_element_by_id("role-description").send_keys(Description)
        time.sleep(4)
    
    def roleSave(self): 
        wait =WebDriverWait(driver,80)
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div[1]/div/div[3]/div[2]/div/header/div/button[2]")))
        wait.until(EC.element_to_be_clickable((By.XPATH,"html/body/div[1]/div/div[3]/div[2]/div/header/div/button[2]")))
        elemm=driver.find_element_by_xpath("html/body/div[1]/div/div[3]/div[2]/div/header/div/button[2]")
        driver.execute_script("arguments[0].click()", elemm)
        time.sleep(3)
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div[4]/div/div/div[2]/div[2]/button[1]")))
        wait.until(EC.element_to_be_clickable((By.XPATH,"html/body/div[4]/div/div/div[2]/div[2]/button[1]")))
        elemm=driver.find_element_by_xpath("html/body/div[4]/div/div/div[2]/div[2]/button[1]")
        driver.execute_script("arguments[0].click()", elemm)
        time.sleep(4)
        
    def roleSearch(self,RoleName):
        wait =WebDriverWait(driver,80)
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div/div/div[3]/div[2]/div/div/ul/li[2]/a")))
        driver.find_element_by_xpath("html/body/div[1]/div/div[3]/div[2]/div/div/ul/li[2]/a").click()
        wait.until(EC.visibility_of_element_located((By.ID,"search-roles")))
        driver.find_element_by_id("search-roles").send_keys(RoleName)
        wait.until(EC.visibility_of_element_located((By.XPATH,"//table/tbody/tr/td[.='"+RoleName+"']/span/a")))
        time.sleep(4)
        if driver.find_element_by_xpath("html/body/div[1]/div/div[3]/div[2]/div/div/div/div/div[2]/div/table/tbody/tr/td[1]/span/a").text == RoleName :
                                       
            
            print("Craeted Role  is Displayed")
        else:
            print "CraetedRole is not displayed"
            raise Exception